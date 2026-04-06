#!/usr/bin/env python3
"""Analyze MedBrains_Features.xlsx to find next modules to build."""

import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook("/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx", read_only=True, data_only=True)

print("=" * 100)
print("MEDBRAINS FEATURE TRACKER — MODULE ANALYSIS")
print("=" * 100)

# Global aggregation
all_modules = defaultdict(lambda: {"Done": 0, "Partial": 0, "Pending": 0, "In Progress": 0, "Deferred": 0, "Other": 0, "Total": 0, "Sheet": ""})

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    if not rows:
        continue

    # Find header row
    header_row = None
    for i, row in enumerate(rows):
        if row and any(str(cell).strip().lower() == "status" for cell in row if cell):
            header_row = i
            break

    if header_row is None:
        print(f"\n--- Sheet: {sheet_name} --- (no Status column found, skipping)")
        continue

    headers = [str(c).strip() if c else "" for c in rows[header_row]]

    # Find column indices
    status_col = None
    module_col = None
    submodule_col = None
    feature_col = None
    priority_col = None

    for j, h in enumerate(headers):
        hl = h.lower()
        if hl == "status":
            status_col = j
        elif hl == "module":
            module_col = j
        elif "sub" in hl and "module" in hl:
            submodule_col = j
        elif hl == "feature":
            feature_col = j
        elif hl == "priority":
            priority_col = j

    if status_col is None or module_col is None:
        print(f"\n--- Sheet: {sheet_name} --- (missing Status or Module column, skipping)")
        continue

    # Collect data
    sheet_stats = defaultdict(lambda: {"Done": 0, "Partial": 0, "Pending": 0, "In Progress": 0, "Deferred": 0, "Other": 0, "Total": 0})
    sheet_total = {"Done": 0, "Partial": 0, "Pending": 0, "In Progress": 0, "Deferred": 0, "Other": 0, "Total": 0}

    current_module = None

    for row in rows[header_row + 1:]:
        if not row or all(c is None for c in row):
            continue

        # Get module (carry forward if blank — module headers span multiple rows)
        mod = str(row[module_col]).strip() if row[module_col] else None
        if mod and mod.lower() not in ("none", ""):
            current_module = mod

        if not current_module:
            continue

        # Get status
        status_raw = str(row[status_col]).strip() if status_col < len(row) and row[status_col] else None
        if not status_raw or status_raw.lower() in ("none", "", "status"):
            continue

        # Get feature text to confirm it's a real feature row
        feat = str(row[feature_col]).strip() if feature_col is not None and feature_col < len(row) and row[feature_col] else None
        if not feat or feat.lower() in ("none", "", "feature"):
            continue

        # Normalize status
        status = status_raw.strip()
        if status in ("Done", "Partial", "Pending", "In Progress", "Deferred"):
            sheet_stats[current_module][status] += 1
            sheet_total[status] += 1
        else:
            sheet_stats[current_module]["Other"] += 1
            sheet_total["Other"] += 1

        sheet_stats[current_module]["Total"] += 1
        sheet_total["Total"] += 1

        # Global aggregation
        key = f"{current_module}"
        all_modules[key]["Sheet"] = sheet_name
        if status in ("Done", "Partial", "Pending", "In Progress", "Deferred"):
            all_modules[key][status] += 1
        else:
            all_modules[key]["Other"] += 1
        all_modules[key]["Total"] += 1

    # Print sheet summary
    print(f"\n{'=' * 100}")
    print(f"SHEET: {sheet_name}")
    print(f"{'=' * 100}")
    print(f"  Sheet totals: Done={sheet_total['Done']}, Partial={sheet_total['Partial']}, "
          f"Pending={sheet_total['Pending']}, In Progress={sheet_total['In Progress']}, "
          f"Deferred={sheet_total['Deferred']}, Other={sheet_total['Other']}, Total={sheet_total['Total']}")
    print()

    # Sort modules by Pending count descending
    sorted_modules = sorted(sheet_stats.items(), key=lambda x: x[1]["Pending"], reverse=True)

    print(f"  {'Module':<45} {'Done':>5} {'Partial':>8} {'Pending':>8} {'InProg':>7} {'Defer':>6} {'Total':>6}")
    print(f"  {'-'*45} {'-'*5} {'-'*8} {'-'*8} {'-'*7} {'-'*6} {'-'*6}")

    for mod, stats in sorted_modules:
        print(f"  {mod:<45} {stats['Done']:>5} {stats['Partial']:>8} {stats['Pending']:>8} "
              f"{stats['In Progress']:>7} {stats['Deferred']:>6} {stats['Total']:>6}")


# ===== GLOBAL SUMMARY: Modules with most Pending features =====
print("\n" + "=" * 100)
print("GLOBAL SUMMARY — ALL MODULES SORTED BY PENDING FEATURES (descending)")
print("=" * 100)

sorted_all = sorted(all_modules.items(), key=lambda x: x[1]["Pending"], reverse=True)

print(f"\n  {'Module':<45} {'Sheet':<28} {'Done':>5} {'Part':>5} {'Pend':>5} {'InPr':>5} {'Defr':>5} {'Tot':>5}")
print(f"  {'-'*45} {'-'*28} {'-'*5} {'-'*5} {'-'*5} {'-'*5} {'-'*5} {'-'*5}")

total_done = 0
total_partial = 0
total_pending = 0
total_inprogress = 0
total_deferred = 0
total_all = 0

for mod, stats in sorted_all:
    if stats["Pending"] > 0:
        print(f"  {mod:<45} {stats['Sheet']:<28} {stats['Done']:>5} {stats['Partial']:>5} "
              f"{stats['Pending']:>5} {stats['In Progress']:>5} {stats['Deferred']:>5} {stats['Total']:>5}")
    total_done += stats["Done"]
    total_partial += stats["Partial"]
    total_pending += stats["Pending"]
    total_inprogress += stats["In Progress"]
    total_deferred += stats["Deferred"]
    total_all += stats["Total"]

print(f"\n  {'GRAND TOTAL':<45} {'':<28} {total_done:>5} {total_partial:>5} "
      f"{total_pending:>5} {total_inprogress:>5} {total_deferred:>5} {total_all:>5}")


# ===== Modules with ZERO Done features (not started at all) =====
print("\n" + "=" * 100)
print("MODULES WITH ZERO DONE FEATURES (completely unstarted)")
print("=" * 100)

unstarted = [(mod, stats) for mod, stats in sorted_all if stats["Done"] == 0 and stats["Pending"] > 0]
unstarted.sort(key=lambda x: x[1]["Pending"], reverse=True)

print(f"\n  {'Module':<45} {'Sheet':<28} {'Pend':>5} {'Part':>5} {'Tot':>5}")
print(f"  {'-'*45} {'-'*28} {'-'*5} {'-'*5} {'-'*5}")

for mod, stats in unstarted:
    print(f"  {mod:<45} {stats['Sheet']:<28} {stats['Pending']:>5} {stats['Partial']:>5} {stats['Total']:>5}")


# ===== Admin & Operations sheet detail =====
print("\n" + "=" * 100)
print("ADMIN & OPERATIONS SHEET — DETAILED BREAKDOWN")
print("=" * 100)

admin_modules = [(mod, stats) for mod, stats in sorted_all if stats["Sheet"] == "Admin & Operations"]
admin_modules.sort(key=lambda x: x[1]["Pending"], reverse=True)

print(f"\n  {'Module':<45} {'Done':>5} {'Part':>5} {'Pend':>5} {'InPr':>5} {'Tot':>5}")
print(f"  {'-'*45} {'-'*5} {'-'*5} {'-'*5} {'-'*5} {'-'*5}")

for mod, stats in admin_modules:
    print(f"  {mod:<45} {stats['Done']:>5} {stats['Partial']:>5} {stats['Pending']:>5} "
          f"{stats['In Progress']:>5} {stats['Total']:>5}")

wb.close()
print("\nDone.")

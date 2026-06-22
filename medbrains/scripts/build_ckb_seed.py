#!/usr/bin/env python3
"""Build the Clinical Knowledge Base diagnosis-reference seed CSV.

Source of de-identified clinical knowledge (NOT committed, operator-local):
  <sir>/analysis/non_notifiable_diseases.xlsx  — 225 common OPD diagnoses
      (columns: Dept, Diagnosis, ICD-10, Notifiable, demographics...).
      We take ONLY Dept / Diagnosis / ICD-10 (no patient data).

Plus a curated national IDSP/IHIP notifiable-disease list (public-health
knowledge, embedded below) that drives the statutory reporting alerts.

Output (committed = github-tracked source + backup):
  crates/medbrains-server/src/seed/data/diagnosis_reference.csv
  columns: icd10_code,name,department,is_notifiable,reporting_body,report_timeframe

Usage:
  python3 scripts/build_ckb_seed.py [--src /path/to/sir]
Re-run whenever the source knowledge changes; commit the resulting CSV.
"""

from __future__ import annotations

import argparse
import csv
import os
import sys

# National notifiable diseases (IDSP / IHIP, Epidemic Diseases Act 1897; the
# standard P/L/S surveillance list). Public-health reference knowledge — not
# patient data. (icd10, name, department, reporting_body, timeframe)
IDSP_NOTIFIABLE: list[tuple[str, str, str, str, str]] = [
    ("A00", "Cholera", "Medicine", "IDSP", "24h"),
    ("A01.0", "Typhoid fever", "Medicine", "IDSP", "weekly"),
    ("A03", "Shigellosis (bacillary dysentery)", "Medicine", "IDSP", "weekly"),
    ("A05", "Acute diarrhoeal disease / food poisoning", "Medicine", "IDSP", "24h"),
    ("A15", "Tuberculosis (pulmonary)", "Medicine", "Ni-kshay/RNTCP", "24h"),
    ("A19", "Miliary tuberculosis", "Medicine", "Ni-kshay/RNTCP", "24h"),
    ("A20", "Plague", "Medicine", "IDSP", "24h"),
    ("A22", "Anthrax", "Medicine", "IDSP", "24h"),
    ("A27", "Leptospirosis", "Medicine", "IDSP", "24h"),
    ("A30", "Leprosy", "Skin", "NLEP", "weekly"),
    ("A33", "Tetanus neonatorum", "Paediatrics", "IDSP", "24h"),
    ("A35", "Tetanus", "Medicine", "IDSP", "24h"),
    ("A36", "Diphtheria", "Paediatrics", "IDSP", "24h"),
    ("A37", "Whooping cough (pertussis)", "Paediatrics", "IDSP", "24h"),
    ("A39", "Meningococcal disease", "Medicine", "IDSP", "24h"),
    ("A80", "Acute flaccid paralysis / poliomyelitis", "Paediatrics", "IDSP", "24h"),
    ("A82", "Rabies", "Medicine", "IDSP", "24h"),
    ("A83.0", "Japanese encephalitis", "Medicine", "IDSP", "24h"),
    ("A90", "Dengue fever", "Medicine", "IDSP", "24h"),
    ("A91", "Dengue haemorrhagic fever", "Medicine", "IDSP", "24h"),
    ("A92.0", "Chikungunya", "Medicine", "IDSP", "weekly"),
    ("A95", "Yellow fever", "Medicine", "IDSP", "24h"),
    ("B05", "Measles", "Paediatrics", "IDSP", "24h"),
    ("B06", "Rubella", "Paediatrics", "IDSP", "weekly"),
    ("B15", "Acute hepatitis A", "Medicine", "IDSP", "weekly"),
    ("B16", "Acute hepatitis B", "Medicine", "IDSP", "weekly"),
    ("B17.1", "Acute hepatitis C", "Medicine", "IDSP", "weekly"),
    ("B19", "Acute viral hepatitis, unspecified", "Medicine", "IDSP", "weekly"),
    ("B50", "Plasmodium falciparum malaria", "Medicine", "IDSP/NVBDCP", "24h"),
    ("B51", "Plasmodium vivax malaria", "Medicine", "IDSP/NVBDCP", "24h"),
    ("B54", "Malaria, unspecified", "Medicine", "IDSP/NVBDCP", "24h"),
    ("B74", "Filariasis", "Medicine", "NVBDCP", "weekly"),
    ("J09", "Influenza due to novel virus (incl. H1N1)", "Medicine", "IDSP", "24h"),
    ("J10", "Influenza, seasonal", "Medicine", "IDSP", "weekly"),
    ("U07.1", "COVID-19", "Medicine", "IDSP/IHIP", "24h"),
    ("W54.0", "Animal bite (rabies risk)", "Casualty", "IDSP", "24h"),
]


# Curated public CDS dosing reference for common drugs — standard adult values
# from the WHO EML / NLEM 2022 / BNF (public reference knowledge, not patient
# data). (generic, max_dose_per_day, max_single_dose, renal_egfr_threshold,
# renal_rule, hepatic_caution, pregnancy_category)
CURATED_CDS: list[tuple[str, str, str, str, str, str, str]] = [
    ("paracetamol", "4000 mg", "1000 mg", "", "", "Reduce dose in chronic liver disease", "B"),
    ("ibuprofen", "2400 mg", "800 mg", "30", "Avoid if eGFR < 30", "", "C"),
    ("diclofenac", "150 mg", "50 mg", "30", "Avoid if eGFR < 30", "", "C"),
    ("aceclofenac", "200 mg", "100 mg", "30", "Avoid if eGFR < 30", "", "C"),
    ("naproxen", "1000 mg", "500 mg", "30", "Avoid if eGFR < 30", "", "C"),
    ("aspirin", "4000 mg", "1000 mg", "", "", "", "C"),
    ("amoxicillin", "3000 mg", "1000 mg", "30", "Reduce frequency if eGFR < 30", "", "B"),
    ("amoxicillin_clav", "3000 mg", "1000 mg", "30", "Reduce frequency if eGFR < 30", "", "B"),
    ("azithromycin", "500 mg", "500 mg", "", "", "Caution in hepatic impairment", "B"),
    ("ciprofloxacin", "1500 mg", "750 mg", "30", "Halve dose if eGFR < 30", "", "C"),
    ("levofloxacin", "750 mg", "750 mg", "50", "Reduce dose if eGFR < 50", "", "C"),
    ("metronidazole", "2000 mg", "800 mg", "", "", "Reduce dose in severe hepatic impairment", "B"),
    ("ceftriaxone", "4000 mg", "2000 mg", "", "", "", "B"),
    ("cefixime", "400 mg", "200 mg", "20", "Reduce dose if eGFR < 20", "", "B"),
    ("metformin", "2550 mg", "1000 mg", "30", "Stop if eGFR < 30 (lactic acidosis)", "", "C"),
    ("glimepiride", "8 mg", "8 mg", "", "", "Caution in hepatic impairment", "C"),
    ("amlodipine", "10 mg", "10 mg", "", "", "", "C"),
    ("telmisartan", "80 mg", "80 mg", "", "", "Caution in biliary obstruction", "D"),
    ("losartan", "100 mg", "100 mg", "", "", "Reduce dose in hepatic impairment", "D"),
    ("atenolol", "100 mg", "100 mg", "35", "Reduce dose if eGFR < 35", "", "D"),
    ("atorvastatin", "80 mg", "80 mg", "", "", "Contraindicated in active liver disease", "X"),
    ("prednisolone", "60 mg", "60 mg", "", "", "", "C"),
    ("pantoprazole", "80 mg", "40 mg", "", "", "Max 20 mg/day in severe hepatic impairment", "B"),
    ("omeprazole", "40 mg", "40 mg", "", "", "Max 20 mg/day in hepatic impairment", "C"),
    ("ranitidine", "300 mg", "150 mg", "50", "Halve dose if eGFR < 50", "", "B"),
    ("furosemide", "80 mg", "40 mg", "", "", "", "C"),
    ("amitriptyline", "150 mg", "75 mg", "", "", "Caution in hepatic impairment", "C"),
    ("tramadol", "400 mg", "100 mg", "30", "Increase dosing interval if eGFR < 30", "Caution in hepatic impairment", "C"),
    ("gentamicin", "5 mg", "", "60", "Reduce dose / extend interval if eGFR < 60", "", "D"),
    ("vancomycin", "2000 mg", "1000 mg", "60", "Adjust by levels if eGFR < 60", "", "C"),
]


def build_drug_formulary(src: str, out_dir: str) -> None:
    """Emit drug_formulary.csv = curated public CDS reference, enriched with
    paediatric mg/kg + hepatic/pregnancy notes from the local pharmacology.py."""
    import re

    cols = [
        "generic_name", "inn_name", "atc_code", "max_dose_per_day", "max_single_dose",
        "dose_per_kg", "renal_adjust_egfr_threshold", "renal_adjust_rule",
        "hepatic_caution", "pregnancy_category", "brands", "is_nlem",
    ]
    # Start from the curated reference. These are WHO-EML / NLEM-2022 essential
    # medicines → government-supply / free-drug-scheme drugs (is_nlem = true).
    by_generic: dict[str, dict] = {}
    for g, maxd, maxs, rt, rr, hep, preg in CURATED_CDS:
        by_generic[g] = {
            "generic_name": g, "inn_name": g, "atc_code": "",
            "max_dose_per_day": maxd, "max_single_dose": maxs, "dose_per_kg": "",
            "renal_adjust_egfr_threshold": rt, "renal_adjust_rule": rr,
            "hepatic_caution": hep, "pregnancy_category": preg,
            "brands": "", "is_nlem": "true",
        }

    # Enrich from pharmacology.py (paediatric mg/kg, hepatic/pregnancy notes).
    if os.path.isfile(os.path.join(src, "pharmacology.py")):
        sys.path.insert(0, src)
        try:
            import pharmacology  # type: ignore

            for key, d in pharmacology.DRUGS.items():
                generic = key.split("_")[0].strip().lower()
                row = by_generic.setdefault(
                    generic,
                    {c: "" for c in cols} | {"generic_name": generic, "inn_name": generic},
                )
                notes = str(d.get("notes") or "")
                if not row["hepatic_caution"] and re.search(r"LFT|hepat|liver", notes, re.I):
                    row["hepatic_caution"] = notes
                if not row["pregnancy_category"]:
                    if re.search(r"avoid in pregnan", notes, re.I):
                        row["pregnancy_category"] = "D"
                    elif re.search(r"pregnan", notes, re.I):
                        row["pregnancy_category"] = "C"
                paedia = d.get("paedia") or {}
                m = re.match(r"([\d.]+)\s*(mcg|mg|g)/kg", str(paedia.get("dose", "")))
                if m and not row["dose_per_kg"]:
                    row["dose_per_kg"] = f"{m.group(1)} {m.group(2)}"
                # brand name(s) from the product label parentheses; NLEM essential.
                row["is_nlem"] = "true"
                brand_m = re.search(r"\(([^)0-9][^)]*)\)", str(d.get("name") or ""))
                if brand_m and not row.get("brands"):
                    brand = re.sub(r"[\d./%]+.*$", "", brand_m.group(1)).strip()
                    if brand:
                        row["brands"] = brand
        except Exception as exc:  # noqa: BLE001
            print(f"WARN: could not import pharmacology.py: {exc}", file=sys.stderr)
    else:
        print("WARN: pharmacology.py not found — curated CDS only.", file=sys.stderr)

    out_path = os.path.join(out_dir, "drug_formulary.csv")
    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(cols)
        for g in sorted(by_generic):
            w.writerow([by_generic[g][c] for c in cols])
    print(f"Wrote {len(by_generic)} drug rows → {out_path}")


def _parse_range(text: str) -> tuple[str, str]:
    """'13.0–17.0 g/dL' → ('13.0','17.0'); returns ('','') if not a numeric range."""
    import re

    m = re.search(r"(-?[\d.]+)\s*[–\-]\s*(-?[\d.]+)", str(text or ""))
    return (m.group(1), m.group(2)) if m else ("", "")


# Pregnancy + elderly normal ranges are NOT in the source matrix; curate the
# well-published shifts (WHO/CDC/IADPSG/BNF) for high-value analytes. Matched by
# substring token against the analyte name. (low, high) — empty string = no shift.
# Pregnancy: physiologic dilution/leukocytosis; Elderly (≥65): age-adjusted upper.
CURATED_LAB_BANDS: list[tuple[tuple[str, ...], str, str, str, str]] = [
    # tokens,                       preg_lo, preg_hi, eld_lo, eld_hi
    (("haemoglobin", "hemoglobin"), "11.0", "14.0", "12.0", "16.0"),
    (("creatinine",),               "0.4", "0.8", "0.7", "1.4"),
    (("platelet",),                 "115", "400", "150", "400"),
    (("tsh",),                      "0.1", "4.0", "0.4", "7.0"),
    (("leucocyte", "leukocyte", "wbc", "tlc"), "6.0", "16.0", "4.0", "11.0"),
    (("fasting",),                  "70", "92", "70", "110"),
    (("urea", "bun"),               "7", "30", "10", "50"),
    (("uric",),                     "2.5", "5.8", "3.5", "8.5"),
]


def _curated_bands(analyte: str) -> tuple[str, str, str, str]:
    a = analyte.lower()
    for tokens, plo, phi, elo, ehi in CURATED_LAB_BANDS:
        if any(t in a for t in tokens):
            return plo, phi, elo, ehi
    return "", "", "", ""


def build_lab_reference(src: str, out_dir: str) -> None:
    """Emit lab_reference.csv from Lab_Value_Matrix[Reference_Ranges] —
    analyte reference ranges + critical thresholds (real de-identified data).
    Pregnancy/elderly bands are a curated overlay (not in the source matrix)."""
    import json

    xlsx = os.path.join(src, "analysis", "Lab_Value_Matrix.xlsx")
    # normal_low/high default to Adult-M; per-age/sex bands follow.
    cols = [
        "test", "analyte", "unit", "normal_low", "normal_high",
        "critical_low", "critical_high", "category",
        "neonate_low", "neonate_high", "infant_low", "infant_high",
        "child_low", "child_high", "adult_m_low", "adult_m_high",
        "adult_f_low", "adult_f_high", "pregnancy_low", "pregnancy_high",
        "elderly_low", "elderly_high",
    ]
    out_path = os.path.join(out_dir, "lab_reference.csv")
    if not os.path.isfile(xlsx):
        print(f"WARN: {xlsx} not found — skipping lab reference.", file=sys.stderr)
        return

    import openpyxl

    cat_map: dict[str, str] = {}
    cat_path = os.path.join(src, "test_category_map.json")
    if os.path.isfile(cat_path):
        cat_map = {k.lower(): v for k, v in json.load(open(cat_path)).items()}

    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb["Reference_Ranges"]
    data = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() for c in data[0]]

    def col(name: str) -> int:
        return hdr.index(name) if name in hdr else -1

    ti, ai, ui = col("Test"), col("Analyte"), col("Unit")
    cl, ch = col("Critical Low"), col("Critical High")
    bands = {
        "neonate": col("Normal: Neonate"),
        "infant": col("Normal: Infant"),
        "child": col("Normal: Child"),
        "adult_m": col("Normal: Adult-M"),
        "adult_f": col("Normal: Adult-F"),
        "pregnancy": col("Normal: Pregnancy"),
        "elderly": col("Normal: Elderly"),
    }

    def band_range(r: tuple, key: str) -> tuple[str, str]:
        idx = bands[key]
        return _parse_range(r[idx]) if idx >= 0 else ("", "")

    rows: list[list[str]] = []
    seen: set[str] = set()
    for r in data[1:]:
        analyte = str(r[ai] or "").strip()
        test = str(r[ti] or "").strip()
        if not analyte or analyte.lower() in seen:
            continue
        seen.add(analyte.lower())
        am_low, am_high = band_range(r, "adult_m")
        c_low = _parse_range(f"{r[cl]}–{r[cl]}")[0] if cl >= 0 and r[cl] else ""
        c_high = _parse_range(f"{r[ch]}–{r[ch]}")[0] if ch >= 0 and r[ch] else ""
        category = cat_map.get(test.lower()) or cat_map.get(analyte.lower()) or "Pathology"
        row = [test, analyte, str(r[ui] or "").strip(), am_low, am_high, c_low, c_high, category]
        for key in ("neonate", "infant", "child", "adult_m", "adult_f"):
            lo, hi = band_range(r, key)
            row.extend([lo, hi])
        # Pregnancy/elderly: prefer the source column if present, else curated overlay.
        plo, phi = band_range(r, "pregnancy")
        elo, ehi = band_range(r, "elderly")
        c_plo, c_phi, c_elo, c_ehi = _curated_bands(analyte)
        row.extend([plo or c_plo, phi or c_phi, elo or c_elo, ehi or c_ehi])
        rows.append(row)

    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(cols)
        w.writerows(rows)
    crit = sum(1 for r in rows if r[5] or r[6])
    preg = sum(1 for r in rows if r[18] or r[20])
    print(
        f"Wrote {len(rows)} lab analytes ({crit} critical, age/sex bands, "
        f"{preg} with pregnancy/elderly overlay) → {out_path}"
    )


# Curated public dangerous drug/ingredient combinations (well-established
# interactions + IV admixture incompatibilities). (ingredient_a, ingredient_b,
# severity, mechanism)
INCOMPATIBILITIES: list[tuple[str, str, str, str]] = [
    ("warfarin", "aspirin", "major", "Additive bleeding risk"),
    ("warfarin", "ibuprofen", "major", "Additive bleeding risk + GI"),
    ("warfarin", "diclofenac", "major", "Additive bleeding risk + GI"),
    ("ramipril", "spironolactone", "major", "Hyperkalaemia"),
    ("enalapril", "spironolactone", "major", "Hyperkalaemia"),
    ("telmisartan", "spironolactone", "major", "Hyperkalaemia"),
    ("losartan", "ramipril", "moderate", "Dual RAAS blockade — hyperkalaemia / AKI"),
    ("ciprofloxacin", "tizanidine", "major", "↑ tizanidine — hypotension/sedation"),
    ("clarithromycin", "atorvastatin", "major", "Rhabdomyolysis (CYP3A4)"),
    ("clarithromycin", "simvastatin", "major", "Rhabdomyolysis (CYP3A4)"),
    ("ceftriaxone", "calcium", "major", "Ceftriaxone–calcium precipitate (IV, esp. neonates)"),
    ("phenytoin", "dextrose", "major", "Phenytoin precipitates in dextrose — saline only"),
    ("methotrexate", "trimethoprim", "major", "Additive antifolate — marrow suppression"),
    ("digoxin", "amiodarone", "major", "↑ digoxin level — toxicity"),
    ("metformin", "contrast", "moderate", "Lactic acidosis — hold around iodinated contrast"),
    ("tramadol", "ondansetron", "moderate", "Serotonin syndrome risk"),
    ("amlodipine", "simvastatin", "moderate", "↑ simvastatin — limit to 20 mg"),
    ("potassium", "spironolactone", "major", "Hyperkalaemia"),
    ("ibuprofen", "ramipril", "moderate", "NSAID blunts ACE-i + AKI ('triple whammy' w/ diuretic)"),
]

# Common combination products whose ingredients aren't parseable from the name.
KNOWN_INGREDIENTS: dict[str, list[str]] = {
    "amoxicillin_clav": ["amoxicillin", "clavulanic acid"],
    "cotrimoxazole": ["sulfamethoxazole", "trimethoprim"],
    "piperacillin_tazo": ["piperacillin", "tazobactam"],
}


def _clean_ingredient(token: str) -> str:
    import re

    t = re.sub(r"\([^)]*\)", "", token)  # drop brand/strength parens
    t = re.sub(r"\b\d+[\d./]*\s*(mg|mcg|g|ml|iu|%)?\b", "", t, flags=re.I)  # drop doses
    return t.strip(" .").lower()


def build_drug_ingredients(src: str, out_dir: str) -> None:
    """Emit drug_ingredients.csv (generic → active ingredient) by parsing the
    pharmacology.py product names for combinations, and ingredient_incompatibility.csv."""
    import re

    pairs: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    if os.path.isfile(os.path.join(src, "pharmacology.py")):
        sys.path.insert(0, src)
        try:
            import pharmacology  # type: ignore

            form_prefix = re.compile(
                r"^(t|tab|cap|c|syp|syr|inj|neb|mdi|oint|cr|gel|drops?|susp|sol|spray)\.?\s+",
                re.I,
            )
            for key, d in pharmacology.DRUGS.items():
                generic = key.split("_")[0].strip().lower()
                if generic in KNOWN_INGREDIENTS:
                    ingredients = KNOWN_INGREDIENTS[generic]
                elif key in KNOWN_INGREDIENTS:
                    ingredients = KNOWN_INGREDIENTS[key]
                else:
                    name = form_prefix.sub("", str(d.get("name") or "").strip())
                    name = re.sub(r"\([^)]*\)", "", name)
                    parts = [p for p in re.split(r"\s*\+\s*", name) if p.strip()]
                    ingredients = [_clean_ingredient(p) for p in parts] or [generic]
                for ing in ingredients:
                    if ing and (generic, ing) not in seen:
                        seen.add((generic, ing))
                        pairs.append((generic, ing))
        except Exception as exc:  # noqa: BLE001
            print(f"WARN: ingredients from pharmacology.py: {exc}", file=sys.stderr)

    ing_path = os.path.join(out_dir, "drug_ingredients.csv")
    with open(ing_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["generic_name", "ingredient"])
        w.writerows(sorted(pairs))
    print(f"Wrote {len(pairs)} drug→ingredient rows → {ing_path}")

    inc_path = os.path.join(out_dir, "ingredient_incompatibility.csv")
    with open(inc_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["ingredient_a", "ingredient_b", "severity", "mechanism"])
        # store the pair ordered so lookups are symmetric.
        for a, b, sev, mech in INCOMPATIBILITIES:
            lo, hi = sorted([a.lower(), b.lower()])
            w.writerow([lo, hi, sev, mech])
    print(f"Wrote {len(INCOMPATIBILITIES)} incompatibility pairs → {inc_path}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--src",
        default=os.path.expanduser("~/Projects/sir"),
        help="Path to the local de-identified knowledge folder (sir).",
    )
    args = parser.parse_args()

    here = os.path.dirname(os.path.abspath(__file__))
    repo = os.path.dirname(here)
    out_dir = os.path.join(repo, "crates", "medbrains-server", "src", "seed", "data")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "diagnosis_reference.csv")

    rows: list[tuple[str, str, str, str, str, str]] = []
    seen: set[str] = set()

    # 1) Curated notifiable list first (so it wins on icd10 collisions).
    for icd, name, dept, body, tf in IDSP_NOTIFIABLE:
        key = icd.strip().upper()
        if key in seen:
            continue
        seen.add(key)
        rows.append((key, name, dept, "true", body, tf))

    # 2) Real de-identified OPD diagnoses (non-notifiable) from the xlsx.
    xlsx = os.path.join(args.src, "analysis", "non_notifiable_diseases.xlsx")
    if os.path.exists(xlsx):
        import openpyxl  # local dependency, operator machine only

        wb = openpyxl.load_workbook(xlsx, read_only=True)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip() for c in data[0]]
        di, ni, ii = hdr.index("Dept"), hdr.index("Diagnosis"), hdr.index("ICD-10")
        for r in data[1:]:
            icd = str(r[ii] or "").strip().upper()
            name = str(r[ni] or "").strip()
            dept = str(r[di] or "").strip()
            if not icd or not name or icd in seen:
                continue
            seen.add(icd)
            rows.append((icd, name, dept, "false", "", ""))
    else:
        print(f"WARN: {xlsx} not found — emitting notifiable list only.", file=sys.stderr)

    rows.sort(key=lambda x: x[0])
    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["icd10_code", "name", "department", "is_notifiable", "reporting_body", "report_timeframe"])
        w.writerows(rows)

    notifiable = sum(1 for r in rows if r[3] == "true")
    print(f"Wrote {len(rows)} rows ({notifiable} notifiable) → {out_path}")

    build_drug_formulary(args.src, out_dir)
    build_lab_reference(args.src, out_dir)
    build_drug_ingredients(args.src, out_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

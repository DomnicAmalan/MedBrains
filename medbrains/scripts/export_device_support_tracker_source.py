from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


ROOT = Path("/Users/apple/Projects/MedBrains")
WORKBOOK_PATH = ROOT / "MedBrains_Features.xlsx"
OUTPUT_DIR = ROOT / "medbrains" / "outputs" / "2026-04-22-device-support-tracker"
OUTPUT_JSON = OUTPUT_DIR / "device_support_tracker_source.json"

KEYWORDS = re.compile(
    r"(device|biometric|scanner|printer|print|barcode|qr code|rfid|ct\b|mri|x-ray|xray|ultrasound|"
    r"dicom|pacs|hl7|fhir|instrument|analyzer|icu|ventilator|monitor|gps|cctv|telemedicine|wearable|"
    r"kiosk|queue|tv|display|token|modality|camera|iris|fingerprint)",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class CategoryProfile:
    domain: str
    device_family: str
    interaction_mode: str
    protocols: str
    integration_pattern: str
    manual_checklist: str
    config_inputs: str
    fallback_strategy: str
    simulator_strategy: str
    ai_assist: str
    next_build_step: str


PROTOCOL_PATTERNS = [
    {
        "protocol_family": "DICOM / PACS",
        "used_for": "CT, MRI, X-ray, ultrasound, modality worklist, image archive",
        "transport": "TCP/IP",
        "real_interaction": "DICOM association, worklist query, study send/retrieve",
        "configuration_surface": "AE title, host, port, modality code, worklist rules",
        "manual_artifacts": "conformance statement, AE title setup, SOP classes, sample studies",
        "no_hardware_fallback": "DICOM file upload, manual report entry, PACS import queue",
    },
    {
        "protocol_family": "HL7 v2 / ASTM",
        "used_for": "Lab analyzers, patient/result exchange, interface engine workflows",
        "transport": "TCP, serial, file drop",
        "real_interaction": "ORU/ORM messages or ASTM records via gateway/agent",
        "configuration_surface": "message mapping, delimiters, host/port, order codes, result codes",
        "manual_artifacts": "interface manual, sample messages, serial settings, code dictionary",
        "no_hardware_fallback": "CSV import, manual result entry, batch upload",
    },
    {
        "protocol_family": "Vendor SDK / Driver",
        "used_for": "Biometric readers, smart-card devices, access-control panels",
        "transport": "USB, serial, LAN, local service",
        "real_interaction": "Call vendor APIs or device service and normalize outputs",
        "configuration_surface": "SDK install, device ID, capture mode, timeout, quality thresholds",
        "manual_artifacts": "SDK docs, driver version, callback format, error codes",
        "no_hardware_fallback": "OTP, PIN, manual override, demographic verification",
    },
    {
        "protocol_family": "OS Print / ZPL / ESC-POS",
        "used_for": "A4 printers, thermal printers, barcode and wristband printers",
        "transport": "OS spooler, USB, LAN",
        "real_interaction": "Render template then send PDF/label commands through print agent",
        "configuration_surface": "printer name, paper size, template, command mode, density",
        "manual_artifacts": "printer language manual, label dimensions, sample output",
        "no_hardware_fallback": "PDF download, browser print, central print station",
    },
    {
        "protocol_family": "Browser / Camera / HID",
        "used_for": "QR/barcode scanners, kiosk check-in, webcam scanning",
        "transport": "Keyboard wedge, camera API, browser runtime",
        "real_interaction": "Read scanned payload and resolve patient/token/entity",
        "configuration_surface": "camera permission, scan mode, checksum, prefix/suffix trimming",
        "manual_artifacts": "scanner configuration guide, expected prefixes, test codes",
        "no_hardware_fallback": "manual code entry, mobile camera, search + confirm",
    },
    {
        "protocol_family": "WebSocket / Display Client",
        "used_for": "Queue TVs, bed boards, signage monitors, kiosk dashboards",
        "transport": "HTTP/WebSocket",
        "real_interaction": "Display client subscribes to live queue/bed events",
        "configuration_surface": "screen mode, location binding, refresh policy, language/theme",
        "manual_artifacts": "display sizing notes, device autoplay/kiosk setup",
        "no_hardware_fallback": "browser tab, desktop screen, nurse station monitor",
    },
]


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def is_feature_row(row: tuple[object, ...]) -> bool:
    feature = clean(row[3] if len(row) > 3 else "")
    return bool(feature)


def classify_feature(text: str, module: str, submodule: str, sheet: str) -> CategoryProfile:
    lowered = " ".join([sheet, module, submodule, text]).lower()

    if any(term in lowered for term in ["ct", "mri", "x-ray", "xray", "ultrasound", "pacs", "dicom", "modality"]):
        return CategoryProfile(
            domain="Radiology & Imaging",
            device_family="Imaging Modality / PACS",
            interaction_mode="DICOM worklist, study archive, report sync",
            protocols="DICOM, vendor PACS API, file import",
            integration_pattern="Connector or gateway using modality/PACS configuration",
            manual_checklist="DICOM conformance statement, AE title, port, SOP classes, sample study",
            config_inputs="AE title, PACS host/port, modality type, worklist filters, body-part codes",
            fallback_strategy="Manual order/reporting + DICOM file upload + PACS import queue",
            simulator_strategy="Use sample DICOM studies and mocked worklist responses",
            ai_assist="Parse modality manuals, extract AE titles/protocol terms, draft config checklist",
            next_build_step="Create DICOM connector templates and PACS onboarding forms",
        )

    if any(term in lowered for term in ["biometric", "fingerprint", "iris", "face", "attendance", "access control", "smart card", "aadhaar"]):
        return CategoryProfile(
            domain="Biometrics & Access",
            device_family="Biometric Reader / Access Device",
            interaction_mode="SDK/service call for enroll, capture, match, or access event read",
            protocols="Vendor SDK, RD service, USB/LAN driver, REST callback",
            integration_pattern="Local agent or vendor SDK wrapper with normalized events",
            manual_checklist="SDK version, capture API, match API, error codes, device IDs, certification",
            config_inputs="device model, driver version, service URL, timeout, quality score threshold",
            fallback_strategy="OTP, PIN, manual override, demographic verification",
            simulator_strategy="Replay captured templates/events in a sandbox service",
            ai_assist="Read SDK docs/manuals and generate capture-flow configs and field mappings",
            next_build_step="Define biometric adapter contract and fallback policy per workflow",
        )

    if any(term in lowered for term in ["lab equipment", "instrument", "analyzer", "astm", "hl7", "lis"]):
        return CategoryProfile(
            domain="Lab Instruments",
            device_family="Analyzer / Instrument Interface",
            interaction_mode="Order/result message exchange through gateway",
            protocols="HL7 v2, ASTM, serial, TCP, file drop",
            integration_pattern="Instrument gateway translating device payloads into lab events",
            manual_checklist="Interface manual, sample ASTM/HL7 messages, serial settings, code map",
            config_inputs="host, port, baud rate, analyzer codes, specimen mapping, retry policy",
            fallback_strategy="CSV import, manual result entry, technician validation queue",
            simulator_strategy="Replay sample result files/messages and validate mappings",
            ai_assist="Parse interface manuals and suggest message mappings and code transforms",
            next_build_step="Build instrument connector registry and sample-message test harness",
        )

    if any(term in lowered for term in ["printer", "print", "wristband", "label", "receipt"]):
        return CategoryProfile(
            domain="Print, Scan & ID",
            device_family="Printer / Wristband / Label Device",
            interaction_mode="Render document/label and send via spooler or printer language",
            protocols="OS print spooler, ZPL, EPL, ESC-POS, PDF",
            integration_pattern="Print agent plus template engine",
            manual_checklist="Printer language manual, label size, DPI, margins, sample output",
            config_inputs="printer name, tray/profile, template, paper size, density/orientation",
            fallback_strategy="PDF download, browser print, central print desk",
            simulator_strategy="Virtual printer profile and PDF snapshot regression tests",
            ai_assist="Extract label dimensions and command examples from manuals/templates",
            next_build_step="Add printer profile registry and per-template routing rules",
        )

    if any(term in lowered for term in ["barcode", "qr code", "scanner", "rfid", "scan", "kiosk"]):
        return CategoryProfile(
            domain="Print, Scan & ID",
            device_family="Scanner / Wristband / RFID Reader",
            interaction_mode="Read scanned identifier and resolve entity in workflow",
            protocols="HID keyboard wedge, browser camera, RFID SDK",
            integration_pattern="Client-side capture with server-side identifier resolution",
            manual_checklist="Scanner programming guide, prefix/suffix settings, RFID sample payload",
            config_inputs="capture mode, scan prefix trimming, code symbology, reader ID",
            fallback_strategy="Manual ID entry, mobile camera, patient search + confirm",
            simulator_strategy="Mock scan payloads and camera captures in test flows",
            ai_assist="Turn scanner manuals into config presets and validation checks",
            next_build_step="Ship scan profile presets for barcode, QR, and RFID workflows",
        )

    if any(term in lowered for term in ["queue", "tv", "display", "token", "monitor", "signage"]):
        return CategoryProfile(
            domain="Queue, Kiosk & Displays",
            device_family="Display Client / Queue Board",
            interaction_mode="WebSocket or polling client renders operational state",
            protocols="HTTP, WebSocket, Android TV/browser runtime",
            integration_pattern="Software-native display client with location binding",
            manual_checklist="Screen resolution, autoplay setup, kiosk/TV browser restrictions",
            config_inputs="screen mode, department, token source, refresh theme, fallback content",
            fallback_strategy="Browser tab on any monitor or nursing station screen",
            simulator_strategy="Mock queue events and test device profiles in browser",
            ai_assist="Generate screen presets and recommend layout from screen specs",
            next_build_step="Treat displays as configurable clients, not blocked hardware work",
        )

    if any(term in lowered for term in ["ventilator", "hemodynamic", "bedside", "icu", "monitoring", "monitor"]):
        return CategoryProfile(
            domain="Bedside & ICU Devices",
            device_family="Monitor / Ventilator / Bedside Device",
            interaction_mode="Periodic feed import, gateway polling, or manual transcription",
            protocols="Vendor feed, HL7, serial, TCP, CSV export",
            integration_pattern="Gateway or nurse-verified import with audit trail",
            manual_checklist="Data export format, sampling interval, units, alarm semantics, port setup",
            config_inputs="device model, feed endpoint, polling interval, unit normalization, alarm map",
            fallback_strategy="Structured manual charting and periodic upload review",
            simulator_strategy="Replay exported device logs and normalize into vitals streams",
            ai_assist="Map device parameter names/units from manuals into canonical vitals fields",
            next_build_step="Define bedside telemetry adapter schema and validation rules",
        )

    if any(term in lowered for term in ["gps", "ambulance", "cctv", "security camera", "camera"]):
        return CategoryProfile(
            domain="Facilities, Security & Transport",
            device_family="GPS / Camera / Transport Device",
            interaction_mode="Read vendor API or telematics feed",
            protocols="REST API, RTSP, MQTT, vendor SDK",
            integration_pattern="External API connector with metadata sync",
            manual_checklist="API docs, auth method, polling/webhook support, latency expectations",
            config_inputs="device ID, API key, refresh interval, zone mapping, event types",
            fallback_strategy="Manual location/status update workflow",
            simulator_strategy="Replay sample GPS points/events through the connector",
            ai_assist="Summarize API manuals and draft field maps and alert rules",
            next_build_step="Add telematics/camera connector inventory and event schema",
        )

    if any(term in lowered for term in ["telemedicine", "wearable", "webcam", "stethoscope"]):
        return CategoryProfile(
            domain="Telemedicine & Wearables",
            device_family="Remote Consult / Wearable Device",
            interaction_mode="SDK or browser/API capture into remote care workflow",
            protocols="WebRTC, BLE, REST API, vendor SDK",
            integration_pattern="Session service plus optional device-specific plugin",
            manual_checklist="SDK docs, browser permissions, device pairing steps, sample payload",
            config_inputs="device model, pairing mode, sampling rate, session binding, consent path",
            fallback_strategy="Video-only consult plus manual vitals entry",
            simulator_strategy="Synthetic wearable streams and mocked session devices",
            ai_assist="Parse manuals and suggest pairing steps, parameter mappings, and QA scenarios",
            next_build_step="Separate video consult baseline from optional peripheral integrations",
        )

    return CategoryProfile(
        domain="Other Hardware & Integrations",
        device_family="General Device / Gateway",
        interaction_mode="Depends on vendor protocol or export format",
        protocols="Vendor API, serial, file drop, manual import",
        integration_pattern="Connector template plus manual fallback",
        manual_checklist="Protocol document, sample payload, auth/setup, error handling",
        config_inputs="connection details, format, identity mapping, retry rules",
        fallback_strategy="Manual workflow + import queue",
        simulator_strategy="Mock connector responses with fixture payloads",
        ai_assist="Extract config fields and mappings from manuals into templates",
        next_build_step="Classify device and choose gateway vs SDK vs native path",
    )


def support_status_for(rows: list[dict[str, str]], profile: CategoryProfile) -> str:
    counts = Counter(row["product_status"] for row in rows)
    if profile.domain == "Queue, Kiosk & Displays":
        return "Configurable"
    if counts.get("Done", 0) >= max(2, len(rows) // 2):
        return "Partial"
    if counts.get("Partial", 0) > 0:
        return "Partial"
    if any(row["product_status"] == "Pending" for row in rows):
        return "Planned"
    return "Research"


def priority_rank(value: str) -> int:
    match value:
        case "P0":
            return 0
        case "P1":
            return 1
        case "P2":
            return 2
        case "P3":
            return 3
        case _:
            return 9


def highest_priority(rows: list[dict[str, str]]) -> str:
    priorities = [row["priority"] for row in rows if row["priority"]]
    return min(priorities, key=priority_rank) if priorities else ""


def build_registry_entries(feature_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    grouped: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in feature_rows:
        grouped[(row["domain"], row["device_family"])].append(row)

    entries: list[dict[str, str]] = []
    for (domain, device_family), rows in sorted(grouped.items()):
        profile = classify_feature(rows[0]["feature"], rows[0]["module"], rows[0]["sub_module"], rows[0]["sheet"])
        status_counts = Counter(row["product_status"] for row in rows)
        related_modules = sorted({row["module"] for row in rows if row["module"]})
        entries.append(
            {
                "domain": domain,
                "device_family": device_family,
                "support_status": support_status_for(rows, profile),
                "priority": highest_priority(rows),
                "feature_count": len(rows),
                "done_count": status_counts.get("Done", 0),
                "partial_count": status_counts.get("Partial", 0),
                "pending_count": status_counts.get("Pending", 0),
                "interaction_mode": profile.interaction_mode,
                "protocols": profile.protocols,
                "integration_pattern": profile.integration_pattern,
                "manual_checklist": profile.manual_checklist,
                "config_inputs": profile.config_inputs,
                "fallback_strategy": profile.fallback_strategy,
                "simulator_strategy": profile.simulator_strategy,
                "ai_assist": profile.ai_assist,
                "next_build_step": profile.next_build_step,
                "related_modules": ", ".join(related_modules),
            }
        )
    return entries


def build_sheet_rows(feature_rows: list[dict[str, str]], domain: str) -> list[dict[str, str]]:
    return [
        row
        for row in feature_rows
        if row["domain"] == domain
    ]


def main() -> None:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)

    feature_rows: list[dict[str, str]] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            if not is_feature_row(row):
                continue

            module = clean(row[1] if len(row) > 1 else "")
            sub_module = clean(row[2] if len(row) > 2 else "")
            feature = clean(row[3] if len(row) > 3 else "")
            source = clean(row[4] if len(row) > 4 else "")
            priority = clean(row[5] if len(row) > 5 else "")
            product_status = clean(row[6] if len(row) > 6 else "")
            rfc_ref = clean(row[7] if len(row) > 7 else "")
            web = clean(row[8] if len(row) > 8 else "")
            mobile = clean(row[9] if len(row) > 9 else "")
            tv = clean(row[10] if len(row) > 10 else "")

            text = " | ".join([sheet_name, module, sub_module, feature])
            if not KEYWORDS.search(text):
                continue

            profile = classify_feature(feature, module, sub_module, sheet_name)
            feature_rows.append(
                {
                    "sheet": sheet_name,
                    "module": module or "Unknown",
                    "sub_module": sub_module,
                    "feature": feature,
                    "source": source,
                    "priority": priority,
                    "product_status": product_status or "Pending",
                    "rfc_ref": rfc_ref,
                    "web": web,
                    "mobile": mobile,
                    "tv": tv,
                    "domain": profile.domain,
                    "device_family": profile.device_family,
                    "interaction_mode": profile.interaction_mode,
                    "protocols": profile.protocols,
                    "integration_pattern": profile.integration_pattern,
                    "manual_checklist": profile.manual_checklist,
                    "config_inputs": profile.config_inputs,
                    "fallback_strategy": profile.fallback_strategy,
                    "simulator_strategy": profile.simulator_strategy,
                    "ai_assist": profile.ai_assist,
                    "next_build_step": profile.next_build_step,
                }
            )

    feature_rows.sort(
        key=lambda row: (
            row["domain"],
            priority_rank(row["priority"]),
            row["module"],
            row["sub_module"],
            row["feature"],
        )
    )

    registry_entries = build_registry_entries(feature_rows)
    support_status_counts = Counter(entry["support_status"] for entry in registry_entries)
    domain_counts = Counter(entry["domain"] for entry in registry_entries)

    manual_intake_rows = []
    for index, entry in enumerate(registry_entries, start=1):
        manual_intake_rows.append(
            {
                "tracker_id": f"DEV-{index:03d}",
                "domain": entry["domain"],
                "device_family": entry["device_family"],
                "target_vendor": "",
                "target_model": "",
                "manual_link": "",
                "protocol_verified": "No",
                "sample_payload_available": "No",
                "sdk_or_driver_captured": "No",
                "simulator_ready": "No",
                "owner": "",
                "notes": entry["manual_checklist"],
            }
        )

    domain_sheets = []
    for domain in sorted({row["domain"] for row in feature_rows}):
        rows = build_sheet_rows(feature_rows, domain)
        domain_sheets.append(
            {
                "name": domain[:31],
                "rows": rows,
            }
        )

    source_payload = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "source_workbook": str(WORKBOOK_PATH),
        "total_hardware_features": len(feature_rows),
        "registry_entry_count": len(registry_entries),
        "support_status_counts": dict(support_status_counts),
        "domain_counts": dict(domain_counts),
        "protocol_patterns": PROTOCOL_PATTERNS,
        "registry_entries": registry_entries,
        "manual_intake_rows": manual_intake_rows,
        "feature_rows": feature_rows,
        "domain_sheets": domain_sheets,
    }

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(source_payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_JSON}")


if __name__ == "__main__":
    main()

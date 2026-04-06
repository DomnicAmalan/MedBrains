import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx', read_only=True, data_only=True)

print("=" * 80)
print("MEDBRAINS FEATURES EXCEL ANALYSIS")
print("=" * 80)

# 1. List all sheet names
print(f"\n{'=' * 80}")
print(f"SHEET NAMES ({len(wb.sheetnames)} sheets)")
print(f"{'=' * 80}")
for i, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    row_count = 0
    for _ in ws.iter_rows(min_row=1):
        row_count += 1
    print(f"  {i:2d}. {name} ({row_count} rows)")

# 2. For each sheet, print column headers
print(f"\n{'=' * 80}")
print("COLUMN HEADERS PER SHEET")
print(f"{'=' * 80}")
for name in wb.sheetnames:
    ws = wb[name]
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)
        break
    print(f"\n  Sheet: '{name}'")
    for j, h in enumerate(headers):
        if h is not None:
            print(f"    Col {get_column_letter(j+1)} ({j+1}): {h}")

# 3. Print first 5 data rows from first 3 sheets
print(f"\n{'=' * 80}")
print("SAMPLE DATA (first 5 rows from first 3 sheets)")
print(f"{'=' * 80}")
for name in wb.sheetnames[:3]:
    ws = wb[name]
    print(f"\n  --- Sheet: '{name}' ---")
    headers = []
    row_num = 0
    for row in ws.iter_rows(values_only=True):
        row_num += 1
        if row_num == 1:
            headers = [str(h) if h else f"Col{i}" for i, h in enumerate(row)]
            continue
        if row_num > 6:  # 1 header + 5 data rows
            break
        # Print each cell with its header
        print(f"\n  Row {row_num - 1}:")
        for h, val in zip(headers, row):
            if val is not None:
                print(f"    {h}: {val}")

# 4. Count total features across all sheets
print(f"\n{'=' * 80}")
print("FEATURE COUNTS PER SHEET")
print(f"{'=' * 80}")
total_features = 0
for name in wb.sheetnames:
    ws = wb[name]
    # Find which column is "Feature" or similar
    headers = []
    feature_col_idx = None
    module_col_idx = None
    submodule_col_idx = None
    status_col_idx = None
    priority_col_idx = None

    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)
        break

    for idx, h in enumerate(headers):
        if h is None:
            continue
        hl = str(h).lower().strip()
        if hl == 'feature' or hl == 'features':
            feature_col_idx = idx
        if hl == 'module':
            module_col_idx = idx
        if 'sub' in hl and 'module' in hl:
            submodule_col_idx = idx
        if hl == 'status':
            status_col_idx = idx
        if hl == 'priority':
            priority_col_idx = idx

    sheet_count = 0
    status_counts = {}
    priority_counts = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_list = list(row)
        has_feature = False
        if feature_col_idx is not None and feature_col_idx < len(row_list):
            if row_list[feature_col_idx] is not None and str(row_list[feature_col_idx]).strip():
                has_feature = True
        elif any(cell is not None and str(cell).strip() for cell in row_list):
            has_feature = True

        if has_feature:
            sheet_count += 1
            if status_col_idx is not None and status_col_idx < len(row_list):
                s = row_list[status_col_idx]
                if s is not None:
                    s = str(s).strip()
                    status_counts[s] = status_counts.get(s, 0) + 1
            if priority_col_idx is not None and priority_col_idx < len(row_list):
                p = row_list[priority_col_idx]
                if p is not None:
                    p = str(p).strip()
                    priority_counts[p] = priority_counts.get(p, 0) + 1

    total_features += sheet_count
    print(f"\n  Sheet: '{name}' -> {sheet_count} features")
    print(f"    Feature col: {'Col ' + str(feature_col_idx+1) if feature_col_idx is not None else 'NOT FOUND'}")
    if status_counts:
        print(f"    Status breakdown: {dict(sorted(status_counts.items()))}")
    if priority_counts:
        print(f"    Priority breakdown: {dict(sorted(priority_counts.items()))}")

print(f"\n{'=' * 80}")
print(f"TOTAL FEATURES ACROSS ALL SHEETS: {total_features}")
print(f"{'=' * 80}")

wb.close()

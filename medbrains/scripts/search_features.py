import openpyxl
import re
from collections import defaultdict

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

# Keywords to search for (case-insensitive)
KEYWORDS = [
    "form", "field", "configuration", "no-code", "builder",
    "dynamic form", "template", "customization", "master",
    "config", "settings", "setup", "customize", "configurable",
    "no code", "nocode", "drag", "drop", "widget"
]

# Compile a single regex pattern for efficiency
# Sort by length descending so longer matches are found first
sorted_kw = sorted(KEYWORDS, key=len, reverse=True)
pattern = re.compile("|".join(re.escape(k) for k in sorted_kw), re.IGNORECASE)

# Special sheets to highlight
SPECIAL_SHEETS = ["Admin & Operations", "Technical Infrastructure", "Onboarding & Setup"]

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

print("=" * 120)
print(f"SHEETS IN WORKBOOK: {wb.sheetnames}")
print("=" * 120)

def detect_columns(ws):
    """Try to find the header row and column indices."""
    header_map = {}
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
        vals = [str(cell.value).strip().lower() if cell.value else "" for cell in row]
        known = {"s.no", "sno", "s no", "module", "sub-module", "submodule", "feature",
                 "source", "priority", "status", "web", "mobile", "tv", "rfc ref",
                 "sub module", "features"}
        matches = sum(1 for v in vals if v in known)
        if matches >= 3:
            header_row = row_idx
            for i, v in enumerate(vals):
                header_map[v] = i
            break
    return header_row, header_map

all_matches = defaultdict(list)
total_matches = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    header_row, header_map = detect_columns(ws)

    if not header_row:
        print(f"\n[!] Sheet '{sheet_name}': Could not detect header row, scanning all cells...")
        for row in ws.iter_rows(values_only=True):
            text = " | ".join(str(c) for c in row if c)
            if pattern.search(text):
                all_matches[sheet_name].append({"raw": text})
                total_matches += 1
        continue

    def get_col(names):
        for n in names:
            if n in header_map:
                return header_map[n]
        return None

    col_module = get_col(["module"])
    col_submodule = get_col(["sub-module", "submodule", "sub module"])
    col_feature = get_col(["feature", "features"])
    col_priority = get_col(["priority"])
    col_status = get_col(["status"])
    col_web = get_col(["web"])
    col_mobile = get_col(["mobile"])
    col_tv = get_col(["tv"])

    sheet_matches = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        vals = list(row)
        search_parts = []
        for ci in [col_module, col_submodule, col_feature]:
            if ci is not None and ci < len(vals) and vals[ci]:
                search_parts.append(str(vals[ci]))

        search_text = " ".join(search_parts)
        if not search_text.strip():
            continue

        if pattern.search(search_text):
            module = str(vals[col_module]).strip() if col_module is not None and col_module < len(vals) and vals[col_module] else ""
            submodule = str(vals[col_submodule]).strip() if col_submodule is not None and col_submodule < len(vals) and vals[col_submodule] else ""
            feature = str(vals[col_feature]).strip() if col_feature is not None and col_feature < len(vals) and vals[col_feature] else ""
            priority = str(vals[col_priority]).strip() if col_priority is not None and col_priority < len(vals) and vals[col_priority] else ""
            status = str(vals[col_status]).strip() if col_status is not None and col_status < len(vals) and vals[col_status] else ""
            web = str(vals[col_web]).strip() if col_web is not None and col_web < len(vals) and vals[col_web] else ""
            mobile = str(vals[col_mobile]).strip() if col_mobile is not None and col_mobile < len(vals) and vals[col_mobile] else ""
            tv = str(vals[col_tv]).strip() if col_tv is not None and col_tv < len(vals) and vals[col_tv] else ""

            matched_kw = list(set(m.group().lower() for m in pattern.finditer(search_text)))

            entry = {
                "row": row_idx,
                "module": module,
                "submodule": submodule,
                "feature": feature,
                "priority": priority,
                "status": status,
                "web": web,
                "mobile": mobile,
                "tv": tv,
                "keywords": matched_kw,
            }
            sheet_matches.append(entry)
            total_matches += 1

    all_matches[sheet_name] = sheet_matches

wb.close()

# Print results
print(f"\nTOTAL MATCHES: {total_matches}")
print("=" * 120)

for sheet_name in wb.sheetnames:
    matches = all_matches.get(sheet_name, [])
    if not matches:
        continue

    is_special = sheet_name in SPECIAL_SHEETS
    marker = " *** SPECIAL FOCUS ***" if is_special else ""

    print(f"\n{'#' * 120}")
    print(f"# SHEET: {sheet_name} ({len(matches)} matches){marker}")
    print(f"{'#' * 120}")

    for i, m in enumerate(matches, 1):
        if "raw" in m:
            print(f"\n  [{i}] RAW: {m['raw']}")
            continue

        print(f"\n  [{i}] Row {m['row']}")
        print(f"      Module:     {m['module']}")
        print(f"      Sub-Module: {m['submodule']}")
        print(f"      Feature:    {m['feature']}")
        print(f"      Priority:   {m['priority']}")
        print(f"      Status:     {m['status']}")
        print(f"      Platforms:  Web={m['web']}  Mobile={m['mobile']}  TV={m['tv']}")
        print(f"      Keywords:   {', '.join(m['keywords'])}")

# Summary section
print(f"\n\n{'=' * 120}")
print("SUMMARY BY SHEET")
print(f"{'=' * 120}")
for sheet_name in wb.sheetnames:
    matches = all_matches.get(sheet_name, [])
    count = len(matches)
    if count > 0:
        marker = " <-- SPECIAL FOCUS" if sheet_name in SPECIAL_SHEETS else ""
        print(f"  {sheet_name}: {count} matches{marker}")

# Keyword frequency
print(f"\n{'=' * 120}")
print("KEYWORD FREQUENCY ACROSS ALL MATCHES")
print(f"{'=' * 120}")
kw_counts = defaultdict(int)
for sheet_matches in all_matches.values():
    for m in sheet_matches:
        if "keywords" in m:
            for kw in m["keywords"]:
                kw_counts[kw] += 1

for kw, count in sorted(kw_counts.items(), key=lambda x: -x[1]):
    print(f"  {kw}: {count}")


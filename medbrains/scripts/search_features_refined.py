import openpyxl
import re
from collections import defaultdict

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

# Refined keywords with word boundaries to avoid false positives
# Group 1: High relevance (form builder, dynamic forms, no-code, configuration)
HIGH_PATTERNS = [
    (r'\bform\s*builder\b', 'form builder'),
    (r'\bdynamic\s*form\b', 'dynamic form'),
    (r'\bno[- ]?code\b', 'no-code'),
    (r'\bdrag\s*(?:and|&|n)?\s*drop\b', 'drag-and-drop'),
    (r'\bcustom\s*field', 'custom field'),
    (r'\bconfigurable\s*field', 'configurable field'),
    (r'\bfield\s*(?:type|builder|config|definition|template)', 'field config'),
    (r'\btemplate\s*(?:builder|engine|designer|editor|management|creation)', 'template builder'),
    (r'\bwidget', 'widget'),
    (r'\bconsent\s*form', 'consent form'),
    (r'\bform\s*(?:design|template|engine|schema|layout|section|customiz)', 'form design'),
    (r'\bbuilder\b', 'builder'),
]

# Group 2: Configuration & Settings
CONFIG_PATTERNS = [
    (r'\bconfiguration\b', 'configuration'),
    (r'\bconfigurable\b', 'configurable'),
    (r'\bsettings?\b', 'settings'),
    (r'\bsetup\b', 'setup'),
    (r'\bcustomiz(?:e|ation|able)\b', 'customization'),
]

# Group 3: Master data
MASTER_PATTERNS = [
    (r'\bmaster\b', 'master'),
]

ALL_GROUPS = {
    "FORM/BUILDER/NO-CODE": HIGH_PATTERNS,
    "CONFIGURATION/SETTINGS": CONFIG_PATTERNS,
    "MASTER DATA": MASTER_PATTERNS,
}

SPECIAL_SHEETS = ["Admin & Operations", "Technical Infrastructure", "Onboarding & Setup"]

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

def detect_columns(ws):
    header_map = {}
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
        vals = [str(cell.value).strip().lower() if cell.value else "" for cell in row]
        known = {"s.no", "sno", "s no", "module", "sub-module", "submodule", "feature",
                 "source", "priority", "status", "web", "mobile", "tv", "rfc ref",
                 "sub module", "features"}
        matches = sum(1 for v in vals if v in known)
        if matches >= 3:
            header_row = row_idx
            for i, v in enumerate(vals):
                header_map[v] = i
            break
    return header_row, header_map

# Collect results organized by group
results_by_group = defaultdict(list)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    header_row, header_map = detect_columns(ws)
    if not header_row:
        continue

    def get_col(names):
        for n in names:
            if n in header_map:
                return header_map[n]
        return None

    col_module = get_col(["module"])
    col_submodule = get_col(["sub-module", "submodule", "sub module"])
    col_feature = get_col(["feature", "features"])
    col_priority = get_col(["priority"])
    col_status = get_col(["status"])
    col_web = get_col(["web"])
    col_mobile = get_col(["mobile"])
    col_tv = get_col(["tv"])

    # Track current module/submodule for rows that inherit from headers
    current_module = ""
    current_submodule = ""

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        vals = list(row)
        
        raw_module = str(vals[col_module]).strip() if col_module is not None and col_module < len(vals) and vals[col_module] else ""
        raw_submodule = str(vals[col_submodule]).strip() if col_submodule is not None and col_submodule < len(vals) and vals[col_submodule] else ""
        raw_feature = str(vals[col_feature]).strip() if col_feature is not None and col_feature < len(vals) and vals[col_feature] else ""
        raw_priority = str(vals[col_priority]).strip() if col_priority is not None and col_priority < len(vals) and vals[col_priority] else ""
        raw_status = str(vals[col_status]).strip() if col_status is not None and col_status < len(vals) and vals[col_status] else ""
        raw_web = str(vals[col_web]).strip() if col_web is not None and col_web < len(vals) and vals[col_web] else ""
        raw_mobile = str(vals[col_mobile]).strip() if col_mobile is not None and col_mobile < len(vals) and vals[col_mobile] else ""
        raw_tv = str(vals[col_tv]).strip() if col_tv is not None and col_tv < len(vals) and vals[col_tv] else ""

        # Update current module/submodule context
        if raw_module and raw_module != "None":
            current_module = raw_module
        if raw_submodule and raw_submodule != "None":
            current_submodule = raw_submodule

        # Skip header-only rows (no feature text)
        if not raw_feature or raw_feature == "None":
            continue

        search_text = f"{current_module} {current_submodule} {raw_feature}"

        for group_name, patterns in ALL_GROUPS.items():
            matched_labels = []
            for pat, label in patterns:
                if re.search(pat, search_text, re.IGNORECASE):
                    matched_labels.append(label)
            
            if matched_labels:
                entry = {
                    "sheet": sheet_name,
                    "row": row_idx,
                    "module": current_module if raw_module in ("", "None") else raw_module,
                    "submodule": current_submodule if raw_submodule in ("", "None") else raw_submodule,
                    "feature": raw_feature,
                    "priority": raw_priority if raw_priority != "None" else "",
                    "status": raw_status if raw_status != "None" else "",
                    "web": raw_web if raw_web != "None" else "",
                    "mobile": raw_mobile if raw_mobile != "None" else "",
                    "tv": raw_tv if raw_tv != "None" else "",
                    "matched": matched_labels,
                }
                results_by_group[group_name].append(entry)

wb.close()

# ========== PRINT RESULTS ==========

for group_name in ["FORM/BUILDER/NO-CODE", "CONFIGURATION/SETTINGS", "MASTER DATA"]:
    entries = results_by_group.get(group_name, [])
    print(f"\n{'=' * 130}")
    print(f"  GROUP: {group_name} ({len(entries)} features)")
    print(f"{'=' * 130}")
    
    # Sort by sheet, then row
    entries.sort(key=lambda e: (e["sheet"], e["row"]))
    
    current_sheet = ""
    for i, e in enumerate(entries, 1):
        if e["sheet"] != current_sheet:
            current_sheet = e["sheet"]
            special = " *** SPECIAL FOCUS ***" if current_sheet in SPECIAL_SHEETS else ""
            print(f"\n  --- {current_sheet}{special} ---")
        
        print(f"\n    [{i}] Row {e['row']}  |  Priority: {e['priority'] or '?'}  |  Status: {e['status'] or '?'}  |  Web={e['web']}  Mob={e['mobile']}  TV={e['tv']}")
        print(f"         Module:     {e['module']}")
        print(f"         Sub-Module: {e['submodule']}")
        print(f"         Feature:    {e['feature']}")
        print(f"         Matched:    {', '.join(e['matched'])}")

# ========== FINAL SUMMARY ==========
print(f"\n\n{'=' * 130}")
print("FINAL SUMMARY")
print(f"{'=' * 130}")
for group_name in ["FORM/BUILDER/NO-CODE", "CONFIGURATION/SETTINGS", "MASTER DATA"]:
    entries = results_by_group.get(group_name, [])
    print(f"\n  {group_name}: {len(entries)} features")
    # Count by sheet
    by_sheet = defaultdict(int)
    for e in entries:
        by_sheet[e["sheet"]] += 1
    for s, c in sorted(by_sheet.items(), key=lambda x: -x[1]):
        marker = " <-- SPECIAL" if s in SPECIAL_SHEETS else ""
        print(f"    {s}: {c}{marker}")

# Count unique features (some may appear in multiple groups)
all_features = set()
for entries in results_by_group.values():
    for e in entries:
        all_features.add((e["sheet"], e["row"]))
print(f"\n  TOTAL UNIQUE FEATURES: {len(all_features)}")


#!/usr/bin/env python3
"""Update MedBrains_Features.xlsx for Billing Phase 2b completion."""

import openpyxl
import os

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "MedBrains_Features.xlsx")
wb = openpyxl.load_workbook(EXCEL_PATH)

# Billing features are in "Admin & Operations" sheet
ws = None
for name in wb.sheetnames:
    if "admin" in name.lower() and "operations" in name.lower():
        ws = wb[name]
        break

if ws is None:
    print("ERROR: Could not find Admin & Operations sheet")
    exit(1)

# Find Status column index
status_col = None
feature_col = None
for col in range(1, ws.max_column + 1):
    val = str(ws.cell(row=1, column=col).value or "").strip().lower()
    if val == "status":
        status_col = col
    if val == "feature":
        feature_col = col

if status_col is None or feature_col is None:
    print("ERROR: Could not find Status or Feature column")
    exit(1)

# Define updates: { feature_substring: new_status }
# Status corrections for already-built features mis-marked as Pending
corrections = {
    "Interim bill generation for long-stay": "Done",
    "Advance/deposit collection": "Done",
    "Insurance/TPA pre-authorization": "Done",
    "Cashless claim processing": "Done",
    "Claim rejection tracking and re-submission": "Done",
    "Corporate billing & invoicing": "Done",
    "Corporate credit management": "Done",
    "Revenue dashboard": "Done",
    "Outstanding receivables tracking": "Done",
    "Department-wise revenue analysis": "Done",
}

# New completions from Phase 2b
new_done = {
    "IPD billing (room charges": "Done",
    "Final bill auto-compilation": "Done",
    "Duplicate bill generation": "Done",
    "GST/tax management": "Done",
    "TPA-wise rate negotiation": "Done",
    "Doctor-wise revenue attribution": "Done",
    "Bad debt identification and write-off": "Done",
    "Audit trail for all financial transactions": "Done",
    "End-of-day cash closing": "Done",
    "Radiology billing": "Done",
    "Insurance panel (smart panel": "Done",
}

new_partial = {
    "ER billing (deferred billing": "Partial",
    "Co-pay/deductible calculation": "Partial",
    "Day-end/month-end reconciliation": "Partial",
}

all_updates = {**corrections, **new_done, **new_partial}
updated = 0

for row in range(2, ws.max_row + 1):
    feature = str(ws.cell(row=row, column=feature_col).value or "")
    for pattern, new_status in all_updates.items():
        if pattern.lower() in feature.lower():
            old_status = ws.cell(row=row, column=status_col).value
            ws.cell(row=row, column=status_col).value = new_status
            print(f"  Row {row}: '{feature[:50]}...' {old_status} → {new_status}")
            updated += 1
            break

wb.save(EXCEL_PATH)
print(f"\nUpdated {updated} rows in {ws.title}")
print("Done!")

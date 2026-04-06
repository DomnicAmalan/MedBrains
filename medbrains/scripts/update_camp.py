#!/usr/bin/env python3
"""Update MedBrains_Features.xlsx — mark Camp Management features as Done/Partial."""

import openpyxl
from pathlib import Path

XLSX = Path(__file__).resolve().parent.parent.parent / "MedBrains_Features.xlsx"

# Map feature substrings → target status
FEATURES = {
    "Camp planning": "Done",
    "Camp location": "Done",
    "Camp patient registration": "Done",
    "Camp lab sample": "Done",
    "Camp billing": "Done",
    "Camp reports": "Partial",
    "Follow-up tracking for camp": "Done",
    "Camp-to-hospital patient conversion": "Done",
}

def main():
    wb = openpyxl.load_workbook(str(XLSX))
    updated = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, max_col=20):
            # Feature text is typically in column D (index 3) or column C (index 2)
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for pattern, status in FEATURES.items():
                        if pattern.lower() in cell.value.lower():
                            # Find the Status column — look for a cell with "Status" in header row
                            status_col = None
                            for hcell in ws[1]:
                                if hcell.value and "status" in str(hcell.value).lower():
                                    status_col = hcell.column
                                    break
                            if status_col:
                                target = ws.cell(row=cell.row, column=status_col)
                                old = target.value
                                target.value = status
                                print(f"  [{ws.title}] Row {cell.row}: '{cell.value[:50]}...' → {status} (was: {old})")
                                updated += 1
                            break  # Only match first pattern per cell

    if updated:
        wb.save(str(XLSX))
        print(f"\nUpdated {updated} features in {XLSX.name}")
    else:
        print("No matching features found — check feature names in Excel")

if __name__ == "__main__":
    main()

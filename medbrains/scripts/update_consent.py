#!/usr/bin/env python3
"""Update MedBrains_Features.xlsx for Consent Management & Medico-Legal module."""

import os
from openpyxl import load_workbook

XLSX = os.path.join(os.path.dirname(__file__), "..", "..", "MedBrains_Features.xlsx")

# Row number -> new_status (from the investigation above)
ROW_UPDATES = {
    # 24.1 Consent Types & Generation
    32: "Done",    # General/surgical/anesthesia/blood transfusion consent
    33: "Done",    # HIV testing, high-risk, research consent (already Done)
    34: "Done",    # Informed refusal documentation
    35: "Done",    # DAMA/LAMA consent
    36: "Done",    # Photography/video/teaching consent
    37: "Done",    # ABDM/DPDP data sharing consent
    38: "Done",    # DNR/advance directive
    39: "Partial", # Multi-language consent (template JSONB ready; auto-translate deferred)
    40: "Partial", # Read-aloud (flag on template; TTS engine deferred)
    41: "Done",    # Organ donation (THOA)
    # 24.2 Digital Signature & Verification
    43: "Partial", # Aadhaar e-Sign (metadata fields; API deferred)
    44: "Partial", # Biometric (metadata fields; device integration deferred)
    45: "Partial", # Digital pen signature (metadata fields; capture UI deferred)
    46: "Partial", # Video consent (metadata fields; recording infra deferred)
    47: "Done",    # Consent verification hard-block
    48: "Done",    # Consent revocation
}


def main():
    wb = load_workbook(XLSX)
    ws = wb["Specialty & Academic"]

    # Status column is G = index 6
    STATUS_COL = 6

    updated = 0
    for row_num, new_status in ROW_UPDATES.items():
        row = ws[row_num]
        feat = str(row[3].value or "").strip()
        old = row[STATUS_COL].value
        if old == new_status:
            print(f"  Row {row_num}: ALREADY '{new_status}' — '{feat[:55]}'")
            continue
        row[STATUS_COL].value = new_status
        print(f"  Row {row_num}: {old} -> {new_status} — '{feat[:55]}'")
        updated += 1

    # 24.3 Medico-Legal Documents — need to find rows 50-56
    for row_num in range(50, 57):
        row = ws[row_num]
        feat = str(row[3].value or "").strip()
        if not feat:
            continue

        feat_lower = feat.lower()
        if "mlc registration" in feat_lower or "police intimation" in feat_lower:
            new = "Done"
        elif "wound certificate" in feat_lower or "body diagram" in feat_lower:
            new = "Done"
        elif "pocso" in feat_lower:
            new = "Done"
        elif "age estimation" in feat_lower:
            new = "Partial"
        elif "death certificate" in feat_lower:
            new = "Partial"
        elif "court summons" in feat_lower:
            new = "Partial"
        elif "medico-legal opinion" in feat_lower:
            new = "Partial"
        else:
            continue

        old = row[STATUS_COL].value
        if old == new:
            print(f"  Row {row_num}: ALREADY '{new}' — '{feat[:55]}'")
            continue
        row[STATUS_COL].value = new
        print(f"  Row {row_num}: {old} -> {new} — '{feat[:55]}'")
        updated += 1

    wb.save(XLSX)
    print(f"\nDone — {updated} features updated in {XLSX}")


if __name__ == "__main__":
    main()

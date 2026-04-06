#!/usr/bin/env python3
"""
Update Front Office & Reception feature statuses in MedBrains_Features.xlsx.

DONE features (8):
- Appointment booking (slot/token configurable)
- Walk-in vs scheduled patient differentiation
- Token generation with estimated wait time
- Priority queue for disabled, elderly, pregnant, emergency referral
- Visitor registration with photo, ID verification, and contact
- Visitor pass generation (time-limited, ward-specific, QR code)
- Visiting hours enforcement per ward type
- Maximum visitor count per patient per slot

PARTIAL features (6):
- Queue display on waiting area screens (config stored, TV app needed)
- SMS/WhatsApp notification 3 patients before turn (schema ready, notification service deferred)
- Special categories (legal counsel, religious, VIP) (visitor_category enum, protocol workflow deferred)
- Visitor management system (core CRUD done, gate integration deferred)
- Parking management (not in scope, needs separate subsystem)
- Lab/radiology/pharmacy queues (queue priority framework built, per-dept integration deferred)

ALREADY DONE (2):
- Real-time doctor availability display
- Outbreak mode visitor restriction
"""

import openpyxl
import os
import re

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "MedBrains_Features.xlsx",
)


def find_status_col(ws):
    """Find the Status column index from the header row."""
    for cell in ws[1]:
        try:
            if str(cell.value or "").strip().lower() == "status":
                return cell.column
        except AttributeError:
            continue
    return None


def word_boundary_match(keyword, text):
    """Match keyword with word boundaries for short keywords, substring for longer."""
    if len(keyword) <= 5:
        pattern = r"\b" + re.escape(keyword) + r"\b"
        return bool(re.search(pattern, text, re.IGNORECASE))
    return keyword.upper() in text.upper()


def main():
    print(f"Opening workbook: {WORKBOOK_PATH}")
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    changes = []

    # ═══════════════════════════════════════════════════════════════════
    # Search all sheets for Front Office / Queue / Visitor features
    # ═══════════════════════════════════════════════════════════════════

    MODULE_KEYWORDS = [
        "Front Office",
        "Queue Management",
        "Visitor Management",
        "visitor registration",
        "visitor pass",
        "visiting hours",
        "enquiry desk",
        "reception desk",
    ]

    DONE_KEYWORDS = [
        "appointment booking",
        "walk-in vs scheduled",
        "token generation with estimated wait",
        "priority queue for disabled",
        "visitor registration with photo",
        "visitor pass generation",
        "visiting hours enforcement",
        "maximum visitor count",
    ]

    PARTIAL_KEYWORDS = [
        "queue display on waiting",
        "SMS/WhatsApp notification",
        "sms notification",
        "whatsapp notification",
        "special categories",
        "legal counsel",
        "visitor management system",
        "parking management",
        "lab sample collection queue",
        "radiology queue",
        "pharmacy queue",
        "lab/radiology/pharmacy",
    ]

    SKIP_STATUSES_FOR_DONE = {"done"}
    SKIP_STATUSES_FOR_PARTIAL = {"done", "partial"}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sc = find_status_col(ws)
        if not sc:
            continue

        for r in range(2, ws.max_row + 1):
            module_val = str(ws.cell(row=r, column=2).value or "")
            submod_val = str(ws.cell(row=r, column=3).value or "")
            feature_val = str(ws.cell(row=r, column=4).value or "")

            if not feature_val.strip():
                continue

            combined = f"{module_val} {submod_val} {feature_val}"

            is_module_related = any(
                word_boundary_match(kw, combined) for kw in MODULE_KEYWORDS
            )

            if not is_module_related:
                continue

            old_val = ws.cell(row=r, column=sc).value or ""
            old_str = str(old_val).strip().lower()

            # Check Done keywords
            if any(word_boundary_match(kw, combined) for kw in DONE_KEYWORDS):
                if old_str not in SKIP_STATUSES_FOR_DONE:
                    ws.cell(row=r, column=sc).value = "Done"
                    changes.append(
                        f"  {sheet_name} row {r}: {old_val!r} -> 'Done' "
                        f"({feature_val[:80]})"
                    )
                continue

            # Check Partial keywords
            if any(word_boundary_match(kw, combined) for kw in PARTIAL_KEYWORDS):
                if old_str not in SKIP_STATUSES_FOR_PARTIAL:
                    ws.cell(row=r, column=sc).value = "Partial"
                    changes.append(
                        f"  {sheet_name} row {r}: {old_val!r} -> 'Partial' "
                        f"({feature_val[:80]})"
                    )
                continue

            # Module-related but not matched — report
            print(
                f"  [SKIPPED] {sheet_name} row {r}: "
                f"'{feature_val[:80]}' (status: {old_val!r})"
            )

    # ═══════════════════════════════════════════════════════════════════
    # Save
    # ═══════════════════════════════════════════════════════════════════
    wb.save(WORKBOOK_PATH)
    wb.close()

    print(f"\n{'=' * 70}")
    print(f"Updated {len(changes)} Front Office feature statuses:")
    print(f"{'=' * 70}")
    for c in changes:
        print(c)

    if not changes:
        print("  (no changes needed — all statuses already up to date)")

    done_count = sum(1 for c in changes if "-> 'Done'" in c)
    partial_count = sum(1 for c in changes if "-> 'Partial'" in c)
    print(f"\nSummary: {done_count} marked Done, {partial_count} marked Partial")


if __name__ == "__main__":
    main()

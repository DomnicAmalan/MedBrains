#!/usr/bin/env python3
"""
Update ICU / Critical Care feature statuses in MedBrains_Features.xlsx.

DONE features:
- ICU flowsheets / hourly vitals documentation
- ICU I/O (intake/output) tracking
- Ventilator settings / ventilator mode recording
- ABG (arterial blood gas) recording
- APACHE II / APACHE IV scoring
- SOFA scoring
- GCS scoring
- ICU device tracking / line tracking
- Bundle compliance checks (CLABSIs, CAUTI)
- ICU nutrition management (enteral/parenteral)
- NICU / Neonatal records
- ICU severity scoring (PRISM, SNAPPE, RASS, CAM-ICU)

PARTIAL features:
- ICU medication infusion tracking (basic JSON field, no full pump integration)
- ICU trend charting / graphing (data collected, no frontend charts yet)
- ICU bed management / real-time census (uses IPD bed system, no dedicated ICU board)
- Ventilator weaning protocols (records data but no protocol engine)
- Alarm management / device integration (no physical device integration)
- ICU handoff / shift report (no dedicated handoff workflow)
- Mortality prediction / outcomes analysis (stores predicted_mortality but no analytics engine)
- Phototherapy management (NICU) (basic tracking, no automated dosing)
"""

import openpyxl
import os
import re

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "MedBrains_Features.xlsx",
)

STATUS_COL = 7  # Column G = Status


def find_status_col(ws):
    """Find the Status column index from the header row."""
    for cell in ws[1]:
        try:
            if str(cell.value or "").strip().lower() == "status":
                return cell.column
        except AttributeError:
            # MergedCell — skip
            continue
    return None


def word_boundary_match(keyword, text):
    """Match keyword with word boundaries to avoid substring false positives.

    Short keywords (<=4 chars like ICU, CCU, ABG, GCS, SOFA, RASS) use
    word-boundary regex. Longer keywords use simple substring matching.
    """
    if len(keyword) <= 5:
        pattern = r'\b' + re.escape(keyword) + r'\b'
        return bool(re.search(pattern, text, re.IGNORECASE))
    return keyword.upper() in text.upper()


def main():
    print(f"Opening workbook: {WORKBOOK_PATH}")
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    changes = []

    # ═══════════════════════════════════════════════════════════════════
    # CLINICAL SHEET — ICU Section (rows 227-255) + IPD ICU rows
    # Precise row-based updates (no keyword ambiguity).
    # ═══════════════════════════════════════════════════════════════════
    ws = wb["Clinical"]
    sc = find_status_col(ws)
    if not sc:
        print("ERROR: No Status column found in Clinical sheet")
        return

    # ── 4.1 ICU Documentation ──
    done_clinical = {
        # ICU flowsheets / hourly vitals + I/O tracking
        229: "Hourly ICU flowsheet (vitals, ventilator settings, drugs, I/O)",
        # Ventilator settings / mode recording + ABG recording
        230: "Ventilator parameter tracking (mode, FiO2, PEEP, TV, RR, ABG correlation)",
        # RASS sedation scoring (part of severity scoring)
        233: "Sedation scoring (RASS/Richmond) and sedation vacation documentation",
        # ICU nutrition management (enteral/parenteral)
        234: "Nutrition tracking (enteral/parenteral with calorie calculation)",
    }

    partial_clinical = {
        # Hemodynamic monitoring — no physical device integration
        231: "Hemodynamic monitoring integration (alarm/device integration not built)",
        # Infusion pump tracking — basic JSON field, no pump integration
        232: "Infusion pump tracking (basic JSON field, no full pump integration)",
    }

    # ── 4.2 Scoring & Outcomes ──
    done_clinical.update({
        # APACHE II/IV scoring
        236: "APACHE II/IV score calculation",
        # SOFA scoring
        237: "SOFA score daily tracking",
        # GCS scoring
        238: "GCS tracking with trend",
        # PRISM (PICU), SNAPPE (NICU) severity scoring
        239: "PRISM score (PICU), SNAPPE (NICU)",
    })

    partial_clinical.update({
        # Mortality prediction — stores predicted_mortality but no analytics engine
        240: "Predicted mortality vs actual outcome comparison (no analytics engine)",
        # ICU LOS tracking — data collected, no frontend trend charting
        241: "ICU length of stay and readmission tracking (data collected, no charts yet)",
    })

    # ── 4.3 Bundle Compliance & Device Tracking ──
    done_clinical.update({
        # Central line bundle compliance (CLABSI)
        243: "Central line bundle compliance (insertion checklist, daily necessity review)",
        # Central line days tracking (device tracking)
        244: "Central line days tracking (auto-calculated from insertion date)",
        # Urinary catheter days tracking (CAUTI)
        245: "Urinary catheter days tracking with daily necessity review",
        # Ventilator days tracking with VAP bundle
        246: "Ventilator days tracking with VAP bundle compliance",
        # Restraint documentation (bundle compliance)
        247: "Restraint documentation with 4-hour review cycle",
    })

    partial_clinical.update({
        # Device-related infection rate — stores data but no auto-calculation analytics
        248: "Device-related infection rate auto-calculation (stores data, no analytics engine)",
    })

    # ── 4.4 NICU Specific ──
    done_clinical.update({
        # NICU / Neonatal records — vital signs
        250: "Neonatal vital signs (with normative ranges by gestational age)",
        # Breast milk management (part of NICU records)
        251: "Breast milk management (labeling, storage, administration)",
        # Neonatal sepsis screening
        253: "Neonatal sepsis screening protocol",
        # Mother-baby matching (NICU security)
        254: "Mother-baby identification matching (Code Pink security)",
    })

    partial_clinical.update({
        # Phototherapy — basic tracking, no automated dosing
        252: "Phototherapy monitoring and bilirubin trending (basic tracking, no automated dosing)",
        # NHSP — no physical device integration
        255: "Newborn Hearing Screening Program (NHSP) integration (no device integration)",
    })

    # ── IPD rows that reference ICU (in Clinical sheet) ──
    partial_clinical.update({
        # ICU bed management — uses IPD bed system, no dedicated ICU board
        136: "ICU bed management (uses IPD bed system, no dedicated ICU board)",
        # NICU/PICU bed management — uses IPD bed system, no dedicated board
        137: "NICU/PICU bed management (uses IPD bed system, no dedicated board)",
    })

    done_clinical.update({
        # ICU care bundle compliance — built as part of ICU bundle compliance
        153: "ICU care bundle compliance",
        # Ventilator settings & monitoring — built in ICU documentation
        154: "Ventilator settings & monitoring",
        # ICU nursing flowsheet — built as hourly flowsheet
        182: "ICU nursing flowsheet (hourly documentation)",
    })

    partial_clinical.update({
        # NICU nursing — neonatal assessment done, phototherapy partial
        183: "NICU nursing — neonatal assessment, feeding chart, phototherapy log (partial phototherapy)",
    })

    # Apply changes to Clinical sheet
    for row_num, desc in done_clinical.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() != "done":
            ws.cell(row=row_num, column=sc).value = "Done"
            changes.append(f"  Clinical row {row_num}: {old_val!r} -> 'Done' ({desc})")

    for row_num, desc in partial_clinical.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() not in ("done", "partial"):
            ws.cell(row=row_num, column=sc).value = "Partial"
            changes.append(f"  Clinical row {row_num}: {old_val!r} -> 'Partial' ({desc})")

    # ═══════════════════════════════════════════════════════════════════
    # KEYWORD-BASED SEARCH — remaining sheets
    # Search for ICU/NICU/Critical Care features across all other sheets
    # and apply status where appropriate.
    # Uses word-boundary matching for short keywords to avoid false
    # positives (e.g. "CCU" in "occupancy", "CAUTI" in "precautions").
    # ═══════════════════════════════════════════════════════════════════

    # Keywords to identify ICU-related rows (checked with word boundaries)
    ICU_KEYWORDS = [
        "ICU", "NICU", "PICU", "CCU", "HDU",
        "Critical Care", "Intensive Care",
        "Ventilator", "APACHE", "SOFA",
        "GCS", "ABG", "CLABSI", "CAUTI",
        "Flowsheet", "Phototherapy",
    ]

    # Sheets already handled above (Clinical) — skip for keyword search
    SKIP_SHEETS = {"Clinical"}

    # Features to mark Done by keyword matching (checked with word boundaries)
    DONE_KEYWORDS = [
        "icu flowsheet",
        "hourly icu",
        "hourly charting",
        "intake/output",
        "i/o tracking",
        "ventilator setting",
        "ventilator parameter",
        "ventilator mode",
        "abg recording",
        "abg correlation",
        "arterial blood gas",
        "apache ii",
        "apache iv",
        "apache score",
        "sofa score",
        "sofa daily",
        "gcs tracking",
        "gcs scoring",
        "device tracking",
        "line tracking",
        "line bundle",
        "central line bundle",
        "catheter days",
        "central line days",
        "bundle compliance",
        "care bundle",
        "clabsi",
        "cauti",
        "nutrition tracking",
        "enteral/parenteral",
        "neonatal vital",
        "neonatal assessment",
        "breast milk",
        "sepsis screening",
        "mother-baby",
        "code pink",
        "sedation scoring",
        "rass",
        "prism score",
        "snappe",
        "restraint documentation",
    ]

    # Features to mark Partial by keyword matching (checked with word boundaries)
    PARTIAL_KEYWORDS = [
        "infusion pump",
        "infusion tracking",
        "medication infusion",
        "trend chart",
        "trend analysis",
        "graphing",
        "icu bed manage",
        "icu bed occupancy",
        "real-time bed",
        "bed census",
        "ventilator weaning",
        "weaning protocol",
        "alarm manage",
        "device integration",
        "hemodynamic",
        "icu handoff",
        "icu handover",
        "shift report",
        "shift handover",
        "mortality predict",
        "predicted mortality",
        "outcome comparison",
        "outcomes analysis",
        "phototherapy monitor",
        "phototherapy manage",
        "bilirubin trending",
        "infection rate",
        "icu length of stay",
        "readmission track",
        "hearing screening",
        "nicu nursing",
    ]

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws2 = wb[sheet_name]
        sc2 = find_status_col(ws2)
        if not sc2:
            continue

        for r in range(2, ws2.max_row + 1):
            # Read module (col 2), sub-module (col 3), feature (col 4)
            module_val = str(ws2.cell(row=r, column=2).value or "")
            submod_val = str(ws2.cell(row=r, column=3).value or "")
            feature_val = str(ws2.cell(row=r, column=4).value or "")

            # Skip header/section rows (no feature text)
            if not feature_val.strip():
                continue

            # Check if this row is ICU-related (word-boundary matching)
            combined = f"{module_val} {submod_val} {feature_val}"
            is_icu_related = any(
                word_boundary_match(kw, combined) for kw in ICU_KEYWORDS
            )

            if not is_icu_related:
                continue

            old_val = ws2.cell(row=r, column=sc2).value or ""
            old_str = str(old_val).strip().lower()

            # Check Done keywords (word-boundary matching)
            if any(word_boundary_match(kw, combined) for kw in DONE_KEYWORDS):
                if old_str != "done":
                    ws2.cell(row=r, column=sc2).value = "Done"
                    changes.append(
                        f"  {sheet_name} row {r}: {old_val!r} -> 'Done' "
                        f"({feature_val[:80]})"
                    )
                continue

            # Check Partial keywords (word-boundary matching)
            if any(word_boundary_match(kw, combined) for kw in PARTIAL_KEYWORDS):
                if old_str not in ("done", "partial"):
                    ws2.cell(row=r, column=sc2).value = "Partial"
                    changes.append(
                        f"  {sheet_name} row {r}: {old_val!r} -> 'Partial' "
                        f"({feature_val[:80]})"
                    )
                continue

            # ICU-related but not matched by Done/Partial keywords — report but leave as-is
            print(
                f"  [SKIPPED] {sheet_name} row {r}: "
                f"'{feature_val[:80]}' (status: {old_val!r})"
            )

    # ═══════════════════════════════════════════════════════════════════
    # Save
    # ═══════════════════════════════════════════════════════════════════
    wb.save(WORKBOOK_PATH)
    wb.close()

    print(f"\n{'=' * 70}")
    print(f"Updated {len(changes)} ICU/Critical Care feature statuses:")
    print(f"{'=' * 70}")
    for c in changes:
        print(c)

    if not changes:
        print("  (no changes needed — all statuses already up to date)")

    # Summary counts
    done_count = sum(1 for c in changes if "-> 'Done'" in c)
    partial_count = sum(1 for c in changes if "-> 'Partial'" in c)
    print(f"\nSummary: {done_count} marked Done, {partial_count} marked Partial")


if __name__ == "__main__":
    main()

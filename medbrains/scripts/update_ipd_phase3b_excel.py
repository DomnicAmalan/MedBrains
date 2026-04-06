#!/usr/bin/env python3
"""Mark IPD Phase 3b features as Done in MedBrains_Features.xlsx."""

import openpyxl
from pathlib import Path

EXCEL_PATH = Path(__file__).resolve().parent.parent.parent / "MedBrains_Features.xlsx"

# IPD Phase 3b feature rows to mark Done (1-indexed rows in the Clinical sheet)
# These are the 15 software-buildable Partial features now completed.
ROWS_TO_MARK_DONE = [
    170,  # Chemo administration (ABAC check)
    178,  # Blood transfusion checklist & reaction
    179,  # Incident reporting (anonymous)
    180,  # Restraint monitoring log (30-min checks)
    181,  # Elopement risk assessment
    184,  # OT surgical safety checklist (WHO)
    185,  # Dialysis nursing (pre/intra/post)
    186,  # Endoscopy nursing (Aldrete score)
    208,  # Surgeon caseload analysis
    209,  # Anesthesia complication tracking
    219,  # Death summary auto-generation
    221,  # Birth certificate generation
    222,  # Transfer summary (inter-hospital)
    223,  # Dept-wise billing config at discharge
    224,  # Billing threshold control
]

SHEET_NAME = "Clinical"
STATUS_COL = 7  # Column G = Status


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        return

    ws = wb[SHEET_NAME]
    updated = 0

    for row_num in ROWS_TO_MARK_DONE:
        cell = ws.cell(row=row_num, column=STATUS_COL)
        old_val = cell.value
        cell.value = "Done"
        print(f"  Row {row_num}: '{old_val}' -> 'Done'")
        updated += 1

    wb.save(EXCEL_PATH)
    print(f"\nUpdated {updated} rows to 'Done' in sheet '{SHEET_NAME}'.")
    print(f"Saved: {EXCEL_PATH}")


if __name__ == "__main__":
    main()

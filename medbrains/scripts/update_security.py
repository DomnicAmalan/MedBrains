#!/usr/bin/env python3
"""Update MedBrains_Features.xlsx to mark Security Department features."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook("/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx")

# Security features are in the "Admin & Operations" sheet
# Find the sheet
sheet_name = None
for name in wb.sheetnames:
    if "Admin" in name and "Oper" in name:
        sheet_name = name
        break
    if "admin" in name.lower() and "oper" in name.lower():
        sheet_name = name
        break

if not sheet_name:
    # Try to find any sheet with security features
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "Security" in cell.value:
                    sheet_name = name
                    break
            if sheet_name:
                break
        if sheet_name:
            break

if not sheet_name:
    print("Could not find Admin & Operations sheet. Listing all sheets:")
    for name in wb.sheetnames:
        print(f"  - {name}")
    exit(1)

ws = wb[sheet_name]
print(f"Found sheet: {sheet_name}")

# Feature mapping: feature substring -> new status
FEATURE_STATUS = {
    "Physical Access": "Partial",          # Hardware deferred
    "Zone definitions": "Done",
    "Access log for each zone": "Done",
    "Access card provisioning": "Done",
    "After-hours access": "Done",
    "Camera inventory": "Done",
    "Incident-linked video": "Partial",    # NVR integration deferred
    "Recording retention": "Done",
    "Incident reporting": "Done",
    "Infant RFID": "Partial",             # Hardware deferred
    "Zone alert if infant": "Partial",    # Sensor deferred
    "Code Pink activation": "Done",
    "Absconder": "Done",
    "Wander guard": "Partial",            # Sensor deferred
    "Code Blue": "Done",
    "Code Red": "Done",
    "Code Pink": "Done",
    "Code Silver": "Done",
    "Code Orange": "Done",
    "Code Black": "Done",
    "Code Yellow": "Done",
    "Post-event debrief": "Done",
    "mass casualty": "Done",
}

# Find the Status column (usually column F or look for "Status" header)
status_col = None
feature_col = None
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col).value
    if header and "Status" in str(header):
        status_col = col
    if header and "Feature" in str(header):
        feature_col = col

if not status_col or not feature_col:
    print(f"Scanning headers in row 1:")
    for col in range(1, ws.max_column + 1):
        print(f"  Col {col}: {ws.cell(row=1, column=col).value}")
    print("Could not find Status or Feature columns by header. Trying default positions...")
    feature_col = 4  # Column D is usually Feature
    status_col = 6   # Column F is usually Status

print(f"Feature column: {feature_col}, Status column: {status_col}")

updated = 0
in_security_section = False

for row in range(2, ws.max_row + 1):
    # Check if we're in the Security Department section
    # Look at Module or Sub-Module column for "Security"
    for col in range(1, min(5, ws.max_column + 1)):
        cell_val = ws.cell(row=row, column=col).value
        if cell_val and isinstance(cell_val, str):
            if "Security" in cell_val and ("Department" in cell_val or "Module" in cell_val or col <= 2):
                in_security_section = True
            elif cell_val.strip() and col <= 2 and "Security" not in cell_val:
                # New module section, stop
                if in_security_section:
                    in_security_section = False

    feature = ws.cell(row=row, column=feature_col).value
    if not feature or not isinstance(feature, str):
        continue

    for keyword, new_status in FEATURE_STATUS.items():
        if keyword.lower() in feature.lower():
            current = ws.cell(row=row, column=status_col).value
            if current != new_status:
                ws.cell(row=row, column=status_col).value = new_status
                print(f"  Row {row}: '{feature[:60]}...' -> {new_status} (was: {current})")
                updated += 1
            break

print(f"\nUpdated {updated} feature statuses")
wb.save("/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx")
print("Saved MedBrains_Features.xlsx")

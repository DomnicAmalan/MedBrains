#!/usr/bin/env python3
"""Add CMS & Blog features sheet to MedBrains_Features.xlsx."""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "MedBrains_Features.xlsx")

# Styling constants (matching existing sheets)
MODULE_FONT = Font(bold=True, size=11, color="1F4E79")
SUBMODULE_FONT = Font(bold=True, size=10, color="2E75B6")
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
MODULE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBMODULE_FILL = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

COLUMNS = [
    ("S.No", 6),
    ("Module", 20),
    ("Sub-Module", 25),
    ("Feature", 60),
    ("Source", 12),
    ("Priority", 10),
    ("Status", 12),
    ("RFC Ref", 20),
    ("Web", 8),
    ("Mobile", 8),
    ("TV", 8),
]

MODULE = "CMS & Blog"
SOURCE = "RFC-CMS"
PRIORITY = "P2"
STATUS = "Pending"
RFC_REF = "RFC-CMS"

# Features organised by sub-module
FEATURES = [
    {"type": "module", "text": MODULE},
    {"type": "submodule", "text": "Post Management"},
    "Create/edit blog posts with markdown editor",
    "Post status workflow (draft/review/approved/published/archived)",
    "Content types (article, quick post, opinion, interview, case study)",
    "Featured posts management",
    "Feature image upload with alt text and caption",
    "Auto-generate reading time",
    "Post revision history and version tracking",
    "Post slug auto-generation and editing",
    "Post excerpt/summary field",
    "Medical review workflow for clinical content",
    "Submit post for review",
    "Approve/reject post review",
    "Publish post with workflow enforcement",
    "Schedule post for future publication",
    "Archive/unarchive posts",
    {"type": "submodule", "text": "Categories"},
    "Category CRUD with hierarchical parent-child",
    "Category slug, description, color, icon",
    "Medical review requirement flag per category",
    "Category sort order management",
    {"type": "submodule", "text": "Tags"},
    "Tag CRUD with slug",
    "Tag description field",
    "Bulk tag management",
    {"type": "submodule", "text": "Media Library"},
    "Media upload (images, documents)",
    "Media gallery with grid/list view",
    "Image preview and metadata display",
    "Alt text management for accessibility",
    "Media delete with orphan check",
    {"type": "submodule", "text": "Authors"},
    "Author profile management (bio, credentials, avatar)",
    "Author social links (website, Twitter, LinkedIn)",
    "Author slug for public URLs",
    "Map CMS authors to HMS user accounts",
    "Author role assignment (admin, editor, author, reviewer)",
    {"type": "submodule", "text": "Subscribers"},
    "Newsletter subscriber list",
    "Email subscription with confirmation flow",
    "Unsubscribe handling",
    "Export subscribers to CSV",
    "Subscriber management (delete, filter)",
    {"type": "submodule", "text": "Analytics"},
    "Dashboard with total views, posts, subscribers",
    "Per-post view analytics",
    "Top posts ranking",
    "Views over time chart",
    "Referrer tracking",
    "Geographic (country) breakdown",
    {"type": "submodule", "text": "SEO"},
    "Meta title and description per post",
    "Open Graph image per post",
    "Canonical URL support",
    "Noindex flag per post",
    "Sitemap generation",
    "RSS feed generation",
    {"type": "submodule", "text": "Content API"},
    "Public API with API key authentication",
    "Paginated post listing with filters",
    "Single post by slug endpoint",
    "Posts by category endpoint",
    "Posts by tag endpoint",
    "Author listing and detail endpoints",
    "Full-text search via PostgreSQL tsvector",
    "Newsletter subscribe endpoint (public)",
    "Page view tracking endpoint",
    {"type": "submodule", "text": "API Keys"},
    "API key generation with SHA-256 hashing",
    "API key list and revoke",
    "Key prefix display (masked)",
    "Key expiration date",
    "Per-key permission scoping",
    {"type": "submodule", "text": "CMS Client Package"},
    "TypeScript API client for Astro site",
    "Type-safe content types (ContentPost, ContentAuthor, etc.)",
    "Ghost-compatible type aliases for migration",
    "Environment-based API URL configuration",
    {"type": "submodule", "text": "Astro Site Integration"},
    "Update imports from Ghost client to CMS client",
    "Update type references across Astro files",
    "Remove static content files (posts/authors)",
    "Configure Astro to fetch from MedBrains API",
]


def main():
    wb = load_workbook(EXCEL_PATH)

    # Remove existing sheet if re-running
    if "CMS & Blog" in wb.sheetnames:
        del wb["CMS & Blog"]

    ws = wb.create_sheet("CMS & Blog")

    # Column widths
    for i, (_, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Header row
    for i, (name, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    sno = 0
    feature_count = 0

    for entry in FEATURES:
        if isinstance(entry, dict):
            if entry["type"] == "module":
                # Module header row
                ws.cell(row=row, column=2, value=entry["text"]).font = MODULE_FONT
                for col in range(1, 12):
                    ws.cell(row=row, column=col).fill = MODULE_FILL
                    ws.cell(row=row, column=col).border = THIN_BORDER
                row += 1
            elif entry["type"] == "submodule":
                # Submodule header row
                ws.cell(row=row, column=3, value=entry["text"]).font = SUBMODULE_FONT
                for col in range(1, 12):
                    ws.cell(row=row, column=col).fill = SUBMODULE_FILL
                    ws.cell(row=row, column=col).border = THIN_BORDER
                row += 1
        else:
            # Feature row
            sno += 1
            feature_count += 1
            values = [sno, MODULE, "", entry, SOURCE, PRIORITY, STATUS, RFC_REF, "Y", "N", "N"]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
                cell.alignment = WRAP_ALIGN
            row += 1

    wb.save(EXCEL_PATH)
    print(f"Done! Added 'CMS & Blog' sheet with {feature_count} features to {EXCEL_PATH}")


if __name__ == "__main__":
    main()

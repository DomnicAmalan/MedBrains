#!/usr/bin/env python3
"""Add Educational Content & Gamification features to MedBrains_Features.xlsx.

Adds 37 new P2 features across two sheets:
- Specialty & Academic → Micro Website (18 features: Content Studio, Video Library, Social Media, Patient Education Portal)
- Mobile Apps → Patient App (19 features: Wellness Tracking, Gamification, Social Challenges)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

# Styling constants (matching project convention)
MODULE_FONT = Font(bold=True, size=11, color="1F4E79")
SUBMODULE_FONT = Font(bold=True, size=10, color="2E75B6")
MODULE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBMODULE_FILL = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")

RFC_REF = "RFC-MODULE-patient-engagement"

# --- Sheet 1: Specialty & Academic → Micro Website (18 features) ---
SPECIALTY_FEATURES = [
    {"type": "module", "text": "Micro Website"},

    {"type": "submodule", "text": "Content Studio"},
    {"feature": "Doctor/staff content authorship — rich text editor for articles & blogs with author profile", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Video upload & management — MP4/WebM support with titles, tags, thumbnails, transcripts", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Medical content review workflow — draft → peer review → medical director approval → publish", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Content categorization & tagging — specialty, condition, target audience, content type taxonomy", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Multi-language content support — create content in multiple languages with linked translations", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Medical disclaimer management — auto-attach regulatory disclaimers to all educational content", "web": "Y", "mobile": "Y", "tv": "N"},

    {"type": "submodule", "text": "Video Library"},
    {"feature": "Video hosting with adaptive streaming — HLS/DASH encoding for bandwidth optimization", "web": "Y", "mobile": "Y", "tv": "Y"},
    {"feature": "Video chapters & timestamps — clickable chapter navigation within videos", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Auto-generated captions & subtitles — speech-to-text transcription for accessibility", "web": "Y", "mobile": "Y", "tv": "Y"},
    {"feature": "Video analytics dashboard — views, watch time, completion rate, drop-off points per video", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "Social Media"},
    {"feature": "Social media post composer — create posts for YouTube, Instagram, Facebook, LinkedIn from CMS", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Social media publishing integration — auto-post to platforms via OAuth 2.0 APIs", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Social media analytics dashboard — likes, shares, comments, reach aggregated per platform", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "Patient Education Portal"},
    {"feature": "Public health education library — searchable & filterable article + video library, SEO-optimized", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Personalized content recommendations — suggest content based on diagnoses, visits, demographics", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Condition-specific education bundles — curated content packs for common conditions (e.g., Diabetes 101)", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Patient education assignment — doctor assigns specific content to patient during consultation", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Educational content push notifications — notify patients of new relevant content", "web": "N", "mobile": "Y", "tv": "N"},
]

# --- Sheet 2: Mobile Apps → Patient App (19 features) ---
MOBILE_FEATURES = [
    {"type": "module", "text": "Patient App"},

    {"type": "submodule", "text": "Wellness Tracking"},
    {"feature": "Wellness goals setup — configure targets for weight, steps, BP, blood sugar, medication, appointments", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Manual health metric logging — daily entry for weight, BP, sugar, steps, sleep, water intake", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Wearable device integration — sync data from Apple Health, Google Fit, Fitbit, Garmin APIs", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Medication adherence tracking — mark doses as taken, configurable reminders, missed dose tracking", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Appointment compliance tracking — kept vs missed appointments, follow-up adherence percentage", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Progress dashboard — charts for goal progress, weekly/monthly trends, averages with color coding", "web": "N", "mobile": "Y", "tv": "N"},

    {"type": "submodule", "text": "Gamification"},
    {"feature": "Points system — earn points for medication adherence, appointment attendance, monthly health milestones", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Achievement badges — unlock badges for milestones (7-day streak, 30-day streak, all goals met, steps target)", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Leaderboard (friends & family) — opt-in friendly competition on wellness points with privacy controls", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Weekly/monthly challenges — hospital-created wellness challenges with bonus point rewards", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Reward redemption — redeem accumulated points for discounts, free checkups, pharmacy vouchers", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Streak tracking — track consecutive days of goal completion with visual streak badges & recovery grace", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Level system — Bronze / Silver / Gold / Platinum tiers based on total lifetime points", "web": "N", "mobile": "Y", "tv": "N"},

    {"type": "submodule", "text": "Social Challenges"},
    {"feature": "Challenge invitations — invite friends & family members to join wellness challenges via link/SMS", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Challenge group chat — private messaging within challenge groups for encouragement & tips", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Public challenge feed — opt-in display of anonymized achievements & milestones", "web": "N", "mobile": "Y", "tv": "Y"},
    {"feature": "Doctor/hospital-led challenges — staff-created public wellness challenges with enrollment tracking", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Wellness analytics dashboard (admin) — engagement metrics, participation rates, drop-off tracking", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Privacy controls — granular sharing settings, default-private, explicit opt-in for social features", "web": "Y", "mobile": "Y", "tv": "N"},
]


def add_features_to_sheet(ws, features, rfc_ref):
    """Append features to a worksheet following the project styling convention."""
    last_row = ws.max_row

    # Find current max S.No
    max_sno = 0
    for row in range(2, last_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, (int, float)):
            max_sno = max(max_sno, int(val))

    current_sno = max_sno
    current_row = last_row + 2  # Leave a blank row separator

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for item in features:
        if item.get("type") == "module":
            # Module header row — col B
            ws.cell(row=current_row, column=2, value=item["text"])
            ws.cell(row=current_row, column=2).font = MODULE_FONT
            for col in range(1, 12):
                ws.cell(row=current_row, column=col).fill = MODULE_FILL
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1

        elif item.get("type") == "submodule":
            # Sub-module header row — col C
            ws.cell(row=current_row, column=3, value=item["text"])
            ws.cell(row=current_row, column=3).font = SUBMODULE_FONT
            for col in range(1, 12):
                ws.cell(row=current_row, column=col).fill = SUBMODULE_FILL
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1

        else:
            # Feature row
            current_sno += 1
            ws.cell(row=current_row, column=1, value=current_sno)      # A: S.No
            # col B (Module) and col C (Sub-Module) left blank for features
            ws.cell(row=current_row, column=4, value=item["feature"])   # D: Feature
            ws.cell(row=current_row, column=5, value="MedBrains")      # E: Source
            ws.cell(row=current_row, column=6, value="P2")             # F: Priority
            ws.cell(row=current_row, column=7, value="Pending")        # G: Status
            ws.cell(row=current_row, column=8, value=rfc_ref)          # H: RFC Ref
            ws.cell(row=current_row, column=9, value=item.get("web", "N"))     # I: Web
            ws.cell(row=current_row, column=10, value=item.get("mobile", "N")) # J: Mobile
            ws.cell(row=current_row, column=11, value=item.get("tv", "N"))     # K: TV

            for col in range(1, 12):
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

            current_row += 1

    return current_sno - max_sno


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # --- Sheet 1: Specialty & Academic ---
    ws1 = wb["Specialty & Academic"]
    count1 = add_features_to_sheet(ws1, SPECIALTY_FEATURES, RFC_REF)
    print(f"Specialty & Academic: added {count1} features")

    # --- Sheet 2: Mobile Apps ---
    ws2 = wb["Mobile Apps"]
    count2 = add_features_to_sheet(ws2, MOBILE_FEATURES, RFC_REF)
    print(f"Mobile Apps: added {count2} features")

    wb.save(EXCEL_PATH)
    total = count1 + count2
    print(f"\nDone! Total {total} features added across 2 sheets.")


if __name__ == "__main__":
    main()

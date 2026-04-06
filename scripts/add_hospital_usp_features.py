#!/usr/bin/env python3
"""Add Hospital Marketing USP features to MedBrains_Features.xlsx.

Adds 80 new P1 features across 5 sheets in 10 marketing USP groups:
- Multi-Hospital & Vendors → Medical Tourism & International Patients (12)
- Clinical → Health Packages & Preventive Care (10) + Pre-Visit Digital Journey (4)
- Admin & Operations → Doctor Profile & Reputation (8) + Cost Transparency (8)
                        + Referral & Loyalty (8) + Concierge & Premium (8)
                        + Patient Outcomes & Quality Showcase (6)
- Specialty & Academic → Specialty Care Programs (10)
- Technical Infrastructure → AI Chatbot & Virtual Assistant (6)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

# Styling constants (matching project convention)
MODULE_FONT = Font(bold=True, size=11, color="1F4E79")
SUBMODULE_FONT = Font(bold=True, size=10, color="2E75B6")
MODULE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBMODULE_FILL = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")

RFC_REF = "RFC-MODULE-hospital-usp"

# =============================================================================
# Sheet 1: Multi-Hospital & Vendors → Medical Tourism & International Patients
# =============================================================================
MULTI_HOSPITAL_FEATURES = [
    {"type": "module", "text": "Medical Tourism & International Patients"},

    {"type": "submodule", "text": "International Patient Services"},
    {"feature": "Multi-language patient concierge portal — English, Arabic, Russian, Chinese, French, German interface", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Visa invitation letter generation — hospital letterhead with treatment plan, cost estimate, doctor details", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Airport pickup/drop coordination — vehicle assignment, driver details, estimated arrival notification", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "International patient coordinator assignment — dedicated communication channel per patient", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Travel & accommodation booking assistance — partner hotel listing with proximity, pricing, availability", "web": "Y", "mobile": "Y", "tv": "N"},

    {"type": "submodule", "text": "International Billing & Insurance"},
    {"feature": "Multi-currency billing — USD, AED, EUR, GBP, RUB with live exchange rates and currency conversion", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "International insurance processing — BUPA, Allianz, Aetna, Cigna claim formats and direct settlement", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "Medical Tourism Operations"},
    {"feature": "Interpreter/translator service booking — per language per appointment with availability calendar", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Medical visa documentation package — treatment summary, cost breakdown, doctor credentials bundle", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "International patient outcomes portfolio — success stories filterable by nationality and procedure", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Post-return telehealth follow-up — cross-timezone scheduling for international patient aftercare", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Medical tourism package builder — treatment + hospital stay + hotel + travel as bundled pricing", "web": "Y", "mobile": "N", "tv": "N"},
]

# =============================================================================
# Sheet 2: Clinical → Health Packages & Preventive Care + Pre-Visit Digital Journey
# =============================================================================
CLINICAL_FEATURES = [
    # --- Health Packages & Preventive Care (10 features) ---
    {"type": "module", "text": "Health Packages & Preventive Care"},

    {"type": "submodule", "text": "Package Management"},
    {"feature": "Health checkup package builder — configurable test bundles with pricing (Executive, Master, Cardiac, Diabetic, etc.)", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Package catalog display — public-facing listing with test details, price, duration, preparation instructions", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Online package purchase — patient self-books and pays for health checkup packages via web/app", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Corporate health screening packages — bulk employee packages with group pricing and scheduling", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Package appointment scheduling — auto-assign lab, radiology, and doctor slots across single or multi-day visit", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "Preventive Care & Wellness"},
    {"feature": "Package report consolidation — single summary report with all results, doctor commentary, and recommendations", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Preventive care reminders — annual checkup nudges based on age, gender, family history, risk profile", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Wellness subscription plans — monthly/quarterly monitoring subscriptions (diabetes care, cardiac care, etc.)", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Package comparison tool — side-by-side comparison of different health packages with test overlap highlighting", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Package conversion analytics — views to bookings to completions to follow-up revenue tracking for marketing ROI", "web": "Y", "mobile": "N", "tv": "N"},

    # --- Pre-Visit Digital Journey (4 features) ---
    {"type": "module", "text": "Pre-Visit Digital Journey"},

    {"type": "submodule", "text": "Digital Check-In & Journey"},
    {"feature": "QR code check-in — scan at hospital entrance, auto-check-in, queue assigned, directions sent to phone", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Online pre-registration — complete demographics, insurance, consent forms digitally before hospital arrival", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Real-time wait time display — estimated wait by department shown on website and app before patient leaves home", "web": "Y", "mobile": "Y", "tv": "Y"},
    {"feature": "Appointment day journey tracker — patient sees step-by-step progress (check-in, consult, lab, pharmacy, billing)", "web": "Y", "mobile": "Y", "tv": "N"},
]

# =============================================================================
# Sheet 3: Admin & Operations → Doctor Profile + Cost Transparency + Referral
#           + Concierge + Patient Outcomes
# =============================================================================
ADMIN_OPS_FEATURES = [
    # --- Doctor Profile & Reputation Showcase (8 features) ---
    {"type": "module", "text": "Doctor Profile & Reputation Showcase"},

    {"type": "submodule", "text": "Doctor Public Profiles"},
    {"feature": "Doctor public profile page — photo, qualifications, specialties, experience, languages, hospital affiliations", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Doctor availability calendar — real-time slot display integrated with appointment booking widget", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Doctor video introduction — 60-second intro video on profile for patient trust-building", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Patient ratings & reviews per doctor — verified visit-based reviews with star ratings and written feedback", "web": "Y", "mobile": "Y", "tv": "N"},

    {"type": "submodule", "text": "Doctor Analytics & Discovery"},
    {"feature": "Doctor awards & publications — certifications, conference presentations, research papers, media features", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Doctor success metrics — procedure count, years of experience, patient satisfaction score (anonymized)", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Doctor search & discovery — filter/sort by specialty, language, gender, availability, ratings, location", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Doctor referral network map — visualize cross-specialty referral relationships and top referral sources", "web": "Y", "mobile": "N", "tv": "N"},

    # --- Cost Transparency & Estimation (8 features) ---
    {"type": "module", "text": "Cost Transparency & Estimation"},

    {"type": "submodule", "text": "Patient-Facing Cost Tools"},
    {"feature": "Treatment cost estimator — patient selects procedure + insurance + room type for estimated cost range", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Public tariff card — standard price list for common procedures, surgeries, packages (NABH requirement)", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Insurance coverage calculator — patient selects insurer + plan for covered amount vs out-of-pocket estimate", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "EMI/installment payment options — split payment plans for high-cost treatments via banking partners", "web": "Y", "mobile": "Y", "tv": "N"},

    {"type": "submodule", "text": "Cost Analytics & Reconciliation"},
    {"feature": "Room category cost comparison — General / Semi-Private / Private / Deluxe / Suite pricing with amenity details", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Advance deposit calculator — estimated deposit based on procedure, expected stay, and insurance coverage", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Bill vs estimate reconciliation — show patients how actual charges compared to initial estimate with variance explanation", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Cost transparency dashboard — average costs by procedure for management benchmarking and public reporting", "web": "Y", "mobile": "N", "tv": "N"},

    # --- Patient Referral & Loyalty Programs (8 features) ---
    {"type": "module", "text": "Patient Referral & Loyalty Programs"},

    {"type": "submodule", "text": "Referral Programs"},
    {"feature": "Patient referral program — existing patients refer friends/family with trackable referral codes and links", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Referral reward system — automatic discounts, cashback, or loyalty points for successful referral conversions", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Doctor-to-doctor referral tracking — which external physicians refer patients, referral volume and revenue attribution", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "Loyalty & VIP"},
    {"feature": "Loyalty tier system — Bronze/Silver/Gold/Platinum tiers based on visit frequency and cumulative spend", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Tier-based benefits — priority appointments, room upgrades, complimentary services escalating by tier", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Loyalty points accumulation & redemption — earn on bills, redeem for services, pharmacy, or partner offers", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "VIP patient identification — flagged across all touchpoints (reception, nursing, billing) for premium treatment", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Referral & loyalty analytics — referral conversion rate, loyalty program ROI, tier distribution, churn by tier", "web": "Y", "mobile": "N", "tv": "N"},

    # --- Concierge & Premium Services (8 features) ---
    {"type": "module", "text": "Concierge & Premium Services"},

    {"type": "submodule", "text": "Concierge Operations"},
    {"feature": "Patient concierge assignment — dedicated coordinator for premium/international patients with task tracking", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Premium room booking — suite/deluxe room selection with photo gallery, amenity list, real-time availability", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Priority appointment scheduling — VIP fast-track queue bypass for consultations, diagnostics, procedures", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Executive lounge & waiting area — premium waiting zone access with complimentary refreshments and Wi-Fi", "web": "Y", "mobile": "Y", "tv": "N"},

    {"type": "submodule", "text": "Premium Service Management"},
    {"feature": "Personalized care plan communication — daily schedule, upcoming tests, diet delivered via app/SMS/WhatsApp", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Concierge task dashboard — coordinator tracks all patient requests, service delivery SLAs, escalations", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Guest & attendant services — extra bed booking, meal ordering, parking reservation for patient companions", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Premium service billing — separate line items for concierge add-ons with transparent pricing", "web": "Y", "mobile": "N", "tv": "N"},

    # --- Patient Outcomes & Quality Showcase (6 features) ---
    {"type": "module", "text": "Patient Outcomes & Quality Showcase"},

    {"type": "submodule", "text": "Outcomes Display"},
    {"feature": "Procedure success rate dashboard — publicly shareable outcomes by department and procedure type", "web": "Y", "mobile": "N", "tv": "Y"},
    {"feature": "Patient satisfaction scorecards — NPS and department-wise satisfaction scores with trend graphs", "web": "Y", "mobile": "N", "tv": "Y"},
    {"feature": "Readmission rate display — low readmission rates as quality indicator with national benchmark comparison", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "Quality Badges & Awards"},
    {"feature": "Average length of stay benchmarks — ALOS by procedure vs national/regional standards", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Accreditation & quality badge showcase — NABH/JCI/NABL/ISO badges with verification links and validity dates", "web": "Y", "mobile": "Y", "tv": "Y"},
    {"feature": "Clinical excellence awards display — department-level recognitions, certifications, and media features", "web": "Y", "mobile": "Y", "tv": "N"},
]

# =============================================================================
# Sheet 4: Specialty & Academic → Specialty Care Programs
# =============================================================================
SPECIALTY_FEATURES = [
    {"type": "module", "text": "Specialty Care Programs"},

    {"type": "submodule", "text": "Bundled Care Packages"},
    {"feature": "Maternity care package — antenatal to postnatal bundled program with pricing, doctor team, and amenities", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Pediatric wellness program — vaccination schedules, growth tracking, milestone alerts, well-child visit bundles", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Elderly care program — chronic disease management, fall prevention, home visit packages, caregiver support", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Cardiac rehabilitation program — post-surgery recovery plan with monitored exercise, diet, and follow-up", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Diabetes management program — continuous monitoring, diet counseling, medication optimization, HbA1c tracking", "web": "Y", "mobile": "Y", "tv": "N"},

    {"type": "submodule", "text": "Specialized Treatment Programs"},
    {"feature": "Orthopedic & sports medicine program — injury assessment, rehab protocol, return-to-activity planning", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Mental wellness program — counseling packages, stress management workshops, therapy session bundles", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Bariatric & weight management program — pre-assessment to post-surgery lifestyle plan with nutritionist", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Fertility & IVF program — consultation to treatment bundled package with success rate display and counseling", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Oncology support program — treatment planning, second opinion, palliative care, survivorship follow-up", "web": "Y", "mobile": "Y", "tv": "N"},
]

# =============================================================================
# Sheet 5: Technical Infrastructure → AI Chatbot & Virtual Assistant
# =============================================================================
TECH_INFRA_FEATURES = [
    {"type": "module", "text": "AI Chatbot & Virtual Assistant"},

    {"type": "submodule", "text": "Chatbot Channels"},
    {"feature": "Website AI chatbot — FAQ answers, doctor search, appointment booking, department directions, operating hours", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "WhatsApp AI assistant — conversational appointment booking, report status queries, bill inquiries in natural language", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Symptom checker — guided questionnaire suggesting appropriate department and specialist based on symptoms", "web": "Y", "mobile": "Y", "tv": "N"},

    {"type": "submodule", "text": "Chatbot Intelligence"},
    {"feature": "Smart appointment routing — chatbot triages urgency level and auto-routes to correct department/doctor", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Chatbot-to-human handoff — seamless escalation from bot to live agent with conversation context preserved", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Chatbot analytics dashboard — conversation volume, resolution rate, top query categories, handoff rate, CSAT", "web": "Y", "mobile": "N", "tv": "N"},
]


def add_features_to_sheet(ws, features, rfc_ref):
    """Append features to a worksheet following the project styling convention.

    Returns total feature count added.
    """
    last_row = ws.max_row

    # Find current max S.No
    max_sno = 0
    for row in range(2, last_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, (int, float)):
            max_sno = max(max_sno, int(val))

    current_sno = max_sno
    current_row = last_row + 2  # Leave a blank row separator

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    feature_count = 0

    for item in features:
        if item.get("type") == "module":
            # Module header row — col B
            ws.cell(row=current_row, column=2, value=item["text"])
            ws.cell(row=current_row, column=2).font = MODULE_FONT
            for col in range(1, 12):
                ws.cell(row=current_row, column=col).fill = MODULE_FILL
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1

        elif item.get("type") == "submodule":
            # Sub-module header row — col C
            ws.cell(row=current_row, column=3, value=item["text"])
            ws.cell(row=current_row, column=3).font = SUBMODULE_FONT
            for col in range(1, 12):
                ws.cell(row=current_row, column=col).fill = SUBMODULE_FILL
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1

        else:
            # Feature row
            current_sno += 1
            feature_count += 1

            ws.cell(row=current_row, column=1, value=current_sno)       # A: S.No
            # col B (Module) and col C (Sub-Module) left blank for features
            ws.cell(row=current_row, column=4, value=item["feature"])    # D: Feature
            ws.cell(row=current_row, column=5, value="MedBrains")       # E: Source
            ws.cell(row=current_row, column=6, value="P1")              # F: Priority
            ws.cell(row=current_row, column=7, value="Pending")         # G: Status
            ws.cell(row=current_row, column=8, value=rfc_ref)           # H: RFC Ref
            ws.cell(row=current_row, column=9, value=item.get("web", "N"))     # I: Web
            ws.cell(row=current_row, column=10, value=item.get("mobile", "N")) # J: Mobile
            ws.cell(row=current_row, column=11, value=item.get("tv", "N"))     # K: TV

            for col in range(1, 12):
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

            current_row += 1

    return feature_count


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # --- Sheet 1: Multi-Hospital & Vendors ---
    ws1 = wb["Multi-Hospital & Vendors"]
    count1 = add_features_to_sheet(ws1, MULTI_HOSPITAL_FEATURES, RFC_REF)
    print(f"Multi-Hospital & Vendors: added {count1} features (Medical Tourism)")

    # --- Sheet 2: Clinical ---
    ws2 = wb["Clinical"]
    count2 = add_features_to_sheet(ws2, CLINICAL_FEATURES, RFC_REF)
    print(f"Clinical: added {count2} features (Health Packages + Pre-Visit Journey)")

    # --- Sheet 3: Admin & Operations ---
    ws3 = wb["Admin & Operations"]
    count3 = add_features_to_sheet(ws3, ADMIN_OPS_FEATURES, RFC_REF)
    print(f"Admin & Operations: added {count3} features (Doctor Profile + Cost + Referral + Concierge + Outcomes)")

    # --- Sheet 4: Specialty & Academic ---
    ws4 = wb["Specialty & Academic"]
    count4 = add_features_to_sheet(ws4, SPECIALTY_FEATURES, RFC_REF)
    print(f"Specialty & Academic: added {count4} features (Specialty Care Programs)")

    # --- Sheet 5: Technical Infrastructure ---
    ws5 = wb["Technical Infrastructure"]
    count5 = add_features_to_sheet(ws5, TECH_INFRA_FEATURES, RFC_REF)
    print(f"Technical Infrastructure: added {count5} features (AI Chatbot & Virtual Assistant)")

    wb.save(EXCEL_PATH)

    total = count1 + count2 + count3 + count4 + count5
    print(f"\nDone! Total {total} features added across 5 sheets (all Pending).")


if __name__ == "__main__":
    main()

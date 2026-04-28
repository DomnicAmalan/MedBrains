"""Append CRDT / Terraform / EKS infra features to MedBrains_Features.xlsx.

Source: RFC-INFRA-2026-001 (CRDT Sync, Terraform, EKS).
Target sheet: 'Technical Infrastructure' — appends a new chapter with module + sub-module headers.
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

WB_PATH = "MedBrains_Features.xlsx"
SHEET = "Technical Infrastructure"

MODULE_FILL = PatternFill("solid", fgColor="D6E4F0")
SUBMOD_FILL = PatternFill("solid", fgColor="E9EFF7")
BOLD = Font(bold=True)
THIN = Side(border_style="thin", color="B0B7BD")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

CHAPTERS: list[dict] = [
    {
        "title": "11. OFFLINE-FIRST CRDT SYNC",
        "module": "Offline Sync",
        "groups": [
            ("11.1 CRDT Engine (Loro)", [
                ("Loro CRDT engine — Rust core + WASM bindings", "RFC-INFRA-001 §A.2", "P1", "Pending", "Y", "Y", "N"),
                ("Tiered data model (T1 server-auth / T2 CRDT / T3 commit-gate)", "RFC-INFRA-001 §A.1", "P1", "Pending", "Y", "Y", "N"),
                ("medbrains-crdt-core crate (Loro wrapper, container schemas)", "RFC-INFRA-001 §A.2", "P1", "Pending", "Y", "Y", "N"),
                ("medbrains-crdt-codec crate (protobuf framing)", "RFC-INFRA-001 §A.2", "P1", "Pending", "Y", "Y", "N"),
                ("medbrains-crdt-policy crate (per-entity merge policies)", "RFC-INFRA-001 §A.2", "P1", "Pending", "Y", "Y", "N"),
                ("Schema versioning + bidirectional migrators (M_n→n+1)", "RFC-INFRA-001 §A.4", "P1", "Pending", "Y", "Y", "N"),
                ("Dual-write window enforcement (30 days)", "RFC-INFRA-001 §A.4", "P2", "Pending", "Y", "Y", "N"),
            ]),
            ("11.2 Sync Protocol", [
                ("WebSocket primary transport (wss://sync.<region>.medbrains.health)", "RFC-INFRA-001 §A.4", "P1", "Pending", "Y", "Y", "N"),
                ("HTTP long-poll fallback for restrictive networks", "RFC-INFRA-001 §A.4", "P2", "Pending", "Y", "Y", "N"),
                ("Protobuf SyncEnvelope wire format", "RFC-INFRA-001 §A.4", "P1", "Pending", "Y", "Y", "N"),
                ("Version vector + delta sync with chunked CatchUp", "RFC-INFRA-001 §A.4", "P1", "Pending", "Y", "Y", "N"),
                ("Backpressure via credit window (64 frames)", "RFC-INFRA-001 §A.4", "P2", "Pending", "Y", "Y", "N"),
                ("Per-device mTLS cert (per-tenant CA)", "RFC-INFRA-001 §A.8", "P1", "Pending", "Y", "Y", "N"),
                ("Replay protection via VV monotonicity", "RFC-INFRA-001 §A.8", "P1", "Pending", "Y", "Y", "N"),
            ]),
            ("11.3 Conflict Resolution Policies", [
                ("Vital signs — append-only multi-writer LoroList", "RFC-INFRA-001 §A.5", "P1", "Pending", "Y", "Y", "N"),
                ("Nursing notes — append-only with soft-retract", "RFC-INFRA-001 §A.5", "P1", "Pending", "Y", "Y", "N"),
                ("Care plan tasks — op-based MovableTree", "RFC-INFRA-001 §A.5", "P2", "Pending", "Y", "Y", "N"),
                ("Provisional Rx (T3) — CRDT until sign, then freeze", "RFC-INFRA-001 §A.5", "P2", "Pending", "Y", "Y", "N"),
                ("MAR (Medication Admin Record) — T1 outbox replay", "RFC-INFRA-001 §A.5", "P1", "Pending", "Y", "Y", "N"),
                ("Chat — LWW per message id", "RFC-INFRA-001 §A.5", "P3", "Pending", "Y", "Y", "N"),
                ("Vitals chart annotations — sub-map per vital op", "RFC-INFRA-001 §A.5", "P3", "Pending", "Y", "Y", "N"),
                ("Compile-time exhaustive policy enum (no LWW default)", "RFC-INFRA-001 §A.5", "P1", "Pending", "Y", "N", "N"),
            ]),
            ("11.4 Edge Node (medbrains-edge)", [
                ("medbrains-edge Rust binary (axum + sqlite + loro)", "RFC-INFRA-001 §A.6", "P1", "Pending", "N", "N", "N"),
                ("LAN sync hub with mDNS discovery", "RFC-INFRA-001 §A.6", "P1", "Pending", "N", "N", "N"),
                ("Outbox WAL queue for T1 writes during WAN outage", "RFC-INFRA-001 §A.6", "P1", "Pending", "N", "N", "N"),
                ("Read-cache for pinned Aurora subset (drug catalog, demographics)", "RFC-INFRA-001 §A.6", "P1", "Pending", "N", "N", "N"),
                ("Idempotency-key replay on reconnect", "RFC-INFRA-001 §A.6", "P1", "Pending", "N", "N", "N"),
                ("Device pairing + per-device cert issue", "RFC-INFRA-001 §A.6", "P2", "Pending", "N", "N", "N"),
                ("Fleet management via FluxCD on single-node k3s", "RFC-INFRA-001 §A.6", "P2", "Pending", "N", "N", "N"),
                ("UPS-backed hardware spec (NUC, 8GB RAM, 256GB SSD, dual-NIC)", "RFC-INFRA-001 §A.6", "P2", "Pending", "N", "N", "N"),
                ("Daily encrypted snapshot to regional S3 (Restic)", "RFC-INFRA-001 §C.8", "P2", "Pending", "N", "N", "N"),
            ]),
            ("11.5 Audit Trail Integrity", [
                ("Per-tenant Merkle hash-chained audit log", "RFC-INFRA-001 §A.7", "P1", "Pending", "Y", "Y", "N"),
                ("Ed25519-signed chain heads", "RFC-INFRA-001 §A.7", "P1", "Pending", "Y", "Y", "N"),
                ("Server-side anchor validation + RDS append", "RFC-INFRA-001 §A.7", "P1", "Pending", "Y", "N", "N"),
                ("Fork detection + device quarantine + SIEM alert", "RFC-INFRA-001 §A.7", "P2", "Pending", "Y", "N", "N"),
            ]),
            ("11.6 Local CRDT Storage", [
                ("Web — IndexedDB via idb (Loro snapshots + op log)", "RFC-INFRA-001 §A.3", "P1", "Pending", "Y", "N", "N"),
                ("Mobile — SQLite Loro store coexisting with WatermelonDB", "RFC-INFRA-001 §A.3", "P1", "Pending", "N", "Y", "N"),
                ("Edge — SQLite per-tenant op log with 30-day hot ring buffer", "RFC-INFRA-001 §A.3", "P1", "Pending", "N", "N", "N"),
                ("S3 cold-archive of edge op logs", "RFC-INFRA-001 §A.3", "P3", "Pending", "N", "N", "N"),
                ("TanStack Query adapter materialising Loro views", "RFC-INFRA-001 §A.3", "P1", "Pending", "Y", "N", "N"),
            ]),
        ],
    },
    {
        "title": "12. INFRASTRUCTURE AS CODE (TERRAFORM)",
        "module": "IaC",
        "groups": [
            ("12.1 Terragrunt Layout", [
                ("Terragrunt wrapper over vanilla Terraform", "RFC-INFRA-001 §B.1", "P1", "Pending", "N", "N", "N"),
                ("Reusable TF modules (vpc/eks/aurora/kms/iam-irsa/s3/...)", "RFC-INFRA-001 §B.2", "P1", "Pending", "N", "N", "N"),
                ("live/global/ stack (Route53, IAM org, ECR)", "RFC-INFRA-001 §B.2", "P1", "Pending", "N", "N", "N"),
                ("live/regions/ap-south-1/ (primary, India)", "RFC-INFRA-001 §B.2", "P1", "Pending", "N", "N", "N"),
                ("live/regions/ap-southeast-1/ (Singapore, V2)", "RFC-INFRA-001 §B.2", "P3", "Pending", "N", "N", "N"),
                ("live/regions/me-south-1/ (Bahrain, V3)", "RFC-INFRA-001 §B.2", "P3", "Pending", "N", "N", "N"),
            ]),
            ("12.2 Single-Command Deploy", [
                ("make deploy ENV=prod REGION=ap-south-1", "RFC-INFRA-001 §B.3", "P1", "Pending", "N", "N", "N"),
                ("make deploy-all ENV=prod (parallel multi-region)", "RFC-INFRA-001 §B.3", "P2", "Pending", "N", "N", "N"),
                ("make plan / make destroy with ENV gating", "RFC-INFRA-001 §B.3", "P1", "Pending", "N", "N", "N"),
                ("CI workflow running terragrunt run-all plan on PRs", "RFC-INFRA-001 §B.3", "P2", "Pending", "N", "N", "N"),
            ]),
            ("12.3 State Management", [
                ("Per-region S3 bucket for tfstate (DPDP residency)", "RFC-INFRA-001 §B.5", "P1", "Pending", "N", "N", "N"),
                ("Per-region DynamoDB lock table", "RFC-INFRA-001 §B.5", "P1", "Pending", "N", "N", "N"),
                ("Cross-stack reads via terraform_remote_state", "RFC-INFRA-001 §B.5", "P1", "Pending", "N", "N", "N"),
                ("State encryption with per-region KMS CMK", "RFC-INFRA-001 §B.5", "P1", "Pending", "N", "N", "N"),
            ]),
            ("12.4 Secrets Management", [
                ("AWS Secrets Manager as source of truth", "RFC-INFRA-001 §B.6", "P1", "Pending", "N", "N", "N"),
                ("External Secrets Operator (ESO) on EKS", "RFC-INFRA-001 §B.6", "P1", "Pending", "N", "N", "N"),
                ("Per-region ClusterSecretStore with IRSA", "RFC-INFRA-001 §B.6", "P1", "Pending", "N", "N", "N"),
                ("Rotation lambdas for DB creds", "RFC-INFRA-001 §B.6", "P2", "Pending", "N", "N", "N"),
                ("KMS envelope encryption for all secrets at rest", "RFC-INFRA-001 §B.6", "P1", "Pending", "N", "N", "N"),
            ]),
            ("12.5 Network Topology", [
                ("Per-region VPC (3 AZs, public/private/db subnets)", "RFC-INFRA-001 §B.4", "P1", "Pending", "N", "N", "N"),
                ("PrivateLink VPC endpoint to Aurora", "RFC-INFRA-001 §B.7", "P1", "Pending", "N", "N", "N"),
                ("Global auth plane in primary region (cross-region replicas)", "RFC-INFRA-001 §B.7", "P2", "Pending", "N", "N", "N"),
                ("Route53 latency routing for app endpoints", "RFC-INFRA-001 §B.7", "P2", "Pending", "N", "N", "N"),
                ("ACM certs per-region + us-east-1 wildcard for CloudFront", "RFC-INFRA-001 §B.4", "P1", "Pending", "N", "N", "N"),
            ]),
            ("12.6 Per-Tenant Isolation (runtime)", [
                ("Tenant onboarding via INSERT (no terraform apply)", "RFC-INFRA-001 §B.8", "P1", "Pending", "N", "N", "N"),
                ("medbrains-tenant-operator (Rust controller)", "RFC-INFRA-001 §B.8", "P2", "Pending", "N", "N", "N"),
                ("Per-tenant Kyverno PolicyException via ApplicationSet", "RFC-INFRA-001 §B.8", "P2", "Pending", "N", "N", "N"),
                ("Per-tenant edge-node CA cert issuance", "RFC-INFRA-001 §B.8", "P2", "Pending", "N", "N", "N"),
                ("Per-tenant CloudWatch log group", "RFC-INFRA-001 §B.8", "P3", "Pending", "N", "N", "N"),
            ]),
        ],
    },
    {
        "title": "13. KUBERNETES / EKS PLATFORM",
        "module": "K8s Platform",
        "groups": [
            ("13.1 EKS Cluster", [
                ("EKS 1.31 with Bottlerocket AMI", "RFC-INFRA-001 §C.1", "P1", "Pending", "N", "N", "N"),
                ("Karpenter autoscaling (replaces Cluster Autoscaler)", "RFC-INFRA-001 §C.1", "P1", "Pending", "N", "N", "N"),
                ("System NodePool (on-demand, AZ-spread)", "RFC-INFRA-001 §C.1", "P1", "Pending", "N", "N", "N"),
                ("App on-demand NodePool (m6i/m7i)", "RFC-INFRA-001 §C.1", "P1", "Pending", "N", "N", "N"),
                ("App spot NodePool (stateless workers)", "RFC-INFRA-001 §C.1", "P2", "Pending", "N", "N", "N"),
                ("GPU NodePool (V2 — radiology AI)", "RFC-INFRA-001 §C.1", "P3", "Pending", "N", "N", "N"),
                ("KMS-encrypted EKS secrets envelope", "RFC-INFRA-001 §C.5", "P1", "Pending", "N", "N", "N"),
            ]),
            ("13.2 Cluster Add-ons", [
                ("AWS Load Balancer Controller", "RFC-INFRA-001 §C.2", "P1", "Pending", "N", "N", "N"),
                ("External-DNS (Route53 sync from Ingress)", "RFC-INFRA-001 §C.2", "P1", "Pending", "N", "N", "N"),
                ("Cert-Manager for internal mTLS certs", "RFC-INFRA-001 §C.2", "P1", "Pending", "N", "N", "N"),
                ("Argo CD (GitOps reconciliation)", "RFC-INFRA-001 §C.2", "P1", "Pending", "N", "N", "N"),
                ("Argo ApplicationSet (per-region/per-env fan-out)", "RFC-INFRA-001 §C.2", "P1", "Pending", "N", "N", "N"),
                ("Argo Rollouts (canary + blue-green)", "RFC-INFRA-001 §C.2", "P2", "Pending", "N", "N", "N"),
                ("Kyverno (policy as YAML)", "RFC-INFRA-001 §C.2", "P1", "Pending", "N", "N", "N"),
                ("Falco (runtime security)", "RFC-INFRA-001 §C.2", "P2", "Pending", "N", "N", "N"),
                ("Velero (S3 backup of cluster state + PVs)", "RFC-INFRA-001 §C.2", "P1", "Pending", "N", "N", "N"),
                ("KEDA (event-driven autoscaling on NATS lag)", "RFC-INFRA-001 §C.2", "P2", "Pending", "N", "N", "N"),
                ("metrics-server (HPA backend)", "RFC-INFRA-001 §C.2", "P1", "Pending", "N", "N", "N"),
            ]),
            ("13.3 Networking & Service Mesh", [
                ("AWS VPC CNI + Cilium chaining", "RFC-INFRA-001 §C.4", "P1", "Pending", "N", "N", "N"),
                ("Cilium Service Mesh (mTLS via SPIFFE)", "RFC-INFRA-001 §C.4", "P1", "Pending", "N", "N", "N"),
                ("Hubble flow logs", "RFC-INFRA-001 §C.4", "P2", "Pending", "N", "N", "N"),
                ("Default-deny NetworkPolicy per namespace", "RFC-INFRA-001 §C.4", "P1", "Pending", "N", "N", "N"),
                ("L7 path-aware policies (per-service allow-list)", "RFC-INFRA-001 §C.4", "P1", "Pending", "N", "N", "N"),
                ("FQDN egress allowlist (Cilium FQDN policy)", "RFC-INFRA-001 §C.4", "P2", "Pending", "N", "N", "N"),
                ("L7 tenant header continuity validation", "RFC-INFRA-001 §C.5", "P1", "Pending", "N", "N", "N"),
                ("Per-tenant rate limits (Envoy ratelimit service)", "RFC-INFRA-001 §C.5", "P2", "Pending", "N", "N", "N"),
            ]),
            ("13.4 Workload Placement", [
                ("medbrains-server (Axum) Deployment + HPA + PDB", "RFC-INFRA-001 §C.3", "P1", "Pending", "N", "N", "N"),
                ("medbrains-tenant-operator Deployment", "RFC-INFRA-001 §C.3", "P2", "Pending", "N", "N", "N"),
                ("Web frontend on S3 + CloudFront with OAC", "RFC-INFRA-001 §C.3", "P1", "Pending", "Y", "N", "N"),
                ("YottaDB StatefulSet (3 replicas, EBS gp3, anti-affinity)", "RFC-INFRA-001 §C.3", "P1", "Pending", "N", "N", "N"),
                ("NATS JetStream StatefulSet (3 replicas)", "RFC-INFRA-001 §C.3", "P1", "Pending", "N", "N", "N"),
                ("Aurora PostgreSQL 16 (managed RDS, NOT in cluster)", "RFC-INFRA-001 §C.3", "P1", "Pending", "N", "N", "N"),
                ("ElastiCache Redis 7 (managed, NOT in cluster)", "RFC-INFRA-001 §C.3", "P1", "Pending", "N", "N", "N"),
                ("MSK (Kafka) — V2 if event volume > 50k msg/s", "RFC-INFRA-001 §C.3", "P3", "Pending", "N", "N", "N"),
            ]),
            ("13.5 Security & Compliance", [
                ("Pod Security Admission — restricted profile", "RFC-INFRA-001 §C.5", "P1", "Pending", "N", "N", "N"),
                ("Cosign image signing in CI", "RFC-INFRA-001 §C.5", "P1", "Pending", "N", "N", "N"),
                ("Kyverno verifyImages admission policy", "RFC-INFRA-001 §C.5", "P1", "Pending", "N", "N", "N"),
                ("IRSA per ServiceAccount (zero static AWS creds)", "RFC-INFRA-001 §C.5", "P1", "Pending", "N", "N", "N"),
                ("KMS at rest (EBS, Aurora, S3, Secrets Manager)", "RFC-INFRA-001 §C.5", "P1", "Pending", "N", "N", "N"),
                ("EKS audit logs to CloudWatch + Loki forwarding", "RFC-INFRA-001 §C.5", "P1", "Pending", "N", "N", "N"),
                ("TLS 1.3 minimum at ALB (HIPAA/DPDP)", "RFC-INFRA-001 §C.5", "P1", "Pending", "N", "N", "N"),
            ]),
            ("13.6 Observability Stack", [
                ("Prometheus (in-cluster + AMP remote-write)", "RFC-INFRA-001 §C.6", "P1", "Pending", "N", "N", "N"),
                ("Loki with S3 backend (90d hot / 1y archive)", "RFC-INFRA-001 §C.6", "P1", "Pending", "N", "N", "N"),
                ("Tempo with S3 backend (30d, tail-based sampling)", "RFC-INFRA-001 §C.6", "P1", "Pending", "N", "N", "N"),
                ("Grafana dashboards via GitOps ConfigMap", "RFC-INFRA-001 §C.6", "P1", "Pending", "N", "N", "N"),
                ("RED/USE dashboards per service", "RFC-INFRA-001 §C.6", "P2", "Pending", "N", "N", "N"),
                ("Business KPI dashboards (admissions/hr, lab TAT)", "RFC-INFRA-001 §C.6", "P2", "Pending", "N", "N", "N"),
                ("Alertmanager → PagerDuty (sev1/2) + Slack (sev3)", "RFC-INFRA-001 §C.6", "P1", "Pending", "N", "N", "N"),
                ("OpenTelemetry Collector (DaemonSet + Deployment)", "RFC-INFRA-001 §C.6", "P1", "Pending", "N", "N", "N"),
                ("tracing-opentelemetry in Rust app", "RFC-INFRA-001 §C.6", "P1", "Pending", "N", "N", "N"),
            ]),
            ("13.7 CI/CD Pipeline", [
                ("GitHub Actions build + test + ECR push", "RFC-INFRA-001 §C.7", "P1", "Pending", "N", "N", "N"),
                ("Argo CD pull-based deployment", "RFC-INFRA-001 §C.7", "P1", "Pending", "N", "N", "N"),
                ("Image promotion via auto-PR (dev) + manual approval (prod)", "RFC-INFRA-001 §C.7", "P2", "Pending", "N", "N", "N"),
                ("Canary rollout (10% → analyse → 50% → 100%)", "RFC-INFRA-001 §C.7", "P2", "Pending", "N", "N", "N"),
                ("Per-region staged rollout (pause one, continue others)", "RFC-INFRA-001 §C.7", "P2", "Pending", "N", "N", "N"),
            ]),
            ("13.8 DR & Backup", [
                ("Aurora PITR 35d + cross-region snapshots every 6h", "RFC-INFRA-001 §C.8", "P1", "Pending", "N", "N", "N"),
                ("YottaDB Velero hourly + EBS daily snapshots", "RFC-INFRA-001 §C.8", "P1", "Pending", "N", "N", "N"),
                ("S3 versioning + CRR for audit-archive and uploads", "RFC-INFRA-001 §C.8", "P1", "Pending", "N", "N", "N"),
                ("Velero hourly backup of K8s objects to S3", "RFC-INFRA-001 §C.8", "P1", "Pending", "N", "N", "N"),
                ("Edge node daily Restic snapshot to regional S3", "RFC-INFRA-001 §C.8", "P2", "Pending", "N", "N", "N"),
                ("Quarterly DR game day per region", "RFC-INFRA-001 §C.8", "P2", "Pending", "N", "N", "N"),
                ("Documented restore runbooks (Aurora, YottaDB, K8s)", "RFC-INFRA-001 §C.8", "P1", "Pending", "N", "N", "N"),
            ]),
        ],
    },
]


def main() -> None:
    wb = load_workbook(WB_PATH)
    ws = wb[SHEET]

    row = ws.max_row + 2
    sno = 1

    for chap in CHAPTERS:
        ws.cell(row=row, column=1, value=chap["title"]).font = BOLD
        ws.cell(row=row, column=1).fill = MODULE_FILL
        for c in range(1, 12):
            ws.cell(row=row, column=c).fill = MODULE_FILL
        row += 1

        for submod_title, features in chap["groups"]:
            ws.cell(row=row, column=1, value=f"  {submod_title}").font = BOLD
            for c in range(1, 12):
                ws.cell(row=row, column=c).fill = SUBMOD_FILL
            row += 1

            submod_short = submod_title.split(" ", 1)[1] if " " in submod_title else submod_title

            for feat, src, prio, status, web, mob, tv in features:
                values = [sno, chap["module"], submod_short, feat, src, prio, status, "RFC-INFRA-001", web, mob, tv]
                for col, val in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = BORDER
                    cell.alignment = WRAP
                row += 1
                sno += 1

    wb.save(WB_PATH)
    print(f"Added {sno - 1} infra features to '{SHEET}' sheet ending at row {row - 1}.")


if __name__ == "__main__":
    main()

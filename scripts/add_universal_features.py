#!/usr/bin/env python3
"""Add Universal/Cross-Cutting App features to MedBrains_Features.xlsx.

Adds 84 features across 4 sheets:
- Technical Infrastructure → Universal App Platform (58 features: 15 Done + 43 Pending)
- IT Security & Infrastructure → Accessibility & WCAG (5) + Clinical Safety UX (5)
- Mobile Apps → Universal Mobile (12 features)
- TV Displays & Queue → Universal TV (4 features)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

# Styling constants (matching project convention)
MODULE_FONT = Font(bold=True, size=11, color="1F4E79")
SUBMODULE_FONT = Font(bold=True, size=10, color="2E75B6")
MODULE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBMODULE_FILL = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")

RFC_REF = "RFC-MODULE-universal-platform"

# =============================================================================
# Sheet 1: Technical Infrastructure → Universal App Platform (58 features)
# =============================================================================
TECH_INFRA_FEATURES = [
    {"type": "module", "text": "Universal App Platform"},

    # --- Already Built (15 Done) ---
    {"type": "submodule", "text": "Built-In Platform Features"},
    {"feature": "Global command palette (Cmd+K) — search pages, navigate modules, run quick actions with keyboard", "status": "Done", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Toast notification system — success/error/warning/info toasts with auto-close and icon", "status": "Done", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Internationalization framework — 13 translation namespaces with react-i18next + HttpBackend", "status": "Done", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Multi-measurement unit conversion — kg↔lb, cm↔in, °C↔°F per locale with backend metric storage", "status": "Done", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Visual form builder — drag-drop field canvas, property editor, conditional logic, versioning", "status": "Done", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Dashboard builder — widget palette, drag-drop layout, data binding, module-scoped views", "status": "Done", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Screen builder — dynamic page generator with zone-based layout (form, table, kanban, calendar)", "status": "Done", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Integration builder — visual pipeline editor, node-based ETL, trigger/action/transform nodes", "status": "Done", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Role-based permission system — 111 permission codes, page guards, element-level visibility hooks", "status": "Done", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Paginated data table — custom render columns, loading skeletons, empty state, sticky header", "status": "Done", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "CSV bulk import/export — upload master data, download lists to CSV", "status": "Done", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Print template builder — configurable hospital-branded print layouts for invoices/reports", "status": "Done", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Collapsible sidebar navigation — pin/unpin, hover-expand, mobile burger, data-driven from config", "status": "Done", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Session management with proactive token refresh — 13-min refresh cycle, auto-logout on expiry", "status": "Done", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Clinical event broadcast — cross-module event provider for admission/discharge/prescription events", "status": "Done", "web": "Y", "mobile": "N", "tv": "N"},

    # --- Group 1: Navigation & Discovery (5 Pending) ---
    {"type": "submodule", "text": "Navigation & Discovery"},
    {"feature": "Enhanced global search — fuzzy search across patients, encounters, orders, documents (beyond nav-only Spotlight)", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Recent items panel — last 20 viewed patients, orders, reports with one-click re-access", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Favorites/bookmarks — pin any entity (patient, report, page) to personal quick-access panel", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Deep link sharing — copy shareable URL for any record/entity to clipboard for team sharing", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Patient-aware breadcrumbs — context breadcrumbs showing patient→encounter→order hierarchy", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},

    # --- Group 2: Notification Center (6 Pending) ---
    {"type": "submodule", "text": "Notification Center"},
    {"feature": "Notification center UI (bell icon) — persistent notification history with unread count, grouped by module", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Per-user notification preferences — choose channel (in-app, email, SMS, push) per event type", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Three-tier alert severity — critical (immediate+audible), urgent (in-tray), routine (daily digest)", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Do-not-disturb mode — suppress non-critical notifications during procedures/consultations", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Alert escalation chain — unacknowledged critical alerts auto-escalate to supervisor after timeout", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Alert fatigue mitigation — deduplicate repeated alerts, batch non-urgent, enforce clinical specificity", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},

    # --- Group 3: Productivity & Shortcuts (6 Pending) ---
    {"type": "submodule", "text": "Productivity & Shortcuts"},
    {"feature": "Keyboard shortcuts framework — configurable hotkeys for common actions across all modules", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Keyboard shortcut reference overlay — Cmd+? opens full shortcuts cheat sheet", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Auto-save form drafts — save form state every 30s to browser storage, recover on return", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Unsaved changes guard — prompt before navigating away from forms with pending edits", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Clinical context preservation — maintain form state, scroll position, filters across patient switches", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Smart phrases — type short abbreviation to auto-insert pre-written clinical text blocks (Epic SmartPhrases pattern)", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},

    # --- Group 4: Personalization & Theming (5 Pending) ---
    {"type": "submodule", "text": "Personalization & Theming"},
    {"feature": "Dark mode toggle — system-preference-aware dark theme with manual override, persisted per user", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "Y"},
    {"feature": "High-contrast mode — WCAG AAA compliant theme for visual accessibility", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Configurable table columns — show/hide/reorder columns in any DataTable, save as named view", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Saved filter presets — save and share frequently used filter combinations per module", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Custom landing page — user chooses default homepage (dashboard, patient list, OPD queue, etc.)", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},

    # --- Group 5: Collaboration (4 Pending) ---
    {"type": "submodule", "text": "Collaboration"},
    {"feature": "In-context comments — add notes/comments on any entity (patient, order, invoice, equipment)", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "@mention colleagues — tag staff in comments, auto-generates notification to mentioned user", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Presence indicators — see who else is viewing the same patient/record to avoid edit conflicts", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Contextual task assignment — assign follow-up tasks to colleagues from any module record", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},

    # --- Group 6: Universal Data Operations (5 Pending) ---
    {"type": "submodule", "text": "Universal Data Operations"},
    {"feature": "Universal list export — export any DataTable to Excel, CSV, or PDF with current filters applied", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Universal entity print — print any detail page with hospital-branded header/footer", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Bulk multi-select actions — checkbox column in DataTable + batch toolbar (print, export, status change)", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Side-by-side record comparison — compare two records of same type (patients, invoices, test results)", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Inline clinical calculators — BMI, GFR, drug dosage, unit conversion accessible from any page", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},

    # --- Group 7: Real-Time Platform (4 Pending) ---
    {"type": "submodule", "text": "Real-Time Platform"},
    {"feature": "WebSocket infrastructure — live data push for queues, beds, alerts without polling", "status": "Pending", "web": "Y", "mobile": "N", "tv": "Y"},
    {"feature": "Real-time queue updates — live OPD/ER queue position changes without page refresh", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "Y"},
    {"feature": "Real-time bed board — live occupancy and status changes pushed to all viewers", "status": "Pending", "web": "Y", "mobile": "N", "tv": "Y"},
    {"feature": "Real-time notification delivery — push notifications to browser without polling", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},

    # --- Group 8: Progressive Web App (4 Pending) ---
    {"type": "submodule", "text": "Progressive Web App"},
    {"feature": "PWA installable — install on desktop/mobile with app icon, splash screen, standalone window", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Offline shell — cached navigation and last-viewed data when network unavailable", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Background sync — queue form submissions when offline, auto-submit on reconnect", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Service worker push notifications — receive alerts even when browser tab is closed", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},

    # --- Group 9: Audit & Activity (4 Pending) ---
    {"type": "submodule", "text": "Audit & Activity"},
    {"feature": "Personal activity feed — chronological log of own actions across all modules", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Entity timeline — visual timeline showing all events on any entity (patient→visits→labs→billing)", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Change history diff — show what changed, by whom, when on any editable record", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Active sessions dashboard — view sessions across devices, force logout from other devices", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
]

# =============================================================================
# Sheet 2: IT Security & Infrastructure → Accessibility & WCAG + Clinical Safety UX (10 features)
# =============================================================================
IT_SECURITY_FEATURES = [
    # --- Group 10: Accessibility & WCAG (5 Pending) ---
    {"type": "module", "text": "Accessibility & WCAG"},
    {"type": "submodule", "text": "WCAG 2.1 AA Compliance"},
    {"feature": "Screen reader optimization — ARIA labels, landmark regions, focus management across all pages", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Keyboard-only navigation — full app usable without mouse via Tab/Enter/Escape", "status": "Pending", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "User-adjustable font scaling — 100% to 200% text size without layout breakage", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Color-blind safe palettes — alternative color schemes for deuteranopia, protanopia, tritanopia", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Reduced motion mode — disable animations for users with vestibular disorders", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},

    # --- Group 11: Clinical Safety UX (5 Pending) ---
    {"type": "module", "text": "Clinical Safety UX"},
    {"type": "submodule", "text": "Patient Safety Safeguards"},
    {"feature": "Break-glass emergency access — override authorization with emergency code, mandatory reason, auto-audit log", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Patient identity banner — always-visible bar showing active patient (name, UHID, age, allergies) on clinical pages", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Similar-name patient alert — visual warning when two patients with similar names are accessed in same session", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "High-risk action confirmation — double-check step for dangerous doses, critical medication changes, irreversible orders", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Clinical handoff summary — auto-generated end-of-shift summary of patients under care with pending actions", "status": "Pending", "web": "Y", "mobile": "Y", "tv": "N"},
]

# =============================================================================
# Sheet 3: Mobile Apps → Universal Mobile (12 features)
# =============================================================================
MOBILE_FEATURES = [
    {"type": "module", "text": "Universal Mobile"},

    {"type": "submodule", "text": "Mobile Authentication"},
    {"feature": "Biometric app login — Face ID / fingerprint authentication for quick app unlock", "status": "Pending", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "PIN fallback — 6-digit PIN when biometric unavailable", "status": "Pending", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Auto-lock timeout — configurable auto-lock after inactivity (1/5/15 min)", "status": "Pending", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Secure clipboard — auto-clear copied patient data after 60 seconds", "status": "Pending", "web": "N", "mobile": "Y", "tv": "N"},

    {"type": "submodule", "text": "Mobile Input & Capture"},
    {"feature": "Universal barcode/QR scanner — scan patient wristband, equipment tag, drug barcode from any screen", "status": "Pending", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Camera document capture — photograph prescriptions, wounds, documents with auto-attach to patient", "status": "Pending", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Voice-to-text input — dictate into any text field using device speech recognition", "status": "Pending", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Swipe gesture actions — swipe in lists for quick approve/reject/mark done", "status": "Pending", "web": "N", "mobile": "Y", "tv": "N"},

    {"type": "submodule", "text": "Mobile Offline & Sync"},
    {"feature": "Offline patient cache — cache frequently accessed patient records for offline viewing", "status": "Pending", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Offline form submission — complete clinical forms offline, queue for sync when connected", "status": "Pending", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Sync conflict resolution — handle concurrent edits with merge/overwrite UI", "status": "Pending", "web": "N", "mobile": "Y", "tv": "N"},
    {"feature": "Background data prefetch — pre-download today's patients/appointments on app launch", "status": "Pending", "web": "N", "mobile": "Y", "tv": "N"},
]

# =============================================================================
# Sheet 4: TV Displays & Queue → Universal TV (4 features)
# =============================================================================
TV_FEATURES = [
    {"type": "module", "text": "Universal TV"},

    {"type": "submodule", "text": "TV Platform Infrastructure"},
    {"feature": "WebSocket auto-refresh framework — live content updates on all TV displays without manual intervention", "status": "Pending", "web": "N", "mobile": "N", "tv": "Y"},
    {"feature": "Multi-language content rotation — cycle display through hospital's configured languages per locale", "status": "Pending", "web": "N", "mobile": "N", "tv": "Y"},
    {"feature": "Emergency broadcast takeover — override all TV displays with emergency message (fire, code blue, lockdown)", "status": "Pending", "web": "N", "mobile": "N", "tv": "Y"},
    {"feature": "Ambient mode — clock, date, hospital branding when no active queue or display content", "status": "Pending", "web": "N", "mobile": "N", "tv": "Y"},
]


def add_features_to_sheet(ws, features, rfc_ref):
    """Append features to a worksheet following the project styling convention.

    Returns (total_added, done_count, pending_count).
    """
    last_row = ws.max_row

    # Find current max S.No
    max_sno = 0
    for row in range(2, last_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, (int, float)):
            max_sno = max(max_sno, int(val))

    current_sno = max_sno
    current_row = last_row + 2  # Leave a blank row separator

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    done_count = 0
    pending_count = 0

    for item in features:
        if item.get("type") == "module":
            # Module header row — col B
            ws.cell(row=current_row, column=2, value=item["text"])
            ws.cell(row=current_row, column=2).font = MODULE_FONT
            for col in range(1, 12):
                ws.cell(row=current_row, column=col).fill = MODULE_FILL
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1

        elif item.get("type") == "submodule":
            # Sub-module header row — col C
            ws.cell(row=current_row, column=3, value=item["text"])
            ws.cell(row=current_row, column=3).font = SUBMODULE_FONT
            for col in range(1, 12):
                ws.cell(row=current_row, column=col).fill = SUBMODULE_FILL
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1

        else:
            # Feature row
            current_sno += 1
            status = item.get("status", "Pending")

            ws.cell(row=current_row, column=1, value=current_sno)       # A: S.No
            # col B (Module) and col C (Sub-Module) left blank for features
            ws.cell(row=current_row, column=4, value=item["feature"])    # D: Feature
            ws.cell(row=current_row, column=5, value="MedBrains")       # E: Source
            ws.cell(row=current_row, column=6, value="P1")              # F: Priority
            ws.cell(row=current_row, column=7, value=status)            # G: Status
            ws.cell(row=current_row, column=8, value=rfc_ref)           # H: RFC Ref
            ws.cell(row=current_row, column=9, value=item.get("web", "N"))     # I: Web
            ws.cell(row=current_row, column=10, value=item.get("mobile", "N")) # J: Mobile
            ws.cell(row=current_row, column=11, value=item.get("tv", "N"))     # K: TV

            for col in range(1, 12):
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

            if status == "Done":
                done_count += 1
            else:
                pending_count += 1

            current_row += 1

    total = current_sno - max_sno
    return total, done_count, pending_count


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # --- Sheet 1: Technical Infrastructure ---
    ws1 = wb["Technical Infrastructure"]
    count1, done1, pend1 = add_features_to_sheet(ws1, TECH_INFRA_FEATURES, RFC_REF)
    print(f"Technical Infrastructure: added {count1} features ({done1} Done, {pend1} Pending)")

    # --- Sheet 2: IT Security & Infrastructure ---
    ws2 = wb["IT, Security & Infrastructure"]
    count2, done2, pend2 = add_features_to_sheet(ws2, IT_SECURITY_FEATURES, RFC_REF)
    print(f"IT Security & Infrastructure: added {count2} features ({done2} Done, {pend2} Pending)")

    # --- Sheet 3: Mobile Apps ---
    ws3 = wb["Mobile Apps"]
    count3, done3, pend3 = add_features_to_sheet(ws3, MOBILE_FEATURES, RFC_REF)
    print(f"Mobile Apps: added {count3} features ({done3} Done, {pend3} Pending)")

    # --- Sheet 4: TV Displays & Queue ---
    ws4 = wb["TV Displays & Queue"]
    count4, done4, pend4 = add_features_to_sheet(ws4, TV_FEATURES, RFC_REF)
    print(f"TV Displays & Queue: added {count4} features ({done4} Done, {pend4} Pending)")

    wb.save(EXCEL_PATH)

    total = count1 + count2 + count3 + count4
    total_done = done1 + done2 + done3 + done4
    total_pending = pend1 + pend2 + pend3 + pend4
    print(f"\nDone! Total {total} features added across 4 sheets ({total_done} Done, {total_pending} Pending).")


if __name__ == "__main__":
    main()

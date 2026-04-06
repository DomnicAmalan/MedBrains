#!/usr/bin/env python3
"""Add Sellable USP Features to MedBrains_Features.xlsx.

Adds 105 new P1 features across 5 sheets in 8 USP groups:
- Clinical → Ambient AI (13) + Journey Orchestration (12)
- Admin & Operations → Compliance Academy (14) + Green Hospital (12)
- IT, Security & Infrastructure → Global Platform (16) + Digital Twin (10)
- Technical Infrastructure → API Ecosystem (15)
- Multi-Hospital & Vendors → Benchmarking Network (13)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

# Styling constants (matching project convention)
MODULE_FONT = Font(bold=True, size=11, color="1F4E79")
SUBMODULE_FONT = Font(bold=True, size=10, color="2E75B6")
MODULE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBMODULE_FILL = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")

RFC_REF = "RFC-MODULE-usp-platform"

# =============================================================================
# Sheet: Clinical → Ambient AI (13) + Journey Orchestration (12) = 25 features
# =============================================================================
CLINICAL_FEATURES = [
    # --- USP 3: Ambient Clinical Intelligence ---
    {"type": "module", "text": "Ambient AI"},

    {"type": "submodule", "text": "Voice Documentation"},
    {"feature": "Ambient voice capture — AI listens to patient-doctor conversation and generates structured SOAP notes", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Auto-generated consultation summary — patient-friendly plain-language visit summary from ambient capture", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Voice-to-order extraction — detect medication, lab, imaging orders from conversation and pre-fill order forms", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Multilingual ambient capture — transcribe Hindi, Tamil, Telugu, Arabic, Nepali conversations to English notes", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Voice-activated patient lookup — 'Show me patient's last blood report' hands-free during consultation", "web": "Y", "mobile": "Y", "tv": "N"},

    {"type": "submodule", "text": "AI Clinical Assistance"},
    {"feature": "AI clinical coding — auto-suggest ICD-10/CPT codes from clinical notes with confidence scores", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Drug interaction check from voice — real-time alert when doctor verbally prescribes conflicting medication", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "AI-assisted discharge summary — auto-draft discharge summary from admission-to-discharge clinical data", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Smart template selection — AI selects appropriate clinical note template based on chief complaint", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Clinical decision support nudges — evidence-based suggestions during consultation (screening due, protocol gap)", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "AI differential diagnosis assistant — suggest differentials from symptoms/findings with probability ranking", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Conversation-to-referral — detect referral intent from conversation and auto-create referral with context", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "AI note quality scoring — flag incomplete/inconsistent documentation before sign-off with fix suggestions", "web": "Y", "mobile": "N", "tv": "N"},

    # --- USP 4: Patient Journey Orchestration ---
    {"type": "module", "text": "Journey Orchestration"},

    {"type": "submodule", "text": "Journey Design"},
    {"feature": "Visual journey builder — drag-and-drop care pathway designer for common conditions (surgery, chemo, dialysis)", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Pre-built journey templates — 30+ templates for common pathways (elective surgery, normal delivery, AMI, stroke)", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Auto-task routing — journey steps auto-assigned to correct department/staff as patient progresses", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Cross-department handoff automation — auto-notify next department when current step completes", "web": "Y", "mobile": "Y", "tv": "N"},

    {"type": "submodule", "text": "Journey Monitoring"},
    {"feature": "Journey progress tracker — patient and family can see where they are in the care journey via portal/app", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "SLA monitoring per journey step — alert when any step exceeds expected duration (lab TAT, OT wait, bed allocation)", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Journey deviation alerts — flag when patient pathway diverges from planned route (unexpected ICU transfer, complication)", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Outcome tracking per journey — link patient outcomes to specific pathway versions for continuous improvement", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Patient communication at each step — auto-send status updates, preparation instructions, expected timeline", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Journey analytics — average time per step, bottleneck identification, completion rate by pathway type", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Insurance pre-authorization per journey — auto-trigger PA requests at journey stages requiring authorization", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Family/caregiver journey view — separate portal view for family showing patient progress and next steps", "web": "Y", "mobile": "Y", "tv": "N"},
]

# =============================================================================
# Sheet: Admin & Operations → Compliance Academy (14) + Green Hospital (12) = 26 features
# =============================================================================
ADMIN_FEATURES = [
    # --- USP 1: Staff Compliance Academy ---
    {"type": "module", "text": "Compliance Academy"},

    {"type": "submodule", "text": "Learning Content"},
    {"feature": "Learning module builder — rich content editor for text, images, embedded video lessons", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Pre-built NABH/JCI compliance course library — 50+ modules mapped to accreditation standards", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Department-specific mandatory modules — Pharmacy→NDPS Act, Lab→NABL, Radiology→AERB, Blood Bank→NACO", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Video lesson hosting with progress tracking — resume where left off, minimum watch time enforcement", "web": "Y", "mobile": "Y", "tv": "N"},

    {"type": "submodule", "text": "Learning Paths & Assessment"},
    {"feature": "Role-based learning paths — auto-assign courses by designation (nurse, doctor, pharmacist, admin)", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Interactive quiz & assessment engine — MCQ, true/false, case-based with configurable pass marks", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "New joiner onboarding learning track — auto-assigned on employee creation with deadline tracking", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Annual re-certification workflow — auto-trigger re-assessment 30 days before certification expiry", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "Compliance & Analytics"},
    {"feature": "Regulatory change auto-notification — detect regulation updates and auto-assign refresher modules", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Compliance certification tracker — issue/expiry dates, auto-renewal reminders (BLS, fire safety, POSH)", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Learning gamification — completion points, department leaderboard, monthly top learner recognition", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Audit-ready training compliance reports — per-employee, per-department, per-regulation completion matrix", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Learning analytics dashboard — completion rates, quiz scores, overdue modules, department comparison", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Offline learning mode — download modules for areas with poor connectivity, sync completion on reconnect", "web": "N", "mobile": "Y", "tv": "N"},

    # --- USP 6: Green Hospital & Sustainability ---
    {"type": "module", "text": "Green Hospital"},

    {"type": "submodule", "text": "Resource Monitoring"},
    {"feature": "Energy consumption dashboard — track electricity, diesel, solar, gas usage per department with cost trends", "web": "Y", "mobile": "N", "tv": "Y"},
    {"feature": "Water consumption monitoring — track water usage per department, recycled water percentage, rainwater harvesting", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Solar/renewable energy ROI tracker — monitor solar generation vs consumption with payback period calculation", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "Waste & Emissions"},
    {"feature": "Biomedical waste tracking — categorize, weigh, log waste by type (red/yellow/white/blue) per BMW Rules 2016", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Carbon footprint calculator — estimate Scope 1/2/3 emissions per department, procedure, and supply chain", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "OT & lab chemical waste tracking — track formaldehyde, xylene, cytotoxic drug waste with safe disposal logs", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Paperless adoption tracker — measure paper reduction across departments (digital consent, e-prescriptions, e-reports)", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "Sustainability Compliance"},
    {"feature": "Green procurement scoring — rate vendors on sustainability (recyclable packaging, local sourcing, carbon offset)", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "NABH sustainability compliance checklist — auto-mapped to NABH 5th edition green hospital criteria", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Sustainability KPI dashboard — composite green score with benchmarks against IGBC/GRIHA standards", "web": "Y", "mobile": "N", "tv": "Y"},
    {"feature": "Staff sustainability awareness modules — monthly eco-tips, waste segregation training, energy-saving practices", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "ESG report generator — annual sustainability report with charts for board/investors/accreditation bodies", "web": "Y", "mobile": "N", "tv": "N"},
]

# =============================================================================
# Sheet: IT, Security & Infrastructure → Global Platform (16) + Digital Twin (10) = 26 features
# =============================================================================
IT_FEATURES = [
    # --- USP 2: Multi-Everything Global Platform ---
    {"type": "module", "text": "Global Platform"},

    {"type": "submodule", "text": "Localization Engine"},
    {"feature": "Country-specific regulatory engine — auto-detect applicable laws/bodies based on hospital country and state", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Multi-script patient name storage — Latin + Devanagari + Arabic + Thai script with transliteration", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Country-specific patient ID formats — Aadhaar (India), Emirates ID (UAE), NRIC (Singapore), SSN (US)", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Multi-calendar support — Gregorian + Hijri + Thai Buddhist + Nepali Bikram Sambat with auto-conversion", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Multi-format date/time/number display — DD/MM/YYYY vs MM/DD/YYYY vs YYYY-MM-DD per locale preference", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "RTL (right-to-left) UI support — Arabic/Hebrew interface mirroring for Gulf region deployments", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Multi-timezone scheduling — per-campus timezone with cross-timezone appointment coordination", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Multi-measurement auto-conversion — kg↔lb, cm↔in, °C↔°F, mmol/L↔mg/dL per locale with stored metric", "web": "Y", "mobile": "Y", "tv": "N"},

    {"type": "submodule", "text": "Country-Specific Compliance"},
    {"feature": "Country-specific tax engine — GST (India), VAT (UAE/EU), sales tax (US) with configurable rates per service", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Country-specific drug scheduling — Schedule H/H1/X (India), Schedule 8 (Australia), Class A/B/C (UK)", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Country-specific consent templates — legal requirements vary (India=witness required, US=HIPAA notice)", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Country-specific clinical coding — ICD-10-CM (US) vs ICD-10-AM (Australia) vs ICD-10 WHO (India)", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Country-specific billing formats — NABH format (India), DHA/HAAD format (UAE), CMS-1500 (US)", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Country-specific emergency protocols — MLC reporting (India), mandatory reporting laws per jurisdiction", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Regional holiday calendar — auto-populate public holidays per country for leave/scheduling modules", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Deployment region selector — data residency compliance (data stays in-country per GDPR/DPDP/PDPA)", "web": "Y", "mobile": "N", "tv": "N"},

    # --- USP 5: Hospital Digital Twin & Simulation ---
    {"type": "module", "text": "Digital Twin"},

    {"type": "submodule", "text": "Live Visualization"},
    {"feature": "Real-time hospital digital twin — live 3D visualization of bed occupancy, staff positions, equipment status", "web": "Y", "mobile": "N", "tv": "Y"},
    {"feature": "Historical replay — replay any past day's operations to identify root causes of delays or incidents", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "What-If Simulation"},
    {"feature": "What-if capacity simulation — model impact of adding beds, closing wards, or changing staffing ratios", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Disaster scenario simulation — simulate mass casualty, pandemic surge, power outage on hospital operations", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Patient flow simulation — model bottleneck impact of changing admission/discharge policies", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "OT scheduling simulation — optimize surgical block allocation by simulating different configurations", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Staffing optimization simulator — model shift patterns and nurse-to-patient ratios against patient acuity", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Equipment utilization simulation — predict impact of adding/removing ventilators, monitors, or imaging machines", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Financial impact modeling — simulate revenue impact of tariff changes, new services, or insurance panel changes", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Simulation report generator — export scenario comparison reports for board/management decision-making", "web": "Y", "mobile": "N", "tv": "N"},
]

# =============================================================================
# Sheet: Technical Infrastructure → API Ecosystem (15 features)
# =============================================================================
TECH_INFRA_FEATURES = [
    # --- USP 7: Open API & Integration Ecosystem ---
    {"type": "module", "text": "API Ecosystem"},

    {"type": "submodule", "text": "Core API Platform"},
    {"feature": "RESTful API for every module — auto-generated OpenAPI 3.1 docs with interactive Swagger UI", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Webhook event system — subscribe to events (patient_admitted, lab_result_ready, bill_generated)", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "OAuth 2.0 third-party app authorization — external apps request scoped access with patient consent", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "FHIR R4 native API layer — Patient, Observation, MedicationRequest, DiagnosticReport resources", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Event-driven inter-module communication — NATS JetStream pub/sub for real-time cross-module triggers", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "Pre-Built Connectors"},
    {"feature": "Pre-built Tally connector — auto-push journal entries, invoices, payment receipts to Tally Prime", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Pre-built SAP connector — GL account sync, purchase order exchange, vendor master sync", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Pre-built ABDM connector — ABHA creation, care context linking, health record push (M1/M2/M3)", "web": "Y", "mobile": "Y", "tv": "N"},
    {"feature": "Pre-built lab instrument adapter — HL7v2 ORM/ORU message parsing for bi-directional LIS interfacing", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "Developer Platform"},
    {"feature": "Integration marketplace UI — browse, install, configure third-party connectors from admin panel", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "API usage dashboard — request counts, error rates, latency per consumer with rate limiting controls", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Bulk import/export engine — CSV, JSON, FHIR Bundle for patient data migration across systems", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Scheduled job framework — configurable cron-like triggers for EOD reports, data sync, cleanup tasks", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "SDKs for popular languages — Python, JavaScript, .NET client libraries auto-generated from OpenAPI spec", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Sandbox environment — isolated test tenant with sample data for third-party developers to build against", "web": "Y", "mobile": "N", "tv": "N"},
]

# =============================================================================
# Sheet: Multi-Hospital & Vendors → Benchmarking Network (13 features)
# =============================================================================
MULTI_HOSPITAL_FEATURES = [
    # --- USP 8: Hospital Benchmarking & Peer Network ---
    {"type": "module", "text": "Benchmarking Network"},

    {"type": "submodule", "text": "KPI Sharing & Comparison"},
    {"feature": "Anonymized KPI sharing — hospitals opt-in to share metrics (ALOS, infection rate, mortality) with peer network", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Peer comparison dashboards — compare your hospital's KPIs against anonymized peers by size/type/region", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "National/regional percentile ranking — see where your hospital stands (top 10%, median, bottom quartile)", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Department-level benchmarking — compare ED wait times, OT utilization, lab TAT against peer averages", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Custom peer group creation — define peer group by bed count, specialty mix, location, accreditation status", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "Clinical & Financial Benchmarks"},
    {"feature": "Accreditation readiness score — auto-calculated NABH/JCI readiness percentage with gap identification", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Revenue per bed benchmarking — compare revenue metrics against similar-sized hospitals in same region", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Staffing ratio benchmarking — nurse-to-patient ratio, doctor-to-bed ratio vs recommended standards", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Infection rate benchmarking — HAI rates compared to NHSN (US) or INICC (India) published benchmarks", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Cost-per-case analysis — compare treatment costs by DRG/procedure against regional averages", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "Intelligence & Reporting"},
    {"feature": "Best-practice recommendations — AI-generated suggestions based on top-performing peers' configurations", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Benchmark trend alerts — auto-notify when your metric falls below peer average for 3 consecutive months", "web": "Y", "mobile": "N", "tv": "N"},
    {"feature": "Benchmarking reports for board/management — exportable PDF/PPT with peer comparison charts", "web": "Y", "mobile": "N", "tv": "N"},
]


def add_features_to_sheet(ws, features, rfc_ref):
    """Append features to a worksheet following the project styling convention."""
    last_row = ws.max_row

    # Find current max S.No
    max_sno = 0
    for row in range(2, last_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, (int, float)):
            max_sno = max(max_sno, int(val))

    current_sno = max_sno
    current_row = last_row + 2  # Leave a blank row separator

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for item in features:
        if item.get("type") == "module":
            # Module header row — col B
            ws.cell(row=current_row, column=2, value=item["text"])
            ws.cell(row=current_row, column=2).font = MODULE_FONT
            for col in range(1, 12):
                ws.cell(row=current_row, column=col).fill = MODULE_FILL
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1

        elif item.get("type") == "submodule":
            # Sub-module header row — col C
            ws.cell(row=current_row, column=3, value=item["text"])
            ws.cell(row=current_row, column=3).font = SUBMODULE_FONT
            for col in range(1, 12):
                ws.cell(row=current_row, column=col).fill = SUBMODULE_FILL
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1

        else:
            # Feature row
            current_sno += 1
            ws.cell(row=current_row, column=1, value=current_sno)      # A: S.No
            # col B (Module) and col C (Sub-Module) left blank for features
            ws.cell(row=current_row, column=4, value=item["feature"])   # D: Feature
            ws.cell(row=current_row, column=5, value="MedBrains")      # E: Source
            ws.cell(row=current_row, column=6, value="P1")             # F: Priority
            ws.cell(row=current_row, column=7, value="Pending")        # G: Status
            ws.cell(row=current_row, column=8, value=rfc_ref)          # H: RFC Ref
            ws.cell(row=current_row, column=9, value=item.get("web", "N"))     # I: Web
            ws.cell(row=current_row, column=10, value=item.get("mobile", "N")) # J: Mobile
            ws.cell(row=current_row, column=11, value=item.get("tv", "N"))     # K: TV

            for col in range(1, 12):
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

            current_row += 1

    return current_sno - max_sno


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # --- Sheet 1: Clinical ---
    ws1 = wb["Clinical"]
    count1 = add_features_to_sheet(ws1, CLINICAL_FEATURES, RFC_REF)
    print(f"Clinical: added {count1} features (Ambient AI + Journey Orchestration)")

    # --- Sheet 2: Admin & Operations ---
    ws2 = wb["Admin & Operations"]
    count2 = add_features_to_sheet(ws2, ADMIN_FEATURES, RFC_REF)
    print(f"Admin & Operations: added {count2} features (Compliance Academy + Green Hospital)")

    # --- Sheet 3: IT, Security & Infrastructure ---
    ws3 = wb["IT, Security & Infrastructure"]
    count3 = add_features_to_sheet(ws3, IT_FEATURES, RFC_REF)
    print(f"IT, Security & Infrastructure: added {count3} features (Global Platform + Digital Twin)")

    # --- Sheet 4: Technical Infrastructure ---
    ws4 = wb["Technical Infrastructure"]
    count4 = add_features_to_sheet(ws4, TECH_INFRA_FEATURES, RFC_REF)
    print(f"Technical Infrastructure: added {count4} features (API Ecosystem)")

    # --- Sheet 5: Multi-Hospital & Vendors ---
    ws5 = wb["Multi-Hospital & Vendors"]
    count5 = add_features_to_sheet(ws5, MULTI_HOSPITAL_FEATURES, RFC_REF)
    print(f"Multi-Hospital & Vendors: added {count5} features (Benchmarking Network)")

    wb.save(EXCEL_PATH)
    total = count1 + count2 + count3 + count4 + count5
    print(f"\nDone! Total {total} features added across 5 sheets.")


if __name__ == "__main__":
    main()

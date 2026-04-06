#!/usr/bin/env python3
"""Analyze MedBrains_Features.xlsx to find modules with the most pending features."""

from collections import defaultdict
from openpyxl import load_workbook

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

# Modules we consider already complete/built
COMPLETED_MODULES = {
    "patient management", "opd", "outpatient (opd)", "outpatient department (opd)",
    "lab/lis", "lab", "laboratory", "laboratory/lis",
    "pharmacy", "billing", "ipd", "inpatient (ipd)", "inpatient department (ipd)",
    "radiology", "blood bank", "icu", "intensive care unit",
    "onboarding & setup", "onboarding", "setup",
    "navigation/access control", "navigation", "access control",
    "integration hub", "integration", "integrations",
    "dashboard", "dashboards",
    "form builder", "screen builder",
}

def normalize_module(name):
    """Normalize module name for comparison."""
    if not name:
        return None
    return str(name).strip()

def is_completed(module_name):
    """Check if a module is in our completed list."""
    if not module_name:
        return False
    lower = module_name.lower().strip()
    for c in COMPLETED_MODULES:
        if c in lower or lower in c:
            return True
    return False

def main():
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # Track per-module stats: {module_name: {sheet, total, pending, done, in_progress, partial}}
    module_stats = defaultdict(lambda: {
        "sheets": set(),
        "total": 0,
        "pending": 0,
        "done": 0,
        "in_progress": 0,
        "partial": 0,
        "other": 0,
    })

    print("=" * 90)
    print("MEDBRAINS FEATURE ANALYSIS — PENDING MODULES")
    print("=" * 90)
    print()

    # First, let's understand the sheet structure
    print(f"Workbook has {len(wb.sheetnames)} sheets: {wb.sheetnames}")
    print()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row — look for "Module" column
        header_row = None
        header_idx = None
        module_col = None
        status_col = None
        feature_col = None
        submodule_col = None

        for i, row in enumerate(rows[:10]):  # Check first 10 rows for header
            if row is None:
                continue
            for j, cell in enumerate(row):
                if cell and str(cell).strip().lower() == "module":
                    header_row = i
                    module_col = j
                if cell and str(cell).strip().lower() == "status":
                    status_col = j
                if cell and str(cell).strip().lower() == "feature":
                    feature_col = j
                if cell and str(cell).strip().lower() in ("sub-module", "submodule", "sub module"):
                    submodule_col = j
            if module_col is not None:
                break

        if module_col is None:
            print(f"  Sheet '{sheet_name}': No 'Module' column found, skipping")
            continue

        print(f"  Sheet '{sheet_name}': Module col={module_col}, Status col={status_col}, "
              f"Feature col={feature_col}, {len(rows)-header_row-1} data rows")

        # Process data rows
        current_module = None
        for row in rows[header_row + 1:]:
            if not row or all(c is None for c in row):
                continue

            # Get module name (carry forward if blank — some sheets merge cells)
            raw_module = row[module_col] if module_col < len(row) else None
            if raw_module:
                current_module = normalize_module(raw_module)
            module_name = current_module

            if not module_name:
                continue

            # Get status
            status = ""
            if status_col is not None and status_col < len(row) and row[status_col]:
                status = str(row[status_col]).strip().lower()

            # Get feature text to confirm this is a real feature row
            feature_text = None
            if feature_col is not None and feature_col < len(row):
                feature_text = row[feature_col]

            # Skip rows that look like headers or empty features
            if not feature_text and not status:
                # Could be a module/submodule header row — skip
                continue

            stats = module_stats[module_name]
            stats["sheets"].add(sheet_name)
            stats["total"] += 1

            if status in ("done", "complete", "completed"):
                stats["done"] += 1
            elif status in ("in progress", "in-progress", "wip"):
                stats["in_progress"] += 1
            elif status in ("partial", "partially done"):
                stats["partial"] += 1
            elif status in ("pending", "") or not status:
                stats["pending"] += 1
            else:
                stats["other"] += 1

    print()
    print("=" * 90)
    print("ALL MODULES — SORTED BY PENDING COUNT (DESCENDING)")
    print("=" * 90)
    print()
    print(f"{'Module':<45} {'Total':>6} {'Pend':>6} {'Done':>6} {'WIP':>6} {'Part':>6} {'Sheet(s)'}")
    print("-" * 120)

    sorted_modules = sorted(module_stats.items(), key=lambda x: x[1]["pending"], reverse=True)
    for mod_name, stats in sorted_modules:
        completed_marker = " [BUILT]" if is_completed(mod_name) else ""
        sheets_str = ", ".join(sorted(stats["sheets"]))
        print(f"{(mod_name + completed_marker):<45} {stats['total']:>6} {stats['pending']:>6} "
              f"{stats['done']:>6} {stats['in_progress']:>6} {stats['partial']:>6}  {sheets_str}")

    # Now filter to unbuilt modules
    print()
    print("=" * 90)
    print("TOP UNBUILT / LEAST-BUILT MODULES (excluding completed)")
    print("=" * 90)
    print()

    unbuilt = [
        (name, stats) for name, stats in sorted_modules
        if not is_completed(name) and stats["pending"] > 0
    ]

    # Sort by: most pending first, then least done
    unbuilt.sort(key=lambda x: (-x[1]["pending"], x[1]["done"]))

    print(f"{'#':<4} {'Module':<45} {'Total':>6} {'Pend':>6} {'Done':>6} {'WIP':>6} {'% Done':>7}  {'Sheet(s)'}")
    print("-" * 120)

    for i, (mod_name, stats) in enumerate(unbuilt[:20], 1):
        pct_done = (stats["done"] / stats["total"] * 100) if stats["total"] > 0 else 0
        sheets_str = ", ".join(sorted(stats["sheets"]))
        print(f"{i:<4} {mod_name:<45} {stats['total']:>6} {stats['pending']:>6} "
              f"{stats['done']:>6} {stats['in_progress']:>6} {pct_done:>6.1f}%  {sheets_str}")

    # Summary
    total_features = sum(s["total"] for s in module_stats.values())
    total_pending = sum(s["pending"] for s in module_stats.values())
    total_done = sum(s["done"] for s in module_stats.values())
    total_wip = sum(s["in_progress"] for s in module_stats.values())

    print()
    print("=" * 90)
    print("OVERALL SUMMARY")
    print("=" * 90)
    print(f"  Total modules found:    {len(module_stats)}")
    print(f"  Total features:         {total_features}")
    print(f"  Done:                   {total_done} ({total_done/total_features*100:.1f}%)")
    print(f"  In Progress:            {total_wip} ({total_wip/total_features*100:.1f}%)")
    print(f"  Pending:                {total_pending} ({total_pending/total_features*100:.1f}%)")
    print(f"  Unbuilt modules (with pending features, excl. completed): {len(unbuilt)}")
    print()

    # Recommendation
    print("=" * 90)
    print("RECOMMENDATION — NEXT MODULE TO BUILD")
    print("=" * 90)
    print()
    if unbuilt:
        top = unbuilt[:5]
        for i, (name, stats) in enumerate(top, 1):
            pct = stats["done"] / stats["total"] * 100 if stats["total"] > 0 else 0
            print(f"  {i}. {name}")
            print(f"     {stats['pending']} pending features, {stats['done']} done ({pct:.0f}%), "
                  f"total {stats['total']}")
            print(f"     Sheets: {', '.join(sorted(stats['sheets']))}")
            print()

    wb.close()


if __name__ == "__main__":
    main()

"""
Assign 'Apps' column to every data row in MedBrains_Features.xlsx.

Apps reflect realtime hospital deployment surfaces — each app is a distinct
ship target with role/location-specific UI:

  Web        : main React SPA (role-gated views)
  Desktop-*  : Tauri builds bound to physical location/hardware
  Mobile-*   : React Native role-specific apps
  TV-*       : Android TV location/audience displays

Output: comma-separated app list in column 12 of every data row.
Default writes to MedBrains_Features.PROPOSAL.xlsx for review.
Pass --apply to update MedBrains_Features.xlsx in place.
"""
from __future__ import annotations
import argparse
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict
from functools import lru_cache
import re
import shutil

SRC = 'MedBrains_Features.xlsx'
DST = 'MedBrains_Features.PROPOSAL.xlsx'

# ---------------------------------------------------------------- canonical apps
APPS = {
    # Web (single SPA, role-gated)
    'Web': 'Main browser SPA — clinicians, admin, all staff role-gated views',

    # Desktop (Tauri, location/hardware-bound)
    'Desktop-Kiosk':         'Patient self-service kiosk (lobby/OPD/IPD entry)',
    'Desktop-Workstation':   'Staff workstation w/ printers (reception/billing/pharmacy counter/lab/OT/ward)',
    'Desktop-DeviceBridge':  'Headless bridge for analyzer HL7, DICOM router, MGPS sensors, modality',

    # Mobile (RN, role-specific)
    'Mobile-Doctor':       'Physician — rounds, orders, e-Rx, notes',
    'Mobile-Nurse':        'Nursing — vitals, eMAR, rounds, audits',
    'Mobile-Patient':      'Patient self-service — appointments, records, payments',
    'Mobile-Phlebo':       'Phlebotomy/sample collection (in-house + home)',
    'Mobile-Pharmacist':   'Pharmacy counter / ward dispensing / cycle count',
    'Mobile-LabTech':      'Lab tech bench — sample receipt, analyzer link, results',
    'Mobile-Radiologist':  'Radiologist — reading, voice report, image review',
    'Mobile-Admin':        'CXO/manager — KPIs, approvals, alerts',
    'Mobile-BME':          'Biomedical engineer (CMMS) — assets, calibration, tickets',
    'Mobile-Housekeeping': 'Cleaning/turnaround/laundry',
    'Mobile-Security':     'Security guards — visitor pass, incident, code response',
    'Mobile-Driver':       'Ambulance driver / runner — dispatch, navigation, logs',
    'Mobile-Dietitian':    'Diet planning, kitchen rounds, tray verification',
    'Mobile-Student':      'Academic ERP — med student/intern logs',
    'Mobile-Faculty':      'Faculty — student supervision, attendance',

    # TV (Android TV, location/audience-specific)
    'TV-Queue':       'OPD waiting hall — token board',
    'TV-DoctorRoom':  'Small display outside consult room',
    'TV-Lab':         'Lab sample collection waiting',
    'TV-Pharmacy':    'Pharmacy counter waiting',
    'TV-Radiology':   'Modality waiting (X-ray/CT/MRI/USG)',
    'TV-Billing':     'Billing counter waiting',
    'TV-Emergency':   'ER triage board + code alert broadcast',
    'TV-Ward':        'Nursing station / IPD ward dashboard',
    'TV-ICU':         'ICU wall display + family info screen',
    'TV-OT':          'OT corridor status for waiting families',
    'TV-Wayfinding':  'Lobby/corridor directional signage',
    'TV-Cafeteria':   'Menu / diet display',
    'TV-Education':   'Patient education content (bedside, waiting)',
    'TV-Donor':       'Blood bank donor wall + drives',
    'TV-Notice':      'General hospital noticeboard',
}

MOBILE_PREFIX = 'Mobile-'
TV_PREFIX     = 'TV-'
DESKTOP_PREFIX= 'Desktop-'

# helpers ---------------------------------------------------------------------
def lower(s): return (s or '').lower()


@lru_cache(maxsize=4096)
def _kw_pattern(needle: str) -> re.Pattern:
    """Compile a keyword to a regex with word boundaries for short literals
    to avoid 'upi' matching 'occUPIed', 'rx ' matching 'thoRX ', etc.
    """
    # If keyword contains a space or punctuation, treat as multi-word phrase — plain substring is fine
    if any(c in needle for c in ' -&/'):
        return re.compile(re.escape(needle))
    # Short alphabetic literal — require word boundary
    return re.compile(r'(?<![a-z])' + re.escape(needle) + r'(?![a-z])')


def has(text, *needles):
    return any(_kw_pattern(n).search(text) for n in needles)

# ---------------------------------------------------------------- classifier
def classify(sheet, mod, sub, feat, web, mob, tv):
    """Return sorted list of canonical app codes for a single feature row."""
    mod_l  = lower(mod)
    sub_l  = lower(sub)
    feat_l = lower(feat)
    apps = set()

    # =========================================================== sheet-specific
    # Mobile Apps sheet — module column is authoritative
    if sheet == 'Mobile Apps':
        if 'doctor app' in mod_l:   apps.add('Mobile-Doctor')
        elif 'patient app' in mod_l: apps.add('Mobile-Patient')
        elif 'phlebo' in mod_l:      apps.add('Mobile-Phlebo')
        elif 'universal mobile' in mod_l:
            # shared infra — all mobile apps
            apps.update([a for a in APPS if a.startswith(MOBILE_PREFIX)])
        elif 'i18n' in mod_l:
            apps.update([a for a in APPS if a.startswith(MOBILE_PREFIX)])
        # add Web if W=Y on this row (some mobile features have a web admin counterpart)
        if web == 'Y': apps.add('Web')
        return _finalize(apps, web, mob, tv)

    # TV Displays & Queue sheet — sub-module is authoritative
    if sheet == 'TV Displays & Queue':
        if 'token generation' in sub_l or 'kiosk' in sub_l:
            apps.update(['Desktop-Kiosk', 'TV-Queue'])
        if 'opd display' in sub_l:        apps.add('TV-Queue')
        if 'doctor room' in sub_l:        apps.add('TV-DoctorRoom')
        if 'lab queue' in sub_l:          apps.add('TV-Lab')
        if 'pharmacy queue' in sub_l:     apps.add('TV-Pharmacy')
        if 'radiology queue' in sub_l:    apps.add('TV-Radiology')
        if 'er & ipd' in sub_l or 'er &' in sub_l:
            apps.update(['TV-Emergency', 'TV-Ward', 'TV-ICU'])
        if 'billing & analytics' in sub_l or 'billing' in sub_l:
            apps.add('TV-Billing')
        # Hardware & Signage — refine per feature, only blanket all TV-* for infra rows
        if 'hardware' in sub_l or 'signage' in sub_l:
            if has(feat_l, 'mobile','push','app'):
                apps.add('Mobile-Patient')  # mobile push for queue position
            elif has(feat_l, 'wayfinding','directional','navigation'):
                apps.add('TV-Wayfinding')
            elif has(feat_l, 'cms','template','content management',
                            'multi-display','cluster','grouping','schedule','playlist'):
                apps.update([a for a in APPS if a.startswith(TV_PREFIX)])
            elif has(feat_l, 'kiosk'):
                apps.update({'Desktop-Kiosk','TV-Queue'})
            elif has(feat_l, 'token print','token printer','thermal printer'):
                apps.update({'Desktop-Kiosk','TV-Queue'})
            else:
                # Hardware row touching display infra without specifics — keep narrow
                apps.add('TV-Queue')
        if 'i18n' in (mod_l + sub_l):
            apps.update([a for a in APPS if a.startswith(TV_PREFIX)])
        if has(feat_l, 'kiosk'):     apps.add('Desktop-Kiosk')
        if has(feat_l, 'wayfinding','directional','navigation board'):
            apps.add('TV-Wayfinding')
        if has(feat_l, 'cafeteria','menu'):
            apps.add('TV-Cafeteria')
        if has(feat_l, 'education','health tips'):
            apps.add('TV-Education')
        if has(feat_l, 'donor wall','blood drive'):
            apps.add('TV-Donor')
        if web == 'Y' or has(feat_l, 'cms','config','admin'):
            apps.add('Web')
        return _finalize(apps, web, mob, tv)

    # Printing & Forms sheet — every print job touches Desktop-Workstation
    if sheet == 'Printing & Forms':
        apps.add('Web')                  # designer/template/audit
        apps.add('Desktop-Workstation')  # the actual printer
        # token/queue printing also Desktop-Kiosk + TV-Queue board
        if has(feat_l, 'token', 'kiosk'):
            apps.update(['Desktop-Kiosk', 'TV-Queue'])
        # wristband — also Mobile-Nurse (scan at bedside)
        if has(feat_l, 'wristband','barcode wristband'):
            apps.add('Mobile-Nurse')
        # patient education printouts
        if has(feat_l, 'pictogram','take-home','discharge instructions','patient education'):
            apps.add('Mobile-Patient')
        # signage / branding printing
        if has(feat_l, 'signage','branding','poster','banner'):
            apps.update([a for a in APPS if a.startswith(TV_PREFIX)])
        # MGPS / O2 daily log → BME mobile
        if has(feat_l, 'mgps','o2 purity','medical gas'):
            apps.add('Mobile-BME')
        # rx pictograms / take-home med list — also Mobile-Pharmacist
        if has(feat_l, 'pharmacy label','drug label','dispense label'):
            apps.add('Mobile-Pharmacist')
        return _finalize(apps, web, mob, tv)

    # CMS & Blog sheet — Web admin + public site, sometimes mobile-patient news feed
    if sheet == 'CMS & Blog':
        apps.add('Web')
        if has(feat_l, 'mobile','app','push'):
            apps.add('Mobile-Patient')
        return _finalize(apps, web, mob, tv)

    # Onboarding & Setup — Web admin only (system setup wizard)
    if sheet == 'Onboarding & Setup':
        apps.add('Web')
        if has(feat_l, 'mobile','app'):
            apps.update(['Mobile-Doctor','Mobile-Nurse','Mobile-Patient'])
        if has(feat_l, 'kiosk'):  apps.add('Desktop-Kiosk')
        if has(feat_l, 'tv','display','signage'):
            apps.update([a for a in APPS if a.startswith(TV_PREFIX)])
        return _finalize(apps, web, mob, tv)

    # Technical Infrastructure — primary surfaces only (don't blanket-add all apps)
    if sheet == 'Technical Infrastructure':
        if web == 'Y':  apps.add('Web')
        if mob == 'Y':
            if 'offline sync' in mod_l or has(feat_l, 'offline','crdt','sync','watermelondb'):
                apps.update({'Mobile-Doctor','Mobile-Nurse','Mobile-Patient','Mobile-Phlebo'})
            elif has(feat_l, 'push notification','fcm','apns','silent push'):
                apps.update({'Mobile-Doctor','Mobile-Nurse','Mobile-Patient'})
            elif has(feat_l, 'i18n','locale','rtl','translation'):
                apps.update({'Mobile-Doctor','Mobile-Nurse','Mobile-Patient'})
            elif has(feat_l, 'biometric','face id','fingerprint'):
                apps.update({'Mobile-Doctor','Mobile-Nurse'})
            elif has(feat_l, 'whatsapp','sms','email'):
                apps.update({'Mobile-Patient'})
            else:
                apps.update({'Mobile-Doctor','Mobile-Patient'})
        if tv == 'Y':
            if has(feat_l, 'queue','token'):
                apps.add('TV-Queue')
            elif has(feat_l, 'realtime','websocket','live'):
                apps.update({'TV-Queue','TV-Ward','TV-ICU'})
            elif has(feat_l, 'i18n','locale','translation'):
                apps.update({'TV-Queue','TV-Ward'})
            else:
                apps.add('TV-Queue')
        if 'device integration' in mod_l or has(feat_l, 'hl7','dicom','astm','modality'):
            apps.add('Desktop-DeviceBridge')
        return _finalize(apps, web, mob, tv)

    # ============================================== Clinical / Diagnostics / Admin / Specialty / IT
    # PER-ROW mode — no module category baseline.
    # Only sub-module + feature text + W/M/T flags drive the assignment.
    _per_row_scan(apps, sub_l, feat_l, web, mob, tv)

    # safety net — if nothing matched at all, fall back to Web only (when W=Y)
    if not apps and web == 'Y':
        apps.add('Web')

    return _finalize(apps, web, mob, tv)


def _module_base(mod_l, sub_l, feat_l, sheet):
    """Default app set per known module name across non-special sheets."""
    # ---- Clinical
    if mod_l == 'opd':
        return {'Web','Mobile-Doctor','Mobile-Patient','TV-Queue','TV-DoctorRoom'}
    if mod_l == 'ipd':
        s = {'Web','Mobile-Doctor','Mobile-Nurse','TV-Ward'}
        if 'operation theatre' in sub_l or 'ot' in sub_l:
            s.update({'TV-OT','Desktop-Workstation'})
        if 'discharge' in sub_l:
            s.update({'Desktop-Workstation','Mobile-Patient'})
        if 'bed' in sub_l or 'ward' in sub_l:
            s.update({'TV-Ward','Mobile-Housekeeping'})
        if 'medication' in sub_l or 'emar' in sub_l:
            s.add('Mobile-Pharmacist')
        return s
    if mod_l == 'icu':
        return {'Web','Mobile-Doctor','Mobile-Nurse','TV-ICU'}
    if mod_l == 'patient mgmt':
        return {'Web','Mobile-Patient','Desktop-Kiosk','Desktop-Workstation'}
    if mod_l == 'ai documentation':
        return {'Web','Mobile-Doctor'}
    if mod_l == 'bedside portal':
        return {'Web','TV-Education','Mobile-Patient'}
    if mod_l == 'care view':
        return {'Web','Mobile-Doctor','Mobile-Nurse','TV-Ward'}
    if mod_l == 'chronic care':
        return {'Web','Mobile-Doctor','Mobile-Patient'}
    if mod_l == 'infection surv.':
        return {'Web','Mobile-Nurse'}
    if mod_l == 'order sets':
        return {'Web','Mobile-Doctor'}
    if mod_l == 'retrospective':
        return {'Web'}

    # ---- Diagnostics & Support
    if mod_l == 'ambulance':
        return {'Web','Mobile-Driver','Mobile-Nurse'}
    if mod_l == 'blood bank':
        return {'Web','Mobile-LabTech','Desktop-Workstation','TV-Donor'}
    if mod_l == 'cssd':
        return {'Web','Mobile-Nurse','Desktop-Workstation'}
    if mod_l == 'diet & kitchen':
        return {'Web','Mobile-Dietitian','Mobile-Nurse','Desktop-Workstation','TV-Cafeteria'}
    if mod_l == 'genetics':
        return {'Web','Mobile-Doctor','Mobile-LabTech'}
    if mod_l == 'inventory':
        return {'Web','Mobile-Pharmacist','Mobile-BME','Desktop-Workstation'}
    if mod_l == 'laboratory':
        s = {'Web','Mobile-LabTech','Mobile-Phlebo','Mobile-Doctor','TV-Lab'}
        if 'sample' in sub_l: s.add('Desktop-Workstation')
        if 'b2b' in sub_l or 'referral' in sub_l: s.add('Mobile-Admin')
        if 'results' in sub_l or 'report' in sub_l:
            s.update({'Mobile-Doctor','Mobile-Patient'})
        return s
    if mod_l == 'pharmacy':
        return {'Web','Mobile-Pharmacist','Desktop-Workstation','TV-Pharmacy'}
    if mod_l == 'radiology':
        return {'Web','Mobile-Radiologist','Mobile-Doctor','Desktop-DeviceBridge','TV-Radiology'}
    if mod_l == 'rx benefits':
        return {'Web','Mobile-Patient'}

    # ---- Admin & Operations
    if mod_l == 'bme/cmms':
        return {'Web','Mobile-BME'}
    if mod_l == 'billing':
        return {'Web','Desktop-Workstation','Mobile-Patient','Mobile-Admin','TV-Billing'}
    if mod_l == 'camp mgmt':
        return {'Web','Mobile-Doctor','Mobile-Nurse','Desktop-Workstation'}
    if mod_l == 'case mgmt':
        return {'Web','Mobile-Doctor','Mobile-Nurse'}
    if mod_l == 'facilities':
        return {'Web','Mobile-Housekeeping','Mobile-BME'}
    if mod_l == 'front office':
        return {'Web','Desktop-Kiosk','Desktop-Workstation','TV-Wayfinding','TV-Queue'}
    if mod_l == 'hr':
        return {'Web','Mobile-Admin'}
    if mod_l == 'housekeeping':
        return {'Web','Mobile-Housekeeping','TV-Ward'}
    if mod_l == 'infection ctrl':
        return {'Web','Mobile-Nurse'}
    if mod_l == 'insurance':
        return {'Web','Mobile-Patient','Mobile-Admin'}
    if mod_l == 'mrd':
        return {'Web','Desktop-Workstation'}
    if mod_l == 'occ health':
        return {'Web','Mobile-Nurse','Mobile-Doctor'}
    if mod_l == 'quality':
        return {'Web','Mobile-Admin','Mobile-Nurse'}
    if mod_l == 'scheduling':
        return {'Web','Mobile-Doctor','Mobile-Nurse','Mobile-Admin'}
    if mod_l == 'security dept':
        return {'Web','Mobile-Security','Desktop-Workstation','TV-Emergency'}
    if mod_l == 'utilization':
        return {'Web','Mobile-Admin'}

    # ---- Specialty & Academic
    if mod_l == 'academic erp':
        return {'Web','Mobile-Student','Mobile-Faculty','Mobile-Admin'}
    if mod_l == 'clinical trials':
        return {'Web','Mobile-Doctor','Mobile-Patient'}
    if mod_l == 'communication':
        return {'Web','Mobile-Doctor','Mobile-Nurse','Mobile-Patient'}
    if mod_l == 'consent':
        return {'Web','Mobile-Patient','Desktop-Workstation'}
    if mod_l == 'correctional':
        return {'Web','Mobile-Nurse','Mobile-Security'}
    if mod_l == 'emergency':
        return {'Web','Mobile-Doctor','Mobile-Nurse','Mobile-Driver','TV-Emergency'}
    if mod_l == 'home care':
        return {'Web','Mobile-Nurse','Mobile-Doctor','Mobile-Patient'}
    if mod_l == 'home dialysis':
        return {'Web','Mobile-Nurse','Mobile-Patient','Desktop-DeviceBridge'}
    if mod_l == 'home infusion':
        return {'Web','Mobile-Nurse','Mobile-Patient'}
    if mod_l == 'hospice':
        return {'Web','Mobile-Nurse','Mobile-Doctor','Mobile-Patient'}
    if mod_l == 'hospital@home':
        return {'Web','Mobile-Doctor','Mobile-Nurse','Mobile-Patient'}
    if mod_l == 'micro website':
        return {'Web','Mobile-Patient'}
    if mod_l == 'post-acute':
        return {'Web','Mobile-Doctor','Mobile-Nurse'}
    if mod_l == 'specialty':
        return {'Web','Mobile-Doctor','Mobile-Nurse'}
    if mod_l == 'telemedicine':
        return {'Web','Mobile-Doctor','Mobile-Patient'}
    if mod_l == 'transplant':
        return {'Web','Mobile-Doctor','Mobile-Nurse','Mobile-LabTech'}
    if mod_l == 'urology':
        return {'Web','Mobile-Doctor','Mobile-Nurse'}

    # ---- IT, Security & Infrastructure
    if mod_l in ('ai comms','ai ops','ai governance'):
        return {'Web','Mobile-Admin'} if mod_l == 'ai governance' else {'Web'}
    if mod_l == 'admin':
        return {'Web','Mobile-Admin'}
    if mod_l == 'analytics builder':
        return {'Web','Mobile-Admin'}
    if mod_l == 'audit':
        return {'Web'}
    if mod_l == 'crm':
        return {'Web','Mobile-Patient','Mobile-Admin'}
    if mod_l == 'care transitions':
        return {'Web','Mobile-Nurse','Mobile-Doctor'}
    if mod_l == 'command center':
        return {'Web','Mobile-Admin','TV-Ward','TV-ICU'}
    if mod_l == 'compliance':
        return {'Web','Mobile-Admin'}
    if mod_l == 'dashboards':
        return {'Web','Mobile-Admin'}
    if mod_l == 'data migration':
        return {'Web'}
    if mod_l == 'data quality':
        return {'Web','Mobile-Admin'}
    if mod_l == 'device integration':
        return {'Web','Desktop-DeviceBridge'}
    if mod_l == 'digital front door':
        return {'Web','Desktop-Kiosk','Mobile-Patient','TV-Wayfinding'}
    if mod_l == 'eod digest':
        return {'Web','Mobile-Admin','Mobile-Doctor'}
    if mod_l == 'encryption':
        return {'Web'}
    if mod_l == 'external platforms':
        return {'Web'}
    if mod_l == 'incentive':
        return {'Web','Mobile-Doctor','Mobile-Admin'}
    if mod_l == 'inflow analytics':
        return {'Web','Mobile-Admin'}
    if mod_l == 'integrations':
        return {'Web','Desktop-DeviceBridge'}
    if mod_l == 'network':
        return {'Web'}
    if mod_l == 'onboarding':
        return {'Web'}
    if mod_l == 'population health':
        return {'Web','Mobile-Admin'}
    if mod_l == 'sdoh':
        return {'Web','Mobile-Nurse'}
    if mod_l == 'security':
        return {'Web','Mobile-Security'}
    if mod_l == 'stock disposal':
        return {'Web','Mobile-Pharmacist','Mobile-Housekeeping'}
    if mod_l == 'tat':
        return {'Web','Mobile-Admin'}
    if mod_l == 'wearables':
        return {'Web','Mobile-Patient'}
    if mod_l == 'workflow':
        return {'Web'}

    # ---- Regulatory & Compliance
    if sheet == 'Regulatory & Compliance':
        s = {'Web'}
        if mod_l == 'nabh' or mod_l == 'jci':
            s.update({'Mobile-Admin','Mobile-Nurse'})
        if mod_l == 'geospatial':
            s.update({'Mobile-Patient'})
        if mod_l == 'abdm' or mod_l == 'aasandhu' in mod_l:
            s.update({'Mobile-Patient'})
        if mod_l == 'pcpndt':
            s.update({'Desktop-Workstation'})
        if mod_l == 'aerb':
            s.update({'Mobile-BME','Mobile-Radiologist'})
        if mod_l == 'cdsco':
            s.update({'Mobile-Pharmacist'})
        if mod_l == 'fire':
            s.update({'Mobile-Security','TV-Emergency'})
        return s

    # ---- Multi-Hospital & Vendors
    if sheet == 'Multi-Hospital & Vendors':
        s = {'Web','Mobile-Admin'}
        if mod_l == 'ext lab':       s.update({'Mobile-LabTech','Mobile-Phlebo','Desktop-DeviceBridge'})
        if mod_l == 'ext pharmacy':  s.update({'Mobile-Pharmacist'})
        if mod_l == 'ext services':  s.update({'Mobile-Nurse'})
        if mod_l == 'teleradiology': s.update({'Mobile-Radiologist','Desktop-DeviceBridge'})
        if mod_l == 'vendors':       s.update({'Mobile-BME'})
        return s

    # default: just Web
    return {'Web'}


def _per_row_scan(apps, sub_l, feat_l, web, mob, tv):
    """Per-row classifier — scan Sub-Module + Feature text together.

    Treats Sub-Module as row-level signal (it varies row-to-row).
    No Module-level category default — each row stands on its own text.
    """
    t = (sub_l or '') + ' :: ' + (feat_l or '')

    # Web is the universal terminal — if W=Y in this row, always include Web
    if web == 'Y':
        apps.add('Web')

    # ------------------------------------------------------------ DESKTOP
    if has(t, 'kiosk','self-service','self check-in','walk-in kiosk',
              'walk-in token','walk-in registration','self-registration','self check'):
        apps.add('Desktop-Kiosk')
    if has(t, 'print','printer','wristband','label print','token print','thermal',
              'barcode print','receipt','sticker','copy print',
              'bedside terminal','bedside tablet','bedside device','bedside portal'):
        apps.add('Desktop-Workstation')
    if has(t, 'hl7','astm','dicom','modality','plc','sensor','iot gateway',
              'analyzer interface','device integration','mgps sensor','bms integration',
              'biometric attendance device'):
        apps.add('Desktop-DeviceBridge')

    # ------------------------------------------------------------ MOBILE roles
    # Doctor — consultation, prescription, clinical notes, rounds, orders, AI clinical aid, specialty pathways
    if has(t, 'doctor','physician','consultant','consultation','prescription','e-prescription',
              'e-rx','soap','progress note','round note','clinical note','rx','differential diagnosis',
              'discharge summary','order set','clinical decision','e-mar review',
              'ai clinical coding','ai differential','ai-drafted','ai-assisted discharge',
              'ai consultation','ambient capture','consultation summary',
              'family history','medical history','past history','allergies','immunization',
              'icd-10','cpt code','snomed','procedure code',
              'package appointment','specialist referral','inter-department referral',
              'referral order','referral letter','outgoing referral',
              'stemi','door-to-balloon','door to balloon','cath lab','hemodynamic',
              'stent','transplant','candidate evaluation','multi-disciplinary',
              'tumor board','mdt meeting','case conference','grand rounds',
              'palliative','hospice enrollment','prognosis','symptom assessment','dnr',
              'comfort care plan','advance directive','transplant committee',
              'home visit doctor','home consult','tele-icu','clinical pathway',
              'cdss','clinical guideline','treatment protocol','protocol-driven order',
              'usg','ultrasound','sonography','doppler','echo report','ecg interpret',
              'mammography','mammogram','ct scan','mri scan','pet scan','pet-ct',
              'fluoroscopy','contrast study','angiogram','bronchoscopy','colonoscopy',
              'endoscopy','laparoscopy','arthroscopy',
              'pregnancy verification','pregnancy screen','beta hcg',
              'audiometric','hearing test','vision test','optometry test',
              'pulmonary function','spirometry','treadmill test','holter monitor',
              'eeg ','emg ','nerve conduction',
              'visit schedule trial','protocol-driven visit','crf entry',
              'return-to-work clearance','fitness for duty','medical board',
              'occupational fitness','pre-employment medical',
              'substance abuse','addiction medicine','de-addiction',
              'prior auth','prior authorization detection','pa required',
              'rounding worklist','round signoff','round documentation',
              'admit order','admission order','transfer order','er to ipd transfer',
              'opd worklist','consultation queue','my patients today',
              'critical lab alert','panic value','panic value alert','crit value',
              'procedure ordering','order procedure','procedure entry','procedure plan',
              'pediatric dosage','weight-based dosage','dose calculator',
              'dental chart','tooth chart','orthodontic','cdt coding','endodontic',
              'periodontic','prosthodontic','dental procedure',
              'pathway deviation','journey deviation','clinical pathway deviation',
              'hereditary risk','genetic counseling','brca screening','lynch syndrome',
              'familial hypercholesterolemia','pedigree chart','genetic risk',
              'wound certificate','injury certificate','medical certificate generation',
              'fitness certificate','disability certificate',
              'video consent capture','video consent recording','recorded consent',
              'critical patient identification','sepsis screen','sepsis alert',
              'autism screen','developmental screen','growth chart','growth percentile',
              'immunization schedule','vaccine due','vaccine catch-up'):
        apps.add('Mobile-Doctor')

    # Nurse — vitals, eMAR, rounds, audits, bedside, shift handover, ward, IC, home, scoring
    if has(t, 'nurs','vitals','emar','medication administration','medication admin',
              'bedside','rounds','triage','fall risk','hand hygiene','pressure ulcer',
              'wound care','iv','intake output','i&o','catheter','restraint','isolation precaution',
              'shift handover','sbar','shift report','handover note','shift conclusion',
              'critical patient flag','incoming shift','call bell','bed turnaround order',
              'patient deterioration',
              'audit checklist','floor audit','infection control audit','5moments audit',
              'ssi audit','clabsi audit','vap audit','cauti audit','hai surveillance',
              'occupational hazard','needle stick','sharp injury','exposure protocol',
              'incident report card','medication error report','near miss report',
              'home visit','home care','wound photo','dressing change',
              'physiotherapy','occupational therapy','speech therapy','enteral feed nurse',
              'home nursing','community nursing',
              'gcs','glasgow coma','coma scale','aldrete','aldrete score','braden',
              'mews','news2','early warning','pain score','pain assessment',
              'i/o chart','fluid balance','iv site','iv start','iv removal',
              'sedation monitoring','recovery monitoring','post-op recovery',
              'enteral feed','parenteral feed','tube feed','rt feed',
              'chemical indicator','sterile pack','sterilization indicator',
              'autoclave load','cssd nurse','sterilizer load',
              'risk assessment','frailty score','morse fall','humpty dumpty',
              'iv access','peripheral line','central line','picc line',
              'catheter care','drain output','colostomy care','tracheostomy care',
              'contact precaution','droplet precaution','airborne precaution',
              'bowles indicator','helix test','bowie dick test',
              'restraint use','restraint monitoring','restraint reassessment',
              'point-of-care testing','poct ','glucose monitoring','blood sugar check',
              'admission form','admission checklist','admission process','admission worklist',
              'pending task','task carryover','task handoff','shift carryover',
              'cross-department handoff','handoff automation','clinical handoff',
              'postnatal care','antenatal care','partogram','labour monitoring','labor monitoring',
              'maternity nursing','obstetric nursing','breastfeeding support','lactation',
              'flash sterilization','flash autoclave','load release','sterilizer release',
              'cssd worklist','sterile pack issue','sterile pack receive',
              'screening tool','risk screening','admission screening',
              'pediatric nursing','neonatal care','kangaroo mother','nicu nursing'):
        apps.add('Mobile-Nurse')

    # Patient self-service — appointments, records, payments, app, patient-facing content
    if has(t, 'patient app','patient portal','patient self','book appointment',
              'appointment booking','health record','phr','pay bill','upi','wallet',
              'rate hospital','patient feedback','patient survey','prescription refill request',
              'patient education','patient consent','family member linking','patient-facing',
              'patient timeline','health locker','abdm health','medication reminder','wellness',
              'patient-friendly','for patient','for referral','for the patient','patient view',
              'discharge instructions','take-home','adherence','drug-o-gram','medication timeline',
              'live queue','wait time','wait time display','show on app','on website and app',
              'on website','family members from bedside','video call with family',
              'message nurse','request assistance','bedside terminal',
              'video consult','telemedicine','tele-consult',
              'sms gateway','whatsapp','push notification','appointment reminder',
              'report ready','bill generated','follow-up reminder','medication compliance',
              'patient signature','e-sign','informed consent','signed consent','consent signing',
              'patient communication','outbound sms','outbound notification',
              'ivr','voice call patient','call patient',
              'apple health','google fit','fitbit','garmin','wearable','smartwatch',
              'bluetooth bp','bluetooth glucose','glucometer integration','pulse oximeter',
              'weight scale integration','step count','heart rate variability',
              'sleep tracking','activity tracking','symptom checker','self-triage',
              'appointment rescheduling','recurring appointment','recurring booking',
              'reschedule appointment','cancel appointment','reschedule booking',
              'health package','wellness package','online package purchase',
              'package booking','package payment','online payment',
              'health card','membership card','loyalty program',
              'feedback survey','patient nps','complaint submission',
              'video consent','recorded consent','consent video'):
        apps.add('Mobile-Patient')

    # Phlebotomist — sample collection
    if has(t, 'phlebotom','sample collect','home collection','sample pickup',
              'sample handover','chain of custody','runner app'):
        apps.add('Mobile-Phlebo')

    # Pharmacist — dispensing, formulary, inventory, stewardship, implant/stent
    if has(t, 'pharmac','dispens','formulary','drug interaction','prescription verify',
              'medication reconciliation','medication recon','dose check','aware','lasa',
              'narcotic register','controlled drug register','crash cart','indent','par level',
              'dead stock','slow-moving','fast-moving','non-moving','obsolete stock',
              'batch tracking','lot tracking','batch number','expiry tracking','fefo',
              'consignment stock','vendor-owned','implant tracking','stent tracking',
              'inventory turnover','reorder point','min-max','abc analysis','ved analysis',
              'sap analysis','stock ledger','grn ','gate receipt','indent generation',
              'stock adjustment','stock count','cycle count','physical verification',
              'pharmacy worklist','dispense worklist','ward indent','iv admixture',
              'compounding','drug formulary','formulary master','atc code','aware classification',
              'antimicrobial stewardship','ams ','antibiotic time-out','de-escalation',
              'multi-store','multi-location store','satellite store','sub-store','main store',
              'store management','warehouse management','warehouse worklist',
              'manual result entry','result entry worklist','autoverification result',
              'near-expiry','expiry alert','expiry tracking','expiring soon','expiry dashboard',
              'return/unused drug','unused drug return','drug return','credit note pharmacy',
              'purchase order','po generation','po approval','grn entry','goods receipt',
              'purchase vs consumption','consumption trend','consumption analysis',
              'cycle count','physical stock','stock take','stock verification',
              'drug substitution','generic substitution','therapeutic substitution'):
        apps.add('Mobile-Pharmacist')

    # Lab tech — bench, analyzer link, accession, result entry
    if has(t, 'lab tech','laboratory tech','sample receipt','sample accession',
              'analyzer worklist','autoverif','quality control lab','qc rule',
              'westgard','reagent','calibrator','levey-jennings',
              'manual result entry','result entry worklist','test result entry',
              'lab bench worklist','specimen worklist','panel verification',
              'pre-analytical','post-analytical','tat lab'):
        apps.add('Mobile-LabTech')

    # Radiologist — PACS, reading, voice report, modality-specific reads
    if has(t, 'radiologist','pacs','dicom viewer','radiology read','image review',
              'voice report','structured report','hanging protocol','mammograph','reporting workstation',
              'usg report','sonography report','doppler report','echo report',
              'ct report','mri report','x-ray report','plain film report',
              'mammography report','pet-ct report','nuclear scan report',
              'image annotation','image measurement','window leveling',
              'teleradiology read','remote reporting','radiology approval'):
        apps.add('Mobile-Radiologist')

    # Admin / CXO / Manager / HR / Quality leadership
    if has(t, 'ceo dashboard','board pack','board report','exec dashboard','executive',
              'admin dashboard','kpi','mis ','management info','revenue dashboard',
              'occupancy dashboard','strategic','tariff','rate card management','vendor approval',
              'departmental p&l','employee master','duty roster','staff attendance',
              'leave management','payroll','appraisal','performance review','recruitment',
              'credentialing','employee onboarding','hr policy','disciplinary',
              'quality indicator','quality dashboard','clinical indicator','operational indicator',
              'patient safety indicator','nabh dashboard','jci dashboard','accreditation',
              'incident report','sentinel event','rca','root cause analysis',
              'insurance panel','tpa management','pre-authorization','claims management',
              'financial mis','revenue analytics','collection efficiency','margin tracking',
              'overtime calculation','overtime claim','shift differential','attendance regularization',
              'leave balance','leave approval','leave application',
              'referral commission','commission management','consultant payout',
              'doctor incentive','doctor payout','panel doctor payment',
              'ar aging','collection report','daily collection','monthly collection',
              'cash flow','expense tracking','budget vs actual','financial close',
              'contract management','vendor master','vendor onboard','vendor payment cycle',
              'complaint trend','grievance trend','feedback analytics',
              'patient experience score','nps trend','satisfaction trend',
              'department p&l','department performance','staff productivity',
              'utilization report','occupancy report','revenue per bed',
              'cost center','profit center','financial reconciliation',
              'staff scheduling overview','shift coverage','staffing gap'):
        apps.add('Mobile-Admin')

    # BME — biomedical engineer (CMMS), medical gas plant, generator/UPS/HVAC
    if has(t, 'biomedical','bme','asset tag','calibration','preventive maintenance','pms ',
              'breakdown ticket','equipment maintenance','equipment service','condemnation',
              'amc tracking','equipment downtime','spare part',
              'equipment master','equipment database','equipment categor','equipment lifecycle',
              'asset register','asset tagging','asset master','asset categor',
              'medical device register','medical equipment','equipment audit','equipment commissioning',
              'warranty tracking','warranty management','infusion pump service','ventilator service',
              'amc ','cmc ','mgps','o2 plant','lmo tank','manifold','cylinder','gas plant',
              'generator','dg set','ups maintenance','hvac','bms ','chiller','boiler','rwh ','stp ',
              'pm history','preventive maintenance history','maintenance history',
              'water consumption','energy consumption','utility monitoring','utility consumption',
              'recycled water','rainwater harvesting','solar power','solar generation',
              'wte ','waste to energy','etp ','effluent treatment',
              'electrical maintenance','plumbing maintenance','civil maintenance',
              'crash trolley check','equipment inspection','fire equipment check',
              'fire safety audit','electrical safety audit',
              'asset depreciation','equipment depreciation','useful life tracking'):
        apps.add('Mobile-BME')

    # Housekeeping — cleaning, linen, turnaround, sanitation, 5S
    if has(t, 'housekeep','cleaning','bed turnaround','linen','laundry',
              'pest control','soiled','deep clean','terminal clean',
              'sanitation','room cleaning','floor cleaning','area cleaning',
              'cleaning schedule','cleaning audit','cleaning staff','janitor',
              'disinfect','5s ','sweep','mop','waste segregation','bmw segreg'):
        apps.add('Mobile-Housekeeping')

    # Security guard — physical security only (not software RBAC/ABAC)
    if has(t, 'visitor pass','visitor mgmt','panic button','duress','security guard',
              'patrol','perimeter','cctv','incident response','badge print',
              'gate pass','vehicle pass','frisking','intrusion',
              'physical access','door access','card access','biometric door',
              'zone definition','security zone','restricted zone','semi-restricted',
              'guard duty','fire drill','disaster drill','code response',
              'lockdown drill','emergency evacuation','crowd control'):
        apps.add('Mobile-Security')

    # Driver — ambulance, runner, dispatch
    if has(t, 'ambulance','driver','dispatch','route to patient','vehicle tracking',
              'emergency vehicle','runner'):
        apps.add('Mobile-Driver')

    # Dietitian — diet plan, kitchen rounds
    if has(t, 'dietitian','diet plan','nutrition counsel','calorie','meal plan',
              'therapeutic diet','feed schedule','enteral feed','parenteral feed','rt feed'):
        apps.add('Mobile-Dietitian')

    # Student / Intern / Faculty / Academic
    if has(t, 'student','intern','postgrad','resident','duty roster student',
              'medical college','attendance student','curriculum','rotation log','elog book',
              'cbme','competency mapping','formative assessment','summative assessment',
              'hostel','mess management','mess account','accommodation allocation',
              'library','e-resource','book issue','book return','journal access',
              'inflibnet','nlist','digital library',
              'thesis','dissertation','research proposal','plagiarism check',
              'examination','exam schedule','internal assessment','university exam',
              'clinical posting','rotation schedule','log book','case log','procedure log',
              'student grievance','student attendance','class attendance'):
        apps.add('Mobile-Student')
    if has(t, 'faculty','teacher','supervisor signoff','assessor','examiner',
              'thesis review','clinical posting plan','syllabus management',
              'faculty appraisal','faculty profile','faculty assignment','phd guide',
              'guide allocation','mentor assignment','academic supervision',
              'conference','abstract submission','cme certificate','cme credit',
              'medical education','peer review academic','journal club'):
        apps.add('Mobile-Faculty')

    # ------------------------------------------------------------ TV displays
    if has(t, 'token board','token display','queue display','queue token',
              'live queue','opd queue','opd display','waiting display',
              'walk-in token','token generation','token call','call next token'):
        apps.add('TV-Queue')
    if has(t, 'doctor room display','consultation room display','outside room',
              'next-token','room indicator'):
        apps.add('TV-DoctorRoom')
    if has(t, 'lab queue','lab waiting','lab token','sample collection queue',
              'phlebotomy queue','phlebotomy waiting'):
        apps.add('TV-Lab')
    if has(t, 'pharmacy queue','pharmacy waiting','pharmacy token'):
        apps.add('TV-Pharmacy')
    if has(t, 'radiology queue','modality queue','x-ray queue','ct queue','mri queue'):
        apps.add('TV-Radiology')
    if has(t, 'billing queue','billing counter display','billing waiting'):
        apps.add('TV-Billing')
    if has(t, 'er display','er triage board','triage board','code red','code blue',
              'code pink','code black','mass casualty board','disaster declaration'):
        apps.add('TV-Emergency')
    if has(t, 'ward dashboard','nursing station display','ipd ward display',
              'bed status board','ward board','bed occupancy display'):
        apps.add('TV-Ward')
    if has(t, 'icu wall','icu family display','icu visitor info','icu dashboard display'):
        apps.add('TV-ICU')
    if has(t, 'ot status','ot board','ot family display','surgical schedule board','ot corridor'):
        apps.add('TV-OT')
    if has(t, 'wayfinding','directional signage','indoor nav','lobby signage',
              'floor map','navigation display'):
        apps.add('TV-Wayfinding')
    if has(t, 'cafeteria','canteen display','menu display','meal display'):
        apps.add('TV-Cafeteria')
    if has(t, 'patient education','health tips','wellness video','education video',
              'bedside education','condition-specific video','post-op care video','medication guide video'):
        apps.add('TV-Education')
    if has(t, 'donor wall','blood drive display','donor appreciation'):
        apps.add('TV-Donor')
    if has(t, 'noticeboard','hospital news display','event display','nabh visit prep board'):
        apps.add('TV-Notice')

    # ------------------------------------------------------------ Telemedicine / shared
    if has(t, 'telemedicine','video consultation','video consult','tele-consult'):
        apps.update({'Mobile-Doctor','Mobile-Patient'})

    # ------------------------------------------------------------ WEB triggers
    # Any of these very-common admin/clinical-Web patterns add Web
    if has(t, 'dashboard','report','config','setting','admin','manage','search',
              'list','export','audit','register','calendar','workflow','approval',
              'integration','api','log','history','view','enrollment','onboard','migration',
              'designer','builder','template','master','tariff','panel','document control',
              'compliance','indicator','quality indicator','nabh','jci','rule engine','ml model'):
        apps.add('Web')

    # ------------------------------------------------------------ Sub-Module-level signal (per-row)
    # Use has() for word-boundary matching to avoid 'icu' matching 'currICUlum' etc.
    if has(sub_l, 'opd') and has(sub_l, 'consult'):
        apps.update({'Mobile-Doctor','Web'})
    if has(sub_l, 'opd') and has(sub_l, 'appointment'):
        apps.update({'Mobile-Doctor','Mobile-Patient','TV-Queue','Web'})
    if has(sub_l, 'ward & bed','bed mgmt','room & bed'):
        apps.update({'Mobile-Nurse','TV-Ward','Mobile-Housekeeping','Web'})
    if has(sub_l, 'operation theatre','ot') or sub_l.startswith('ot '):
        apps.update({'Mobile-Doctor','Mobile-Nurse','TV-OT','Desktop-Workstation','Web'})
    if has(sub_l, 'icu','nicu','picu'):
        apps.update({'Mobile-Doctor','Mobile-Nurse','TV-ICU','Web'})
    if has(sub_l, 'discharge'):
        apps.update({'Desktop-Workstation','Mobile-Patient','Web'})
    if has(sub_l, 'sample mgmt','sample management'):
        apps.update({'Mobile-Phlebo','Mobile-LabTech','Desktop-Workstation','Web'})
    if (has(sub_l, 'results','reporting')) and has(sub_l, 'lab','laboratory'):
        apps.update({'Web','Mobile-Doctor','Mobile-Patient'})
    if has(sub_l, 'dispens','dispensing'):
        apps.update({'Mobile-Pharmacist','Desktop-Workstation','TV-Pharmacy','Web'})
    if has(sub_l, 'formulary'):
        apps.update({'Mobile-Pharmacist','Web'})
    if has(sub_l, 'reading') and has(sub_l, 'rad','radiology'):
        apps.update({'Mobile-Radiologist','Web'})
    if has(sub_l, 'triage'):
        apps.update({'Mobile-Nurse','Mobile-Doctor','TV-Emergency','Web'})
    if has(sub_l, 'mass casualty'):
        apps.update({'Mobile-Doctor','Mobile-Nurse','Mobile-Driver','TV-Emergency','Web'})
    if has(sub_l, 'insurance') and has(sub_l, 'verif','verification'):
        apps.update({'Web','Mobile-Patient'})
    if has(sub_l, 'audit trail'):
        apps.add('Web')
    if has(sub_l, 'kitchen'):
        apps.update({'Mobile-Dietitian','Desktop-Workstation','Web'})
    if has(sub_l, 'donor'):
        apps.update({'Mobile-LabTech','TV-Donor','Web'})
    if has(sub_l, 'visitor'):
        apps.update({'Mobile-Security','Desktop-Workstation','Web'})


def _keyword_refine(apps, feat_l, sub_l, web, mob, tv):
    """Apply universal keyword rules — add apps if specific tech/role hinted."""
    # printing
    if has(feat_l, 'print','printer','wristband','label','barcode label','token print'):
        apps.add('Desktop-Workstation')
    # kiosk
    if has(feat_l, 'kiosk','self-service','self check-in'):
        apps.add('Desktop-Kiosk')
    # device bridge
    if has(feat_l, 'hl7','astm','dicom','modality','plc','sensor stream','iot gateway'):
        apps.add('Desktop-DeviceBridge')
    # phlebo home-collection
    if has(feat_l, 'home collection','phlebotom','sample pickup'):
        apps.add('Mobile-Phlebo')
    # ambulance / driver
    if has(feat_l, 'ambulance','driver','dispatch','route to patient'):
        apps.add('Mobile-Driver')
    # nurse-specific
    if has(feat_l, 'emar','vitals capture','medication administration','barcode wristband scan',
                  'nursing rounds','bedside scan','fall risk audit','hand hygiene audit'):
        apps.add('Mobile-Nurse')
    # doctor-specific
    if has(feat_l, 'e-prescription','consultation note','soap note','round note',
                  'doctor app','progress note'):
        apps.add('Mobile-Doctor')
    # patient self-service
    if has(feat_l, 'patient app','patient self-service','patient portal',
                  'health record','book appointment','pay bill','upi','rate hospital'):
        apps.add('Mobile-Patient')
    # admin/exec
    if has(feat_l, 'ceo dashboard','board pack','exec dashboard','kpi','board report'):
        apps.add('Mobile-Admin')
    # security
    if has(feat_l, 'visitor pass','code red','code blue','code pink','code black',
                  'panic button','duress','incident response'):
        apps.add('Mobile-Security')
        apps.add('TV-Emergency')
    # housekeeping
    if has(feat_l, 'housekeeping','cleaning','bed turnaround','linen','laundry'):
        apps.add('Mobile-Housekeeping')
    # BME
    if has(feat_l, 'asset tag','calibration','breakdown ticket','preventive maintenance',
                  'pms log','biomedical'):
        apps.add('Mobile-BME')
    # radiology
    if has(feat_l, 'pacs','dicom viewer','radiology read','radiologist'):
        apps.add('Mobile-Radiologist')
    # OT
    if has(feat_l, 'ot schedule','ot status','operation theatre','surgical pause','time-out'):
        apps.add('TV-OT')
    # wayfinding
    if has(feat_l, 'wayfinding','indoor nav','directional signage','floor plan'):
        apps.add('TV-Wayfinding')
        apps.add('Mobile-Patient')
    # cafeteria / diet
    if has(feat_l, 'cafeteria','menu display','meal tray'):
        apps.add('TV-Cafeteria')
    # patient education
    if has(feat_l, 'patient education','health tips','wellness video'):
        apps.add('TV-Education')
    # blood bank
    if has(feat_l, 'donor','blood drive'):
        apps.add('TV-Donor')
    # noticeboard / events
    if has(feat_l, 'notice','event poster','hospital news'):
        apps.add('TV-Notice')
    # telemedicine
    if has(feat_l, 'telemedicine','video consultation','video consult'):
        apps.update({'Mobile-Doctor','Mobile-Patient'})
    # academic
    if has(feat_l, 'student','intern','postgrad','medical college','attendance','curriculum'):
        apps.update({'Mobile-Student','Mobile-Faculty'})
    # dietitian
    if has(feat_l, 'dietitian','diet plan','nutrition counsel'):
        apps.add('Mobile-Dietitian')
    # lab tech
    if has(feat_l, 'analyzer','sample receipt','sample accession','lab worklist'):
        apps.add('Mobile-LabTech')
    # pharmacist
    if has(feat_l, 'dispense','formulary','drug interaction','prescription verify','ward dispens'):
        apps.add('Mobile-Pharmacist')
    # ER — narrow keywords only (avoid matching "emergency contact", "infection emergency", etc)
    if has(feat_l, 'triage','er triage','mass casualty','code yellow','code orange',
                  'code red','code blue','disaster declaration','emergency code',
                  'er dashboard','er display'):
        apps.add('TV-Emergency')
    # OPD waiting display
    if has(feat_l, 'queue token','live queue','token board'):
        apps.add('TV-Queue')


def _section_default(sheet, ctx_mod, ctx_sub, web, mob, tv):
    """Fallback for rows whose Module col is empty — infer from section context.

    Excel uses section header rows (Module='Patient App', no Feature) followed
    by data rows with Module=None.  We carry the last non-empty Module forward.
    """
    cm = lower(ctx_mod)
    cs = lower(ctx_sub)

    if sheet == 'Mobile Apps':
        if 'patient app' in cm:        return {'Mobile-Patient'}
        if 'doctor app' in cm:         return {'Mobile-Doctor'}
        if 'phlebo' in cm:             return {'Mobile-Phlebo'}
        if 'universal mobile' in cm:   return {a for a in APPS if a.startswith(MOBILE_PREFIX)}
        if 'i18n' in cm:               return {a for a in APPS if a.startswith(MOBILE_PREFIX)}
        return {'Mobile-Patient'} if mob == 'Y' else set()

    if sheet == 'TV Displays & Queue':
        if tv == 'Y':
            return {a for a in APPS if a.startswith(TV_PREFIX)}
        return set()

    return set()


def _finalize(apps, web, mob, tv):
    """Strip apps disallowed by existing Y/N flags; sort canonical order."""
    apps = set(apps)
    # Honor existing column constraints — if explicitly N, remove
    if web == 'N':
        apps.discard('Web')
    if mob == 'N':
        apps = {a for a in apps if not a.startswith(MOBILE_PREFIX)}
    if tv == 'N':
        apps = {a for a in apps if not a.startswith(TV_PREFIX)}
    # if all three N (pure infra), and apps still empty → mark
    if not apps:
        return '(infra — no UI)'
    # canonical order
    order = ['Web','Desktop-Kiosk','Desktop-Workstation','Desktop-DeviceBridge'] + \
            sorted(a for a in APPS if a.startswith(MOBILE_PREFIX)) + \
            sorted(a for a in APPS if a.startswith(TV_PREFIX))
    return ', '.join(a for a in order if a in apps)


# ---------------------------------------------------------------- main
def main():
    parser = argparse.ArgumentParser(description="Assign MedBrains app targets to feature rows.")
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Update MedBrains_Features.xlsx in place instead of writing proposal workbook.",
    )
    parser.add_argument(
        "--src",
        default=SRC,
        help="Source workbook path. Defaults to MedBrains_Features.xlsx.",
    )
    parser.add_argument(
        "--dst",
        default=DST,
        help="Proposal output path when --apply is not used.",
    )
    parser.add_argument(
        "--sheet",
        action="append",
        default=None,
        help="Limit processing to specified sheet name(s). Repeat for multiple. Default: all sheets.",
    )
    parser.add_argument(
        "--ignore-flags",
        action="store_true",
        help="Ignore existing W/M/T flags during classification (assign apps purely by scenario). "
             "Also overwrites W/M/T columns to match assigned Apps.",
    )
    args = parser.parse_args()

    output = args.src if args.apply else args.dst
    if not args.apply:
        shutil.copyfile(args.src, output)

    wb = load_workbook(output)

    distribution = Counter()
    per_sheet_stats = defaultdict(lambda: {'rows': 0, 'infra': 0, 'apps_per_row': []})

    target_sheets = set(args.sheet) if args.sheet else None
    for sn in wb.sheetnames:
        if target_sheets and sn not in target_sheets:
            continue
        ws = wb[sn]
        # Add Apps header (col 12) — replace anything there
        ws.cell(row=1, column=12, value='Apps').font = Font(bold=True)
        ws.cell(row=1, column=12).fill = PatternFill(
            'solid', fgColor='1F4332')
        ws.cell(row=1, column=12).font = Font(bold=True, color='FFFFFF')
        ws.cell(row=1, column=12).alignment = Alignment(
            horizontal='left', vertical='center', wrap_text=True)
        ws.column_dimensions['L'].width = 55

        # carry-forward section context (top-down scan)
        ctx_mod = None
        ctx_sub = None
        for r in range(2, ws.max_row+1):
            raw_mod = ws.cell(row=r, column=2).value
            raw_sub = ws.cell(row=r, column=3).value
            if raw_mod: ctx_mod = raw_mod
            if raw_sub: ctx_sub = raw_sub
            feat = ws.cell(row=r, column=4).value
            if not feat:
                continue
            mod  = raw_mod or ctx_mod
            sub  = raw_sub or ctx_sub
            web  = ws.cell(row=r, column=9).value
            mob  = ws.cell(row=r, column=10).value
            tv   = ws.cell(row=r, column=11).value
            # When --ignore-flags, hide the existing N flags so classifier doesn't strip
            cls_w = 'Y' if args.ignore_flags else web
            cls_m = 'Y' if args.ignore_flags else mob
            cls_t = 'Y' if args.ignore_flags else tv
            apps_str = classify(sn, mod, sub, feat, cls_w, cls_m, cls_t)
            if apps_str == '(infra — no UI)':
                fallback = _section_default(sn, ctx_mod, ctx_sub, cls_w, cls_m, cls_t)
                if fallback:
                    apps_str = _finalize(fallback, cls_w, cls_m, cls_t)
            cell = ws.cell(row=r, column=12, value=apps_str)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.font = Font(name='Inter Tight', size=10)
            # When --ignore-flags, sync W/M/T columns from Apps assignment
            if args.ignore_flags and apps_str != '(infra — no UI)':
                new_w = 'Y' if 'Web' in apps_str.split(', ') else 'N'
                new_m = 'Y' if MOBILE_PREFIX in apps_str else 'N'
                new_t = 'Y' if TV_PREFIX in apps_str else 'N'
                ws.cell(row=r, column=9, value=new_w)
                ws.cell(row=r, column=10, value=new_m)
                ws.cell(row=r, column=11, value=new_t)
            per_sheet_stats[sn]['rows'] += 1
            if apps_str == '(infra — no UI)':
                per_sheet_stats[sn]['infra'] += 1
            else:
                count = apps_str.count(',') + 1
                per_sheet_stats[sn]['apps_per_row'].append(count)
                for a in apps_str.split(', '):
                    distribution[a] += 1

    wb.save(output)

    # report
    print(f'\n=== WROTE: {output} ===\n')
    print('=== APP COVERAGE (row count per app) ===')
    for a, c in distribution.most_common():
        print(f'  {c:5d}  {a}')
    print('\n=== PER-SHEET STATS ===')
    for sn in wb.sheetnames:
        s = per_sheet_stats[sn]
        if s['rows'] == 0: continue
        avg = (sum(s['apps_per_row'])/len(s['apps_per_row'])) if s['apps_per_row'] else 0
        print(f'  {sn:35s} rows={s["rows"]:4d} infra={s["infra"]:3d} avg-apps-per-row={avg:.1f}')


if __name__ == '__main__':
    main()

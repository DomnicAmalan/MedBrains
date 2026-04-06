#!/usr/bin/env python3
"""Comprehensive audit of MedBrains_Features.xlsx"""

import openpyxl
from collections import defaultdict

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

COMPLETED_MODULES = {
    "Onboarding & Setup",
    "Patient Management",
    "OPD",
    "Lab/LIS", "Lab / LIS", "Laboratory", "Lab",
    "Pharmacy",
    "Billing",
    "IPD",
    "ICU",
    "CSSD",
    "Diet & Kitchen", "Diet Kitchen", "Diet & Kitchen Management",
    "Emergency", "Emergency Medicine",
    "Blood Bank",
    "Radiology",
    "Quality Management", "Quality",
    "Infection Control",
    "Procurement", "Procurement & Inventory",
    "Front Office & Reception", "Front Office", "Reception",
    "HR & Staff Management", "HR", "Human Resources",
    "Consent Management", "Consent",
    "Facilities Management", "Facilities", "FMS",
    "Camp Management", "Camp",
    "Documents & Printing", "Printing & Forms", "Documents",
}

# Normalize module name for completed check
def is_completed_module(name):
    if not name:
        return False
    n = name.strip()
    if n in COMPLETED_MODULES:
        return True
    # Fuzzy match
    nl = n.lower()
    for cm in COMPLETED_MODULES:
        if cm.lower() in nl or nl in cm.lower():
            return True
    return False

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

# Collect all data
grand_total = {"Done": 0, "Partial": 0, "Pending": 0, "In Progress": 0, "Deferred": 0, "Other": 0, "Total": 0}

# For flagging
in_progress_flags = []  # (sheet, module, submodule, feature)
completed_pending_flags = []  # (sheet, module, submodule, feature)

SEP = "-" * 140

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    
    if len(rows) < 2:
        print(f"\n{'=' * 140}")
        print(f"SHEET: {sheet_name} (empty or header only)")
        continue
    
    # Find header row and column indices
    header_row = None
    header_idx = None
    for i, row in enumerate(rows):
        if row and any(str(c).strip().lower() in ("feature", "features", "feature name", "feature description") for c in row if c):
            header_row = row
            header_idx = i
            break
        # Also check for S.No as header indicator
        if row and any(str(c).strip().lower() in ("s.no", "s.no.", "sno", "sr.no", "sr no") for c in row if c):
            header_row = row
            header_idx = i
            break
    
    if header_row is None:
        # Try first row as header
        header_row = rows[0]
        header_idx = 0
    
    # Map column names to indices
    col_map = {}
    for j, cell in enumerate(header_row):
        if cell:
            key = str(cell).strip().lower()
            col_map[key] = j
    
    # Find relevant columns
    module_col = None
    submodule_col = None
    feature_col = None
    status_col = None
    
    for key, idx in col_map.items():
        if key in ("module", "module name"):
            module_col = idx
        elif key in ("sub-module", "sub module", "submodule", "sub-module name"):
            submodule_col = idx
        elif key in ("feature", "features", "feature name", "feature description"):
            feature_col = idx
        elif key in ("status",):
            status_col = idx
    
    print(f"\n{'=' * 140}")
    print(f"SHEET: {sheet_name}")
    print(f"  Columns found: Module={module_col}, Sub-Module={submodule_col}, Feature={feature_col}, Status={status_col}")
    print(f"  Header: {[str(c)[:30] if c else '' for c in header_row]}")
    
    if feature_col is None and status_col is None:
        print("  ** Could not identify Feature/Status columns, skipping **")
        continue
    
    # Parse data rows
    # Structure: sheet -> module -> submodule -> list of (feature, status)
    sheet_data = defaultdict(lambda: defaultdict(list))
    
    current_module = "(No Module)"
    current_submodule = "(No Sub-Module)"
    
    for i, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if not row or all(c is None for c in row):
            continue
        
        # Get module
        if module_col is not None and module_col < len(row) and row[module_col]:
            val = str(row[module_col]).strip()
            if val and val.lower() not in ("none", ""):
                current_module = val
        
        # Get submodule
        if submodule_col is not None and submodule_col < len(row) and row[submodule_col]:
            val = str(row[submodule_col]).strip()
            if val and val.lower() not in ("none", ""):
                current_submodule = val
        elif submodule_col is not None:
            pass  # Keep current
        
        # Get feature
        feature = ""
        if feature_col is not None and feature_col < len(row) and row[feature_col]:
            feature = str(row[feature_col]).strip()
        
        # Get status
        status = "Pending"
        if status_col is not None and status_col < len(row) and row[status_col]:
            s = str(row[status_col]).strip()
            if s and s.lower() not in ("none", ""):
                status = s
        
        if not feature or feature.lower() in ("none", ""):
            continue
        
        # Skip if it looks like a sub-header row (no feature text, just module/submodule)
        if len(feature) < 2:
            continue
        
        sheet_data[current_module][current_submodule].append((feature, status))
    
    # Print results for this sheet
    sheet_totals = {"Done": 0, "Partial": 0, "Pending": 0, "In Progress": 0, "Deferred": 0, "Other": 0, "Total": 0}
    
    print(f"\n  {'Module':<40} {'Sub-Module':<35} {'Total':>6} {'Done':>6} {'Partial':>8} {'Pending':>8} {'InProg':>7} {'Defer':>6} {'Other':>6}")
    print(f"  {SEP}")
    
    for module in sorted(sheet_data.keys()):
        module_totals = {"Done": 0, "Partial": 0, "Pending": 0, "In Progress": 0, "Deferred": 0, "Other": 0, "Total": 0}
        
        for submodule in sorted(sheet_data[module].keys()):
            entries = sheet_data[module][submodule]
            counts = {"Done": 0, "Partial": 0, "Pending": 0, "In Progress": 0, "Deferred": 0, "Other": 0}
            
            for feat, stat in entries:
                s = stat.strip()
                if s in counts:
                    counts[s] += 1
                else:
                    # Normalize
                    sl = s.lower()
                    if sl == "done":
                        counts["Done"] += 1
                    elif sl == "partial":
                        counts["Partial"] += 1
                    elif sl == "pending":
                        counts["Pending"] += 1
                    elif sl in ("in progress", "in-progress", "inprogress", "wip"):
                        counts["In Progress"] += 1
                    elif sl in ("deferred", "defer"):
                        counts["Deferred"] += 1
                    else:
                        counts["Other"] += 1
                
                # Flag in-progress
                if s.lower() in ("in progress", "in-progress", "inprogress", "wip"):
                    in_progress_flags.append((sheet_name, module, submodule, feat, stat))
                
                # Flag completed module with pending
                if s.lower() == "pending" and is_completed_module(module):
                    completed_pending_flags.append((sheet_name, module, submodule, feat))
            
            total = sum(counts.values())
            
            sm_display = submodule[:33] if len(submodule) > 33 else submodule
            m_display = module[:38] if len(module) > 38 else module
            
            print(f"  {m_display:<40} {sm_display:<35} {total:>6} {counts['Done']:>6} {counts['Partial']:>8} {counts['Pending']:>8} {counts['In Progress']:>7} {counts['Deferred']:>6} {counts['Other']:>6}")
            
            for k in counts:
                module_totals[k] = module_totals.get(k, 0) + counts[k]
            module_totals["Total"] = module_totals.get("Total", 0) + total
        
        # Module subtotal
        mt = module_totals
        m_display = f"  >> {module}"[:40]
        print(f"  {m_display:<40} {'** SUBTOTAL **':<35} {mt['Total']:>6} {mt['Done']:>6} {mt['Partial']:>8} {mt['Pending']:>8} {mt['In Progress']:>7} {mt['Deferred']:>6} {mt['Other']:>6}")
        print(f"  {'':40} {'':35} {'':>6} {'':>6} {'':>8} {'':>8} {'':>7} {'':>6} {'':>6}")
        
        for k in mt:
            sheet_totals[k] = sheet_totals.get(k, 0) + mt[k]
    
    # Sheet total
    st = sheet_totals
    print(f"  {SEP}")
    print(f"  {'SHEET TOTAL':<40} {'':<35} {st['Total']:>6} {st['Done']:>6} {st['Partial']:>8} {st['Pending']:>8} {st['In Progress']:>7} {st['Deferred']:>6} {st['Other']:>6}")
    
    for k in st:
        grand_total[k] = grand_total.get(k, 0) + st[k]

# Grand total
print(f"\n\n{'=' * 140}")
print("GRAND TOTAL ACROSS ALL SHEETS")
print(f"{'=' * 140}")
gt = grand_total
print(f"  Total Features: {gt['Total']}")
print(f"  Done:           {gt['Done']} ({gt['Done']/max(gt['Total'],1)*100:.1f}%)")
print(f"  Partial:        {gt['Partial']} ({gt['Partial']/max(gt['Total'],1)*100:.1f}%)")
print(f"  Pending:        {gt['Pending']} ({gt['Pending']/max(gt['Total'],1)*100:.1f}%)")
print(f"  In Progress:    {gt['In Progress']} ({gt['In Progress']/max(gt['Total'],1)*100:.1f}%)")
print(f"  Deferred:       {gt['Deferred']} ({gt['Deferred']/max(gt['Total'],1)*100:.1f}%)")
print(f"  Other:          {gt['Other']} ({gt['Other']/max(gt['Total'],1)*100:.1f}%)")

# Flag: In Progress features
print(f"\n\n{'=' * 140}")
print(f"FLAG: Features marked 'In Progress' ({len(in_progress_flags)} total)")
print(f"{'=' * 140}")
if in_progress_flags:
    for sheet, mod, sub, feat, stat in in_progress_flags:
        print(f"  [{sheet}] {mod} > {sub}")
        print(f"    Feature: {feat[:120]}")
        print(f"    Status:  {stat}")
        print()
else:
    print("  None found.")

# Flag: Completed modules with Pending features
print(f"\n{'=' * 140}")
print(f"FLAG: Completed modules with PENDING features ({len(completed_pending_flags)} total)")
print(f"{'=' * 140}")
if completed_pending_flags:
    # Group by module
    by_module = defaultdict(list)
    for sheet, mod, sub, feat in completed_pending_flags:
        by_module[mod].append((sheet, sub, feat))
    
    for mod in sorted(by_module.keys()):
        items = by_module[mod]
        print(f"\n  MODULE: {mod} ({len(items)} pending features)")
        print(f"  {'-' * 120}")
        for sheet, sub, feat in items:
            print(f"    [{sheet}] {sub:<35} | {feat[:85]}")
else:
    print("  None found.")

wb.close()
print(f"\n\nAudit complete.")

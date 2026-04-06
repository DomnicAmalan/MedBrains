#!/usr/bin/env python3
"""Extract ALL billing-related features from MedBrains_Features.xlsx across all sheets."""

import openpyxl
from collections import defaultdict

XLSX_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"


def find_billing_features():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    # We'll collect features grouped by status
    features_by_status = defaultdict(list)
    total_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=False))
        if not rows:
            continue

        # Find header row — look for a row containing "Module" or "Feature"
        header_row = None
        header_idx = {}
        for row_num, row in enumerate(rows, start=1):
            cell_values = [str(c.value).strip().lower() if c.value else "" for c in row]
            if "module" in cell_values or "feature" in cell_values:
                header_row = row_num
                for col_idx, val in enumerate(cell_values):
                    if val:
                        header_idx[val] = col_idx
                break

        if header_row is None:
            continue

        # Map column names (handle variations)
        def get_col(names):
            """Return column index for first matching name."""
            for n in names:
                if n in header_idx:
                    return header_idx[n]
            return None

        col_sno = get_col(["s.no", "s. no", "sno", "no", "sr", "sr.", "#"])
        col_module = get_col(["module"])
        col_submodule = get_col(["sub-module", "submodule", "sub module"])
        col_feature = get_col(["feature", "features", "feature name", "feature description"])
        col_status = get_col(["status"])
        col_priority = get_col(["priority", "p"])
        col_web = get_col(["web"])
        col_mobile = get_col(["mobile", "mob"])
        col_tv = get_col(["tv"])
        col_source = get_col(["source", "src"])
        col_rfc = get_col(["rfc ref", "rfc", "rfc reference"])

        if col_feature is None and col_module is None:
            continue

        # Track current module/submodule for merged cells or header rows
        current_module = ""
        current_submodule = ""

        for row_num, row in enumerate(rows, start=1):
            if row_num <= header_row:
                continue

            def cell_val(col_idx):
                if col_idx is None or col_idx >= len(row):
                    return ""
                v = row[col_idx].value
                return str(v).strip() if v is not None else ""

            module = cell_val(col_module) or current_module
            submodule = cell_val(col_submodule) or current_submodule
            feature = cell_val(col_feature)
            status = cell_val(col_status)
            priority = cell_val(col_priority)
            web = cell_val(col_web)
            mobile = cell_val(col_mobile)
            tv = cell_val(col_tv)

            # Update current tracking
            if cell_val(col_module):
                current_module = cell_val(col_module)
            if cell_val(col_submodule):
                current_submodule = cell_val(col_submodule)

            # Skip empty rows or rows that are just headers/section dividers
            if not feature or feature.lower() in ("feature", "features", "feature name"):
                continue

            # Check if this feature is billing-related
            billing_keywords = [
                "billing", "invoice", "invoicing", "payment", "receipt",
                "charge", "tariff", "rate plan", "discount",
                "refund", "credit note", "gst", "tpa",
                "insurance claim", "cashier", "revenue",
                "financial", "price", "pricing",
            ]

            text_to_search = f"{module} {submodule} {feature}".lower()

            is_billing_module = "billing" in module.lower()
            is_billing_keyword = any(kw in text_to_search for kw in billing_keywords)

            if is_billing_module or is_billing_keyword:
                record = {
                    "sheet": sheet_name,
                    "row": row_num,
                    "module": module,
                    "submodule": submodule,
                    "feature": feature,
                    "status": status if status else "Pending",
                    "priority": priority if priority else "-",
                    "web": web if web else "-",
                    "mobile": mobile if mobile else "-",
                    "tv": tv if tv else "-",
                }
                features_by_status[record["status"]].append(record)
                total_count += 1

    wb.close()

    # Print results grouped by status
    status_order = ["Done", "Partial", "In Progress", "Pending", "Deferred"]
    all_statuses = list(features_by_status.keys())
    for s in all_statuses:
        if s not in status_order:
            status_order.append(s)

    print("=" * 140)
    print(f"BILLING-RELATED FEATURES IN MedBrains_Features.xlsx — TOTAL: {total_count}")
    print("=" * 140)

    for status in status_order:
        if status not in features_by_status:
            continue
        items = features_by_status[status]
        print(f"\n{'─' * 140}")
        print(f"  STATUS: {status.upper()}  ({len(items)} features)")
        print(f"{'─' * 140}")
        print(
            f"  {'#':<4} {'Sheet':<28} {'Row':<5} {'Module':<22} {'Sub-Module':<24} "
            f"{'Pri':<5} {'Web':<4} {'Mob':<4} {'TV':<4} Feature"
        )
        print(
            f"  {'—'*4} {'—'*28} {'—'*5} {'—'*22} {'—'*24} "
            f"{'—'*5} {'—'*4} {'—'*4} {'—'*4} {'—'*60}"
        )

        for i, r in enumerate(items, 1):
            feat_text = r["feature"]
            if len(feat_text) > 90:
                feat_text = feat_text[:87] + "..."
            print(
                f"  {i:<4} {r['sheet']:<28} {r['row']:<5} {r['module']:<22} {r['submodule']:<24} "
                f"{r['priority']:<5} {r['web']:<4} {r['mobile']:<4} {r['tv']:<4} {feat_text}"
            )

    # Summary
    print(f"\n{'=' * 140}")
    print("SUMMARY BY STATUS:")
    for status in status_order:
        if status in features_by_status:
            print(f"  {status:<15} {len(features_by_status[status]):>4} features")
    print(f"  {'TOTAL':<15} {total_count:>4} features")
    print(f"{'=' * 140}")

    # Also group by sheet
    by_sheet = defaultdict(list)
    for items in features_by_status.values():
        for r in items:
            by_sheet[r["sheet"]].append(r)

    print("\nBY SHEET:")
    for sheet, items in sorted(by_sheet.items()):
        status_counts = defaultdict(int)
        for r in items:
            status_counts[r["status"]] += 1
        counts_str = ", ".join(f"{s}: {c}" for s, c in sorted(status_counts.items()))
        print(f"  {sheet:<30} {len(items):>3} features  ({counts_str})")


if __name__ == "__main__":
    find_billing_features()

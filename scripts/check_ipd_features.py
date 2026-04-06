import openpyxl

wb = openpyxl.load_workbook('/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx', read_only=True, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        continue

    header_row = None
    for i, row in enumerate(rows):
        if row and any(str(c).strip().lower() == 'status' for c in row if c):
            header_row = i
            break

    if header_row is None:
        continue

    headers = [str(c).strip().lower() if c else '' for c in rows[header_row]]

    cols = {}
    for j, h in enumerate(headers):
        if h == 'status': cols['status'] = j
        elif h == 'module': cols['module'] = j
        elif h in ('sub-module', 'sub_module', 'submodule', 'sub module'): cols['sub'] = j
        elif h == 'feature': cols['feature'] = j
        elif h == 'web': cols['web'] = j
        elif h == 'mobile': cols['mobile'] = j
        elif h == 'tv': cols['tv'] = j

    if 'status' not in cols or 'module' not in cols:
        continue

    current_module = None
    current_sub = None
    found = False

    for row in rows[header_row + 1:]:
        if not row or all(c is None for c in row):
            continue

        mod = row[cols['module']] if cols['module'] < len(row) else None
        if mod and str(mod).strip():
            current_module = str(mod).strip()

        sub = row[cols.get('sub', -1)] if cols.get('sub', -1) < len(row) else None
        if sub and str(sub).strip():
            current_sub = str(sub).strip()

        feature = row[cols.get('feature', -1)] if cols.get('feature', -1) < len(row) else None
        status = row[cols.get('status', -1)] if cols.get('status', -1) < len(row) else None

        if not current_module or not feature:
            continue

        mod_lower = current_module.lower()
        if 'ipd' in mod_lower or 'inpatient' in mod_lower or 'nursing' in mod_lower or 'ot' in mod_lower or 'operation theatre' in mod_lower or 'surgical' in mod_lower:
            if not found:
                print(f"\n=== Sheet: {sheet_name} ===")
                found = True

            status_str = str(status).strip() if status else 'Pending'
            web = row[cols.get('web', -1)] if cols.get('web', -1) < len(row) else None
            mobile = row[cols.get('mobile', -1)] if cols.get('mobile', -1) < len(row) else None

            print(f"[{status_str:>10}] {current_module} > {current_sub or '---'} > {str(feature).strip()}")

wb.close()

import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook('/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx', read_only=True, data_only=True)

module_stats = defaultdict(lambda: {'Done': 0, 'Partial': 0, 'Pending': 0, 'In Progress': 0, 'Total': 0})

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        continue

    # Find header row
    header_row = None
    for i, row in enumerate(rows):
        if row and any(str(c).strip().lower() == 'status' for c in row if c):
            header_row = i
            break

    if header_row is None:
        continue

    headers = [str(c).strip().lower() if c else '' for c in rows[header_row]]

    status_col = None
    module_col = None
    sub_module_col = None
    feature_col = None

    for j, h in enumerate(headers):
        if h == 'status':
            status_col = j
        elif h == 'module':
            module_col = j
        elif h in ('sub-module', 'sub_module', 'submodule', 'sub module'):
            sub_module_col = j
        elif h == 'feature':
            feature_col = j

    if status_col is None or module_col is None:
        continue

    current_module = None
    for row in rows[header_row + 1:]:
        if not row or all(c is None for c in row):
            continue

        mod = row[module_col] if module_col < len(row) else None
        if mod and str(mod).strip():
            current_module = str(mod).strip()

        status = row[status_col] if status_col < len(row) else None
        feature = row[feature_col] if feature_col and feature_col < len(row) else None

        if not current_module or not feature:
            continue

        status_str = str(status).strip() if status else 'Pending'

        module_stats[current_module]['Total'] += 1
        if status_str == 'Done':
            module_stats[current_module]['Done'] += 1
        elif status_str == 'Partial':
            module_stats[current_module]['Partial'] += 1
        elif status_str == 'In Progress':
            module_stats[current_module]['In Progress'] += 1
        else:
            module_stats[current_module]['Pending'] += 1

wb.close()

# Sort by pending count descending
sorted_modules = sorted(module_stats.items(), key=lambda x: (-x[1]['Pending'], x[0]))

print(f"{'Module':<45} {'Done':>5} {'Part':>5} {'Pend':>5} {'Total':>6}")
print('-' * 72)
for mod, stats in sorted_modules:
    if stats['Pending'] > 0 or stats['Done'] == 0:
        print(f"{mod:<45} {stats['Done']:>5} {stats['Partial']:>5} {stats['Pending']:>5} {stats['Total']:>6}")

print(f"\n--- FULLY DONE or PARTIAL (no pending) ---")
for mod, stats in sorted(module_stats.items()):
    if stats['Pending'] == 0 and stats['Done'] > 0:
        print(f"{mod:<45} {stats['Done']:>5} {stats['Partial']:>5} {stats['Pending']:>5} {stats['Total']:>6}")

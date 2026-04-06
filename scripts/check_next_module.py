#!/usr/bin/env python3
"""
Comprehensive MedBrains feature analysis script.

Reads MedBrains_Features.xlsx and provides:
1. Per-sheet feature counts by status
2. Modules with mostly "Pending" features
3. Modules with "In Progress" status
4. Priority ordering for next module to build
5. Detailed feature list for the top candidate module
"""

import os
from collections import defaultdict
from openpyxl import load_workbook

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "MedBrains_Features.xlsx")


def read_features(wb):
    """Read all features from all sheets, returning structured data."""
    all_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=False))

        if not rows:
            all_data[sheet_name] = {"features": [], "header_row": None}
            continue

        # Find header row (look for row containing "Feature" or "S.No" or "Module")
        header_row_idx = None
        header_map = {}
        for idx, row in enumerate(rows):
            vals = [str(c.value).strip().lower() if c.value else "" for c in row]
            if "feature" in vals or "s.no" in vals or "module" in vals:
                header_row_idx = idx
                for col_idx, val in enumerate(vals):
                    if val:
                        header_map[val] = col_idx
                break

        if header_row_idx is None:
            # Try first row as header
            header_row_idx = 0
            vals = [str(c.value).strip().lower() if c.value else "" for c in rows[0]]
            for col_idx, val in enumerate(vals):
                if val:
                    header_map[val] = col_idx

        # Find column indices
        def find_col(names):
            for name in names:
                if name in header_map:
                    return header_map[name]
            return None

        sno_col = find_col(["s.no", "s.no.", "sno", "no", "no."])
        module_col = find_col(["module"])
        submodule_col = find_col(["sub-module", "sub module", "submodule"])
        feature_col = find_col(["feature", "features"])
        status_col = find_col(["status"])
        priority_col = find_col(["priority"])
        web_col = find_col(["web"])
        mobile_col = find_col(["mobile"])
        tv_col = find_col(["tv"])

        features = []
        current_module = None
        current_submodule = None

        for row in rows[header_row_idx + 1:]:
            vals = [c.value for c in row]

            # Skip completely empty rows
            if all(v is None or str(v).strip() == "" for v in vals):
                continue

            # Get cell values
            sno = str(vals[sno_col]).strip() if sno_col is not None and sno_col < len(vals) and vals[sno_col] else ""
            module = str(vals[module_col]).strip() if module_col is not None and module_col < len(vals) and vals[module_col] else ""
            submodule = str(vals[submodule_col]).strip() if submodule_col is not None and submodule_col < len(vals) and vals[submodule_col] else ""
            feature = str(vals[feature_col]).strip() if feature_col is not None and feature_col < len(vals) and vals[feature_col] else ""
            status = str(vals[status_col]).strip() if status_col is not None and status_col < len(vals) and vals[status_col] else ""
            priority = str(vals[priority_col]).strip() if priority_col is not None and priority_col < len(vals) and vals[priority_col] else ""
            web = str(vals[web_col]).strip() if web_col is not None and web_col < len(vals) and vals[web_col] else ""
            mobile = str(vals[mobile_col]).strip() if mobile_col is not None and mobile_col < len(vals) and vals[mobile_col] else ""
            tv = str(vals[tv_col]).strip() if tv_col is not None and tv_col < len(vals) and vals[tv_col] else ""

            # Track current module/submodule from merged/header rows
            if module and module.lower() not in ("none", "nan", ""):
                current_module = module
            if submodule and submodule.lower() not in ("none", "nan", ""):
                current_submodule = submodule

            # Skip rows that are just module/submodule headers (no feature text)
            if not feature or feature.lower() in ("none", "nan", ""):
                continue

            # Normalize status
            status_normalized = status.lower().strip() if status else "pending"
            if status_normalized in ("none", "nan", ""):
                status_normalized = "pending"

            # Map to standard statuses
            status_map = {
                "done": "Done",
                "partial": "Partial",
                "in progress": "In Progress",
                "pending": "Pending",
                "deferred": "Deferred",
            }
            status_final = status_map.get(status_normalized, status.title() if status else "Pending")

            features.append({
                "sno": sno,
                "module": current_module or "",
                "submodule": current_submodule or "",
                "feature": feature,
                "status": status_final,
                "priority": priority,
                "web": web,
                "mobile": mobile,
                "tv": tv,
            })

        all_data[sheet_name] = {
            "features": features,
            "header_map": header_map,
        }

    return all_data


def print_section(title, char="="):
    width = 100
    print(f"\n{char * width}")
    print(f"  {title}")
    print(f"{char * width}")


def main():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    all_data = read_features(wb)

    total_features = 0
    total_by_status = defaultdict(int)
    module_status_map = defaultdict(lambda: defaultdict(int))
    module_features = defaultdict(list)
    module_sheet_map = {}  # module -> sheet name

    # =========================================================================
    # 1. Per-sheet feature counts by status
    # =========================================================================
    print_section("1. FEATURE COUNTS BY STATUS (PER SHEET)")

    for sheet_name, data in all_data.items():
        features = data["features"]
        if not features:
            continue

        status_counts = defaultdict(int)
        for f in features:
            status_counts[f["status"]] += 1
            total_by_status[f["status"]] += 1
            total_features += 1

            # Track by module
            mod = f["module"]
            if mod:
                module_status_map[mod][f["status"]] += 1
                module_features[mod].append(f)
                if mod not in module_sheet_map:
                    module_sheet_map[mod] = sheet_name

        sheet_total = sum(status_counts.values())
        done_pct = (status_counts.get("Done", 0) / sheet_total * 100) if sheet_total > 0 else 0

        print(f"\n  Sheet: {sheet_name} ({sheet_total} features, {done_pct:.1f}% done)")
        print(f"  {'Status':<15} {'Count':>6} {'Percentage':>10}")
        print(f"  {'-'*35}")
        for status in ["Done", "Partial", "In Progress", "Pending", "Deferred"]:
            count = status_counts.get(status, 0)
            pct = (count / sheet_total * 100) if sheet_total > 0 else 0
            if count > 0:
                print(f"  {status:<15} {count:>6} {pct:>9.1f}%")
        # Print any non-standard statuses
        for status, count in sorted(status_counts.items()):
            if status not in ["Done", "Partial", "In Progress", "Pending", "Deferred"]:
                pct = (count / sheet_total * 100) if sheet_total > 0 else 0
                print(f"  {status:<15} {count:>6} {pct:>9.1f}%")

    print(f"\n  {'='*50}")
    print(f"  GRAND TOTAL: {total_features} features")
    for status in ["Done", "Partial", "In Progress", "Pending", "Deferred"]:
        count = total_by_status.get(status, 0)
        pct = (count / total_features * 100) if total_features > 0 else 0
        if count > 0:
            print(f"    {status:<15} {count:>6} ({pct:.1f}%)")
    for status, count in sorted(total_by_status.items()):
        if status not in ["Done", "Partial", "In Progress", "Pending", "Deferred"]:
            pct = (count / total_features * 100) if total_features > 0 else 0
            print(f"    {status:<15} {count:>6} ({pct:.1f}%)")

    # =========================================================================
    # 2. Modules with mostly "Pending" features
    # =========================================================================
    print_section("2. MODULES WITH MOSTLY PENDING FEATURES (>= 50%)")

    pending_modules = []
    for mod, statuses in sorted(module_status_map.items()):
        total = sum(statuses.values())
        pending = statuses.get("Pending", 0)
        pct = (pending / total * 100) if total > 0 else 0
        if pct >= 50:
            pending_modules.append((mod, total, pending, pct))

    pending_modules.sort(key=lambda x: -x[2])  # Sort by pending count descending

    print(f"\n  {'Module':<45} {'Total':>6} {'Pending':>8} {'%Pend':>7} {'Done':>5} {'Part':>5}")
    print(f"  {'-'*80}")
    for mod, total, pending, pct in pending_modules:
        done = module_status_map[mod].get("Done", 0)
        partial = module_status_map[mod].get("Partial", 0)
        print(f"  {mod:<45} {total:>6} {pending:>8} {pct:>6.1f}% {done:>5} {partial:>5}")

    # =========================================================================
    # 3. Modules with "In Progress" status
    # =========================================================================
    print_section("3. MODULES WITH 'IN PROGRESS' FEATURES")

    in_progress_modules = []
    for mod, statuses in sorted(module_status_map.items()):
        ip_count = statuses.get("In Progress", 0)
        if ip_count > 0:
            total = sum(statuses.values())
            in_progress_modules.append((mod, total, ip_count))

    if in_progress_modules:
        print(f"\n  {'Module':<45} {'Total':>6} {'In Progress':>12}")
        print(f"  {'-'*65}")
        for mod, total, ip in sorted(in_progress_modules, key=lambda x: -x[2]):
            print(f"  {mod:<45} {total:>6} {ip:>12}")

        # List the actual in-progress features
        print(f"\n  In-Progress Feature Details:")
        print(f"  {'-'*90}")
        for mod, _, _ in sorted(in_progress_modules, key=lambda x: -x[2]):
            for f in module_features[mod]:
                if f["status"] == "In Progress":
                    sub = f["submodule"][:25] if f["submodule"] else ""
                    feat = f["feature"][:55] if f["feature"] else ""
                    print(f"  [{mod[:20]}] {sub:<25} {feat}")
    else:
        print("\n  No modules currently have 'In Progress' features.")

    # =========================================================================
    # 4. Priority ordering for next modules to build
    # =========================================================================
    print_section("4. PRIORITY ORDERING -- NEXT MODULES TO BUILD")

    # Calculate per module
    module_scores = []
    for mod, statuses in module_status_map.items():
        total = sum(statuses.values())
        done = statuses.get("Done", 0)
        partial = statuses.get("Partial", 0)
        pending = statuses.get("Pending", 0)
        in_progress = statuses.get("In Progress", 0)
        deferred = statuses.get("Deferred", 0)

        done_pct = ((done + partial) / total * 100) if total > 0 else 0
        pending_pct = (pending / total * 100) if total > 0 else 0

        sheet = module_sheet_map.get(mod, "")

        module_scores.append({
            "module": mod,
            "sheet": sheet,
            "total": total,
            "done": done,
            "partial": partial,
            "in_progress": in_progress,
            "pending": pending,
            "deferred": deferred,
            "done_pct": done_pct,
            "pending_pct": pending_pct,
        })

    # Group by completion status
    fully_pending = [m for m in module_scores if m["done"] == 0 and m["partial"] == 0 and m["in_progress"] == 0 and m["pending"] > 0]
    partially_done = [m for m in module_scores if (m["done"] > 0 or m["partial"] > 0) and m["pending"] > 0]
    fully_done_or_no_pending = [m for m in module_scores if m["pending"] == 0]

    # Modules not yet started
    print(f"\n  A. MODULES NOT YET STARTED (0% done, all pending):")
    print(f"  {'Module':<45} {'Sheet':<35} {'Pending':>8}")
    print(f"  {'-'*90}")
    for m in sorted(fully_pending, key=lambda x: -x["pending"]):
        print(f"  {m['module']:<45} {m['sheet']:<35} {m['pending']:>8}")

    # Modules partially done
    print(f"\n  B. MODULES PARTIALLY DONE (have remaining pending features):")
    print(f"  {'Module':<45} {'Done':>5} {'Part':>5} {'Pend':>5} {'Total':>6} {'Done%':>6}")
    print(f"  {'-'*75}")
    for m in sorted(partially_done, key=lambda x: x["done_pct"]):
        print(f"  {m['module']:<45} {m['done']:>5} {m['partial']:>5} {m['pending']:>5} {m['total']:>6} {m['done_pct']:>5.0f}%")

    # Fully done
    print(f"\n  C. MODULES COMPLETED (no pending features):")
    print(f"  {'Module':<45} {'Done':>5} {'Part':>5} {'Defer':>5}")
    print(f"  {'-'*65}")
    for m in sorted(fully_done_or_no_pending, key=lambda x: x["module"]):
        if m["done"] > 0 or m["partial"] > 0:
            print(f"  {m['module']:<45} {m['done']:>5} {m['partial']:>5} {m['deferred']:>5}")

    # =========================================================================
    # 5. Top candidate module features
    # =========================================================================
    print_section("5. RECOMMENDED NEXT MODULE & ALL ITS FEATURES")

    # Define priority tiers based on RFC and clinical workflow dependencies
    tier_map = {
        # P1 - Core clinical (should already be done)
        "Patient Management": 1,
        "OPD": 1,
        "Billing & Revenue": 1,
        # P2 - Essential support (should already be done)
        "Laboratory / LIS": 2,
        "Pharmacy": 2,
        "IPD / Inpatient": 2,
        # P3 - Clinical extensions
        "Nursing Module": 3,
        "Operation Theatre (OT)": 3,
        "Physiotherapy & Rehabilitation": 3,
        # P4 - Administrative & Operational
        "Housekeeping": 4,
        "Ambulance Services": 4,
        "Mortuary / Body Management": 4,
        "Laundry & Linen": 4,
        "Asset / Equipment Management": 4,
        "Human Resources (HR)": 4,
        # P5 - Specialty
        "Medical College / Academic Module": 5,
        "Psychiatry Module": 5,
        "Patient Experience Module": 5,
    }

    # Find candidates: not-started or barely-started modules
    candidates = []
    for m in module_scores:
        if m["pending"] == 0:
            continue  # skip fully done
        tier = tier_map.get(m["module"], 6)
        # Weight: lower tier is higher priority; within same tier, more pending = more important
        candidates.append((tier, -m["pending"], m["module"], m))

    candidates.sort()

    if candidates:
        # Pick top candidate
        top = candidates[0]
        top_module = top[2]
        top_stats = top[3]

        print(f"\n  TOP CANDIDATE: {top_module}")
        print(f"  Sheet: {module_sheet_map.get(top_module, 'Unknown')}")
        print(f"  Total features: {top_stats['total']}  |  Done: {top_stats['done']}  |  Partial: {top_stats['partial']}  |  Pending: {top_stats['pending']}  |  In Progress: {top_stats['in_progress']}")

        print(f"\n  ALL FEATURES:")
        print(f"  {'#':<5} {'Sub-Module':<32} {'Feature':<50} {'Status':<12} {'Priority':<6}")
        print(f"  {'-'*107}")

        for i, f in enumerate(module_features[top_module], 1):
            submod = (f["submodule"][:30] if f["submodule"] else "")
            feat = (f["feature"][:48] if f["feature"] else "")
            pri = f["priority"] if f["priority"] and f["priority"].lower() not in ("none", "nan", "") else ""
            print(f"  {i:<5} {submod:<32} {feat:<50} {f['status']:<12} {pri:<6}")

        # Show next 10 candidates
        if len(candidates) > 1:
            print(f"\n  NEXT CANDIDATES (in priority order):")
            print(f"  {'Tier':<6} {'Module':<45} {'Pending':>8} {'Done':>5} {'Part':>5}")
            print(f"  {'-'*72}")
            for tier, neg_pend, mod, m in candidates[1:16]:
                print(f"  T{tier:<5} {mod:<45} {m['pending']:>8} {m['done']:>5} {m['partial']:>5}")
    else:
        print("\n  All modules appear to be complete!")

    wb.close()


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
MedBrains Feature Tracker — Remaining Work Analysis
Reads MedBrains_Features.xlsx and reports:
  1. Overall feature counts by status (per sheet + grand total)
  2. Modules/sub-modules with significant Pending features
  3. Recommended next module(s) to build
"""

import openpyxl
from collections import defaultdict

XLSX_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

# Modules known to be completed (from MEMORY.md)
COMPLETED_MODULES = {
    # Normalized lowercase names for matching
    "auth", "patient mgmt", "patient management", "patients",
    "opd", "opd/outpatient", "outpatient",
    "laboratory", "lab", "lab/lis",
    "pharmacy", "pharmacy/dispensing",
    "billing", "billing & finance",
    "ipd", "ipd/inpatient", "inpatient",
    "icu", "icu/critical care", "critical care",
    "cssd", "cssd/sterilization",
    "diet & kitchen", "diet/kitchen", "diet kitchen", "dietary",
    "emergency", "emergency/casualty",
    "radiology", "radiology/imaging",
    "blood bank", "blood bank/transfusion",
    "quality", "quality management", "quality mgmt",
    "infection control", "infection ctrl",
    "procurement", "procurement/supply",
    "hr", "hr/staff", "hr & staff", "staff management", "hr/staff mgmt",
    "front office", "front office/reception", "reception",
    "consent", "consent management",
    "facilities", "facilities management", "facilities mgmt", "fms",
    "camp", "camp management", "camp mgmt",
    "navigation", "settings", "dashboard",
    "security", "onboarding", "setup",
    # Add more aliases as needed
}


def normalize_status(val):
    """Normalize status values to canonical forms."""
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("done", "complete", "completed"):
        return "Done"
    if s in ("partial", "partially done", "partially"):
        return "Partial"
    if s in ("pending", "not started", "todo"):
        return "Pending"
    if s in ("in progress", "in-progress", "wip"):
        return "In Progress"
    if s in ("deferred", "defer", "later"):
        return "Deferred"
    return None  # Not a feature row


def is_module_completed(module_name):
    """Check if a module is in the completed list."""
    if not module_name:
        return False
    return module_name.strip().lower() in COMPLETED_MODULES


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)

    # Track data
    sheet_stats = {}  # sheet_name -> {status: count}
    module_stats = defaultdict(lambda: defaultdict(int))  # (sheet, module) -> {status: count}
    submodule_stats = defaultdict(lambda: defaultdict(int))  # (sheet, module, submodule) -> {status: count}
    grand_total = defaultdict(int)
    total_features = 0

    # Detailed pending tracking
    pending_by_module = []  # list of (sheet, module, submodule, pending_count, total_count, features_list)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_counts = defaultdict(int)
        current_module = None
        current_submodule = None
        sheet_features = 0

        for row in ws.iter_rows(min_row=2, values_only=False):
            cells = [c.value for c in row]

            # Column indices (0-based): 0=S.No, 1=Module, 2=Sub-Module, 3=Feature, 4=Source, 5=Priority, 6=Status
            sno = cells[0] if len(cells) > 0 else None
            module = cells[1] if len(cells) > 1 else None
            submodule = cells[2] if len(cells) > 2 else None
            feature = cells[3] if len(cells) > 3 else None
            status_raw = cells[6] if len(cells) > 6 else None

            # Track current module/submodule from header rows
            if module and not sno and not status_raw:
                # This is a module header row
                current_module = str(module).strip()
                current_submodule = None
                continue

            if submodule and not sno and not status_raw:
                # This is a sub-module header row
                current_submodule = str(submodule).strip()
                continue

            # Update current module/submodule if present in feature row
            if module:
                current_module = str(module).strip()
            if submodule:
                current_submodule = str(submodule).strip()

            status = normalize_status(status_raw)
            if status is None:
                continue

            # This is a feature row
            sheet_features += 1
            total_features += 1
            sheet_counts[status] += 1
            grand_total[status] += 1

            mod_key = current_module or "Unknown"
            sub_key = current_submodule or "General"

            module_stats[(sheet_name, mod_key)][status] += 1
            module_stats[(sheet_name, mod_key)]["_total"] += 1
            submodule_stats[(sheet_name, mod_key, sub_key)][status] += 1
            submodule_stats[(sheet_name, mod_key, sub_key)]["_total"] += 1

        sheet_counts["_total"] = sheet_features
        sheet_stats[sheet_name] = dict(sheet_counts)

    # ========================================
    # REPORT
    # ========================================

    print("=" * 100)
    print("MEDBRAINS FEATURE TRACKER — REMAINING WORK ANALYSIS")
    print("=" * 100)

    # 1. Overall counts
    print("\n" + "=" * 100)
    print("1. OVERALL FEATURE COUNTS BY STATUS")
    print("=" * 100)

    print(f"\n{'Status':<15} {'Count':>8} {'Percentage':>12}")
    print("-" * 40)
    for status in ["Done", "Partial", "In Progress", "Pending", "Deferred"]:
        count = grand_total.get(status, 0)
        pct = (count / total_features * 100) if total_features else 0
        print(f"{status:<15} {count:>8} {pct:>10.1f}%")
    print("-" * 40)
    print(f"{'TOTAL':<15} {total_features:>8} {'100.0%':>12}")

    # 2. Per-sheet breakdown
    print("\n" + "=" * 100)
    print("2. PER-SHEET BREAKDOWN")
    print("=" * 100)
    print(f"\n{'Sheet':<35} {'Total':>6} {'Done':>6} {'Partial':>8} {'InProg':>7} {'Pending':>8} {'Deferred':>9}")
    print("-" * 85)
    for sheet_name in wb.sheetnames:
        s = sheet_stats.get(sheet_name, {})
        total = s.get("_total", 0)
        done = s.get("Done", 0)
        partial = s.get("Partial", 0)
        inprog = s.get("In Progress", 0)
        pending = s.get("Pending", 0)
        deferred = s.get("Deferred", 0)
        print(f"{sheet_name:<35} {total:>6} {done:>6} {partial:>8} {inprog:>7} {pending:>8} {deferred:>9}")

    # 3. Modules with significant Pending features
    print("\n" + "=" * 100)
    print("3. MODULES WITH SIGNIFICANT PENDING FEATURES (sorted by Pending count)")
    print("=" * 100)

    # Build a list of (sheet, module, pending, partial, done, total)
    module_pending_list = []
    for (sheet, mod), counts in module_stats.items():
        pending = counts.get("Pending", 0)
        partial = counts.get("Partial", 0)
        done = counts.get("Done", 0)
        inprog = counts.get("In Progress", 0)
        deferred = counts.get("Deferred", 0)
        total = counts.get("_total", 0)
        completed = is_module_completed(mod)
        module_pending_list.append((sheet, mod, pending, partial, done, inprog, deferred, total, completed))

    module_pending_list.sort(key=lambda x: x[2], reverse=True)

    print(f"\n{'Sheet':<32} {'Module':<28} {'Pend':>5} {'Part':>5} {'Done':>5} {'InPr':>5} {'Def':>4} {'Tot':>5}  {'Built?'}")
    print("-" * 120)
    for sheet, mod, pending, partial, done, inprog, deferred, total, completed in module_pending_list:
        if pending == 0 and partial == 0:
            continue
        built = "YES" if completed else "---"
        marker = "" if completed else " <<<"
        print(f"{sheet:<32} {mod:<28} {pending:>5} {partial:>5} {done:>5} {inprog:>5} {deferred:>4} {total:>5}  {built}{marker}")

    # 4. Unbuilt modules with Pending features
    print("\n" + "=" * 100)
    print("4. UNBUILT MODULES — HIGHEST PENDING (Not yet completed)")
    print("=" * 100)

    unbuilt = [(s, m, p, pa, d, ip, df, t) for s, m, p, pa, d, ip, df, t, c in module_pending_list if not c and p > 0]
    unbuilt.sort(key=lambda x: x[2], reverse=True)

    print(f"\n{'#':>3} {'Sheet':<32} {'Module':<28} {'Pending':>8} {'Total':>6}")
    print("-" * 85)
    for i, (sheet, mod, pending, partial, done, inprog, deferred, total) in enumerate(unbuilt, 1):
        print(f"{i:>3} {sheet:<32} {mod:<28} {pending:>8} {total:>6}")

    # 5. Sub-module detail for top unbuilt modules
    print("\n" + "=" * 100)
    print("5. SUB-MODULE DETAIL FOR TOP UNBUILT MODULES")
    print("=" * 100)

    top_unbuilt_keys = set()
    for sheet, mod, pending, *_ in unbuilt[:15]:
        top_unbuilt_keys.add((sheet, mod))

    for (sheet, mod, sub), counts in sorted(submodule_stats.items(), key=lambda x: (x[0][0], x[0][1], x[0][2])):
        if (sheet, mod) not in top_unbuilt_keys:
            continue
        pending = counts.get("Pending", 0)
        partial = counts.get("Partial", 0)
        done = counts.get("Done", 0)
        total = counts.get("_total", 0)
        if pending == 0:
            continue
        print(f"  [{sheet}] {mod} > {sub}: {pending} Pending / {total} Total (Done={done}, Partial={partial})")

    # 6. Already-built modules with remaining Partial/Pending
    print("\n" + "=" * 100)
    print("6. ALREADY-BUILT MODULES WITH REMAINING PARTIAL/PENDING FEATURES")
    print("=" * 100)

    built_remaining = [(s, m, p, pa, d, ip, df, t) for s, m, p, pa, d, ip, df, t, c in module_pending_list if c and (p > 0 or pa > 0)]
    built_remaining.sort(key=lambda x: x[2] + x[3], reverse=True)

    print(f"\n{'Sheet':<32} {'Module':<28} {'Pending':>8} {'Partial':>8} {'Done':>6} {'Total':>6}")
    print("-" * 90)
    for sheet, mod, pending, partial, done, inprog, deferred, total in built_remaining:
        print(f"{sheet:<32} {mod:<28} {pending:>8} {partial:>8} {done:>6} {total:>6}")

    # 7. Recommendation
    print("\n" + "=" * 100)
    print("7. RECOMMENDED NEXT MODULES TO BUILD")
    print("=" * 100)

    print("""
Based on the analysis above, the recommended next modules are those with:
  - High Pending count
  - NOT yet built (not in COMPLETED_MODULES list)
  - Web platform features (not purely Mobile/TV — those are separate phases)

TOP CANDIDATES (by Pending feature count):
""")

    for i, (sheet, mod, pending, partial, done, inprog, deferred, total) in enumerate(unbuilt[:10], 1):
        completion_pct = (done / total * 100) if total else 0
        print(f"  {i}. {mod} ({sheet})")
        print(f"     Pending: {pending} | Total: {total} | Already Done: {done} ({completion_pct:.0f}%)")
        print()

    wb.close()


if __name__ == "__main__":
    main()

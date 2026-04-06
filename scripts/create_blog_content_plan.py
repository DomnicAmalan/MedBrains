#!/usr/bin/env python3
"""Create MedBrains Blog Content Plan Excel.

Generates a comprehensive blog content calendar with:
- 6 audience-targeted category sheets
- 12-month publishing schedule
- SEO keywords, CTA mapping, MedBrains feature tie-ins
- Summary dashboard sheet
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Blog_Content_Plan.xlsx"

# ── Styling ──────────────────────────────────────────────────────────────────
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
MONTH_FONT = Font(bold=True, size=11, color="1F4E79")
MONTH_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
CAT_FONT = Font(bold=True, size=10, color="2E75B6")
CAT_FILL = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Colors for status
STATUS_COLORS = {
    "Planned": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "In Progress": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    "Published": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}

# ── Column definitions per content sheet ─────────────────────────────────────
COLUMNS = [
    ("S.No", 5),
    ("Month", 10),
    ("Week", 6),
    ("Publish Date", 12),
    ("Blog Title", 55),
    ("Category", 20),
    ("Target Audience", 20),
    ("Format", 14),
    ("SEO Keywords (3-5)", 35),
    ("CTA / Conversion Goal", 25),
    ("MedBrains Feature Tie-In", 30),
    ("Word Count", 10),
    ("Status", 12),
    ("Author/Assigned To", 16),
    ("Notes", 25),
]

# ── Blog categories & colors ────────────────────────────────────────────────
CATEGORIES = {
    "Hospital Leadership": {"color": "4472C4", "audience": "CEOs, CMOs, Hospital Admins"},
    "Patient Education": {"color": "70AD47", "audience": "Patients & Families"},
    "Doctor & Clinical": {"color": "ED7D31", "audience": "Doctors, Nurses, Clinical Staff"},
    "Operations & Efficiency": {"color": "FFC000", "audience": "Hospital Managers, Dept Heads"},
    "Technology & Innovation": {"color": "5B9BD5", "audience": "CIOs, IT Managers, Health Tech"},
    "Vendor & Industry": {"color": "A5A5A5", "audience": "Vendors, Suppliers, Partners"},
    "Regulatory & Compliance": {"color": "BF4B4B", "audience": "Compliance Officers, QA Teams"},
    "Medical Tourism": {"color": "7030A0", "audience": "Intl Patients, Facilitators, Govt"},
}


# =============================================================================
# BLOG CONTENT: 6 sheets, ~156 blogs across 12 months
# =============================================================================

# ── Sheet 1: Hospital Leadership & Strategy (26 blogs) ──────────────────────
HOSPITAL_LEADERSHIP = [
    # Month 1 (Jul 2026)
    {"month": "Jul 2026", "week": "W1", "title": "Why 70% of Indian Hospitals Still Run on Paper — And What It Costs Them",
     "cat": "Hospital Leadership", "format": "Long-form", "keywords": "hospital digitization India, HMS benefits, paper-free hospital, hospital management system ROI",
     "cta": "Book HMS Demo", "feature": "Onboarding & Setup Wizard", "words": 2000},
    {"month": "Jul 2026", "week": "W2", "title": "The 7-Layer Hospital Configuration — How Top Hospitals Manage Multi-Campus Complexity",
     "cat": "Hospital Leadership", "format": "Explainer", "keywords": "multi-campus hospital management, hospital configuration hierarchy, enterprise HMS",
     "cta": "Download Architecture Whitepaper", "feature": "7-Layer Config Architecture", "words": 1800},
    {"month": "Jul 2026", "week": "W3", "title": "Hospital CEO's Guide to NABH Accreditation — What Your Software Must Do",
     "cat": "Hospital Leadership", "format": "Guide", "keywords": "NABH accreditation software, hospital quality management, NABH checklist HMS",
     "cta": "Get NABH Readiness Checklist", "feature": "Quality Management Module", "words": 2500},
    {"month": "Jul 2026", "week": "W4", "title": "How Patient Loyalty Programs Increase Hospital Revenue by 25%",
     "cat": "Hospital Leadership", "format": "Case Study", "keywords": "patient loyalty program hospital, hospital revenue growth, patient retention healthcare",
     "cta": "See Loyalty Module Demo", "feature": "Patient Referral & Loyalty Programs", "words": 1500},

    # Month 2 (Aug 2026)
    {"month": "Aug 2026", "week": "W1", "title": "5 Revenue Streams Hospitals Miss Without a Modern HMS",
     "cat": "Hospital Leadership", "format": "Listicle", "keywords": "hospital revenue optimization, HMS revenue features, hospital billing automation",
     "cta": "Revenue Audit Consultation", "feature": "Billing Module (18-tab)", "words": 1800},
    {"month": "Aug 2026", "week": "W3", "title": "Building a Medical Tourism Brand — The Hospital Software Foundation You Need",
     "cat": "Hospital Leadership", "format": "Strategy", "keywords": "medical tourism India hospital, international patient management, medical tourism software",
     "cta": "Medical Tourism Module Demo", "feature": "Medical Tourism & Intl Patients", "words": 2000},

    # Month 3 (Sep 2026)
    {"month": "Sep 2026", "week": "W1", "title": "Health Checkup Packages — The Most Underutilized Revenue Engine in Hospitals",
     "cat": "Hospital Leadership", "format": "Long-form", "keywords": "health checkup packages hospital, preventive care revenue, wellness package pricing",
     "cta": "Package Builder Demo", "feature": "Health Packages & Preventive Care", "words": 2000},
    {"month": "Sep 2026", "week": "W3", "title": "Concierge Medicine in Indian Hospitals — Premium Services That Patients Will Pay For",
     "cat": "Hospital Leadership", "format": "Trend", "keywords": "concierge medicine India, premium hospital services, VIP patient management",
     "cta": "Premium Services Module Demo", "feature": "Concierge & Premium Services", "words": 1800},

    # Month 4 (Oct 2026)
    {"month": "Oct 2026", "week": "W1", "title": "Hospital Marketing in 2027 — Why Your HMS Is Your Most Powerful Marketing Tool",
     "cat": "Hospital Leadership", "format": "Long-form", "keywords": "hospital marketing strategy, healthcare marketing digital, hospital brand building",
     "cta": "Marketing Feature Tour", "feature": "Micro Website + CRM", "words": 2200},
    {"month": "Oct 2026", "week": "W3", "title": "Doctor Profiles That Convert — How Online Reputation Drives Appointments",
     "cat": "Hospital Leadership", "format": "Data-backed", "keywords": "doctor online profile, hospital appointment conversion, doctor reputation management",
     "cta": "Doctor Profile Module Demo", "feature": "Doctor Profile & Reputation", "words": 1500},

    # Month 5 (Nov 2026)
    {"month": "Nov 2026", "week": "W1", "title": "Cost Transparency — Why Hospitals That Publish Prices Get More Patients",
     "cat": "Hospital Leadership", "format": "Research", "keywords": "hospital cost transparency, healthcare pricing transparency, treatment cost estimator",
     "cta": "Cost Estimator Demo", "feature": "Cost Transparency & Estimation", "words": 2000},
    {"month": "Nov 2026", "week": "W3", "title": "Multi-Hospital Benchmarking — How Leading Chains Compare Performance Across Sites",
     "cat": "Hospital Leadership", "format": "Guide", "keywords": "hospital benchmarking KPI, multi-hospital performance, healthcare analytics dashboard",
     "cta": "Benchmarking Network Demo", "feature": "Benchmarking Network", "words": 1800},

    # Month 6 (Dec 2026)
    {"month": "Dec 2026", "week": "W1", "title": "2027 Hospital Technology Predictions — 10 Trends Every Hospital CEO Should Know",
     "cat": "Hospital Leadership", "format": "Listicle", "keywords": "hospital technology trends 2027, healthcare IT predictions, digital hospital transformation",
     "cta": "Schedule Strategy Call", "feature": "AI Chatbot + Digital Twin", "words": 2500},
    {"month": "Dec 2026", "week": "W3", "title": "The True Cost of Running Multiple Hospital Software Systems vs One Integrated HMS",
     "cat": "Hospital Leadership", "format": "Comparison", "keywords": "integrated HMS vs multiple software, hospital software consolidation, HMS total cost ownership",
     "cta": "TCO Calculator Download", "feature": "Full Platform", "words": 2000},

    # Months 7-12 (Jan-Jun 2027)
    {"month": "Jan 2027", "week": "W1", "title": "Hospital Board Reporting — Dashboards That Actually Help Decision-Making",
     "cat": "Hospital Leadership", "format": "Guide", "keywords": "hospital dashboard KPI, healthcare board reporting, hospital management dashboard",
     "cta": "Dashboard Builder Demo", "feature": "Dashboard Builder", "words": 1800},
    {"month": "Jan 2027", "week": "W3", "title": "IPD Revenue Optimization — From Admission to Discharge, Every Rupee Tracked",
     "cat": "Hospital Leadership", "format": "Deep Dive", "keywords": "IPD billing optimization, hospital inpatient revenue, discharge billing automation",
     "cta": "IPD Module Demo", "feature": "IPD Module (102 features)", "words": 2000},
    {"month": "Feb 2027", "week": "W1", "title": "Emergency Department as Profit Center — Not Just a Compliance Requirement",
     "cat": "Hospital Leadership", "format": "Strategy", "keywords": "emergency department revenue, ER hospital profitability, emergency module HMS",
     "cta": "Emergency Module Demo", "feature": "Emergency Module", "words": 1800},
    {"month": "Feb 2027", "week": "W3", "title": "Hospital Insurance & TPA Management — Stop Losing Money on Claim Rejections",
     "cat": "Hospital Leadership", "format": "Problem-Solution", "keywords": "hospital TPA management, insurance claim rejection hospital, prior authorization software",
     "cta": "Insurance Module Demo", "feature": "Insurance & TPA Module", "words": 1800},
    {"month": "Mar 2027", "week": "W1", "title": "How Specialty Care Programs (Maternity, Cardiac, Diabetes) Build Hospital Brand",
     "cat": "Hospital Leadership", "format": "Case Study", "keywords": "specialty care program hospital, bundled care package healthcare, maternity program hospital",
     "cta": "Specialty Programs Demo", "feature": "Specialty Care Programs", "words": 2000},
    {"month": "Mar 2027", "week": "W3", "title": "Patient Outcomes Dashboard — Showcasing Quality to Attract More Patients",
     "cat": "Hospital Leadership", "format": "Explainer", "keywords": "patient outcomes dashboard, hospital quality showcase, clinical excellence display",
     "cta": "Outcomes Module Demo", "feature": "Patient Outcomes & Quality Showcase", "words": 1500},
    {"month": "Apr 2027", "week": "W1", "title": "From OPD Chaos to Orchestrated Flow — How Smart Queues Transform Patient Experience",
     "cat": "Hospital Leadership", "format": "Before/After", "keywords": "OPD queue management, hospital patient flow, outpatient department optimization",
     "cta": "OPD Module Demo", "feature": "OPD + Front Office", "words": 1800},
    {"month": "Apr 2027", "week": "W3", "title": "Hospital Procurement Savings — Vendor Analytics That Pay for Your HMS",
     "cat": "Hospital Leadership", "format": "ROI Analysis", "keywords": "hospital procurement optimization, vendor performance analytics, hospital supply chain",
     "cta": "Procurement Module Demo", "feature": "Procurement Module", "words": 1500},
    {"month": "May 2027", "week": "W1", "title": "Green Hospital Operations — Sustainability as a Competitive Advantage",
     "cat": "Hospital Leadership", "format": "Trend", "keywords": "green hospital India, sustainable healthcare operations, hospital energy management",
     "cta": "Green Hospital Features Tour", "feature": "Green Hospital Module", "words": 1800},
    {"month": "May 2027", "week": "W3", "title": "Referral Networks — How Doctor-to-Doctor Referrals Build Hospital Volume",
     "cat": "Hospital Leadership", "format": "Data-backed", "keywords": "doctor referral network, hospital referral tracking, physician referral management",
     "cta": "Referral Module Demo", "feature": "Patient Referral & Loyalty", "words": 1500},
    {"month": "Jun 2027", "week": "W1", "title": "Hospital Pre-Visit Digital Journey — Why Patients Choose Digitally-Ready Hospitals",
     "cat": "Hospital Leadership", "format": "Long-form", "keywords": "digital patient journey, hospital pre-registration online, QR check-in hospital",
     "cta": "Digital Journey Demo", "feature": "Pre-Visit Digital Journey", "words": 2000},
    {"month": "Jun 2027", "week": "W3", "title": "Year 1 of HMS Implementation — What to Expect and How to Maximize ROI",
     "cat": "Hospital Leadership", "format": "Guide", "keywords": "HMS implementation timeline, hospital software ROI, HMS deployment best practices",
     "cta": "Implementation Consultation", "feature": "Full Platform", "words": 2500},
]

# ── Sheet 2: Patient Education & Wellness (26 blogs) ────────────────────────
PATIENT_EDUCATION = [
    {"month": "Jul 2026", "week": "W1", "title": "Your First Hospital Visit — What to Bring, What to Expect, and How to Prepare",
     "cat": "Patient Education", "format": "Guide", "keywords": "first hospital visit India, hospital visit preparation, what to bring to hospital",
     "cta": "Download Visit Checklist", "feature": "Pre-Visit Digital Journey", "words": 1200},
    {"month": "Jul 2026", "week": "W2", "title": "Understanding Your Hospital Bill — A Patient's Guide to Healthcare Costs in India",
     "cat": "Patient Education", "format": "Explainer", "keywords": "hospital bill explanation India, healthcare costs breakdown, understand medical bill",
     "cta": "Try Cost Estimator", "feature": "Cost Transparency & Estimation", "words": 1500},
    {"month": "Jul 2026", "week": "W4", "title": "Health Checkup Packages — Which One Is Right for Your Age and Risk Profile?",
     "cat": "Patient Education", "format": "Quiz/Interactive", "keywords": "health checkup package India, annual health screening, preventive health checkup",
     "cta": "Browse Health Packages", "feature": "Health Packages & Preventive Care", "words": 1500},

    {"month": "Aug 2026", "week": "W2", "title": "How to Choose the Right Doctor — 7 Things to Look for Beyond Qualifications",
     "cat": "Patient Education", "format": "Listicle", "keywords": "choose doctor India, find best doctor hospital, doctor selection tips",
     "cta": "Search Our Doctors", "feature": "Doctor Profile & Reputation", "words": 1200},
    {"month": "Aug 2026", "week": "W4", "title": "What Is a Health Insurance Pre-Authorization and Why Does Your Hospital Need It?",
     "cat": "Patient Education", "format": "FAQ", "keywords": "pre-authorization health insurance, hospital insurance approval, cashless hospitalization process",
     "cta": "Check Insurance Coverage", "feature": "Insurance & TPA Module", "words": 1200},

    {"month": "Sep 2026", "week": "W2", "title": "Lab Reports Decoded — What Your Blood Test Numbers Actually Mean",
     "cat": "Patient Education", "format": "Explainer", "keywords": "blood test results explained, lab report normal range, understand CBC report",
     "cta": "View Reports Online", "feature": "Lab Module + Patient App", "words": 1800},
    {"month": "Sep 2026", "week": "W4", "title": "Vaccination Schedule for Children — From Birth to 18 Years (India NIS + Optional)",
     "cat": "Patient Education", "format": "Reference", "keywords": "vaccination schedule India children, immunization chart baby, pediatric vaccination",
     "cta": "Book Vaccination Appointment", "feature": "Pediatric Wellness Program", "words": 2000},

    {"month": "Oct 2026", "week": "W2", "title": "Diabetes Management at Home — Daily Monitoring, Diet, and When to See Your Doctor",
     "cat": "Patient Education", "format": "Guide", "keywords": "diabetes management home India, blood sugar monitoring tips, diabetes diet plan",
     "cta": "Join Diabetes Care Program", "feature": "Diabetes Management Program", "words": 1800},
    {"month": "Oct 2026", "week": "W4", "title": "Coming to India for Medical Treatment — A Complete Guide for International Patients",
     "cat": "Patient Education", "format": "Mega Guide", "keywords": "medical tourism India guide, treatment in India international patient, medical visa India",
     "cta": "Contact International Desk", "feature": "Medical Tourism & Intl Patients", "words": 3000},

    {"month": "Nov 2026", "week": "W2", "title": "What Happens During a Hospital Stay — A Step-by-Step IPD Journey Explained",
     "cat": "Patient Education", "format": "Walkthrough", "keywords": "hospital admission process India, IPD process hospital, hospital stay what to expect",
     "cta": "Pre-Register Online", "feature": "IPD Module + Pre-Visit Journey", "words": 1500},
    {"month": "Nov 2026", "week": "W4", "title": "Pregnancy Care Checklist — Trimester-by-Trimester Guide with Hospital Visit Schedule",
     "cat": "Patient Education", "format": "Checklist", "keywords": "pregnancy care checklist India, antenatal visit schedule, maternity hospital guide",
     "cta": "View Maternity Package", "feature": "Maternity Care Package", "words": 2000},

    {"month": "Dec 2026", "week": "W2", "title": "Emergency Room vs Urgent Care — When to Go Where and What to Expect",
     "cat": "Patient Education", "format": "Decision Guide", "keywords": "emergency room vs urgent care India, when to go ER, hospital emergency department",
     "cta": "Check ER Wait Times", "feature": "Emergency Module + Wait Times", "words": 1200},
    {"month": "Dec 2026", "week": "W4", "title": "Heart Health After 40 — Screenings, Lifestyle Changes, and Warning Signs",
     "cat": "Patient Education", "format": "Health Guide", "keywords": "heart health screening 40, cardiac checkup India, heart disease prevention",
     "cta": "Book Cardiac Checkup", "feature": "Cardiac Rehab Program", "words": 1800},

    {"month": "Jan 2027", "week": "W2", "title": "Mental Health Is Health — When and How to Seek Professional Help",
     "cat": "Patient Education", "format": "Awareness", "keywords": "mental health help India, counseling appointment hospital, therapy session booking",
     "cta": "Book Counseling Session", "feature": "Mental Wellness Program", "words": 1500},
    {"month": "Jan 2027", "week": "W4", "title": "Understanding Radiology — X-Ray, CT, MRI, Ultrasound: Which Scan and Why",
     "cat": "Patient Education", "format": "Explainer", "keywords": "radiology scans explained, MRI vs CT scan, types of medical imaging",
     "cta": "Book Radiology Appointment", "feature": "Radiology Module", "words": 1500},

    {"month": "Feb 2027", "week": "W2", "title": "Patient Rights in Indian Hospitals — What You Can Expect and Demand",
     "cat": "Patient Education", "format": "Know Your Rights", "keywords": "patient rights India hospital, consumer protection healthcare, hospital patient charter",
     "cta": "Read Our Patient Charter", "feature": "Consent Management", "words": 1500},
    {"month": "Feb 2027", "week": "W4", "title": "Post-Surgery Recovery Guide — Orthopedic, Cardiac, and General Surgery Tips",
     "cat": "Patient Education", "format": "Recovery Guide", "keywords": "post surgery recovery tips, hospital discharge care, surgery aftercare India",
     "cta": "Join Rehab Program", "feature": "Specialty Care Programs", "words": 1800},

    {"month": "Mar 2027", "week": "W2", "title": "How to Read Your Hospital Discharge Summary — Every Section Explained",
     "cat": "Patient Education", "format": "Explainer", "keywords": "discharge summary explained, hospital discharge paper, medical report understanding",
     "cta": "Access Reports Online", "feature": "Patient App + Reports", "words": 1200},
    {"month": "Mar 2027", "week": "W4", "title": "Senior Citizen Health — Annual Checkups, Fall Prevention, and Caregiver Support",
     "cat": "Patient Education", "format": "Guide", "keywords": "senior health checkup India, elderly care hospital, fall prevention elderly",
     "cta": "View Elderly Care Program", "feature": "Elderly Care Program", "words": 1800},

    {"month": "Apr 2027", "week": "W2", "title": "Blood Donation FAQ — Who Can Donate, Process, Benefits, and Myths Busted",
     "cat": "Patient Education", "format": "FAQ", "keywords": "blood donation India FAQ, blood bank hospital, who can donate blood",
     "cta": "Schedule Blood Donation", "feature": "Blood Bank Module", "words": 1200},
    {"month": "Apr 2027", "week": "W4", "title": "Telemedicine — How Video Consultations Work and When They're Right for You",
     "cat": "Patient Education", "format": "How-To", "keywords": "telemedicine consultation India, online doctor visit, video consultation hospital",
     "cta": "Book Video Consult", "feature": "Telemedicine Module", "words": 1200},

    {"month": "May 2027", "week": "W2", "title": "Understanding Health Insurance Terms — Copay, Deductible, Sub-Limits Explained Simply",
     "cat": "Patient Education", "format": "Glossary", "keywords": "health insurance terms India, copay meaning, insurance deductible explained",
     "cta": "Insurance Coverage Calculator", "feature": "Insurance Coverage Calculator", "words": 1500},
    {"month": "May 2027", "week": "W4", "title": "Weight Management — Beyond Diets: Medical Programs That Actually Work",
     "cat": "Patient Education", "format": "Program Guide", "keywords": "weight management program hospital, bariatric surgery India, medical weight loss",
     "cta": "Consult Bariatric Team", "feature": "Bariatric & Weight Management", "words": 1500},

    {"month": "Jun 2027", "week": "W2", "title": "Cancer Screening Guide — Which Tests, At What Age, How Often",
     "cat": "Patient Education", "format": "Reference Guide", "keywords": "cancer screening tests India, preventive cancer checkup, early cancer detection",
     "cta": "Book Cancer Screening", "feature": "Oncology Support Program", "words": 2000},
    {"month": "Jun 2027", "week": "W4", "title": "Your Digital Health Record — How ABDM and Health IDs Keep Your Data Safe and Accessible",
     "cat": "Patient Education", "format": "Explainer", "keywords": "ABDM health ID India, digital health record, Ayushman Bharat Digital Mission patient",
     "cta": "Link Your Health ID", "feature": "Patient App + ABDM", "words": 1500},
]

# ── Sheet 3: Doctor & Clinical Excellence (26 blogs) ────────────────────────
DOCTOR_CLINICAL = [
    {"month": "Jul 2026", "week": "W1", "title": "Building Your Online Doctor Profile — Why It Matters More Than Your Visiting Card",
     "cat": "Doctor & Clinical", "format": "Practical Guide", "keywords": "doctor online profile tips, physician personal branding, doctor digital presence",
     "cta": "Set Up Your Profile", "feature": "Doctor Profile & Reputation", "words": 1500},
    {"month": "Jul 2026", "week": "W3", "title": "Clinical Decision Support — How Smart Alerts Prevent Prescription Errors",
     "cat": "Doctor & Clinical", "format": "Evidence-based", "keywords": "clinical decision support system, prescription error prevention, drug interaction alert HMS",
     "cta": "See CDSS in Action", "feature": "Pharmacy + Order Sets", "words": 1800},

    {"month": "Aug 2026", "week": "W1", "title": "Order Sets That Save 15 Minutes Per Patient — Templates Every Department Needs",
     "cat": "Doctor & Clinical", "format": "Practical", "keywords": "clinical order sets templates, hospital order set builder, standing orders hospital",
     "cta": "Order Set Builder Demo", "feature": "Order Sets Module", "words": 1500},
    {"month": "Aug 2026", "week": "W3", "title": "Lab Quality Control — Westgard Rules, Levey-Jennings Charts, and Why They Matter",
     "cat": "Doctor & Clinical", "format": "Technical", "keywords": "Westgard rules laboratory, Levey-Jennings chart QC, lab quality control NABL",
     "cta": "QC Module Demo", "feature": "Lab QC & Compliance", "words": 2000},

    {"month": "Sep 2026", "week": "W1", "title": "Antibiotic Stewardship in Practice — WHO AWaRe Classification and Your Hospital",
     "cat": "Doctor & Clinical", "format": "Clinical", "keywords": "antibiotic stewardship program, WHO AWaRe classification, antimicrobial resistance hospital",
     "cta": "Stewardship Module Demo", "feature": "Pharmacy Stewardship", "words": 2000},
    {"month": "Sep 2026", "week": "W3", "title": "ICU Documentation That Actually Helps — From Ventilator Logs to SOFA Scores",
     "cat": "Doctor & Clinical", "format": "Workflow", "keywords": "ICU documentation software, SOFA score calculation, ventilator management HMS",
     "cta": "ICU Module Demo", "feature": "ICU Module (24 features)", "words": 1800},

    {"month": "Oct 2026", "week": "W1", "title": "NDPS Compliance Made Simple — Register, Track, and Report Without the Paperwork",
     "cat": "Doctor & Clinical", "format": "Compliance", "keywords": "NDPS compliance hospital, narcotic register digital, controlled substance tracking",
     "cta": "NDPS Register Demo", "feature": "Pharmacy NDPS Register", "words": 1500},
    {"month": "Oct 2026", "week": "W3", "title": "Infection Control Surveillance — HAI Tracking That NABH Auditors Love",
     "cat": "Doctor & Clinical", "format": "Best Practice", "keywords": "hospital infection control surveillance, HAI tracking software, NABH infection control",
     "cta": "Infection Control Demo", "feature": "Infection Control Module", "words": 1800},

    {"month": "Nov 2026", "week": "W1", "title": "Patient Consent Management — Digital Workflows That Are Legally Watertight",
     "cat": "Doctor & Clinical", "format": "Legal/Clinical", "keywords": "digital consent management hospital, informed consent software, patient consent workflow",
     "cta": "Consent Module Demo", "feature": "Consent Management Module", "words": 1500},
    {"month": "Nov 2026", "week": "W3", "title": "Critical Value Reporting — From Lab Result to Doctor Notification in Under 10 Minutes",
     "cat": "Doctor & Clinical", "format": "Workflow", "keywords": "critical value reporting lab, lab critical alert system, turnaround time laboratory",
     "cta": "Lab Alerts Demo", "feature": "Lab Critical Alerts", "words": 1500},

    {"month": "Dec 2026", "week": "W1", "title": "OT Scheduling That Actually Works — Surgeon Preferences, Equipment, and Anesthesia Coordination",
     "cat": "Doctor & Clinical", "format": "Operations", "keywords": "OT scheduling software, operation theatre management, surgical scheduling hospital",
     "cta": "OT Module Demo", "feature": "IPD OT Module", "words": 1800},
    {"month": "Dec 2026", "week": "W3", "title": "Telemedicine Best Practices for Doctors — Building Rapport Through a Screen",
     "cat": "Doctor & Clinical", "format": "Tips", "keywords": "telemedicine doctor tips, video consultation best practice, telehealth physician guide",
     "cta": "Enable Telemedicine", "feature": "Telemedicine Module", "words": 1200},

    {"month": "Jan 2027", "week": "W1", "title": "Diet Orders in Hospital — How Standardized Meal Plans Improve Patient Outcomes",
     "cat": "Doctor & Clinical", "format": "Evidence-based", "keywords": "hospital diet management, therapeutic diet plan, diet order system hospital",
     "cta": "Diet Module Demo", "feature": "Diet & Kitchen Module", "words": 1500},
    {"month": "Jan 2027", "week": "W3", "title": "Phlebotomy Queue Management — Reducing Wait Times and Hemolysis Rates",
     "cat": "Doctor & Clinical", "format": "Process", "keywords": "phlebotomy management system, blood collection queue, sample collection hospital",
     "cta": "Phlebotomy Module Demo", "feature": "Lab Phlebotomy Queue", "words": 1200},

    {"month": "Feb 2027", "week": "W1", "title": "Blood Bank Safety Chain — From Donor to Transfusion, Every Step Verified",
     "cat": "Doctor & Clinical", "format": "Safety", "keywords": "blood bank safety hospital, blood transfusion verification, blood component tracking",
     "cta": "Blood Bank Module Demo", "feature": "Blood Bank Module", "words": 1800},
    {"month": "Feb 2027", "week": "W3", "title": "CSSD Sterilization Tracking — Autoclave Logs, Biological Indicators, and Recall",
     "cat": "Doctor & Clinical", "format": "Compliance", "keywords": "CSSD sterilization tracking, autoclave log hospital, surgical instrument sterilization",
     "cta": "CSSD Module Demo", "feature": "CSSD Module", "words": 1500},

    {"month": "Mar 2027", "week": "W1", "title": "Clinical Pathways vs Order Sets — When to Use Which for Better Patient Outcomes",
     "cat": "Doctor & Clinical", "format": "Comparison", "keywords": "clinical pathways vs order sets, standardized care hospital, evidence-based care pathway",
     "cta": "Explore Order Sets", "feature": "Order Sets + Specialty Programs", "words": 1500},
    {"month": "Mar 2027", "week": "W3", "title": "Histopathology and Molecular Reports — Structured Reporting That Oncologists Need",
     "cat": "Doctor & Clinical", "format": "Specialty", "keywords": "histopathology report structured, molecular pathology software, cancer lab report",
     "cta": "Specialized Lab Demo", "feature": "Lab Specialized Reports", "words": 1800},

    {"month": "Apr 2027", "week": "W1", "title": "Nursing Documentation That Saves Time — Templates, Checklists, and Smart Forms",
     "cat": "Doctor & Clinical", "format": "Practical", "keywords": "nursing documentation software, nurse charting template, clinical documentation HMS",
     "cta": "Nursing Module Demo", "feature": "IPD Nursing Documentation", "words": 1500},
    {"month": "Apr 2027", "week": "W3", "title": "Drug Formulary Management — How DTC Committees Can Use Software to Enforce Compliance",
     "cat": "Doctor & Clinical", "format": "Governance", "keywords": "hospital drug formulary, DTC committee software, formulary compliance management",
     "cta": "Formulary Module Demo", "feature": "Pharmacy Formulary", "words": 1500},

    {"month": "May 2027", "week": "W1", "title": "ADR Reporting Made Easy — Pharmacovigilance That Protects Patients and Your License",
     "cat": "Doctor & Clinical", "format": "Regulatory", "keywords": "adverse drug reaction reporting, pharmacovigilance hospital, ADR reporting PvPI",
     "cta": "ADR Module Demo", "feature": "Regulatory ADR Reports", "words": 1500},
    {"month": "May 2027", "week": "W3", "title": "Emergency MLC Documentation — What Doctors Must Record and Why Software Helps",
     "cat": "Doctor & Clinical", "format": "Legal/Clinical", "keywords": "MLC documentation hospital, medico legal case format, emergency MLC register",
     "cta": "Emergency Module Demo", "feature": "Emergency MLC", "words": 1500},

    {"month": "Jun 2027", "week": "W1", "title": "Camp Management — Running Successful Health Camps with Digital Registration and Follow-Up",
     "cat": "Doctor & Clinical", "format": "How-To", "keywords": "health camp management software, medical camp registration, community health camp",
     "cta": "Camp Module Demo", "feature": "Camp Management Module", "words": 1500},
    {"month": "Jun 2027", "week": "W3", "title": "Clinical Audit Made Easy — Quality Indicators That Drive Real Improvement",
     "cat": "Doctor & Clinical", "format": "Quality", "keywords": "clinical audit software hospital, healthcare quality indicators, clinical governance HMS",
     "cta": "Quality Module Demo", "feature": "Quality Management Module", "words": 1800},
]

# ── Sheet 4: Operations & Efficiency (26 blogs) ─────────────────────────────
OPERATIONS = [
    {"month": "Jul 2026", "week": "W2", "title": "Front Office Transformation — From Chaotic Reception to Smooth Patient Flow",
     "cat": "Operations & Efficiency", "format": "Before/After", "keywords": "hospital front office management, patient reception software, hospital queue management",
     "cta": "Front Office Module Demo", "feature": "Front Office & Reception", "words": 1500},
    {"month": "Jul 2026", "week": "W4", "title": "Hospital Inventory That Doesn't Run Out — ABC-VED Analysis and Auto Reorder",
     "cat": "Operations & Efficiency", "format": "Technical", "keywords": "hospital inventory management, ABC VED analysis medical, auto reorder hospital stock",
     "cta": "Inventory Module Demo", "feature": "Inventory Phase 2", "words": 1800},

    {"month": "Aug 2026", "week": "W2", "title": "HR for Hospitals — Shift Management, Leave, and Compliance in One System",
     "cat": "Operations & Efficiency", "format": "Overview", "keywords": "hospital HR management software, nurse shift scheduling, healthcare staff management",
     "cta": "HR Module Demo", "feature": "HR & Staff Management", "words": 1500},
    {"month": "Aug 2026", "week": "W4", "title": "Hospital Pharmacy Operations — From Drug Catalog to Batch Tracking to NDPS Register",
     "cat": "Operations & Efficiency", "format": "Comprehensive", "keywords": "hospital pharmacy management system, drug batch tracking, pharmacy operations software",
     "cta": "Pharmacy Module Demo", "feature": "Pharmacy Module (28 features)", "words": 2000},

    {"month": "Sep 2026", "week": "W2", "title": "Hospital Billing Automation — GST, TDS, Journal Entries Without the Spreadsheet Nightmare",
     "cat": "Operations & Efficiency", "format": "Pain-point", "keywords": "hospital billing automation, GST healthcare services, hospital accounting software",
     "cta": "Billing Module Demo", "feature": "Billing Module Phase 3", "words": 1800},
    {"month": "Sep 2026", "week": "W4", "title": "Facilities Management for Hospitals — MGPS, Fire Safety, Water Quality in One Dashboard",
     "cat": "Operations & Efficiency", "format": "Overview", "keywords": "hospital facilities management, MGPS monitoring software, fire safety hospital compliance",
     "cta": "Facilities Module Demo", "feature": "Facilities Management", "words": 1800},

    {"month": "Oct 2026", "week": "W2", "title": "Biomedical Waste Management — Digital Tracking from Ward to Disposal",
     "cat": "Operations & Efficiency", "format": "Compliance", "keywords": "biomedical waste management hospital, BMW rules 2016, waste segregation tracking hospital",
     "cta": "Waste Management Demo", "feature": "Infection Control + BMW", "words": 1500},
    {"month": "Oct 2026", "week": "W4", "title": "Hospital Housekeeping — Room Turnaround, Linen Management, and Cleaning Schedules",
     "cat": "Operations & Efficiency", "format": "Operational", "keywords": "hospital housekeeping management, room turnaround software, hospital linen tracking",
     "cta": "Housekeeping Module Demo", "feature": "Housekeeping Module", "words": 1200},

    {"month": "Nov 2026", "week": "W2", "title": "Patient Consumable Tracking — Every Syringe, Glove, and Implant Accounted For",
     "cat": "Operations & Efficiency", "format": "Process", "keywords": "patient consumable tracking hospital, implant registry hospital, surgical item billing",
     "cta": "Consumables Module Demo", "feature": "Patient Consumables + Implant Registry", "words": 1500},
    {"month": "Nov 2026", "week": "W4", "title": "Hospital Work Orders — From Complaint to Resolution with SLA Tracking",
     "cat": "Operations & Efficiency", "format": "Workflow", "keywords": "hospital work order system, maintenance management hospital, facility work order tracking",
     "cta": "Work Order Module Demo", "feature": "FMS Work Orders", "words": 1200},

    {"month": "Dec 2026", "week": "W2", "title": "Outsourced Lab Management — Tracking Samples, Results, and TAT with External Labs",
     "cat": "Operations & Efficiency", "format": "Process", "keywords": "outsourced laboratory management, external lab tracking, sample dispatch hospital",
     "cta": "Outsourced Lab Demo", "feature": "Lab Outsourced Orders", "words": 1500},
    {"month": "Dec 2026", "week": "W4", "title": "Hospital Duty Roster — Fair Scheduling, Swap Requests, and On-Call Management",
     "cat": "Operations & Efficiency", "format": "HR Operations", "keywords": "hospital duty roster software, nurse scheduling shift swap, on-call management hospital",
     "cta": "Roster Module Demo", "feature": "HR Duty Roster", "words": 1500},

    {"month": "Jan 2027", "week": "W2", "title": "Purchase Orders to Payment — End-to-End Hospital Procurement Workflow",
     "cat": "Operations & Efficiency", "format": "E2E Guide", "keywords": "hospital purchase order workflow, procurement management hospital, vendor payment tracking",
     "cta": "Procurement Module Demo", "feature": "Procurement + Supplier Payments", "words": 1800},
    {"month": "Jan 2027", "week": "W4", "title": "Hospital Kitchen at Scale — Diet Orders, Meal Prep, and Delivery Tracking for 500+ Beds",
     "cat": "Operations & Efficiency", "format": "Scale Story", "keywords": "hospital kitchen management, diet order software, large hospital meal management",
     "cta": "Diet Kitchen Module Demo", "feature": "Diet & Kitchen Module", "words": 1500},

    {"month": "Feb 2027", "week": "W2", "title": "Compliance Calendar — Never Miss a License Renewal, Inspection, or Filing Deadline",
     "cat": "Operations & Efficiency", "format": "Tool Guide", "keywords": "hospital compliance calendar, regulatory deadline tracking, license renewal hospital",
     "cta": "Compliance Calendar Demo", "feature": "Regulatory Compliance Calendar", "words": 1200},
    {"month": "Feb 2027", "week": "W4", "title": "Bank Reconciliation for Hospitals — Matching Payments Without Manual Effort",
     "cat": "Operations & Efficiency", "format": "Finance", "keywords": "hospital bank reconciliation, healthcare payment matching, hospital finance automation",
     "cta": "Bank Recon Module Demo", "feature": "Billing Bank Reconciliation", "words": 1500},

    {"month": "Mar 2027", "week": "W2", "title": "Hospital Printing & Forms — From Consent Forms to Discharge Summaries, All Branded",
     "cat": "Operations & Efficiency", "format": "Overview", "keywords": "hospital print management, medical form templates, branded hospital documents",
     "cta": "Print Builder Demo", "feature": "Print Template Builder", "words": 1200},
    {"month": "Mar 2027", "week": "W4", "title": "Dead Stock and Near-Expiry Management — Stop Losing Money in Your Store Room",
     "cat": "Operations & Efficiency", "format": "Problem-Solution", "keywords": "hospital dead stock management, near-expiry drug tracking, FEFO hospital pharmacy",
     "cta": "Batch Management Demo", "feature": "Pharmacy Batch & Expiry", "words": 1500},

    {"month": "Apr 2027", "week": "W2", "title": "Hospital Energy Management — Track Consumption, Reduce Costs, Meet Green Standards",
     "cat": "Operations & Efficiency", "format": "Sustainability", "keywords": "hospital energy management system, healthcare energy consumption, green hospital operations",
     "cta": "Energy Module Demo", "feature": "FMS Energy Readings", "words": 1500},
    {"month": "Apr 2027", "week": "W4", "title": "Patient Appointment No-Shows — SMS, WhatsApp Reminders That Actually Reduce Them",
     "cat": "Operations & Efficiency", "format": "Data-backed", "keywords": "reduce patient no-shows, appointment reminder hospital, healthcare SMS reminders",
     "cta": "Communication Module Demo", "feature": "Communication Module", "words": 1200},

    {"month": "May 2027", "week": "W2", "title": "PCPNDT Compliance for Radiology — Digital Forms, Quarterly Reporting, Zero Violations",
     "cat": "Operations & Efficiency", "format": "Regulatory", "keywords": "PCPNDT compliance software, radiology gender determination prevention, PCPNDT form F",
     "cta": "PCPNDT Module Demo", "feature": "Regulatory PCPNDT Forms", "words": 1500},
    {"month": "May 2027", "week": "W4", "title": "Hospital Fire Safety Compliance — Equipment Tracking, Inspections, Drills, and NOC",
     "cat": "Operations & Efficiency", "format": "Safety", "keywords": "hospital fire safety compliance, fire equipment tracking, fire NOC hospital",
     "cta": "Fire Safety Module Demo", "feature": "FMS Fire Safety", "words": 1500},

    {"month": "Jun 2027", "week": "W2", "title": "ERP Integration for Hospitals — Exporting to Tally, SAP Without Manual Data Entry",
     "cat": "Operations & Efficiency", "format": "Integration", "keywords": "hospital ERP integration, Tally hospital billing, SAP healthcare export",
     "cta": "ERP Export Demo", "feature": "Billing ERP Export", "words": 1500},
    {"month": "Jun 2027", "week": "W4", "title": "GSTR Filing for Hospitals — Generating Returns from Your Billing Data Automatically",
     "cat": "Operations & Efficiency", "format": "Tax/Finance", "keywords": "hospital GST filing, GSTR hospital billing, healthcare GST compliance",
     "cta": "GST Module Demo", "feature": "Billing GST & TDS", "words": 1500},
]

# ── Sheet 5: Technology & Vendor (26 blogs) ──────────────────────────────────
TECH_VENDOR = [
    {"month": "Jul 2026", "week": "W1", "title": "Rust in Healthcare — Why We Chose Memory Safety Over Move-Fast-Break-Things",
     "cat": "Technology & Innovation", "format": "Technical", "keywords": "Rust healthcare software, memory safe healthcare, Rust hospital management system",
     "cta": "Tech Architecture Guide", "feature": "Backend Architecture", "words": 2000},
    {"month": "Jul 2026", "week": "W3", "title": "Multi-Tenant Hospital Software — How One Codebase Serves 100 Hospitals Securely",
     "cat": "Technology & Innovation", "format": "Architecture", "keywords": "multi-tenant HMS, hospital software SaaS, row level security healthcare",
     "cta": "Security Whitepaper", "feature": "RLS + Multi-tenancy", "words": 2000},

    {"month": "Aug 2026", "week": "W1", "title": "Hospital API Ecosystem — Why Open APIs Are the Future of Healthcare Integration",
     "cat": "Technology & Innovation", "format": "Vision", "keywords": "hospital API integration, healthcare open API, HMS REST API",
     "cta": "API Documentation", "feature": "API Ecosystem", "words": 1800},
    {"month": "Aug 2026", "week": "W3", "title": "AI Chatbots in Hospitals — Beyond FAQ: Symptom Checking, Appointment Booking, Bill Queries",
     "cat": "Technology & Innovation", "format": "Trend", "keywords": "hospital AI chatbot, healthcare virtual assistant, WhatsApp bot hospital",
     "cta": "Chatbot Module Demo", "feature": "AI Chatbot & Virtual Assistant", "words": 1800},

    {"month": "Sep 2026", "week": "W1", "title": "Hospital Data Security — Encryption, Audit Trails, and Break-Glass Access in Practice",
     "cat": "Technology & Innovation", "format": "Security", "keywords": "hospital data security, healthcare data encryption, audit trail HMS",
     "cta": "Security Features Tour", "feature": "Clinical Safety UX", "words": 2000},
    {"month": "Sep 2026", "week": "W3", "title": "Why We Built with Mantine Instead of Tailwind for Hospital UI",
     "cat": "Technology & Innovation", "format": "Technical", "keywords": "Mantine vs Tailwind healthcare, hospital UI framework, React component library HMS",
     "cta": "View UI Screenshots", "feature": "Frontend Architecture", "words": 1500},

    {"month": "Oct 2026", "week": "W1", "title": "Offline-First Mobile for Hospitals — Why WatermelonDB Powers Our Clinical App",
     "cat": "Technology & Innovation", "format": "Technical", "keywords": "offline first hospital app, WatermelonDB healthcare, mobile clinical app",
     "cta": "Mobile App Demo", "feature": "Mobile Offline", "words": 1500},
    {"month": "Oct 2026", "week": "W3", "title": "How Hospitals Can Evaluate HMS Software — A 50-Point Technical Checklist",
     "cat": "Technology & Innovation", "format": "Checklist", "keywords": "hospital software evaluation checklist, HMS comparison guide, choose hospital management system",
     "cta": "Download Evaluation Checklist", "feature": "Full Platform", "words": 2500},

    {"month": "Nov 2026", "week": "W1", "title": "ABDM Integration — Health IDs, Consent Framework, and Health Record Exchange Explained",
     "cat": "Technology & Innovation", "format": "Integration", "keywords": "ABDM integration hospital, health ID software, Ayushman Bharat Digital Mission HMS",
     "cta": "ABDM Compliance Guide", "feature": "ABDM Integration", "words": 2000},
    {"month": "Nov 2026", "week": "W3", "title": "WebSocket-Driven Hospital Displays — Real-Time Queues Without Constant Refreshing",
     "cat": "Technology & Innovation", "format": "Technical", "keywords": "hospital display system, real-time queue display, WebSocket healthcare",
     "cta": "TV Display Demo", "feature": "TV Displays", "words": 1500},

    {"month": "Dec 2026", "week": "W1", "title": "Hospital LIS Integration — Connecting Lab Analyzers Without Vendor Lock-In",
     "cat": "Technology & Innovation", "format": "Integration", "keywords": "LIS integration hospital, lab analyzer interface, laboratory information system",
     "cta": "LIS Integration Guide", "feature": "Lab Module", "words": 1800},
    {"month": "Dec 2026", "week": "W3", "title": "DICOM and PACS — Making Medical Imaging Work in Your HMS",
     "cat": "Technology & Innovation", "format": "Integration", "keywords": "DICOM PACS integration, medical imaging HMS, radiology PACS hospital",
     "cta": "Radiology Integration Demo", "feature": "Radiology DICOM/PACS", "words": 1800},

    # Vendor & Industry posts
    {"month": "Jan 2027", "week": "W1", "title": "Hospital Vendor Portal Vision — Self-Service PO Tracking, Invoice Submission, Payment Status",
     "cat": "Vendor & Industry", "format": "Product Vision", "keywords": "hospital vendor portal, supplier management healthcare, vendor self-service hospital",
     "cta": "Vendor Module Info", "feature": "Procurement + Vendor Performance", "words": 1500},
    {"month": "Jan 2027", "week": "W3", "title": "Medical Equipment Suppliers — What Hospitals Actually Need from Your After-Sales Service",
     "cat": "Vendor & Industry", "format": "Vendor Guide", "keywords": "medical equipment supplier tips, hospital equipment maintenance, biomedical vendor hospital",
     "cta": "Partner with MedBrains", "feature": "FMS + Condemnations", "words": 1500},

    {"month": "Feb 2027", "week": "W1", "title": "Pharmaceutical Distribution to Hospitals — How Digital POs Transform the Supply Chain",
     "cat": "Vendor & Industry", "format": "Industry", "keywords": "hospital pharmaceutical supply chain, drug distribution hospital, pharma vendor management",
     "cta": "Procurement Module Info", "feature": "Procurement + Pharmacy", "words": 1500},
    {"month": "Feb 2027", "week": "W3", "title": "Hospital Linen & Laundry Vendors — SLA Tracking and Quality Metrics That Matter",
     "cat": "Vendor & Industry", "format": "Niche", "keywords": "hospital laundry management, linen vendor hospital, housekeeping SLA tracking",
     "cta": "Housekeeping Module Info", "feature": "Housekeeping Module", "words": 1200},

    {"month": "Mar 2027", "week": "W1", "title": "Hospital Food Service Partners — Managing Outsourced Kitchen Operations Digitally",
     "cat": "Vendor & Industry", "format": "Vendor Ops", "keywords": "hospital food service management, outsourced hospital kitchen, catering vendor hospital",
     "cta": "Diet Kitchen Module Info", "feature": "Diet & Kitchen Module", "words": 1200},
    {"month": "Mar 2027", "week": "W3", "title": "Healthcare IT Consulting — What Hospitals Need During HMS Implementation",
     "cat": "Vendor & Industry", "format": "Industry", "keywords": "hospital IT consulting, HMS implementation consultant, healthcare digital transformation",
     "cta": "Implementation Partner Program", "feature": "Full Platform", "words": 1500},

    {"month": "Apr 2027", "week": "W1", "title": "Diagnostic Lab Chains — How B2B Lab Management Scales Multi-Location Operations",
     "cat": "Vendor & Industry", "format": "B2B", "keywords": "diagnostic lab chain management, B2B lab software, multi-location laboratory",
     "cta": "Lab B2B Module Demo", "feature": "Lab B2B Clients", "words": 1500},
    {"month": "Apr 2027", "week": "W3", "title": "Medical Tourism Facilitators — What Software Support Hospitals Should Offer You",
     "cat": "Vendor & Industry", "format": "Partner Guide", "keywords": "medical tourism facilitator hospital, medical travel coordinator, international patient coordinator",
     "cta": "Tourism Partner Program", "feature": "Medical Tourism Module", "words": 1500},

    {"month": "May 2027", "week": "W1", "title": "Insurance TPA Integration — What Hospital Software Must Support for Cashless Claims",
     "cat": "Vendor & Industry", "format": "Integration", "keywords": "TPA integration hospital software, cashless claim process, insurance hospital management",
     "cta": "TPA Integration Guide", "feature": "Insurance & TPA Module", "words": 1800},
    {"month": "May 2027", "week": "W3", "title": "Hospital Construction & Design Firms — Technology Infrastructure Your Hospital Needs From Day One",
     "cat": "Vendor & Industry", "format": "Advisory", "keywords": "hospital construction technology, healthcare facility design IT, smart hospital infrastructure",
     "cta": "Tech Planning Consultation", "feature": "Facilities Management", "words": 1500},

    {"month": "Jun 2027", "week": "W1", "title": "Building Healthcare Integrations — HL7 FHIR, ABDM, and the Future of Interoperability",
     "cat": "Technology & Innovation", "format": "Standards", "keywords": "HL7 FHIR hospital, healthcare interoperability India, ABDM FHIR integration",
     "cta": "Integration Documentation", "feature": "API + ABDM", "words": 2000},
    {"month": "Jun 2027", "week": "W3", "title": "Hospital Software RFP Guide — How to Write Requirements That Get You the Right HMS",
     "cat": "Vendor & Industry", "format": "Buyer Guide", "keywords": "hospital software RFP template, HMS procurement guide, hospital management system selection",
     "cta": "Download RFP Template", "feature": "Full Platform", "words": 2500},
]

# ── Sheet 6: Regulatory Deep-Dives + Medical Tourism (26 blogs) ─────────────
REGULATORY_TOURISM = [
    {"month": "Jul 2026", "week": "W2", "title": "NABH Accreditation 2026 — Complete Guide for Hospitals Starting Their Journey",
     "cat": "Regulatory & Compliance", "format": "Mega Guide", "keywords": "NABH accreditation guide 2026, hospital accreditation India, NABH standards requirements",
     "cta": "NABH Compliance Checklist", "feature": "Quality + Regulatory Modules", "words": 3000},
    {"month": "Jul 2026", "week": "W4", "title": "JCI vs NABH — Which Accreditation Should Your Hospital Pursue First?",
     "cat": "Regulatory & Compliance", "format": "Comparison", "keywords": "JCI vs NABH comparison, hospital accreditation choice, international hospital accreditation",
     "cta": "Accreditation Readiness Audit", "feature": "Quality Accreditation Module", "words": 2000},

    {"month": "Aug 2026", "week": "W2", "title": "Clinical Establishments Act 2010 — What Every Hospital Must Comply With",
     "cat": "Regulatory & Compliance", "format": "Legal", "keywords": "Clinical Establishments Act hospital, hospital registration India, healthcare compliance law",
     "cta": "Compliance Calendar Setup", "feature": "Regulatory Compliance Calendar", "words": 1800},
    {"month": "Aug 2026", "week": "W4", "title": "India as a Medical Tourism Destination — Numbers, Specialties, and Competitive Advantages",
     "cat": "Medical Tourism", "format": "Market Overview", "keywords": "India medical tourism statistics, healthcare tourism India, medical travel destination India",
     "cta": "International Patient Desk", "feature": "Medical Tourism Module", "words": 2000},

    {"month": "Sep 2026", "week": "W2", "title": "NABL Accreditation for Hospital Labs — Requirements, Process, and Common Pitfalls",
     "cat": "Regulatory & Compliance", "format": "Guide", "keywords": "NABL accreditation hospital lab, laboratory accreditation India, NABL requirements",
     "cta": "Lab Compliance Demo", "feature": "Lab NABL Documents", "words": 2000},
    {"month": "Sep 2026", "week": "W4", "title": "Medical Tourism for Cardiac Surgery — Why International Patients Choose Indian Hospitals",
     "cat": "Medical Tourism", "format": "Specialty Focus", "keywords": "cardiac surgery medical tourism India, heart surgery abroad, cardiac care India cost",
     "cta": "View Cardiac Packages", "feature": "Cardiac Rehab Program", "words": 1800},

    {"month": "Oct 2026", "week": "W2", "title": "Drugs and Cosmetics Act — Hospital Pharmacy Compliance Essentials",
     "cat": "Regulatory & Compliance", "format": "Legal", "keywords": "Drugs Cosmetics Act hospital pharmacy, pharmacy compliance India, Schedule H H1 X drugs",
     "cta": "Pharmacy Compliance Tour", "feature": "Pharmacy Regulatory", "words": 1800},
    {"month": "Oct 2026", "week": "W4", "title": "Orthopedic Medical Tourism — Joint Replacement and Spine Surgery Packages for International Patients",
     "cat": "Medical Tourism", "format": "Specialty Focus", "keywords": "orthopedic medical tourism India, joint replacement abroad, spine surgery India cost",
     "cta": "View Orthopedic Packages", "feature": "Orthopedic Program + Tourism", "words": 1500},

    {"month": "Nov 2026", "week": "W2", "title": "Biomedical Waste Management Rules 2016 — Digital Compliance for Hospitals",
     "cat": "Regulatory & Compliance", "format": "Compliance", "keywords": "biomedical waste management rules, BMW compliance hospital, waste tracking software hospital",
     "cta": "BMW Tracking Demo", "feature": "Infection Control BMW", "words": 1800},
    {"month": "Nov 2026", "week": "W4", "title": "IVF and Fertility Tourism to India — Setting Up Your Hospital for International Patients",
     "cat": "Medical Tourism", "format": "Specialty Focus", "keywords": "IVF medical tourism India, fertility treatment abroad, IVF cost India international",
     "cta": "View Fertility Packages", "feature": "Fertility & IVF Program", "words": 1500},

    {"month": "Dec 2026", "week": "W2", "title": "Patient Data Privacy in India — IT Act, DPDP Act, and What Your HMS Must Do",
     "cat": "Regulatory & Compliance", "format": "Legal", "keywords": "patient data privacy India, DPDP Act healthcare, health data protection hospital",
     "cta": "Security Compliance Demo", "feature": "Data Security + Audit", "words": 2000},
    {"month": "Dec 2026", "week": "W4", "title": "Dental and Cosmetic Surgery Tourism — Building Digital Packages That Sell Globally",
     "cat": "Medical Tourism", "format": "Specialty Focus", "keywords": "dental tourism India, cosmetic surgery tourism, dental treatment abroad India",
     "cta": "Package Builder Demo", "feature": "Tourism Package Builder", "words": 1500},

    {"month": "Jan 2027", "week": "W2", "title": "AERB Radiation Safety Compliance — What Hospitals with Radiology Must Track",
     "cat": "Regulatory & Compliance", "format": "Niche Compliance", "keywords": "AERB compliance hospital, radiation safety hospital, radiology compliance India",
     "cta": "Radiology Compliance Demo", "feature": "Radiology AERB", "words": 1500},
    {"month": "Jan 2027", "week": "W4", "title": "Medical Tourism Patient Journey — From Online Inquiry to Post-Return Follow-Up",
     "cat": "Medical Tourism", "format": "Journey Map", "keywords": "medical tourism patient journey, international patient workflow, medical travel hospital process",
     "cta": "Tourism Workflow Demo", "feature": "Medical Tourism Full Module", "words": 2000},

    {"month": "Feb 2027", "week": "W2", "title": "Fire Safety Compliance for Hospitals — NOC, Equipment, Drills, and Digital Records",
     "cat": "Regulatory & Compliance", "format": "Safety", "keywords": "hospital fire safety compliance India, fire NOC hospital, fire drill documentation",
     "cta": "Fire Safety Module Demo", "feature": "FMS Fire Safety", "words": 1500},
    {"month": "Feb 2027", "week": "W4", "title": "Eye Surgery Medical Tourism — LASIK, Cataract, and Retinal Procedures for Global Patients",
     "cat": "Medical Tourism", "format": "Specialty Focus", "keywords": "eye surgery medical tourism India, LASIK abroad India, cataract surgery cost India",
     "cta": "View Eye Surgery Packages", "feature": "Tourism Package Builder", "words": 1500},

    {"month": "Mar 2027", "week": "W2", "title": "PNDT Act Compliance — Ultrasound Documentation and Gender Determination Prevention",
     "cat": "Regulatory & Compliance", "format": "Critical Compliance", "keywords": "PNDT Act compliance hospital, ultrasound Form F, gender determination prevention",
     "cta": "PCPNDT Module Demo", "feature": "PCPNDT Forms", "words": 1800},
    {"month": "Mar 2027", "week": "W4", "title": "Weight Loss Surgery Tourism — Bariatric Packages for Middle East and African Patients",
     "cat": "Medical Tourism", "format": "Market Focus", "keywords": "bariatric surgery medical tourism India, weight loss surgery abroad, gastric bypass India cost",
     "cta": "View Bariatric Packages", "feature": "Bariatric Program + Tourism", "words": 1500},

    {"month": "Apr 2027", "week": "W2", "title": "Mental Healthcare Act 2017 — What Hospital Psychiatry Departments Must Document",
     "cat": "Regulatory & Compliance", "format": "Legal", "keywords": "Mental Healthcare Act hospital, psychiatry documentation India, mental health compliance",
     "cta": "Psychiatry Module Info", "feature": "Psychiatry Module", "words": 1800},
    {"month": "Apr 2027", "week": "W4", "title": "Organ Transplant Tourism — Legal, Ethical, and Operational Considerations for Hospitals",
     "cat": "Medical Tourism", "format": "Ethics + Legal", "keywords": "organ transplant tourism India, transplant hospital compliance, THOA Act hospital",
     "cta": "Transplant Compliance Guide", "feature": "Regulatory Module", "words": 2000},

    {"month": "May 2027", "week": "W2", "title": "NMC (National Medical Commission) — New Regulations Every Teaching Hospital Must Know",
     "cat": "Regulatory & Compliance", "format": "Regulatory Update", "keywords": "NMC regulations hospital, National Medical Commission compliance, teaching hospital requirements",
     "cta": "Academic Module Info", "feature": "Medical College Module", "words": 1800},
    {"month": "May 2027", "week": "W4", "title": "Wellness and Ayurveda Tourism — Integrating Traditional Medicine Packages for International Patients",
     "cat": "Medical Tourism", "format": "Specialty Focus", "keywords": "Ayurveda medical tourism India, wellness tourism packages, traditional medicine hospital",
     "cta": "Wellness Package Demo", "feature": "Tourism Package Builder", "words": 1500},

    {"month": "Jun 2027", "week": "W2", "title": "Hospital Accreditation Checklist — 700 Criteria Across 34 Departments Digitized",
     "cat": "Regulatory & Compliance", "format": "Product Story", "keywords": "hospital accreditation checklist digital, NABH checklist software, accreditation management",
     "cta": "Compliance Module Demo", "feature": "Quality + Regulatory Modules", "words": 2000},
    {"month": "Jun 2027", "week": "W4", "title": "Cancer Treatment Tourism — Oncology Programs That Attract International Patients",
     "cat": "Medical Tourism", "format": "Specialty Focus", "keywords": "cancer treatment medical tourism India, oncology abroad India, cancer hospital India international",
     "cta": "View Oncology Program", "feature": "Oncology Support + Tourism", "words": 1800},
]


ALL_SHEETS = [
    ("Hospital Leadership & Strategy", HOSPITAL_LEADERSHIP),
    ("Patient Education & Wellness", PATIENT_EDUCATION),
    ("Doctor & Clinical Excellence", DOCTOR_CLINICAL),
    ("Operations & Efficiency", OPERATIONS),
    ("Technology & Vendor Ecosystem", TECH_VENDOR),
    ("Regulatory & Medical Tourism", REGULATORY_TOURISM),
]


def compute_publish_date(month_str, week_str):
    """Approximate publish date from 'Jul 2026' + 'W1'."""
    months = {
        "Jul 2026": date(2026, 7, 1), "Aug 2026": date(2026, 8, 1),
        "Sep 2026": date(2026, 9, 1), "Oct 2026": date(2026, 10, 1),
        "Nov 2026": date(2026, 11, 1), "Dec 2026": date(2026, 12, 1),
        "Jan 2027": date(2027, 1, 1), "Feb 2027": date(2027, 2, 1),
        "Mar 2027": date(2027, 3, 1), "Apr 2027": date(2027, 4, 1),
        "May 2027": date(2027, 5, 1), "Jun 2027": date(2027, 6, 1),
    }
    base = months.get(month_str, date(2026, 7, 1))
    week_num = int(week_str[1]) - 1  # W1=0, W2=1, W3=2, W4=3
    return base + timedelta(days=week_num * 7)


def create_content_sheet(wb, sheet_name, blogs):
    """Create a content calendar sheet."""
    ws = wb.create_sheet(title=sheet_name)

    # Column widths
    for i, (_, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Header row
    for i, (name, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Data rows
    current_month = None
    row = 2
    sno = 0

    for blog in blogs:
        # Month separator
        if blog["month"] != current_month:
            current_month = blog["month"]
            for col in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = MONTH_FILL
                cell.border = THIN
            ws.cell(row=row, column=2, value=current_month).font = MONTH_FONT
            row += 1

        sno += 1
        pub_date = compute_publish_date(blog["month"], blog["week"])

        values = [
            sno,
            blog["month"],
            blog["week"],
            pub_date,
            blog["title"],
            blog["cat"],
            CATEGORIES.get(blog["cat"], {}).get("audience", ""),
            blog.get("format", "Article"),
            blog.get("keywords", ""),
            blog.get("cta", ""),
            blog.get("feature", ""),
            blog.get("words", 1500),
            "Planned",
            "",
            "",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN
            cell.alignment = WRAP

        # Date format
        ws.cell(row=row, column=4).number_format = "DD-MMM-YYYY"

        # Status color
        ws.cell(row=row, column=13).fill = STATUS_COLORS.get("Planned", PatternFill())

        row += 1

    return sno


def create_summary_sheet(wb, sheet_data):
    """Create a summary/dashboard sheet."""
    ws = wb.create_sheet(title="Dashboard", index=0)

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value="MedBrains Blog Content Plan — 12 Month Calendar")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Subtitle
    ws.merge_cells("A2:H2")
    ws.cell(row=2, column=1, value="Jul 2026 — Jun 2027 | ~3 posts/week | 6 audience categories").font = Font(italic=True, size=10, color="666666")

    # Category summary table
    row = 4
    headers = ["Category", "Target Audience", "Blog Count", "Avg Word Count", "Posts/Month", "Primary CTA"]
    col_widths = [30, 30, 12, 14, 12, 25]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w

    row += 1
    total_blogs = 0
    total_words = 0

    for sheet_name, blogs in sheet_data:
        count = len(blogs)
        total_blogs += count
        avg_words = sum(b.get("words", 1500) for b in blogs) // max(count, 1)
        total_words += sum(b.get("words", 1500) for b in blogs)
        posts_month = round(count / 12, 1)

        # Get primary category
        cats = set(b["cat"] for b in blogs)
        primary_cat = list(cats)[0] if len(cats) == 1 else ", ".join(sorted(cats))
        audience = CATEGORIES.get(list(cats)[0], {}).get("audience", "Mixed")
        cta = "Product Demos + Guides"

        values = [sheet_name, audience, count, avg_words, posts_month, cta]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = THIN
            cell.alignment = WRAP
        row += 1

    # Total row
    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=row, column=3, value=total_blogs).font = Font(bold=True, size=11)
    ws.cell(row=row, column=4, value=total_words // max(total_blogs, 1)).font = Font(bold=True)
    ws.cell(row=row, column=5, value=round(total_blogs / 12, 1)).font = Font(bold=True)
    for i in range(1, 7):
        ws.cell(row=row, column=i).border = THIN

    # Monthly schedule overview
    row += 3
    ws.cell(row=row, column=1, value="Monthly Publishing Schedule").font = Font(bold=True, size=12, color="1F4E79")
    row += 1

    month_headers = ["Month"] + [s[0] for s in sheet_data] + ["Total"]
    for i, h in enumerate(month_headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN
        ws.column_dimensions[get_column_letter(i)].width = max(col_widths[0] if i <= len(col_widths) else 12, len(h) + 4)

    row += 1
    months_order = [
        "Jul 2026", "Aug 2026", "Sep 2026", "Oct 2026", "Nov 2026", "Dec 2026",
        "Jan 2027", "Feb 2027", "Mar 2027", "Apr 2027", "May 2027", "Jun 2027",
    ]

    for m in months_order:
        ws.cell(row=row, column=1, value=m).border = THIN
        month_total = 0
        for col_idx, (_, blogs) in enumerate(sheet_data, 2):
            count = sum(1 for b in blogs if b["month"] == m)
            ws.cell(row=row, column=col_idx, value=count if count > 0 else "").border = THIN
            month_total += count
        ws.cell(row=row, column=len(sheet_data) + 2, value=month_total).border = THIN
        ws.cell(row=row, column=len(sheet_data) + 2).font = Font(bold=True)
        row += 1

    # Publishing guidelines
    row += 2
    ws.cell(row=row, column=1, value="Content Guidelines").font = Font(bold=True, size=12, color="1F4E79")
    row += 1
    guidelines = [
        "Publishing cadence: 3 posts per week (Mon/Wed/Fri), rotating across categories",
        "Word count targets: 1200-1500 (listicles/FAQ), 1500-2000 (guides/explainers), 2000-3000 (mega guides)",
        "SEO: Every post needs 3-5 target keywords, meta description, internal links to 2-3 related posts",
        "CTA: Every post ends with a relevant call-to-action tied to a MedBrains feature or demo",
        "Visuals: Minimum 2 images/screenshots per post, 1 infographic for data-heavy posts",
        "Social: Share each post on LinkedIn (hospital admins), Twitter/X (tech), WhatsApp groups (doctors)",
        "Repurpose: Convert top posts into LinkedIn carousels, short videos, email newsletter segments",
        "Guest posts: Invite hospital administrators and doctors to contribute 1 post/quarter",
        "Metrics: Track page views, time on page, CTA click rate, demo request attribution per post",
        "Review cycle: Draft → Medical review (clinical posts) → SEO review → Publish → Social distribution",
    ]
    for g in guidelines:
        ws.cell(row=row, column=1, value=g).alignment = WRAP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

    return total_blogs


def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Create content sheets
    counts = {}
    for sheet_name, blogs in ALL_SHEETS:
        count = create_content_sheet(wb, sheet_name, blogs)
        counts[sheet_name] = count
        print(f"{sheet_name}: {count} blog posts")

    # Create summary dashboard (inserted at index 0)
    total = create_summary_sheet(wb, ALL_SHEETS)

    wb.save(EXCEL_PATH)
    print(f"\nDone! Created {EXCEL_PATH}")
    print(f"Total: {total} blog posts across 6 categories over 12 months")
    print(f"Average: {total / 12:.1f} posts/month, {total / 52:.1f} posts/week")


if __name__ == "__main__":
    main()

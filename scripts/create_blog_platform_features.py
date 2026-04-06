#!/usr/bin/env python3
"""
Create Healthcare_Blog_Platform_Features.xlsx — feature tracker for HealthPulse
healthcare publishing platform (like The Verge for healthcare).
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style constants ──────────────────────────────────────────────────────
MODULE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBMODULE_FILL = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUMMARY_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
REF_HEADER_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

BOLD = Font(bold=True)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
MODULE_FONT = Font(bold=True, size=11)
SUBMODULE_FONT = Font(bold=True, size=10, italic=True)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COLUMNS = ["S.No", "Module", "Sub-Module", "Feature", "Priority", "Status",
           "Phase", "Platform", "Complexity", "Notes"]
COL_WIDTHS = [6, 22, 22, 55, 10, 10, 10, 12, 12, 40]


def style_header(ws):
    """Apply header styling to row 1."""
    for col_idx, (title, width) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J1"


def add_module_header(ws, row, module_name):
    """Add a module-level header row with blue fill."""
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = MODULE_FILL
        cell.font = MODULE_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP
    ws.cell(row=row, column=2, value=module_name)
    return row + 1


def add_submodule_header(ws, row, submodule_name):
    """Add a sub-module header row with light fill."""
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = SUBMODULE_FILL
        cell.font = SUBMODULE_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP
    ws.cell(row=row, column=3, value=submodule_name)
    return row + 1


def add_feature(ws, row, sno, module, submodule, feature, priority="P1",
                status="Pending", phase="MVP", platform="Web",
                complexity="M", notes=""):
    """Add a single feature row."""
    values = [sno, module, submodule, feature, priority, status,
              phase, platform, complexity, notes]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border = THIN_BORDER
        cell.alignment = WRAP if col_idx in (4, 10) else WRAP_CENTER
    return row + 1


def write_features(ws, features):
    """Write all features for a sheet. features is a list of dicts grouped by module/submodule."""
    style_header(ws)
    row = 2
    sno = 1
    for group in features:
        module = group["module"]
        row = add_module_header(ws, row, module)
        for sub in group["submodules"]:
            submodule = sub["name"]
            row = add_submodule_header(ws, row, submodule)
            for f in sub["features"]:
                row = add_feature(
                    ws, row, sno, module, submodule,
                    f["feature"],
                    f.get("priority", "P1"),
                    f.get("status", "Pending"),
                    f.get("phase", "MVP"),
                    f.get("platform", "Web"),
                    f.get("complexity", "M"),
                    f.get("notes", ""),
                )
                sno += 1
    return sno - 1  # total features


# ═══════════════════════════════════════════════════════════════════════════
# SHEET DEFINITIONS
# ═══════════════════════════════════════════════════════════════════════════

def sheet_platform_core():
    return [
        {
            "module": "Homepage",
            "submodules": [
                {
                    "name": "Hero Section",
                    "features": [
                        {"feature": "Featured story hero with full-bleed image, headline, excerpt, author byline", "priority": "P0", "phase": "MVP", "complexity": "L"},
                        {"feature": "Rotating hero carousel (max 5 stories, editorial-curated)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Breaking news takeover banner with dismiss", "priority": "P1", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Video hero variant for multimedia stories", "priority": "P2", "phase": "V2", "complexity": "L"},
                    ],
                },
                {
                    "name": "Storystream Feed",
                    "features": [
                        {"feature": "Editorially curated storystream feed (not algorithmic) — mixed content types", "priority": "P0", "phase": "MVP", "complexity": "L"},
                        {"feature": "Feed cards: thumbnail, headline, excerpt, category badge, author, timestamp", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Quick post cards (compact, no thumbnail, ~140 char preview)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Infinite scroll with lazy loading (10 items per batch)", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Sticky category filter bar on feed (All / each category tab)", "priority": "P1", "phase": "MVP", "complexity": "M"},
                        {"feature": "Pinned posts — editor can pin 1-3 posts to top of feed", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Content type filter (articles, quick posts, reviews, guides)", "priority": "P2", "phase": "V1.2", "complexity": "S"},
                    ],
                },
                {
                    "name": "Sidebar / Secondary",
                    "features": [
                        {"feature": "Trending stories widget (top 5 by views in 24h)", "priority": "P1", "phase": "MVP", "complexity": "M"},
                        {"feature": "Newsletter signup widget (inline, sticky on scroll)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Popular categories widget with article counts", "priority": "P2", "phase": "V1.1", "complexity": "S"},
                        {"feature": "MedBrains HMS promo banner (lead-gen CTA)", "priority": "P1", "phase": "MVP", "complexity": "S", "notes": "Dual business model — HMS lead generation"},
                        {"feature": "Upcoming events / webinars widget", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                    ],
                },
            ],
        },
        {
            "module": "Navigation",
            "submodules": [
                {
                    "name": "Header Navigation",
                    "features": [
                        {"feature": "Responsive top navbar with logo, category dropdowns, search icon, user menu", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Mega menu for categories — shows subcategories + featured article", "priority": "P1", "phase": "V1.1", "complexity": "L"},
                        {"feature": "Mobile hamburger menu with category accordion", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Sticky header that hides on scroll down, shows on scroll up", "priority": "P1", "phase": "MVP", "complexity": "S"},
                        {"feature": "Search bar expand-on-click with autocomplete suggestions", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                    ],
                },
                {
                    "name": "Footer",
                    "features": [
                        {"feature": "Multi-column footer: categories, about, legal, social links, newsletter", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Medical disclaimer in footer (always visible)", "priority": "P0", "phase": "MVP", "complexity": "S", "notes": "Regulatory requirement"},
                        {"feature": "Back-to-top button", "priority": "P2", "phase": "V1.1", "complexity": "S"},
                    ],
                },
                {
                    "name": "Search",
                    "features": [
                        {"feature": "Full-text search across articles, authors, categories, tags", "priority": "P0", "phase": "MVP", "complexity": "L"},
                        {"feature": "Search results page with filters (date, category, content type, author)", "priority": "P1", "phase": "MVP", "complexity": "M"},
                        {"feature": "Search autocomplete / typeahead suggestions", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Search analytics — track popular queries, zero-result queries", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                    ],
                },
            ],
        },
        {
            "module": "Mobile & PWA",
            "submodules": [
                {
                    "name": "Responsive Design",
                    "features": [
                        {"feature": "Fully responsive layout (mobile-first, 320px to 4K)", "priority": "P0", "phase": "MVP", "complexity": "L"},
                        {"feature": "Touch-optimized interactions (swipe, tap targets ≥44px)", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Adaptive images (srcset, WebP/AVIF with fallback)", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Mobile-specific reading mode (minimal chrome, max content)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                    ],
                },
                {
                    "name": "PWA",
                    "features": [
                        {"feature": "Service worker for offline reading of cached articles", "priority": "P2", "phase": "V1.2", "complexity": "L", "platform": "Web"},
                        {"feature": "Add-to-homescreen prompt", "priority": "P2", "phase": "V1.2", "complexity": "S", "platform": "Web"},
                        {"feature": "Web push notifications for breaking news / newsletter", "priority": "P2", "phase": "V1.2", "complexity": "M", "platform": "Web"},
                    ],
                },
            ],
        },
    ]


def sheet_content_types():
    return [
        {
            "module": "Quick Posts (Micro-Blog)",
            "submodules": [
                {
                    "name": "Quick Post Creation",
                    "features": [
                        {"feature": "Quick post editor — plain text + markdown, max 500 words, minimal UI", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Optional single image/embed attachment", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Category and tag selection (quick select chips)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "One-click publish (skip full editorial workflow for quick posts)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Source URL field — link to external article being commented on", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Quick post thread — chain multiple quick posts into a thread", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                    ],
                },
                {
                    "name": "Quick Post Display",
                    "features": [
                        {"feature": "Compact card layout in feed (no hero image, inline text preview)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Expand-in-place reading (no page navigation needed)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Share buttons (Twitter/X, LinkedIn, WhatsApp, copy link)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                    ],
                },
            ],
        },
        {
            "module": "Long-Form Articles",
            "submodules": [
                {
                    "name": "Rich Text Editor",
                    "features": [
                        {"feature": "WYSIWYG editor with heading levels, bold, italic, lists, blockquotes", "priority": "P0", "phase": "MVP", "complexity": "L", "notes": "Ghost editor or custom"},
                        {"feature": "Inline image insertion with caption, alt text, credit", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Embed support: YouTube, Twitter, Instagram, CodePen, PubMed", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Table insertion and editing", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Custom callout/alert boxes (info, warning, expert opinion, key takeaway)", "priority": "P1", "phase": "MVP", "complexity": "M"},
                        {"feature": "Footnotes / endnotes with reference linking", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Inline citation insertion — link to PubMed, DOI, clinical study", "priority": "P0", "phase": "MVP", "complexity": "M", "notes": "Healthcare credibility essential"},
                        {"feature": "Code snippet embedding (for healthtech/dev articles)", "priority": "P2", "phase": "V1.2", "complexity": "S"},
                        {"feature": "Markdown mode toggle (switch between WYSIWYG and raw markdown)", "priority": "P2", "phase": "V1.1", "complexity": "M"},
                    ],
                },
                {
                    "name": "Article Display",
                    "features": [
                        {"feature": "Full-width hero image with gradient text overlay", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Author byline with avatar, name, credentials, follow button", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Medical reviewer badge — reviewed by Dr. X, MD (with date)", "priority": "P0", "phase": "MVP", "complexity": "S", "notes": "E-E-A-T compliance for Google YMYL"},
                        {"feature": "Dual timestamp: Published date + Last medically reviewed date", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Reading time estimate (words/200 wpm)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Auto-generated table of contents (from H2/H3 headings)", "priority": "P1", "phase": "MVP", "complexity": "M"},
                        {"feature": "Reading progress bar (sticky top)", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Related articles section (3-4 cards, same category/tags)", "priority": "P1", "phase": "MVP", "complexity": "M"},
                        {"feature": "Medical disclaimer banner per article (contextual based on category)", "priority": "P0", "phase": "MVP", "complexity": "S", "notes": "Legal requirement for patient-facing content"},
                        {"feature": "Source citations section at article bottom (numbered references)", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Print-friendly view / save as PDF", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                    ],
                },
            ],
        },
        {
            "module": "Reviews & Buyer's Guides",
            "submodules": [
                {
                    "name": "Reviews",
                    "features": [
                        {"feature": "Review scoring system (0-10 with half-points) per criteria", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Scorecard component — overall score + category scores (features, value, support, ease)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Pros/cons structured section", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Product spec table (structured data for comparison)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Verdict badge (Editor's Choice, Best Value, Best for Small Hospitals)", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Conflict of interest / sponsorship disclosure on reviews", "priority": "P0", "phase": "V1.1", "complexity": "S", "notes": "Trust & transparency"},
                        {"feature": "Review categories: HMS software, medical devices, health apps, EHR systems", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                    ],
                },
                {
                    "name": "Buyer's Guides",
                    "features": [
                        {"feature": "Multi-product comparison table (side-by-side specs + scores)", "priority": "P2", "phase": "V1.2", "complexity": "L"},
                        {"feature": "Product recommendation engine (by hospital size, budget, specialty)", "priority": "P2", "phase": "V2", "complexity": "XL"},
                        {"feature": "Affiliate/CTA links per product (tracked for monetization)", "priority": "P2", "phase": "V1.2", "complexity": "M", "notes": "Revenue stream"},
                        {"feature": "Last updated date with auto-stale warning (>6 months)", "priority": "P2", "phase": "V1.2", "complexity": "S"},
                    ],
                },
            ],
        },
        {
            "module": "Live Blogs",
            "submodules": [
                {
                    "name": "Live Blog Engine",
                    "features": [
                        {"feature": "Live blog page — reverse-chron updates with timestamps", "priority": "P2", "phase": "V1.2", "complexity": "L"},
                        {"feature": "Real-time auto-refresh (new updates appear without page reload)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Pinned key update at top of live blog", "priority": "P2", "phase": "V1.2", "complexity": "S"},
                        {"feature": "Multi-author live blog (different reporters can post updates)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Embed tweets, images, videos inline in live updates", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Use cases: HIMSS live coverage, Union Budget health analysis, NMC updates", "priority": "P2", "phase": "V1.2", "complexity": "S", "notes": "Key differentiator for news coverage"},
                    ],
                },
            ],
        },
        {
            "module": "Multimedia Content",
            "submodules": [
                {
                    "name": "Video",
                    "features": [
                        {"feature": "Video embed support (YouTube, Vimeo, custom player)", "priority": "P1", "phase": "MVP", "complexity": "M"},
                        {"feature": "Video-first article template (video hero + transcript below)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Video series / playlists grouped by topic", "priority": "P2", "phase": "V2", "complexity": "M"},
                    ],
                },
                {
                    "name": "Podcast",
                    "features": [
                        {"feature": "Podcast player embed (inline audio player with show notes)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Podcast series page with episode list and RSS feed", "priority": "P2", "phase": "V2", "complexity": "M"},
                        {"feature": "Transcript auto-display below podcast player", "priority": "P2", "phase": "V2", "complexity": "M"},
                    ],
                },
                {
                    "name": "Visual Content",
                    "features": [
                        {"feature": "Image gallery component (lightbox, swipe on mobile)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Infographic full-width display with zoom", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Interactive data visualization embeds (charts, maps)", "priority": "P2", "phase": "V2", "complexity": "L"},
                        {"feature": "Before/after image slider (for medical device comparisons)", "priority": "P2", "phase": "V2", "complexity": "M"},
                    ],
                },
            ],
        },
        {
            "module": "Case Studies",
            "submodules": [
                {
                    "name": "Case Study Template",
                    "features": [
                        {"feature": "Structured case study template: challenge, solution, results, metrics", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Key metrics highlight boxes (e.g., 30% reduction in wait time)", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Hospital/organization profile sidebar in case studies", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Case study gallery page with filter by hospital type/size/solution", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "MedBrains HMS case study integration — link to demo request", "priority": "P1", "phase": "V1.1", "complexity": "S", "notes": "HMS lead generation"},
                    ],
                },
            ],
        },
    ]


def sheet_categories():
    return [
        {
            "module": "Category System Architecture",
            "submodules": [
                {
                    "name": "Core Category Infrastructure",
                    "features": [
                        {"feature": "Hierarchical category system: parent category → subcategories → tags", "priority": "P0", "phase": "MVP", "complexity": "L"},
                        {"feature": "Category page template: header, description, featured post, feed", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Subcategory landing pages with filtered feed", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Category-specific color coding and icons throughout UI", "priority": "P1", "phase": "MVP", "complexity": "S"},
                        {"feature": "Category badges on all content cards and article pages", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Multi-category assignment (article can belong to 1 primary + secondary categories)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Category RSS feed (per-category syndication)", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Category newsletter segment (auto-send to subscribers of that category)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Admin: create/edit/archive categories without code changes", "priority": "P0", "phase": "MVP", "complexity": "M", "notes": "Scalable — new verticals via admin panel"},
                    ],
                },
                {
                    "name": "Tag System",
                    "features": [
                        {"feature": "Free-form tag creation by editors (with suggested existing tags)", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Tag page with all content tagged with that term", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Tag cloud / popular tags widget", "priority": "P2", "phase": "V1.2", "complexity": "S"},
                        {"feature": "Tag merge/rename tool for editorial cleanup", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Auto-suggested tags based on article content (AI-assisted)", "priority": "P2", "phase": "V2", "complexity": "L"},
                    ],
                },
            ],
        },
        {
            "module": "Core Categories (Launch)",
            "submodules": [
                {
                    "name": "Hospital Leadership & Strategy",
                    "features": [
                        {"feature": "Category: Hospital Leadership & Strategy — CEO/CMO/admin audience", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Subcategories: Strategic Planning, Financial Management, Quality Accreditation, Governance", "priority": "P1", "phase": "MVP", "complexity": "S"},
                        {"feature": "Content types: opinion pieces, case studies, interviews, data analysis", "priority": "P0", "phase": "MVP", "complexity": "S"},
                    ],
                },
                {
                    "name": "Patient Education & Wellness",
                    "features": [
                        {"feature": "Category: Patient Education & Wellness — patient/family audience", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Subcategories: Disease Guides, Wellness Tips, Nutrition, Mental Health, Pregnancy", "priority": "P1", "phase": "MVP", "complexity": "S"},
                        {"feature": "Enhanced medical review requirement (all patient content must be reviewed)", "priority": "P0", "phase": "MVP", "complexity": "M", "notes": "Mandatory for YMYL compliance"},
                        {"feature": "Readability score enforcement (Flesch-Kincaid ≤ 8th grade for patient content)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                    ],
                },
                {
                    "name": "Doctor & Clinical Excellence",
                    "features": [
                        {"feature": "Category: Doctor & Clinical Excellence — physician/nurse audience", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Subcategories: Clinical Guidelines, Case Reports, Research Updates, CME", "priority": "P1", "phase": "MVP", "complexity": "S"},
                        {"feature": "Content types: clinical pearls, evidence summaries, expert opinions", "priority": "P0", "phase": "MVP", "complexity": "S"},
                    ],
                },
                {
                    "name": "Operations & Efficiency",
                    "features": [
                        {"feature": "Category: Operations & Efficiency — hospital manager/dept head audience", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Subcategories: Process Improvement, Supply Chain, Facility Management, Staffing", "priority": "P1", "phase": "MVP", "complexity": "S"},
                        {"feature": "Content types: how-to guides, checklists, templates, benchmarking data", "priority": "P0", "phase": "MVP", "complexity": "S"},
                    ],
                },
                {
                    "name": "Technology & Innovation",
                    "features": [
                        {"feature": "Category: Technology & Innovation — CIO/IT/healthtech audience", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Subcategories: AI in Healthcare, EHR/EMR, Telemedicine, Cybersecurity, Interoperability", "priority": "P1", "phase": "MVP", "complexity": "S"},
                        {"feature": "Product review integration (link to Reviews content type)", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                    ],
                },
                {
                    "name": "Regulatory & Compliance",
                    "features": [
                        {"feature": "Category: Regulatory & Compliance — compliance officer/legal audience", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Subcategories: NABH/JCI, NMC Updates, Legal Rulings, Data Privacy, DPDP Act", "priority": "P1", "phase": "MVP", "complexity": "S"},
                        {"feature": "Regulatory alert badges on time-sensitive compliance content", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                    ],
                },
            ],
        },
        {
            "module": "Specialty Verticals (Growth Phase)",
            "submodules": [
                {
                    "name": "Nursing & Allied Health",
                    "features": [
                        {"feature": "Category: Nursing & Allied Health — nurses, physios, technicians", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Subcategories: Nursing Practice, Physiotherapy, Lab Tech, Radiology Tech", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                    ],
                },
                {
                    "name": "Pharma Industry",
                    "features": [
                        {"feature": "Category: Pharma Industry — pharma companies, distributors", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Subcategories: Drug Launches, Pricing & Policy, Supply Chain, R&D Pipeline", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                    ],
                },
                {
                    "name": "Medical Devices & Equipment",
                    "features": [
                        {"feature": "Category: Medical Devices & Equipment — manufacturers, biomedical engineers", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Subcategories: Device Reviews, Procurement Guides, Regulatory Approvals, Innovation", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                    ],
                },
                {
                    "name": "Health Insurance & TPA",
                    "features": [
                        {"feature": "Category: Health Insurance & TPA — insurers, TPA executives, brokers", "priority": "P2", "phase": "V1.2", "complexity": "S"},
                        {"feature": "Subcategories: Policy Analysis, Claim Trends, Regulatory Changes, Product Reviews", "priority": "P2", "phase": "V1.2", "complexity": "S"},
                    ],
                },
                {
                    "name": "HealthTech & Startups",
                    "features": [
                        {"feature": "Category: HealthTech & Startups — founders, investors, digital health builders", "priority": "P2", "phase": "V1.2", "complexity": "S"},
                        {"feature": "Subcategories: Funding News, Product Launches, Founder Interviews, Market Analysis", "priority": "P2", "phase": "V1.2", "complexity": "S"},
                    ],
                },
                {
                    "name": "Medical Education & Research",
                    "features": [
                        {"feature": "Category: Medical Education & Research — students, academics, researchers", "priority": "P2", "phase": "V1.2", "complexity": "S"},
                        {"feature": "Subcategories: NEET/USMLE, Residency, Research Methods, Conference Reports", "priority": "P2", "phase": "V1.2", "complexity": "S"},
                    ],
                },
            ],
        },
        {
            "module": "Content Discovery",
            "submodules": [
                {
                    "name": "Topic Pages",
                    "features": [
                        {"feature": "Auto-generated topic pages from frequently used tags", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Curated topic hub pages (e.g., 'AI in Healthcare' with intro + curated list)", "priority": "P2", "phase": "V1.2", "complexity": "L"},
                        {"feature": "Topic following — users can follow specific topics for personalized feed", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                    ],
                },
                {
                    "name": "Content Series & Collections",
                    "features": [
                        {"feature": "Content series — multi-part articles linked sequentially (Part 1, 2, 3...)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Editorial collections — curated bundles of related articles", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Series navigation component (prev/next within series)", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                    ],
                },
                {
                    "name": "Related Content Engine",
                    "features": [
                        {"feature": "Related articles algorithm (based on category + tags + recency)", "priority": "P1", "phase": "MVP", "complexity": "M"},
                        {"feature": "\"More from this author\" section on article pages", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "\"Readers also liked\" section (based on reading patterns)", "priority": "P2", "phase": "V2", "complexity": "L"},
                    ],
                },
            ],
        },
    ]


def sheet_author_editorial():
    return [
        {
            "module": "Author Profiles",
            "submodules": [
                {
                    "name": "Author Page",
                    "features": [
                        {"feature": "Public author profile page: photo, bio, credentials, social links", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Author credential display: MD, MBBS, PhD, board certifications", "priority": "P0", "phase": "MVP", "complexity": "S", "notes": "E-E-A-T compliance"},
                        {"feature": "ORCID integration — link to research publications", "priority": "P2", "phase": "V1.2", "complexity": "S"},
                        {"feature": "Author article feed (all content by this author)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Follow author button + follower count", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Author expertise tags (cardiology, health policy, digital health, etc.)", "priority": "P1", "phase": "MVP", "complexity": "S"},
                        {"feature": "Conflict of interest / disclosure statement on author profile", "priority": "P0", "phase": "MVP", "complexity": "S", "notes": "Transparency requirement"},
                    ],
                },
                {
                    "name": "Author Types",
                    "features": [
                        {"feature": "Staff writer profiles (full-time editorial team)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Guest contributor profiles (external experts, doctors)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Medical reviewer profiles (separate role — reviews but doesn't write)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Multi-author articles (co-authored, each author bylined)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Editorial board page — list of all medical reviewers and advisors", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                    ],
                },
            ],
        },
        {
            "module": "Editorial Workflow",
            "submodules": [
                {
                    "name": "Content Pipeline",
                    "features": [
                        {"feature": "Article status workflow: Draft → In Review → Medical Review → Editor Approval → Scheduled → Published", "priority": "P0", "phase": "MVP", "complexity": "L"},
                        {"feature": "Assign reviewer / editor to article", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Inline comments / suggestions on draft (like Google Docs)", "priority": "P1", "phase": "V1.1", "complexity": "L"},
                        {"feature": "Revision history — full diff between versions", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Quick post simplified workflow: Draft → Published (editor bypass for senior writers)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Content scheduling — publish at specific date/time", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Embargo support — article exists but hidden until embargo lifts", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                    ],
                },
                {
                    "name": "Medical Review Process",
                    "features": [
                        {"feature": "Medical review assignment — route article to qualified reviewer by specialty", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Medical review checklist — factual accuracy, source verification, bias check", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Medical review badge auto-applied after approval", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Review expiry — flag articles needing re-review after 12 months", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Category-based review rules (patient content = mandatory, tech news = optional)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                    ],
                },
                {
                    "name": "Editorial Calendar",
                    "features": [
                        {"feature": "Calendar view of scheduled content (month/week view)", "priority": "P1", "phase": "V1.1", "complexity": "L"},
                        {"feature": "Drag-and-drop scheduling on calendar", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Category distribution view — ensure balanced coverage across categories", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Health awareness days integration (World Heart Day, etc.) — suggested content", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Planned vs published tracking", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                    ],
                },
            ],
        },
    ]


def sheet_seo():
    return [
        {
            "module": "Technical SEO",
            "submodules": [
                {
                    "name": "Schema Markup",
                    "features": [
                        {"feature": "Schema.org Article markup on all articles (headline, author, datePublished)", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Schema.org MedicalWebPage markup on patient education content", "priority": "P0", "phase": "MVP", "complexity": "M", "notes": "Critical for Google YMYL trust signals"},
                        {"feature": "Schema.org Person markup for authors (name, credentials, sameAs links)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Schema.org FAQPage markup on FAQ-style articles", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Schema.org Review markup on review content (aggregateRating)", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Schema.org BreadcrumbList for navigation trail", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Schema.org Organization markup on about page", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "JSON-LD injection (not microdata) for all schema", "priority": "P0", "phase": "MVP", "complexity": "S"},
                    ],
                },
                {
                    "name": "Meta & Crawlability",
                    "features": [
                        {"feature": "Auto-generated meta title + description with manual override", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Canonical URL on every page (prevent duplicate content)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "XML sitemap auto-generation (articles, categories, authors, tags)", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "robots.txt configuration", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Open Graph meta tags (og:title, og:description, og:image)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Twitter Card meta tags (summary_large_image)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "hreflang tags for future multilingual support (Hindi, regional)", "priority": "P2", "phase": "V2", "complexity": "M"},
                        {"feature": "Structured URL patterns: /category/subcategory/article-slug", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "301 redirect management for URL changes", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                    ],
                },
                {
                    "name": "Performance",
                    "features": [
                        {"feature": "Core Web Vitals optimization — LCP < 2.5s, FID < 100ms, CLS < 0.1", "priority": "P0", "phase": "MVP", "complexity": "L", "notes": "Google ranking factor"},
                        {"feature": "Static site generation (SSG) for articles via Astro", "priority": "P0", "phase": "MVP", "complexity": "L"},
                        {"feature": "Image optimization pipeline (WebP/AVIF, lazy loading, responsive srcset)", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "CDN caching (Cloudflare) with cache-busting on content updates", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Lighthouse CI — automated score tracking in deploys", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Critical CSS inlining for above-the-fold content", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                    ],
                },
            ],
        },
        {
            "module": "Content SEO",
            "submodules": [
                {
                    "name": "SEO Tools for Writers",
                    "features": [
                        {"feature": "SEO preview panel in editor (Google snippet preview, char counts)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Focus keyword field with density indicator", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Internal linking suggestions (suggest related published articles while writing)", "priority": "P1", "phase": "V1.1", "complexity": "L"},
                        {"feature": "Readability score (Flesch-Kincaid) displayed in editor", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Alt text enforcement — warn if images missing alt text", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Heading structure validator (H1 → H2 → H3, no skips)", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                    ],
                },
                {
                    "name": "Internal Linking",
                    "features": [
                        {"feature": "Auto-suggest internal links based on content similarity", "priority": "P2", "phase": "V1.2", "complexity": "L"},
                        {"feature": "Orphan content report (articles with no internal links pointing to them)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Broken link checker (scan all internal links weekly)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                    ],
                },
            ],
        },
    ]


def sheet_newsletter():
    return [
        {
            "module": "Newsletter System",
            "submodules": [
                {
                    "name": "Newsletter Core",
                    "features": [
                        {"feature": "Email newsletter signup form (name, email, preference checkboxes)", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Double opt-in confirmation email", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "One-click unsubscribe (CAN-SPAM/GDPR compliant)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Newsletter preference center (select categories, frequency)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Welcome email sequence (3-part onboarding drip)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Ghost built-in newsletter (free tier, no extra cost)", "priority": "P0", "phase": "MVP", "complexity": "S", "notes": "Ghost CMS includes newsletter features"},
                    ],
                },
                {
                    "name": "Newsletter Types",
                    "features": [
                        {"feature": "Daily digest — top 5 stories across all categories", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Weekly category newsletters (one per active category)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Breaking news alert email (manual trigger for urgent stories)", "priority": "P1", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Monthly roundup — best of the month, editor picks", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Personalized newsletter — AI-selected based on reading history", "priority": "P2", "phase": "V2", "complexity": "XL"},
                    ],
                },
                {
                    "name": "Newsletter Analytics",
                    "features": [
                        {"feature": "Open rate, click rate, unsubscribe rate per newsletter", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Top clicked links per newsletter issue", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Subscriber growth chart (daily/weekly/monthly)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Segment performance comparison", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                    ],
                },
            ],
        },
        {
            "module": "Membership & Access Tiers",
            "submodules": [
                {
                    "name": "Tier Structure",
                    "features": [
                        {"feature": "Free tier — all content accessible, ad-supported", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Registered tier — free account for bookmarks, following, comments", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Premium tier — ad-free, exclusive content, early access", "priority": "P2", "phase": "V1.2", "complexity": "L"},
                        {"feature": "Stripe integration for paid subscriptions", "priority": "P2", "phase": "V1.2", "complexity": "L"},
                        {"feature": "Ghost membership system (built-in free/paid tiers)", "priority": "P2", "phase": "V1.2", "complexity": "M", "notes": "Ghost has built-in Stripe integration"},
                    ],
                },
                {
                    "name": "User Accounts",
                    "features": [
                        {"feature": "Email/password registration", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Google OAuth login", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "User profile page (name, photo, bio, followed topics/authors)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Account settings (email preferences, password change, delete account)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                    ],
                },
            ],
        },
    ]


def sheet_social_distribution():
    return [
        {
            "module": "Social Sharing",
            "submodules": [
                {
                    "name": "Share Functionality",
                    "features": [
                        {"feature": "Share buttons on every article: Twitter/X, LinkedIn, Facebook, WhatsApp, copy link", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "WhatsApp share with pre-formatted message (title + excerpt + URL)", "priority": "P0", "phase": "MVP", "complexity": "S", "notes": "Critical for Indian audience"},
                        {"feature": "Native share API integration (mobile browsers share sheet)", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Share count display (optional, can be vanity metric)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Text selection share (highlight text → share popup)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                    ],
                },
                {
                    "name": "Social Preview",
                    "features": [
                        {"feature": "Auto-generated social card images (og:image) with article title overlay", "priority": "P1", "phase": "MVP", "complexity": "M"},
                        {"feature": "Custom social card image upload per article", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Twitter Card preview in editor before publishing", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                    ],
                },
            ],
        },
        {
            "module": "Distribution Channels",
            "submodules": [
                {
                    "name": "RSS & Syndication",
                    "features": [
                        {"feature": "Main RSS feed (all content)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Per-category RSS feeds", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Per-author RSS feeds", "priority": "P2", "phase": "V1.2", "complexity": "S"},
                        {"feature": "Full-text RSS (not truncated) for reader apps", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Google News sitemap submission", "priority": "P0", "phase": "MVP", "complexity": "M", "notes": "Critical for news discovery"},
                        {"feature": "Apple News format export", "priority": "P2", "phase": "V2", "complexity": "L"},
                    ],
                },
                {
                    "name": "Social Auto-Posting",
                    "features": [
                        {"feature": "Auto-post to Twitter/X on article publish (with custom text)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Auto-post to LinkedIn company page on publish", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Auto-post to Telegram channel for breaking news", "priority": "P1", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Social posting queue — schedule social shares at optimal times", "priority": "P2", "phase": "V2", "complexity": "L"},
                    ],
                },
                {
                    "name": "Content API & Embeds",
                    "features": [
                        {"feature": "Public content API (read-only) for partners and data consumers", "priority": "P2", "phase": "V1.2", "complexity": "L", "platform": "API"},
                        {"feature": "Embeddable article widget (iframe snippet for partner sites)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "oEmbed support for article URLs", "priority": "P2", "phase": "V2", "complexity": "M"},
                    ],
                },
            ],
        },
    ]


def sheet_ux_personalization():
    return [
        {
            "module": "Personalization",
            "submodules": [
                {
                    "name": "Following System",
                    "features": [
                        {"feature": "Follow categories — add to personalized feed", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Follow authors — add author's content to feed", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Follow tags/topics — fine-grained interest tracking", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "\"My Feed\" page — personalized content based on follows", "priority": "P1", "phase": "V1.1", "complexity": "L"},
                        {"feature": "Onboarding flow — select interests on first visit", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                    ],
                },
                {
                    "name": "Reading Features",
                    "features": [
                        {"feature": "Bookmark / Save for Later (requires account)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Reading history — recently read articles list", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "\"Continue reading\" — resume from where you left off on long articles", "priority": "P2", "phase": "V2", "complexity": "L"},
                        {"feature": "Bookmarks organized in collections/folders", "priority": "P2", "phase": "V2", "complexity": "M"},
                    ],
                },
            ],
        },
        {
            "module": "Reading Experience",
            "submodules": [
                {
                    "name": "Accessibility & Comfort",
                    "features": [
                        {"feature": "Dark mode toggle (system preference + manual override)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Font size adjustment (S / M / L / XL)", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Serif / sans-serif font toggle for articles", "priority": "P2", "phase": "V1.2", "complexity": "S"},
                        {"feature": "WCAG 2.1 AA compliance (contrast, keyboard nav, screen reader)", "priority": "P0", "phase": "MVP", "complexity": "L"},
                        {"feature": "Reduced motion preference respect", "priority": "P1", "phase": "MVP", "complexity": "S"},
                    ],
                },
                {
                    "name": "Comments & Discussion",
                    "features": [
                        {"feature": "Comment section on articles (requires registered account)", "priority": "P2", "phase": "V1.2", "complexity": "L"},
                        {"feature": "Threaded replies (max 2 levels deep)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Comment moderation queue for editors", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Comment upvote / helpful vote", "priority": "P2", "phase": "V2", "complexity": "S"},
                        {"feature": "Expert comment badge (verified medical professional comments)", "priority": "P2", "phase": "V2", "complexity": "M", "notes": "Trust differentiator"},
                        {"feature": "Report comment for misinformation (healthcare-specific moderation)", "priority": "P2", "phase": "V1.2", "complexity": "M", "notes": "Patient safety — moderate medical claims"},
                    ],
                },
            ],
        },
    ]


def sheet_healthcare_specific():
    return [
        {
            "module": "Medical Compliance",
            "submodules": [
                {
                    "name": "Disclaimer System",
                    "features": [
                        {"feature": "Global medical disclaimer footer on every page", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Category-specific disclaimers (patient content gets stronger disclaimer)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Article-level custom disclaimer override (for sensitive topics)", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Emergency disclaimer — \"If you are experiencing a medical emergency, call 112\"", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Drug information disclaimer — \"Do not self-medicate based on this content\"", "priority": "P0", "phase": "MVP", "complexity": "S"},
                    ],
                },
                {
                    "name": "E-E-A-T Compliance",
                    "features": [
                        {"feature": "Author expertise display on every article (credentials, specialization)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Medical reviewer badge with reviewer profile link", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Source citations section with links to peer-reviewed studies", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "About Us page with editorial policy and medical review process", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Editorial standards page — fact-checking, corrections policy", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Corrections/updates log — transparent correction history per article", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                    ],
                },
                {
                    "name": "Data Privacy (DPDP Act)",
                    "features": [
                        {"feature": "Cookie consent banner (DPDP Act / ePrivacy compliant)", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Privacy policy page with data processing details", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "User data export (right to portability)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Account deletion (right to erasure)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "No health data collection beyond reading preferences", "priority": "P0", "phase": "MVP", "complexity": "S", "notes": "Platform is content-only, never collects patient health data"},
                    ],
                },
            ],
        },
        {
            "module": "Healthcare Content Tools",
            "submodules": [
                {
                    "name": "Medical Terminology",
                    "features": [
                        {"feature": "Medical terminology tooltips — hover for plain-language explanation", "priority": "P2", "phase": "V1.2", "complexity": "L"},
                        {"feature": "Glossary page — searchable medical terms dictionary", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "ICD-10 code linking — inline references to diagnosis codes", "priority": "P2", "phase": "V2", "complexity": "M"},
                    ],
                },
                {
                    "name": "Interactive Health Widgets",
                    "features": [
                        {"feature": "BMI calculator embed widget", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Drug interaction checker link/embed (link to external tool)", "priority": "P2", "phase": "V2", "complexity": "M"},
                        {"feature": "Clinical trial search integration (link to CTRI/ClinicalTrials.gov)", "priority": "P2", "phase": "V2", "complexity": "M"},
                        {"feature": "CME quiz component — embed quiz at end of clinical articles", "priority": "P2", "phase": "V2", "complexity": "L", "notes": "Value-add for doctor audience"},
                        {"feature": "Dosage calculator embed (with mandatory disclaimer)", "priority": "P2", "phase": "V2", "complexity": "L"},
                    ],
                },
                {
                    "name": "Health Awareness",
                    "features": [
                        {"feature": "Health awareness calendar integration (WHO health days)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Disease outbreak tracking dashboard (link to IDSP/WHO data)", "priority": "P2", "phase": "V2", "complexity": "L"},
                        {"feature": "Vaccination schedule reference page (India NIS)", "priority": "P2", "phase": "V2", "complexity": "M"},
                    ],
                },
            ],
        },
    ]


def sheet_analytics_monetization():
    return [
        {
            "module": "Analytics",
            "submodules": [
                {
                    "name": "Traffic Analytics",
                    "features": [
                        {"feature": "Page view tracking (privacy-respecting, e.g., Plausible/Fathom or GA4)", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Unique visitor counts (daily, weekly, monthly)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Traffic source breakdown (organic, social, direct, referral, email)", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Geographic distribution of readers", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Device/browser breakdown", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                    ],
                },
                {
                    "name": "Content Performance",
                    "features": [
                        {"feature": "Article performance dashboard: views, time on page, scroll depth, shares", "priority": "P1", "phase": "V1.1", "complexity": "L"},
                        {"feature": "Category performance — views, engagement, growth per category", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Author performance — articles published, total views, avg engagement", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Top performing content report (last 7/30/90 days)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Content decay report — articles with declining traffic (need refresh)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Scroll depth tracking (how far readers actually read)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                    ],
                },
                {
                    "name": "A/B Testing",
                    "features": [
                        {"feature": "A/B test headlines (show different headlines, measure CTR)", "priority": "P2", "phase": "V1.2", "complexity": "L"},
                        {"feature": "A/B test featured images", "priority": "P2", "phase": "V2", "complexity": "M"},
                        {"feature": "A/B test newsletter subject lines", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                    ],
                },
            ],
        },
        {
            "module": "Monetization",
            "submodules": [
                {
                    "name": "Advertising",
                    "features": [
                        {"feature": "Google AdSense integration (auto-placed display ads)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Direct ad sales management — sell ad slots to healthcare vendors", "priority": "P2", "phase": "V1.2", "complexity": "L", "notes": "Higher CPM than programmatic"},
                        {"feature": "Ad placement zones: header banner, sidebar, in-article, footer", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Ad-free experience for premium subscribers", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Healthcare ad policy enforcement (no misleading health claims in ads)", "priority": "P1", "phase": "V1.1", "complexity": "S", "notes": "Editorial integrity"},
                    ],
                },
                {
                    "name": "Sponsored Content",
                    "features": [
                        {"feature": "Sponsored article template with prominent 'Sponsored' badge", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Sponsored content disclosure (compliant with ASCI guidelines)", "priority": "P1", "phase": "V1.1", "complexity": "S", "notes": "ASCI compliance for Indian market"},
                        {"feature": "Sponsor dashboard — view metrics of their sponsored content", "priority": "P2", "phase": "V1.2", "complexity": "L"},
                        {"feature": "Native advertising format (blends with editorial but clearly labeled)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                    ],
                },
                {
                    "name": "MedBrains HMS Lead Generation",
                    "features": [
                        {"feature": "CTA banners on relevant articles (Hospital Leadership, Operations, Technology)", "priority": "P0", "phase": "MVP", "complexity": "S", "notes": "Core business model — HMS lead gen"},
                        {"feature": "Demo request form embedded in technology/operations articles", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Lead attribution tracking — which article drove each demo request", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "MedBrains case study integration (embedded within relevant content)", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                        {"feature": "Exit-intent popup for HMS demo on hospital admin-focused articles", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Lead scoring — track content consumption for lead qualification", "priority": "P2", "phase": "V2", "complexity": "L"},
                        {"feature": "CRM integration (forward leads to sales pipeline)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                    ],
                },
                {
                    "name": "Affiliate & Partnerships",
                    "features": [
                        {"feature": "Affiliate link management (tracking, disclosure)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Partner content hub — branded pages for sponsors/partners", "priority": "P2", "phase": "V2", "complexity": "L"},
                        {"feature": "Webinar/event co-hosting with sponsors", "priority": "P2", "phase": "V2", "complexity": "L"},
                        {"feature": "Revenue reporting dashboard (ads + sponsors + subscriptions + affiliates + leads)", "priority": "P2", "phase": "V1.2", "complexity": "L"},
                    ],
                },
            ],
        },
    ]


def sheet_admin():
    return [
        {
            "module": "CMS Administration",
            "submodules": [
                {
                    "name": "Dashboard",
                    "features": [
                        {"feature": "Admin dashboard — published today, scheduled, drafts in review, total articles", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Quick stats: total subscribers, page views today, trending articles", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Content pipeline view (articles by status across workflow stages)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                    ],
                },
                {
                    "name": "User & Role Management",
                    "features": [
                        {"feature": "Roles: Super Admin, Editor-in-Chief, Section Editor, Writer, Medical Reviewer, Contributor", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Invite writer / contributor by email", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Per-category editor assignment (Section Editor for Technology, etc.)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Activity log — who published/edited/deleted what and when", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                    ],
                },
                {
                    "name": "Content Operations",
                    "features": [
                        {"feature": "Bulk operations — bulk publish, bulk unpublish, bulk delete, bulk tag", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Content import — import from WordPress, Medium, Ghost (JSON)", "priority": "P2", "phase": "V1.2", "complexity": "L"},
                        {"feature": "Content export — full backup as JSON/XML", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Media library — centralized image/file management", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Image upload with auto-compression and format conversion", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Scheduled content auto-publish (cron-based, timezone-aware)", "priority": "P0", "phase": "MVP", "complexity": "M"},
                    ],
                },
            ],
        },
        {
            "module": "Infrastructure",
            "submodules": [
                {
                    "name": "Hosting & Deployment",
                    "features": [
                        {"feature": "Cloudflare Pages hosting (static frontend)", "priority": "P0", "phase": "MVP", "complexity": "M", "notes": "Free tier, global CDN"},
                        {"feature": "Ghost CMS hosted instance (headless API mode)", "priority": "P0", "phase": "MVP", "complexity": "M", "notes": "~$30/month Ghost Pro or self-hosted"},
                        {"feature": "CI/CD pipeline — auto-deploy on push to main", "priority": "P0", "phase": "MVP", "complexity": "M"},
                        {"feature": "Staging environment for content preview before publish", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Automated daily backups (content + database)", "priority": "P0", "phase": "MVP", "complexity": "M"},
                    ],
                },
                {
                    "name": "Monitoring",
                    "features": [
                        {"feature": "Uptime monitoring with alerts (UptimeRobot/Checkly)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Error tracking (Sentry or similar)", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "Performance monitoring — track Core Web Vitals in production", "priority": "P1", "phase": "V1.1", "complexity": "M"},
                        {"feature": "CDN cache hit ratio monitoring", "priority": "P2", "phase": "V1.2", "complexity": "S"},
                    ],
                },
                {
                    "name": "Security",
                    "features": [
                        {"feature": "SSL/TLS certificate (auto via Cloudflare)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "DDoS protection (Cloudflare WAF)", "priority": "P0", "phase": "MVP", "complexity": "S"},
                        {"feature": "Content Security Policy headers", "priority": "P1", "phase": "MVP", "complexity": "S"},
                        {"feature": "Regular dependency vulnerability scanning", "priority": "P1", "phase": "V1.1", "complexity": "S"},
                    ],
                },
            ],
        },
    ]


def sheet_news():
    return [
        {
            "module": "Indian Healthcare News",
            "submodules": [
                {
                    "name": "News Workflow",
                    "features": [
                        {"feature": "Breaking news fast-publish workflow (skip medical review for news)", "priority": "P1", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Breaking news banner on homepage (dismissible, auto-expire)", "priority": "P1", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Priority alert email for breaking healthcare news", "priority": "P1", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Push notification for breaking news (web push)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "News article template — leaner than long-form, focus on facts + sources", "priority": "P1", "phase": "V1.2", "complexity": "M"},
                    ],
                },
                {
                    "name": "Policy & Government",
                    "features": [
                        {"feature": "Healthcare policy tracker page (active policies, pending bills, implementations)", "priority": "P2", "phase": "V1.2", "complexity": "L"},
                        {"feature": "Ayushman Bharat / PMJAY scheme updates section", "priority": "P1", "phase": "V1.2", "complexity": "M"},
                        {"feature": "ABDM (Ayushman Bharat Digital Mission) coverage — ABHA, health records", "priority": "P1", "phase": "V1.2", "complexity": "M"},
                        {"feature": "NMC (National Medical Commission) regulatory updates", "priority": "P1", "phase": "V1.2", "complexity": "M"},
                        {"feature": "NABH accreditation news and guideline changes", "priority": "P1", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Union Budget healthcare impact analysis (annual special coverage)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "State health policy tracker (key states: Maharashtra, Karnataka, TN, Kerala, Delhi)", "priority": "P2", "phase": "V2", "complexity": "L"},
                    ],
                },
                {
                    "name": "Industry News",
                    "features": [
                        {"feature": "Hospital industry news — mergers, expansions, new facilities", "priority": "P1", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Healthcare startup funding news", "priority": "P1", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Court ruling / legal update alerts (Supreme Court, NCDRC health rulings)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Drug price regulation news (NPPA, DPCO updates)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "Medical education news (NEET, NEXT exam, new medical colleges)", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                    ],
                },
                {
                    "name": "News Aggregation & Events",
                    "features": [
                        {"feature": "PIB (Press Information Bureau) health news monitoring", "priority": "P2", "phase": "V2", "complexity": "L"},
                        {"feature": "MoHFW (Ministry of Health) press release tracking", "priority": "P2", "phase": "V2", "complexity": "L"},
                        {"feature": "Live blog for major events (HIMSS India, AHPI Conference, NABH events)", "priority": "P2", "phase": "V1.2", "complexity": "L"},
                        {"feature": "Event calendar — upcoming healthcare conferences, webinars, summits", "priority": "P2", "phase": "V1.2", "complexity": "M"},
                        {"feature": "News ticker / scrolling banner for latest updates", "priority": "P2", "phase": "V2", "complexity": "M"},
                    ],
                },
            ],
        },
    ]


def sheet_mobile_app():
    return [
        {
            "module": "Mobile App (Native)",
            "submodules": [
                {
                    "name": "Core Reading",
                    "features": [
                        {"feature": "Native article reader with optimized typography", "priority": "P2", "phase": "V2", "complexity": "L", "platform": "Mobile"},
                        {"feature": "Offline reading — download articles for offline access", "priority": "P2", "phase": "V2", "complexity": "L", "platform": "Mobile"},
                        {"feature": "Bookmarks sync with web account", "priority": "P2", "phase": "V2", "complexity": "M", "platform": "Mobile"},
                        {"feature": "Follow categories/authors/topics synced with web", "priority": "P2", "phase": "V2", "complexity": "M", "platform": "Mobile"},
                        {"feature": "Dark mode with OLED black option", "priority": "P2", "phase": "V2", "complexity": "M", "platform": "Mobile"},
                    ],
                },
                {
                    "name": "Notifications",
                    "features": [
                        {"feature": "Push notification for breaking news", "priority": "P2", "phase": "V2", "complexity": "M", "platform": "Mobile"},
                        {"feature": "Push notification for followed author new article", "priority": "P2", "phase": "V2", "complexity": "M", "platform": "Mobile"},
                        {"feature": "Notification preferences (category-level control)", "priority": "P2", "phase": "V2", "complexity": "M", "platform": "Mobile"},
                        {"feature": "Silent/digest notification option (bundle into daily digest)", "priority": "P2", "phase": "V2", "complexity": "M", "platform": "Mobile"},
                    ],
                },
                {
                    "name": "Mobile-Specific",
                    "features": [
                        {"feature": "Share extension — share articles from other apps to bookmarks", "priority": "P2", "phase": "V2", "complexity": "M", "platform": "Mobile"},
                        {"feature": "Widget — today's top 3 headlines on home screen", "priority": "P2", "phase": "V2", "complexity": "M", "platform": "Mobile"},
                        {"feature": "Text-to-speech for articles (accessibility + commute listening)", "priority": "P2", "phase": "V2", "complexity": "L", "platform": "Mobile"},
                        {"feature": "Haptic feedback on interactions", "priority": "P2", "phase": "V2", "complexity": "S", "platform": "Mobile"},
                        {"feature": "App Store / Play Store listing optimization (ASO)", "priority": "P2", "phase": "V2", "complexity": "M", "platform": "Mobile"},
                    ],
                },
            ],
        },
    ]


def sheet_categories_reference(wb):
    """Sheet 14: Content Categories Reference — not a feature sheet, a reference."""
    ws = wb.create_sheet("Content Categories Ref")
    headers = ["#", "Category Name", "Target Audience", "Content Types",
               "Posting Frequency", "Example Topics", "Competitor Benchmarks",
               "SEO Opportunity", "Phase"]
    header_widths = [5, 28, 30, 30, 15, 40, 30, 15, 10]

    for col_idx, (title, width) in enumerate(zip(headers, header_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = REF_HEADER_FILL
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    categories = [
        [1, "Hospital Leadership & Strategy", "CEOs, CMOs, Hospital Admins, Board Members",
         "Opinion pieces, Case studies, Interviews, Data analysis",
         "3-4/week", "Hospital growth strategies, NABH prep, Financial management, Leadership interviews",
         "Becker's Hospital Review, Healthcare Finance News", "High", "MVP"],
        [2, "Patient Education & Wellness", "Patients, Families, Caregivers",
         "Guides, How-tos, Explainers, Infographics, Q&A",
         "5-7/week", "Disease guides, Wellness tips, Nutrition, Mental health, Pregnancy",
         "Cleveland Clinic Health Essentials, Healthline, Mayo Clinic", "Very High", "MVP"],
        [3, "Doctor & Clinical Excellence", "Doctors, Nurses, Clinical Staff",
         "Clinical pearls, Evidence summaries, Expert opinions, CME",
         "3-4/week", "Clinical guidelines, Case reports, Research updates, Drug updates",
         "Medscape, NEJM Journal Watch, UpToDate Blog", "High", "MVP"],
        [4, "Operations & Efficiency", "Hospital Managers, Dept Heads, COOs",
         "How-to guides, Checklists, Templates, Benchmarks",
         "2-3/week", "Process improvement, Supply chain, Staffing, Facility management",
         "Becker's, Advisory Board, HFMA", "Medium", "MVP"],
        [5, "Technology & Innovation", "CIOs, IT Managers, Health Tech Professionals",
         "Product reviews, Analysis, Tutorials, Case studies",
         "3-5/week", "AI in healthcare, EHR/EMR, Telemedicine, Cybersecurity, Interoperability",
         "Healthcare IT News, HIMSS, MobiHealthNews", "High", "MVP"],
        [6, "Regulatory & Compliance", "Compliance Officers, Legal Teams, Quality Heads",
         "Analysis, Alerts, Checklists, Explainers",
         "2-3/week", "NABH/JCI updates, NMC regulations, Legal rulings, DPDP Act, PCPNDT",
         "Compliance Today (HCCA), Regulatory Focus", "Medium", "MVP"],
        [7, "Nursing & Allied Health", "Nurses, Physiotherapists, Technicians",
         "Practice guides, Career advice, Skill building, Research",
         "2-3/week", "Nursing practice, Patient safety, Career growth, Skill workshops",
         "American Nurse Journal, Nursing Times", "Medium", "V1.1"],
        [8, "Pharma Industry", "Pharma Cos, Distributors, Drug Manufacturers",
         "Industry news, Analysis, Market reports, Interviews",
         "3-4/week", "Drug launches, Pricing policy, Supply chain, R&D pipeline, CDSCO updates",
         "Pharma Times, ET Healthworld Pharma", "High", "V1.1"],
        [9, "Medical Devices & Equipment", "Manufacturers, Biomedical Engineers, Procurement",
         "Reviews, Buyer's guides, Regulatory, Innovation",
         "2-3/week", "Device reviews, Procurement guides, CDSCO approvals, Innovation showcase",
         "MedTech Dive, Medical Device Network", "Medium", "V1.1"],
        [10, "Health Insurance & TPA", "Insurers, TPA Executives, Brokers",
         "Policy analysis, Claim trends, Regulatory, Product reviews",
         "2/week", "Policy analysis, Claim trends, IRDAI updates, New product launches",
         "Insurance Journal, Policybazaar Blog", "Medium", "V1.2"],
        [11, "HealthTech & Startups", "Founders, Investors, Digital Health Builders",
         "Funding news, Product launches, Founder interviews, Market analysis",
         "3-4/week", "Funding rounds, Product launches, Founder stories, Market maps",
         "Rock Health, CB Insights Healthcare, YourStory Health", "High", "V1.2"],
        [12, "Medical Education & Research", "Students, Academics, Researchers",
         "Study guides, Research methods, Conference reports, Career advice",
         "2/week", "NEET/NEXT prep, Residency guidance, Research methodology, Conference highlights",
         "Student BMJ, Medical Education Online", "Medium", "V1.2"],
        [13, "Indian Healthcare News", "All healthcare stakeholders",
         "Breaking news, Policy updates, Court rulings, Budget analysis",
         "Daily", "Policy changes, ABDM updates, Hospital M&A, Court rulings, Budget impact",
         "ET Healthworld, Mint Health, Business Standard Health", "Very High", "V1.2"],
    ]

    for row_idx, cat in enumerate(categories, 2):
        for col_idx, val in enumerate(cat, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = WRAP


def sheet_tech_stack_reference(wb):
    """Sheet 15: Tech Stack Reference — architecture documentation."""
    ws = wb.create_sheet("Tech Stack Reference")
    headers = ["Layer", "Technology", "Purpose", "Cost (Monthly)", "Notes"]
    header_widths = [18, 22, 45, 15, 45]

    for col_idx, (title, width) in enumerate(zip(headers, header_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = REF_HEADER_FILL
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    stack = [
        ["CMS", "Ghost (Headless)", "Content management, newsletter, membership, API", "$30 (Pro) or $5-10 (self-hosted)", "Headless mode — API only, frontend decoupled"],
        ["Frontend", "Astro 5", "Static site generation, island architecture, fast performance", "Free (SSG)", "Content-first framework, 0 JS by default, partial hydration"],
        ["Frontend UI", "Tailwind CSS 4", "Utility-first styling for content pages", "Free", "Pairs well with Astro, great for content layout"],
        ["Hosting", "Cloudflare Pages", "Static hosting with global CDN, edge functions", "Free tier (unlimited sites)", "Fast deploys, auto-SSL, DDoS protection included"],
        ["Image CDN", "Cloudflare Images / R2", "Image optimization, WebP/AVIF, responsive srcset", "$5-10", "Or use Ghost built-in image processing"],
        ["Search", "Pagefind", "Client-side static search (no server needed)", "Free", "Built into Astro, perfect for static sites"],
        ["Analytics", "Plausible / Fathom", "Privacy-respecting page analytics", "$9-14", "GDPR/DPDP compliant, no cookie consent needed"],
        ["Newsletter", "Ghost Built-in", "Email newsletter with segmentation, automation", "Included in Ghost", "Up to 500 members free, scales with Ghost plan"],
        ["Payments", "Stripe (via Ghost)", "Paid memberships, subscriptions", "2.9% + $0.30/tx", "Ghost has native Stripe integration"],
        ["Comments", "Giscus / Cactus", "Discussion system for articles", "Free", "GitHub-based (Giscus) or matrix-based (Cactus)"],
        ["Forms", "Formspree / Tally", "Contact forms, demo request forms", "Free tier", "For MedBrains HMS lead capture forms"],
        ["Monitoring", "UptimeRobot", "Uptime monitoring with alerts", "Free (50 monitors)", "5-min check interval on free tier"],
        ["Error Tracking", "Sentry", "Frontend error tracking", "Free tier", "Track JS errors in production"],
        ["CI/CD", "GitHub Actions", "Build + deploy pipeline", "Free (public repos)", "Auto-deploy on push to main branch"],
        ["DNS", "Cloudflare DNS", "DNS management, proxy", "Free", "Included with Cloudflare Pages"],
        ["Email (Transactional)", "Mailgun / Postmark", "Transactional emails (welcome, password reset)", "$0-15", "Only needed if Ghost email is insufficient"],
        ["", "", "", "", ""],
        ["TOTAL ESTIMATED", "", "", "$45-80/month", "Scales to 100k+ monthly visitors before needing upgrades"],
    ]

    for row_idx, item in enumerate(stack, 2):
        for col_idx, val in enumerate(item, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = WRAP
        if item[0] == "TOTAL ESTIMATED":
            for col_idx in range(1, 6):
                ws.cell(row=row_idx, column=col_idx).font = BOLD


def sheet_summary(wb, sheet_counts):
    """Summary dashboard sheet."""
    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_properties.tabColor = "FF4472C4"

    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value="HealthPulse — Healthcare Publishing Platform Feature Tracker")
    title_cell.font = Font(bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    sub = ws.cell(row=2, column=1, value="Dual Business: MedBrains HMS Lead Generation + Standalone Media Revenue")
    sub.font = Font(italic=True, size=11, color="666666")
    sub.alignment = Alignment(horizontal="center")

    # ── Feature Counts by Sheet ──
    row = 4
    section_headers = ["Sheet Name", "Feature Count"]
    for col_idx, h in enumerate(section_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = SUMMARY_HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = WRAP_CENTER
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15

    row += 1
    total = 0
    for name, count in sheet_counts:
        ws.cell(row=row, column=1, value=name).border = THIN_BORDER
        ws.cell(row=row, column=2, value=count).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = WRAP_CENTER
        total += count
        row += 1

    # Total row
    total_cell_a = ws.cell(row=row, column=1, value="TOTAL FEATURES")
    total_cell_a.font = BOLD
    total_cell_a.border = THIN_BORDER
    total_cell_b = ws.cell(row=row, column=2, value=total)
    total_cell_b.font = BOLD
    total_cell_b.border = THIN_BORDER
    total_cell_b.alignment = WRAP_CENTER

    # ── Phase Breakdown ──
    row += 2
    phase_header_row = row
    phase_headers = ["Phase", "Description", "Est. Features"]
    ws.column_dimensions["C"].width = 15
    for col_idx, h in enumerate(phase_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = SUMMARY_HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = WRAP_CENTER

    phases = [
        ["MVP", "Launch-ready blog with 6 core categories, articles, quick posts, SEO, medical review", "~90-100"],
        ["V1.1", "Newsletter, membership, reviews, specialty verticals (7-9), editorial calendar", "~70-80"],
        ["V1.2", "News vertical, live blogs, personalization, monetization, comments, buyer's guides", "~70-80"],
        ["V2", "Mobile app, advanced analytics, AI features, podcasts, interactive health tools", "~80-90"],
    ]
    for phase in phases:
        row += 1
        for col_idx, val in enumerate(phase, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = WRAP

    # ── Priority Breakdown ──
    row += 2
    prio_headers = ["Priority", "Meaning"]
    for col_idx, h in enumerate(prio_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = SUMMARY_HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = WRAP_CENTER

    priorities = [
        ["P0", "Must-have for launch — blocking"],
        ["P1", "High priority — needed soon after launch"],
        ["P2", "Nice to have — growth phase features"],
        ["P3", "Future / aspirational"],
    ]
    for p in priorities:
        row += 1
        for col_idx, val in enumerate(p, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = WRAP

    # ── Business Model Summary ──
    row += 2
    biz_cell = ws.cell(row=row, column=1, value="Business Model")
    biz_cell.font = Font(bold=True, size=12, color="2F5496")
    row += 1
    biz_items = [
        "1. MedBrains HMS Lead Generation — CTA banners, demo forms, case studies, lead attribution",
        "2. Display Advertising — Google AdSense + direct healthcare vendor ad sales",
        "3. Sponsored Content — native advertising, sponsor dashboards, ASCI-compliant disclosure",
        "4. Paid Subscriptions — premium ad-free tier via Ghost membership + Stripe",
        "5. Newsletter Sponsorships — sponsored slots in email newsletters",
        "6. Affiliate Revenue — product review affiliate links, buyer's guide CTAs",
        "7. Events & Webinars — co-hosted with healthcare sponsors (V2)",
    ]
    for item in biz_items:
        ws.cell(row=row, column=1, value=item).alignment = WRAP
        row += 1


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    sheet_defs = [
        ("Platform Core", sheet_platform_core),
        ("Content Types & Editor", sheet_content_types),
        ("Categories & Taxonomy", sheet_categories),
        ("Author & Editorial", sheet_author_editorial),
        ("SEO & Discovery", sheet_seo),
        ("Newsletter & Membership", sheet_newsletter),
        ("Social & Distribution", sheet_social_distribution),
        ("UX & Personalization", sheet_ux_personalization),
        ("Healthcare-Specific", sheet_healthcare_specific),
        ("Analytics & Monetization", sheet_analytics_monetization),
        ("Admin & Operations", sheet_admin),
        ("News & Breaking Coverage", sheet_news),
        ("Mobile App (Future)", sheet_mobile_app),
    ]

    sheet_counts = []
    for name, data_fn in sheet_defs:
        ws = wb.create_sheet(name)
        count = write_features(ws, data_fn())
        sheet_counts.append((name, count))
        print(f"  {name}: {count} features")

    # Reference sheets (no features, just data)
    sheet_categories_reference(wb)
    sheet_tech_stack_reference(wb)

    # Summary dashboard (first sheet)
    sheet_summary(wb, sheet_counts)

    # Move summary to first position
    wb.move_sheet("Summary Dashboard", offset=-len(wb.sheetnames) + 1)

    # Save
    output_path = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                               "Healthcare_Blog_Platform_Features.xlsx")
    wb.save(output_path)
    total = sum(c for _, c in sheet_counts)
    print(f"\nSaved: {output_path}")
    print(f"Total features: {total}")
    print(f"Sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()

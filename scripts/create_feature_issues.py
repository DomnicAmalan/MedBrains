#!/usr/bin/env python3
"""Create one GitHub epic issue per module from MedBrains_Features.xlsx.

Each module sheet becomes a single epic issue whose body lists EVERY feature
(Done + Pending) as a checklist item grouped by sub-module — Done checked,
Pending unchecked — so the whole module is reviewable in one place (catching
anything missed). Idempotent: skips a module whose epic issue already exists.

  python3 scripts/create_feature_issues.py            # create missing epics
  python3 scripts/create_feature_issues.py --dry-run  # print, don't create
"""

from __future__ import annotations

import json
import subprocess
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "MedBrains_Features.xlsx"
EPIC_PREFIX = "Epic (features):"
BODY_LIMIT = 60000  # GitHub issue body cap is 65536; stay under


def cell(v) -> str:
    return "" if v is None else str(v).strip()


def existing_epic_titles() -> set[str]:
    out = subprocess.run(
        ["gh", "issue", "list", "--state", "all", "--limit", "500", "--json", "title"],
        capture_output=True,
        text=True,
        cwd=ROOT,
    )
    if out.returncode != 0:
        return set()
    return {i["title"] for i in json.loads(out.stdout or "[]")}


def build_body(name: str, by_sub: dict[str, list[dict]], totals: dict) -> str:
    lines = [
        f"Generated from `MedBrains_Features.xlsx` — every feature in **{name}** as a "
        "reviewable checklist (✅/checked = Done). Review the Done items too in case "
        "something was marked done but is missing or partial.",
        "",
        f"**{totals['all']} features** · {totals['done']} done · {totals['partial']} partial · "
        f"{totals['pending']} pending. Detailed stories + acceptance criteria: "
        f"`medbrains/docs/backlog/stories/`.",
        "",
    ]
    for sub in sorted(by_sub):
        lines.append(f"### {sub}")
        for f in by_sub[sub]:
            box = "[x]" if f["status"].lower() == "done" else "[ ]"
            tags = " ".join(
                t
                for t in (
                    f"`{f['priority']}`" if f["priority"] else "",
                    f"`{f['status']}`" if f["status"] else "",
                    f"`{f['platforms']}`" if f["platforms"] else "",
                )
                if t
            )
            lines.append(f"- {box} {f['feature']} {tags}".rstrip())
        lines.append("")
    body = "\n".join(lines)
    if len(body) > BODY_LIMIT:
        body = body[:BODY_LIMIT] + "\n\n_…truncated; see the story markdown for the full list._\n"
    return body


def main(argv: list[str]) -> int:
    dry = "--dry-run" in argv
    if not XLSX.exists():
        print(f"{XLSX} not found")
        return 1
    existing = set() if dry else existing_epic_titles()
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    created = skipped = 0

    for name in wb.sheetnames:
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            continue
        col = {cell(c).lower(): i for i, c in enumerate(header) if c}

        def g(r, key):
            i = col.get(key)
            return cell(r[i]) if i is not None and i < len(r) else ""

        by_sub: dict[str, list[dict]] = defaultdict(list)
        totals = {"all": 0, "done": 0, "partial": 0, "pending": 0}
        for r in it:
            if not r or not any(x is not None and cell(x) for x in r):
                continue
            feature = g(r, "feature")
            if not feature:
                continue
            status = g(r, "status")
            platforms = ", ".join(
                lbl
                for k, lbl in (("web", "Web"), ("mobile", "Mobile"), ("tv", "TV"))
                if g(r, k).upper() in ("Y", "YES", "✓")
            )
            by_sub[g(r, "sub-module") or "General"].append(
                {"feature": feature, "priority": g(r, "priority"), "status": status, "platforms": platforms}
            )
            totals["all"] += 1
            totals[status.lower()] = totals.get(status.lower(), 0) + 1

        if totals["all"] == 0:
            continue
        title = f"{EPIC_PREFIX} {name}"
        if title in existing:
            print(f"  skip (exists): {title}")
            skipped += 1
            continue
        body = build_body(name, by_sub, totals)
        if dry:
            print(f"  would create: {title}  ({totals['all']} features, {len(body)} chars)")
            created += 1
            continue
        res = subprocess.run(
            ["gh", "issue", "create", "--title", title, "--body", body],
            capture_output=True,
            text=True,
            cwd=ROOT,
        )
        if res.returncode == 0:
            print(f"  created: {title} -> {res.stdout.strip()}")
            created += 1
        else:
            print(f"  FAILED: {title}: {res.stderr.strip()}")

    print(f"\n{'(dry-run) ' if dry else ''}{created} created, {skipped} skipped")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

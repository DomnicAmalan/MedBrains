#!/usr/bin/env python3
"""Create one GitHub issue per feature (story) from MedBrains_Features.xlsx.

~2,800 issues. Paced to dodge GitHub secondary rate limits, resume-safe (a
local state file records created keys), idempotent (also skips titles that
already exist). Each story links to its module epic and gets platform labels.

  python3 scripts/create_story_issues.py             # create all
  python3 scripts/create_story_issues.py --status pending,partial
  python3 scripts/create_story_issues.py --dry-run
"""

from __future__ import annotations

import json
import subprocess
import sys
import time
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "MedBrains_Features.xlsx"
STATE = ROOT / "scripts" / ".story_issues_done.txt"
EPIC_PREFIX = "Epic (features):"
PACE = 1.5          # seconds between creates
BACKOFF = 90        # seconds to wait on a rate-limit error

MODULE_ROLE = {
    "Onboarding & Setup": "hospital admin", "Clinical": "clinician",
    "Diagnostics & Support": "lab/radiology technician", "Admin & Operations": "operations admin",
    "Specialty & Academic": "specialist", "IT, Security & Infrastructure": "system administrator",
    "TV Displays & Queue": "front-desk/display operator", "Printing & Forms": "staff member",
    "Mobile Apps": "mobile app user", "Technical Infrastructure": "platform engineer",
    "Regulatory & Compliance": "compliance officer", "Multi-Hospital & Vendors": "group administrator",
    "CMS & Blog": "content editor",
}
MODULE_TAG = {
    "Onboarding & Setup": "Onboarding", "Clinical": "Clinical",
    "Diagnostics & Support": "Diagnostics", "Admin & Operations": "Admin",
    "Specialty & Academic": "Specialty", "IT, Security & Infrastructure": "IT-Security",
    "TV Displays & Queue": "TV", "Printing & Forms": "Printing", "Mobile Apps": "Mobile",
    "Technical Infrastructure": "Infra", "Regulatory & Compliance": "Regulatory",
    "Multi-Hospital & Vendors": "Multi-Hospital", "CMS & Blog": "CMS",
}
MODULE_REG = {
    "Clinical": "Clinical coding (ICD-10/11, SNOMED) + IPSG safety (2-ID, allergy/LASA, consent).",
    "Diagnostics & Support": "LOINC + critical-value reporting (NABL); DICOM/AERB/PCPNDT for imaging.",
    "Admin & Operations": "Domain norms where relevant (NDPS/Schedule H/INN for pharmacy; GST/CGHS/TPA for billing).",
    "IT, Security & Infrastructure": "RBAC least-privilege, audit trail, encryption at rest; no PHI in logs.",
    "Specialty & Academic": "Specialty statutes (Mental Healthcare Act 2017, MTP/PCPNDT) + NABH consent.",
    "Regulatory & Compliance": "Maps to the applicable norm (NABH/JCI/NDPS/D&C/PNDT) with accreditation evidence.",
}


def cell(v) -> str:
    return "" if v is None else str(v).strip()


def epic_numbers() -> dict[str, int]:
    out = subprocess.run(["gh", "issue", "list", "--state", "all", "--limit", "60",
                          "--search", EPIC_PREFIX, "--json", "number,title"],
                         capture_output=True, text=True, cwd=ROOT)
    nums = {}
    for i in json.loads(out.stdout or "[]"):
        if i["title"].startswith(EPIC_PREFIX):
            nums[i["title"][len(EPIC_PREFIX):].strip()] = i["number"]
    return nums


def existing_titles() -> set[str]:
    out = subprocess.run(["gh", "issue", "list", "--state", "all", "--limit", "5000", "--json", "title"],
                         capture_output=True, text=True, cwd=ROOT)
    if out.returncode != 0:
        return set()
    return {i["title"] for i in json.loads(out.stdout or "[]")}


def labels_for(row_get) -> list[str]:
    out = []
    if row_get("web").upper() in ("Y", "YES", "✓"):
        out.append("platform:web")
    if row_get("mobile").upper() in ("Y", "YES", "✓"):
        out.append("platform:mobile")
    if row_get("tv").upper() in ("Y", "YES", "✓"):
        out.append("platform:tv")
    apps = row_get("apps").lower()
    if "desktop-kiosk" in apps:
        out.append("surface:kiosk")
    if "desktop-devicebridge" in apps:
        out.append("surface:device-bridge")
    if "desktop-workstation" in apps:
        out.append("surface:workstation")
    if "desktop" in apps:
        out.append("platform:desktop")
    return sorted(set(out))


def body_for(module, sub, role, feature, platforms_label, meta, epic) -> str:
    verb = feature[0].lower() + feature[1:] if feature else feature
    ac = [
        f"- [ ] The {role} can {verb} from the relevant screen (loading / empty / error states).",
        "- [ ] Backend: tenant-scoped + RLS, typed errors, `cargo clippy` clean; migration for new entities.",
        "- [ ] Frontend built from the `@/components/ui` seam, permission-gated, TanStack Query, Zod-validated.",
        "- [ ] Permission-gated (`P.<module>.<action>`) at page + element level.",
    ]
    if MODULE_REG.get(module):
        ac.append(f"- [ ] {MODULE_REG[module]}")
    ac.append("- [ ] Audit-logged; tests (CRUD + integration) + `make check-all` pass.")
    return (
        f"> As a **{role}**, I want **{feature.lower()}**.\n\n"
        f"`{meta}`\n\n"
        "**Before coding (per CLAUDE.md):** research the real-world scenario, then expand "
        "these into concrete Given/When/Then business-logic acceptance criteria.\n\n"
        "**Acceptance criteria (starter — module-tailored)**\n" + "\n".join(ac) + "\n\n"
        f"Module: **{module}** › {sub}" + (f" · Part of #{epic}" if epic else "")
    )


def main(argv) -> int:
    dry = "--dry-run" in argv
    statuses = None
    for a in argv:
        if a.startswith("--status"):
            val = a.split("=", 1)[1] if "=" in a else argv[argv.index(a) + 1]
            statuses = {s.strip().lower() for s in val.split(",")}

    done_keys = set(STATE.read_text().splitlines()) if STATE.exists() else set()
    epics = {} if dry else epic_numbers()
    seen_titles = set() if dry else existing_titles()
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    created = skipped = failed = 0

    for name in wb.sheetnames:
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            continue
        col = {cell(c).lower(): i for i, c in enumerate(header) if c}

        def g(r, key):
            i = col.get(key)
            return cell(r[i]) if i is not None and i < len(r) else ""

        role = MODULE_ROLE.get(name, "user")
        tag = MODULE_TAG.get(name, name[:10])
        epic = epics.get(name)

        for r in it:
            if not r:
                continue
            feature = g(r, "feature")
            if not feature:
                continue
            status = g(r, "status")
            if statuses and status.lower() not in statuses:
                continue
            sub = g(r, "sub-module") or "General"
            key = f"{name}|{sub}|{feature}"
            title = f"[{tag}] {feature}"[:250]
            if key in done_keys or title in seen_titles:
                skipped += 1
                continue
            platforms = ", ".join(
                lbl for k, lbl in (("web", "Web"), ("mobile", "Mobile"), ("tv", "TV"))
                if g(r, k).upper() in ("Y", "YES", "✓"))
            meta = " · ".join(p for p in (
                g(r, "priority"), status,
                f"Platforms: {platforms}" if platforms else "",
                f"Source: {g(r, 'source')}" if g(r, "source") else "",
                f"RFC: {g(r, 'rfc ref')}" if g(r, "rfc ref") else "") if p)
            body = body_for(name, sub, role, feature, platforms, meta, epic)
            labels = labels_for(lambda k: g(r, k))

            if dry:
                created += 1
                continue

            args = ["gh", "issue", "create", "--title", title, "--body", body]
            for lab in labels:
                args += ["--label", lab]
            res = subprocess.run(args, capture_output=True, text=True, cwd=ROOT)
            if res.returncode == 0:
                created += 1
                seen_titles.add(title)
                with STATE.open("a") as fh:
                    fh.write(key + "\n")
                if created % 25 == 0:
                    print(f"  …{created} created", flush=True)
                time.sleep(PACE)
            elif "rate limit" in (res.stderr or "").lower() or "was submitted too quickly" in (res.stderr or "").lower():
                print(f"  rate-limited at {created}; backoff {BACKOFF}s", flush=True)
                time.sleep(BACKOFF)
            else:
                failed += 1
                print(f"  FAIL [{tag}] {feature[:50]}: {res.stderr.strip()[:120]}", flush=True)
                time.sleep(PACE)

    print(f"\n{'(dry) ' if dry else ''}{created} created, {skipped} skipped, {failed} failed", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

from __future__ import annotations

import json
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path("/Users/apple/Projects/MedBrains")
SOURCE_XLSX = ROOT / "MedBrains_Features.xlsx"
OUTPUT_JSON = ROOT / "medbrains" / "outputs" / "2026-04-22-test-tracker" / "test_tracker_source.json"


@dataclass(frozen=True)
class CoverageOverride:
    test_status: str
    existing_coverage: str
    next_action: str


COVERAGE_OVERRIDES: dict[str, CoverageOverride] = {
    "quick patient registration form (name, age, gender, phone, address)": CoverageOverride(
        test_status="Done",
        existing_coverage="Playwright patient full + quick registration scenarios",
        next_action="Add negative-path and validation assertions",
    ),
    "patient search (name, phone, uhid, abha, aadhaar)": CoverageOverride(
        test_status="Done",
        existing_coverage="Playwright patient search scenario",
        next_action="Add fuzzy search and UHID-specific assertions",
    ),
    "department creation (code, name, type: clinical/admin/support)": CoverageOverride(
        test_status="Done",
        existing_coverage="Playwright admin department CRUD scenario",
        next_action="Expand into hierarchy and working-hours coverage",
    ),
    "custom role creation with name and description": CoverageOverride(
        test_status="Done",
        existing_coverage="Playwright admin role create/save permissions/delete scenario",
        next_action="Add role metadata edit coverage",
    ),
    "single user creation form (name, email, role, department)": CoverageOverride(
        test_status="Done",
        existing_coverage="Playwright admin doctor user CRUD with department assignment scenario",
        next_action="Add facility assignment assertions",
    ),
    "user management (create, assign role, assign department)": CoverageOverride(
        test_status="Done",
        existing_coverage="API client CRUD coverage + Playwright user CRUD with role and department assignment",
        next_action="Add permission override drawer assertions",
    ),
    "department-wise indent generation with item catalog": CoverageOverride(
        test_status="Done",
        existing_coverage="Playwright indent creation with catalog-backed requisition coverage",
        next_action="Add stock issuance assertions once seeded stock movements are included",
    ),
    "indent approval workflow (dept → store → purchase committee)": CoverageOverride(
        test_status="Done",
        existing_coverage="Playwright indent submit and approve workflow scenario",
        next_action="Extend with rejection and committee-routing variants",
    ),
    "purchase order generation from approved indents": CoverageOverride(
        test_status="Done",
        existing_coverage="Playwright procurement PO creation from linked approved indent plus sidecar verification",
        next_action="Add GRN and receipt completion checks for the same chain",
    ),
}


CURRENT_COVERAGE = [
    {
        "area": "API Client",
        "module": "Security / CERT-In",
        "test_name": "Security incidents contract alignment",
        "status": "Done",
        "notes": "Added regression coverage for /security-incidents path and PATCH updates.",
    },
    {
        "area": "API Client",
        "module": "IT Security",
        "test_name": "CERT-In incident update flows",
        "status": "Done",
        "notes": "Added tests for reportToCertIn, getIncidentUpdates, and addIncidentUpdate.",
    },
    {
        "area": "API Client",
        "module": "IT Security",
        "test_name": "Vulnerability and compliance endpoints",
        "status": "Done",
        "notes": "Added CRUD-style coverage for vulnerabilities and compliance requirement updates.",
    },
    {
        "area": "API Client",
        "module": "IT Security",
        "test_name": "System health and onboarding wizard endpoints",
        "status": "Done",
        "notes": "Added coverage for system-health, backups, onboarding progress, and complete-step.",
    },
    {
        "area": "Playwright UI",
        "module": "Patient Management",
        "test_name": "Patient registration full form",
        "status": "Done",
        "notes": "Scenario covers register, list visibility, and detail drawer open.",
    },
    {
        "area": "Playwright UI",
        "module": "Patient Management",
        "test_name": "Patient quick registration and search",
        "status": "Done",
        "notes": "Scenario covers quick register and list search.",
    },
    {
        "area": "Playwright UI",
        "module": "Admin Settings",
        "test_name": "Department CRUD",
        "status": "Done",
        "notes": "Scenario covers create, edit, and delete on settings departments tab.",
    },
    {
        "area": "Playwright UI",
        "module": "Admin Users",
        "test_name": "Doctor user CRUD with department assignment",
        "status": "Done",
        "notes": "Scenario covers doctor create, department assignment, edit, and delete via /admin/users.",
    },
    {
        "area": "Playwright UI",
        "module": "Admin Roles",
        "test_name": "Role create, permission save, and delete",
        "status": "Done",
        "notes": "Scenario covers create, permission selection save, reopen verification, and delete.",
    },
    {
        "area": "Playwright UI",
        "module": "Indent + Procurement",
        "test_name": "Indent approval to linked PO sidecar chain",
        "status": "Done",
        "notes": "Scenario covers catalog-backed indent creation, approval, PO creation from linked indent, and flow-tracker sidecar verification.",
    },
]


def normalize_status(value: Any) -> str:
    if value is None:
        return "Pending"
    text = str(value).strip()
    return text if text else "Pending"


def default_test_status(product_status: str) -> str:
    normalized = product_status.lower()
    if normalized in {"done", "partial"}:
        return "Planned"
    if normalized in {"deferred"}:
        return "Blocked"
    return "Blocked"


def derive_test_layer(web: Any, mobile: Any, tv: Any) -> str:
    layers: list[str] = []
    if str(web or "").upper() == "Y":
        layers.extend(["API Client", "Playwright UI"])
    if str(mobile or "").upper() == "Y":
        layers.append("Mobile Later")
    if str(tv or "").upper() == "Y":
        layers.append("TV Later")
    if not layers:
        layers.append("Backlog Review")
    return " + ".join(dict.fromkeys(layers))


def derive_test_scope(product_status: str) -> str:
    if product_status.lower() in {"done", "partial"}:
        return "Basic happy-path, CRUD, and regression coverage"
    return "Feature implementation pending before test authoring"


def derive_existing_coverage(product_status: str) -> str:
    if product_status.lower() in {"done", "partial"}:
        return "No dedicated seeded tests yet"
    return "Blocked until feature build progresses"


def derive_next_action(product_status: str) -> str:
    if product_status.lower() in {"done", "partial"}:
        return "Write API client and Playwright happy-path coverage"
    return "Implement or stabilize feature before test work"


def classify_row(row: tuple[Any, ...]) -> dict[str, Any] | None:
    serial = row[0] if len(row) > 0 else None
    module = row[1] if len(row) > 1 else None
    sub_module = row[2] if len(row) > 2 else None
    feature = row[3] if len(row) > 3 else None

    if serial is not None:
        product_status = normalize_status(row[6] if len(row) > 6 else None)
        feature_text = str(feature).strip() if feature else ""
        test_status = default_test_status(product_status)
        existing_coverage = derive_existing_coverage(product_status)
        next_action = derive_next_action(product_status)

        override = COVERAGE_OVERRIDES.get(feature_text.lower())
        if override is not None:
            test_status = override.test_status
            existing_coverage = override.existing_coverage
            next_action = override.next_action

        return {
            "row_type": "feature",
            "s_no": serial,
            "module": module,
            "sub_module": sub_module,
            "feature": feature,
            "source": row[4] if len(row) > 4 else None,
            "priority": row[5] if len(row) > 5 else None,
            "product_status": product_status,
            "rfc_ref": row[7] if len(row) > 7 else None,
            "web": row[8] if len(row) > 8 else None,
            "mobile": row[9] if len(row) > 9 else None,
            "tv": row[10] if len(row) > 10 else None,
            "test_layer": derive_test_layer(row[8] if len(row) > 8 else None, row[9] if len(row) > 9 else None, row[10] if len(row) > 10 else None),
            "test_scope": derive_test_scope(product_status),
            "test_status": test_status,
            "existing_coverage": existing_coverage,
            "next_action": next_action,
        }

    if module and not sub_module and not feature:
        return {"row_type": "module_header", "label": str(module)}
    if sub_module and not feature:
        return {"row_type": "submodule_header", "label": str(sub_module)}
    return None


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)

    sheets_payload: list[dict[str, Any]] = []
    product_counter: Counter[str] = Counter()
    test_counter: Counter[str] = Counter()
    total_features = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[dict[str, Any]] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            classified = classify_row(row)
            if classified is None:
                continue
            rows.append(classified)

            if classified["row_type"] == "feature":
                total_features += 1
                product_counter[classified["product_status"]] += 1
                test_counter[classified["test_status"]] += 1

        sheets_payload.append({"name": sheet_name, "rows": rows})

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(
        json.dumps(
            {
                "generated_at": datetime.now().isoformat(timespec="seconds"),
                "source_workbook": str(SOURCE_XLSX),
                "total_features": total_features,
                "product_status_counts": dict(product_counter),
                "test_status_counts": dict(test_counter),
                "current_coverage": CURRENT_COVERAGE,
                "sheets": sheets_payload,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print(f"Wrote {OUTPUT_JSON}")


if __name__ == "__main__":
    main()

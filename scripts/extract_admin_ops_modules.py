#!/usr/bin/env python3
"""Extract features for remaining Admin & Operations modules from MedBrains_Features.xlsx"""

import openpyxl
import re
import sys

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

# Target modules (case-insensitive patterns)
TARGET_PATTERNS = [
    r"schedul",
    r"utilization",
    r"case\s*(mgmt|management)",
    r"occupational|occ\.?\s*health",
]

# Related/nearby modules to also check
RELATED_PATTERNS = [
    r"resource\s*manage",
    r"appointment",
    r"roster",
    r"shift",
    r"workforce",
    r"manpower",
    r"bed\s*manage",
    r"capacity",
    r"care\s*coord",
    r"discharge\s*plan",
    r"social\s*work",
    r"employee\s*health",
    r"staff\s*health",
    r"work.*comp",
]

def matches_any(text, patterns):
    if not text:
        return False
    text_lower = text.lower().strip()
    for pat in patterns:
        if re.search(pat, text_lower, re.IGNORECASE):
            return True
    return False

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    print("=" * 120)
    print("SHEETS IN WORKBOOK:")
    print("=" * 120)
    for name in wb.sheetnames:
        print(f"  - {name}")
    print()

    # Find Admin & Operations sheet
    target_sheet = None
    for name in wb.sheetnames:
        if "admin" in name.lower() and "oper" in name.lower():
            target_sheet = name
            break

    if not target_sheet:
        print("Could not find 'Admin & Operations' sheet. Searching all sheets...")
        target_sheet = None

    # First, let's look at the Admin & Operations sheet structure
    if target_sheet:
        print(f"=" * 120)
        print(f"EXAMINING SHEET: '{target_sheet}'")
        print(f"=" * 120)
        ws = wb[target_sheet]

        # Print header row
        print(f"\nHeader row (row 1):")
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            print(f"  Col {col}: {val}")

        # Also check row 2 in case headers are there
        print(f"\nRow 2:")
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=2, column=col).value
            if val:
                print(f"  Col {col}: {val}")
        print()

    # Now scan all sheets for target modules
    all_sheets_to_scan = [target_sheet] if target_sheet else wb.sheetnames

    # Also scan all sheets for completeness
    all_sheets_to_scan = wb.sheetnames

    # Collect all unique module names first
    print("=" * 120)
    print("ALL UNIQUE MODULE NAMES ACROSS ALL SHEETS:")
    print("=" * 120)
    all_modules = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Try to find module column (usually col 2)
        for row in range(1, ws.max_row + 1):
            for col in range(1, min(ws.max_column + 1, 5)):
                val = ws.cell(row=row, column=col).value
                if val and isinstance(val, str) and len(val) > 2 and len(val) < 80:
                    # Check if it looks like a module name (not a header, not a feature description)
                    pass
        # Just get col 2 values as module names
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=2).value
            if val and isinstance(val, str) and len(val.strip()) > 0:
                all_modules.add(val.strip())

    for mod in sorted(all_modules):
        if len(mod) < 60:  # Filter out long descriptions
            print(f"  {mod}")
    print()

    # Now extract features for target modules
    print("=" * 120)
    print("TARGET MODULE FEATURES (Scheduling, Utilization, Case Management, Occupational Health)")
    print("=" * 120)

    target_features = []
    related_features = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Determine column mapping from header
        headers = {}
        for col in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=col).value
            if h:
                headers[col] = str(h).strip().lower()

        # Find key columns
        sno_col = None
        mod_col = None
        submod_col = None
        feat_col = None
        status_col = None
        web_col = None
        mobile_col = None
        tv_col = None
        priority_col = None

        for col, h in headers.items():
            if h in ("s.no", "s.no.", "sno", "#", "no", "no."):
                sno_col = col
            elif h in ("module", "module name"):
                mod_col = col
            elif "sub" in h and "module" in h:
                submod_col = col
            elif h in ("feature", "feature name", "features", "feature description"):
                feat_col = col
            elif h in ("status",):
                status_col = col
            elif h in ("web",):
                web_col = col
            elif h in ("mobile",):
                mobile_col = col
            elif h in ("tv",):
                tv_col = col
            elif h in ("priority",):
                priority_col = col

        # Scan rows
        current_module = ""
        current_submodule = ""

        for row in range(2, ws.max_row + 1):
            # Read module (may be merged/empty - carry forward)
            mod_val = ws.cell(row=row, column=mod_col).value if mod_col else None
            if mod_val and isinstance(mod_val, str) and mod_val.strip():
                current_module = mod_val.strip()

            submod_val = ws.cell(row=row, column=submod_col).value if submod_col else None
            if submod_val and isinstance(submod_val, str) and submod_val.strip():
                current_submodule = submod_val.strip()

            feat_val = ws.cell(row=row, column=feat_col).value if feat_col else None
            if not feat_val or not isinstance(feat_val, str) or not feat_val.strip():
                continue

            sno_val = ws.cell(row=row, column=sno_col).value if sno_col else row
            status_val = ws.cell(row=row, column=status_col).value if status_col else ""
            web_val = ws.cell(row=row, column=web_col).value if web_col else ""
            mobile_val = ws.cell(row=row, column=mobile_col).value if mobile_col else ""
            tv_val = ws.cell(row=row, column=tv_col).value if tv_col else ""
            priority_val = ws.cell(row=row, column=priority_col).value if priority_col else ""

            feature_info = {
                "sheet": sheet_name,
                "row": row,
                "sno": sno_val,
                "module": current_module,
                "submodule": current_submodule,
                "feature": feat_val.strip(),
                "status": str(status_val).strip() if status_val else "",
                "priority": str(priority_val).strip() if priority_val else "",
                "web": str(web_val).strip() if web_val else "",
                "mobile": str(mobile_val).strip() if mobile_val else "",
                "tv": str(tv_val).strip() if tv_val else "",
            }

            # Check if module or submodule matches target patterns
            combined = f"{current_module} {current_submodule} {feat_val}"

            if matches_any(current_module, TARGET_PATTERNS) or matches_any(current_submodule, TARGET_PATTERNS):
                target_features.append(feature_info)
            elif matches_any(current_module, RELATED_PATTERNS) or matches_any(current_submodule, RELATED_PATTERNS):
                related_features.append(feature_info)
            elif matches_any(feat_val, TARGET_PATTERNS):
                # Feature text itself mentions target module
                related_features.append(feature_info)

    # Print target features grouped by module
    if target_features:
        current_mod = ""
        current_sub = ""
        for f in target_features:
            if f["module"] != current_mod:
                current_mod = f["module"]
                print(f"\n{'=' * 100}")
                print(f"MODULE: {current_mod}")
                print(f"{'=' * 100}")
                current_sub = ""
            if f["submodule"] != current_sub:
                current_sub = f["submodule"]
                print(f"\n  SUB-MODULE: {current_sub}")
                print(f"  {'-' * 90}")

            status_str = f["status"] if f["status"] else "N/A"
            platform = f"Web={f['web']} Mobile={f['mobile']} TV={f['tv']}"
            print(f"    Row {f['row']:4d} | S.No {f['sno']!s:>6} | Status: {status_str:<12} | {platform}")
            print(f"           | Feature: {f['feature']}")
    else:
        print("\n  *** NO FEATURES FOUND matching target patterns ***")
        print("  This may mean the module names are different. Check the module list above.\n")

    # Print related features
    if related_features:
        print(f"\n{'=' * 120}")
        print("RELATED/NEARBY MODULE FEATURES:")
        print(f"{'=' * 120}")
        current_mod = ""
        current_sub = ""
        for f in related_features:
            if f["module"] != current_mod:
                current_mod = f["module"]
                print(f"\n  MODULE: {current_mod}")
                current_sub = ""
            if f["submodule"] != current_sub:
                current_sub = f["submodule"]
                print(f"    SUB-MODULE: {current_sub}")

            status_str = f["status"] if f["status"] else "N/A"
            platform = f"Web={f['web']} Mobile={f['mobile']} TV={f['tv']}"
            print(f"      Row {f['row']:4d} | S.No {f['sno']!s:>6} | Status: {status_str:<12} | {platform}")
            print(f"             | Feature: {f['feature']}")

    # Summary
    print(f"\n{'=' * 120}")
    print("SUMMARY")
    print(f"{'=' * 120}")
    print(f"Target features found: {len(target_features)}")
    print(f"Related features found: {len(related_features)}")

    if target_features:
        # Count by module and status
        from collections import Counter
        mod_counts = Counter()
        status_counts = Counter()
        for f in target_features:
            mod_counts[f["module"]] += 1
            status_counts[(f["module"], f["status"])] += 1

        print("\nBy Module:")
        for mod, count in mod_counts.most_common():
            print(f"  {mod}: {count} features")
            for (m, s), c in sorted(status_counts.items()):
                if m == mod:
                    print(f"    {s or 'N/A'}: {c}")

    # Also dump the Admin & Operations sheet completely to see all modules
    if target_sheet:
        print(f"\n{'=' * 120}")
        print(f"COMPLETE MODULE LIST FROM '{target_sheet}' SHEET:")
        print(f"{'=' * 120}")
        ws = wb[target_sheet]
        seen_modules = []
        for row in range(2, ws.max_row + 1):
            mod_val = ws.cell(row=row, column=mod_col).value if mod_col else None
            if mod_val and isinstance(mod_val, str) and mod_val.strip():
                mod_name = mod_val.strip()
                if mod_name not in seen_modules:
                    seen_modules.append(mod_name)

        for i, mod in enumerate(seen_modules, 1):
            # Check if it matches target
            is_target = matches_any(mod, TARGET_PATTERNS)
            is_related = matches_any(mod, RELATED_PATTERNS)
            marker = " *** TARGET ***" if is_target else (" [related]" if is_related else "")
            print(f"  {i:3d}. {mod}{marker}")

if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Generate the MedBrains hospital-implementation workbook.

A single Excel workbook with an Index/legend sheet, a README, and one sheet per
master-data set a hospital fills in during onboarding (departments, lab tests,
drug formulary, procedures, charges, beds, doctors, staff, insurers, vendors,
locations). Each sheet has the exact columns the importers/seeds expect, with
required columns marked `*`, per-column notes as cell comments, and sample rows.

Run:  uv run python scripts/generate_implementation_workbooks.py
Out:  implementation-templates/MedBrains-Implementation-Workbook.xlsx
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
REQ_FILL = PatternFill("solid", fgColor="E9EFF7")
TITLE_FONT = Font(bold=True, size=13)
HEAD_FONT = Font(bold=True)
NOTE_FONT = Font(italic=True, color="6F6F6F", size=10)
THIN = Side(style="thin", color="C6C6C6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Each template: (sheet, purpose, import_method, columns, samples)
# column = (name, required, note)
TEMPLATES = [
    (
        "Facility Profile", "Hospital / institution master details", "Admin → Settings → General + Branding",
        [("facility_name", True, "Registered hospital name"),
         ("legal_entity", False, "Owning trust / company name"),
         ("registration_no", False, "Clinical Establishments Act registration no."),
         ("accreditation", False, "NABH / JCI / NABL status"),
         ("gstin", False, "GST identification number"),
         ("pan", False, "PAN of the entity"),
         ("address", True, "Full address"),
         ("city", True, "City"),
         ("state", True, "State (from Geo)"),
         ("pincode", True, "6-digit PIN code"),
         ("phone", True, "Main contact number"),
         ("email", False, "Official email"),
         ("website", False, "Website"),
         ("established_year", False, "Year established"),
         ("bed_capacity", False, "Sanctioned bed count")],
        [["City Care Hospital", "City Care Trust", "CEA/2021/0042", "NABH", "29ABCDE1234F1Z5",
          "ABCDE1234F", "12 MG Road", "Bengaluru", "Karnataka", "560001", "080-12345678",
          "info@citycare.in", "citycare.in", "2010", "250"]],
    ),
    (
        "Licenses & Compliance", "Statutory licenses & certificates", "Admin → Settings → Regulatory",
        [("license_type", True, "Drug License / NABH / NABL / BMW Auth / Fire NOC / "
          "PNDT / Blood Bank / Clinical Establishment …"),
         ("license_number", True, "License / certificate number"),
         ("issuing_authority", False, "Issuing body"),
         ("issue_date", False, "DD/MM/YYYY"),
         ("expiry_date", False, "DD/MM/YYYY (drives renewal alerts)"),
         ("scope", False, "Departments / activities covered")],
        [["Drug License (Form 20/21)", "KA-DL-2021-1234", "Drugs Control Dept",
          "01/04/2021", "31/03/2026", "Pharmacy"],
         ["BMW Authorization", "KSPCB-BMW-5678", "State Pollution Control Board",
          "01/01/2023", "31/12/2025", "Whole facility"]],
    ),
    (
        "Departments", "Clinical & support departments", "Seeded defaults + Admin → Departments",
        [("code", True, "Short unique code, e.g. CARD"),
         ("name", True, "Cardiology"),
         ("department_type", True, "One of: clinical / diagnostic / support / administrative"),
         ("parent_code", False, "Parent department code, if sub-department")],
        [["CARD", "Cardiology", "clinical", ""],
         ["LAB", "Laboratory", "diagnostic", ""]],
    ),
    (
        "Lab Tests", "Laboratory test catalogue", "POST /api/setup/import/lab-catalog (CSV)",
        [("code", True, "Unique test code, e.g. CBC"),
         ("name", True, "Complete Blood Count"),
         ("sample_type", False, "Blood / Serum / Urine …"),
         ("normal_range", False, "Reference range text"),
         ("unit", False, "g/dL, mmol/L …"),
         ("price", False, "Price in INR (number)"),
         ("tat_hours", False, "Turnaround time in hours (integer)")],
        [["CBC", "Complete Blood Count", "Blood", "", "", "250", "4"],
         ["LFT", "Liver Function Test", "Serum", "", "", "600", "6"]],
    ),
    (
        "Drug Formulary", "Pharmacy drug catalogue", "POST /api/setup/import/pharmacy-catalog (CSV)",
        [("code", True, "Unique drug code"),
         ("name", True, "Brand / dispensing name"),
         ("generic_name", False, "Generic name"),
         ("category", False, "Therapeutic category"),
         ("manufacturer", False, "Manufacturer"),
         ("unit", False, "tablet / strip / vial / ml"),
         ("base_price", False, "Purchase price INR"),
         ("tax_percent", False, "GST % (number)"),
         ("reorder_level", False, "Reorder qty (integer)"),
         ("drug_schedule", False, "Schedule H / H1 / X / G / OTC"),
         ("inn_name", False, "WHO INN name"),
         ("atc_code", False, "WHO ATC code"),
         ("mrp", False, "Max retail price INR")],
        [["PARA500", "Paracetamol 500mg", "Paracetamol", "Analgesic", "GSK", "tablet",
          "0.8", "12", "100", "OTC", "paracetamol", "N02BE01", "1.5"]],
    ),
    (
        "Procedures", "Procedure / day-care catalogue", "Seeded defaults + Admin → Procedures",
        [("code", True, "Unique procedure code"),
         ("name", True, "Procedure name"),
         ("category", False, "dressing / minor_surgery / injection / ent …"),
         ("base_price", False, "Price INR"),
         ("duration_minutes", False, "Typical duration (integer)"),
         ("requires_consent", False, "TRUE / FALSE"),
         ("requires_anaesthesia", False, "TRUE / FALSE")],
        [["SUTURE", "Wound suturing", "minor_surgery", "600", "30", "TRUE", "TRUE"]],
    ),
    (
        "Charge Master", "Billable services / charges", "Seeded defaults + Admin → Charge Master",
        [("code", True, "Unique charge code"),
         ("name", True, "Charge name"),
         ("category", False, "registration / consultation / nursing / room …"),
         ("base_price", False, "Price INR"),
         ("tax_percent", False, "GST % (number)")],
        [["REG_NEW", "New patient registration", "registration", "100", "0"],
         ["OPD_GEN", "General OPD consultation", "consultation", "300", "0"]],
    ),
    (
        "Bed Types", "Ward / bed categories", "Seeded defaults + Admin → Bed Types",
        [("code", True, "Unique bed-type code"),
         ("name", True, "General / Semi-private / Private / ICU …"),
         ("daily_rate", False, "Per-day charge INR"),
         ("description", False, "Notes")],
        [["GEN", "General Ward", "1000", "Shared ward"],
         ["ICU", "Intensive Care Unit", "8000", "Critical care"]],
    ),
    (
        "Doctors", "Consultants / doctors", "Admin → Doctors",
        [("full_name", True, "Dr. full name"),
         ("qualification", False, "MBBS, MD …"),
         ("registration_no", False, "Medical council reg. no."),
         ("department_code", False, "Department code from Departments sheet"),
         ("specialization", False, "Specialty"),
         ("phone", False, "Contact number"),
         ("email", False, "Email")],
        [["Dr. A. Sharma", "MBBS, MD", "KMC-12345", "CARD", "Cardiology", "", ""]],
    ),
    (
        "Staff / Users", "System users & roles", "Admin → Users",
        [("full_name", True, "Staff full name"),
         ("email", True, "Login email (unique)"),
         ("role", True, "Built-in role code, e.g. nurse / pharmacist / cashier"),
         ("department_code", False, "Department code"),
         ("designation", False, "Job title"),
         ("phone", False, "Contact number")],
        [["Nurse R. Iyer", "r.iyer@hospital.in", "nurse", "CARD", "Staff Nurse", ""]],
    ),
    (
        "Insurance / TPA", "Insurers, TPAs, corporates", "Admin → Insurance / Corporate",
        [("name", True, "Insurer / TPA / corporate name"),
         ("type", True, "insurer / tpa / corporate"),
         ("contact_person", False, "Contact name"),
         ("phone", False, "Phone"),
         ("email", False, "Email"),
         ("address", False, "Address")],
        [["Star Health", "insurer", "", "", "", ""],
         ["MediAssist", "tpa", "", "", "", ""]],
    ),
    (
        "Vendors / Suppliers", "Procurement vendors", "Admin → Vendors",
        [("name", True, "Vendor name"),
         ("gstin", False, "GST identification number"),
         ("contact_person", False, "Contact name"),
         ("phone", False, "Phone"),
         ("email", False, "Email"),
         ("category", False, "pharma / consumables / equipment …"),
         ("address", False, "Address")],
        [["MedSupply Co", "29ABCDE1234F1Z5", "", "", "", "pharma", ""]],
    ),
    (
        "Locations", "Campus / building / floor / room", "Admin → Locations",
        [("campus", True, "Campus name"),
         ("building", False, "Building name"),
         ("floor", False, "Floor label"),
         ("room", False, "Room / ward name"),
         ("code", False, "Unique location code")],
        [["Main Campus", "Block A", "Ground", "OPD-1", "MC-A-G-OPD1"]],
    ),
    (
        "Beds", "Physical beds", "Admin → Beds / IPD setup",
        [("bed_number", True, "Bed number / label"),
         ("ward", True, "Ward name"),
         ("bed_type_code", True, "Bed-type code from Bed Types sheet"),
         ("status", False, "available / occupied / maintenance")],
        [["A-101", "General Ward A", "GEN", "available"]],
    ),
]


def style_master_sheet(ws, columns, samples):
    for col_idx, (name, required, note) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=f"{name} *" if required else name)
        cell.font = HEAD_FONT
        cell.fill = REQ_FILL if required else HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if note:
            cell.comment = Comment(("Required. " if required else "") + note, "MedBrains")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(34, len(name) + 10))
    for r, row in enumerate(samples, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = NOTE_FONT
            cell.border = BORDER
    ws.freeze_panes = "A2"


def build_index(ws):
    ws["A1"] = "MedBrains — Hospital Implementation Workbook"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Fill one sheet per master-data set. Columns marked * are required. Hover a header for notes."
    ws["A2"].font = NOTE_FONT
    head = ["Sheet", "What it captures", "How it loads into MedBrains"]
    for c, h in enumerate(head, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HEAD_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    for i, (sheet, purpose, how, *_rest) in enumerate(TEMPLATES, start=5):
        for c, val in enumerate((sheet, purpose, how), start=1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for c, w in ((1, 22), (2, 38), (3, 46)):
        ws.column_dimensions[get_column_letter(c)].width = w


def build_readme(ws):
    lines = [
        ("MedBrains Implementation Workbook — instructions", TITLE_FONT),
        ("", None),
        ("1. Use the Index sheet as the legend of every master-data set.", None),
        ("2. Fill each sheet — one row per item. Keep the header row unchanged.", None),
        ("3. Columns marked with * are mandatory. Hover the header cell for the format.", None),
        ("4. Codes must be unique within a sheet and are used to link across sheets", None),
        ("   (e.g. department_code in Doctors must exist in the Departments sheet).", None),
        ("5. Delete the grey sample rows before handing the file back.", None),
        ("6. Lab Tests + Drug Formulary import directly via CSV; the rest are loaded", None),
        ("   through the Admin UI (or come pre-seeded with sensible defaults).", None),
        ("7. Dates DD/MM/YYYY, amounts in INR, booleans as TRUE/FALSE.", None),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
    ws.column_dimensions["A"].width = 90


def main() -> None:
    wb = Workbook()
    build_index(wb.active)
    wb.active.title = "Index"
    build_readme(wb.create_sheet("README"))
    for sheet, _purpose, _how, columns, samples in TEMPLATES:
        # Sheet names can't contain / \ ? * [ ] : and cap at 31 chars.
        safe = sheet.replace(" / ", " & ").replace("/", "&")[:31]
        ws = wb.create_sheet(safe)
        style_master_sheet(ws, columns, samples)

    out_dir = "implementation-templates"
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "MedBrains-Implementation-Workbook.xlsx")
    wb.save(path)
    print(f"wrote {path} — {len(TEMPLATES)} master-data sheets + Index + README")

    write_legend_markdown()


def write_legend_markdown() -> None:
    """Emit the human-readable legend (kept in sync with the workbook)."""
    doc_dir = "medbrains/docs/implementation"
    os.makedirs(doc_dir, exist_ok=True)
    out = [
        "# MedBrains — Implementation Master-Data Legend",
        "",
        "Everything we collect **from the institution side** during setup. The matching",
        "Excel workbook is generated by `scripts/generate_implementation_workbooks.py`",
        "→ `implementation-templates/MedBrains-Implementation-Workbook.xlsx` (one sheet",
        "per set below). Columns marked **\\*** are required.",
        "",
        f"_{len(TEMPLATES)} master-data sets._",
        "",
    ]
    for sheet, purpose, how, columns, _samples in TEMPLATES:
        out.append(f"## {sheet}")
        out.append(f"_{purpose}. Loads via: {how}._")
        out.append("")
        out.append("| Column | Required | Notes |")
        out.append("|---|---|---|")
        for name, required, note in columns:
            out.append(f"| `{name}` | {'**Yes**' if required else 'no'} | {note} |")
        out.append("")
    path = os.path.join(doc_dir, "master-data-templates.md")
    with open(path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(out))
    print(f"wrote {path}")


if __name__ == "__main__":
    main()

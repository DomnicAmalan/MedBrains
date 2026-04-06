#!/usr/bin/env python3
"""Generate Patient Registration regulatory cross-check Excel workbook.

Each field includes specific clause references from regulatory standards
so the UI can show "Required per NABH AAC.2.3" hints.

Output: data/Patient_Registration_Fields.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from pathlib import Path

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="087F5B")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_FONT = Font(bold=True, size=11)
SUBSECTION_FILL = PatternFill("solid", fgColor="E9EFF7")
SUBSECTION_FONT = Font(bold=True, size=10)
MANDATORY_FILL = PatternFill("solid", fgColor="FFF3E0")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def style_row(ws, row, cols, fill=None, font=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if fill:
            cell.fill = fill
        if font:
            cell.font = font


def add_section(ws, row, text, cols):
    ws.cell(row, 1, text)
    style_row(ws, row, cols, SECTION_FILL, SECTION_FONT)
    return row + 1


def add_subsection(ws, row, text, cols):
    ws.cell(row, 1, text)
    style_row(ws, row, cols, SUBSECTION_FILL, SUBSECTION_FONT)
    return row + 1


# ── Sheet 1: Registration Fields ─────────────────────────────────────────────

FIELDS_HEADERS = [
    "S.No",
    "Section",
    "Field Name",
    "DB Column",
    "Data Type",
    "Required",  # Mandatory / Optional / Conditional
    "FHIR R4 Path",
    "NABH Clause",
    "JCI Clause",
    "HIPAA Ref",
    "ABDM/ABHA Ref",
    "NHS/UK Ref",
    "CBAHI/SA Ref",
    "UAE/NABIDH Ref",
    "EU/EHDS Ref",
    "Validation Rule",
    "UI Hint Text",
]

# Each tuple: (section, field_name, db_column, data_type, required,
#              fhir_path, nabh, jci, hipaa, abdm, nhs, cbahi, uae, eu,
#              validation, ui_hint)
FIELDS_DATA = [
    # ── 1. CORE IDENTITY ──
    ("Core Identity", "Prefix / Title", "prefix", "TEXT", "Optional",
     "Patient.name.prefix", "", "", "", "", "", "", "", "",
     "Values: Mr, Mrs, Ms, Dr, Prof, Master, Baby, Mx",
     "Honorific prefix for the patient's name"),

    ("Core Identity", "First Name", "first_name", "TEXT", "Mandatory",
     "Patient.name.given[0]", "AAC.2 — Patient must be identifiable by name",
     "IPSG.1 ME1 — Use at least 2 identifiers (name is primary)",
     "45 CFR 164.514(b) — Name is PHI identifier #1",
     "ABHA Profile: firstName (mandatory)",
     "PDS: Given Name (mandatory for NHS Number allocation)",
     "ESR — Correct Patient Identification requires full name",
     "NABIDH MDS: Patient Name (mandatory)",
     "IPS Patient: name.given (must-support)",
     "Min 1 char, max 100, Unicode, no digits",
     "Required by NABH AAC.2, JCI IPSG.1, ABDM, NHS PDS. Primary patient identifier."),

    ("Core Identity", "Middle Name", "middle_name", "TEXT", "Optional",
     "Patient.name.given[1]", "", "",
     "USCDI v2 — Middle Name added as data element", "ABHA Profile: middleName",
     "", "", "", "IPS Patient: name.given (must-support)",
     "Max 100 chars",
     "Recommended by USCDI v2, ABDM. Improves MPI matching accuracy."),

    ("Core Identity", "Last Name / Surname", "last_name", "TEXT", "Mandatory",
     "Patient.name.family", "AAC.2 — Patient must be identifiable by name",
     "IPSG.1 ME1 — Name as patient identifier",
     "45 CFR 164.514(b) — Name is PHI identifier #1",
     "ABHA Profile: lastName",
     "PDS: Family Name (mandatory for NHS Number allocation)",
     "ESR — Correct Patient Identification",
     "NABIDH MDS: Patient Name (mandatory)",
     "IPS Patient: name.family (must-support, min 1 if no text)",
     "Min 1 char, max 100, Unicode",
     "Required by NABH AAC.2, JCI IPSG.1, ABDM, NHS PDS, IPS."),

    ("Core Identity", "Suffix", "suffix", "TEXT", "Optional",
     "Patient.name.suffix", "", "", "", "", "", "", "", "",
     "Values: Jr, Sr, II, III, PhD",
     "Name suffix for formal records."),

    ("Core Identity", "Full Name (Local Script)", "full_name_local", "TEXT", "Conditional",
     "", "AAC.2 — Regional language documentation",
     "", "", "ABHA: name field may be in Devanagari/regional",
     "", "", "", "",
     "UTF-8, max 200 chars",
     "Required for regional language hospitals per NABH. Stores name in local script (Hindi, Tamil, etc.)"),

    ("Core Identity", "Father's Name", "father_name", "TEXT", "Conditional",
     "", "NABH — Required for minors, insurance claims in India",
     "", "",
     "ABDM/Aadhaar KYC: father's name captured in e-KYC",
     "", "", "", "",
     "Max 100 chars. Required if patient is minor or for insurance.",
     "Required for minors per NABH. Used in Aadhaar e-KYC verification. Common in Indian insurance claims."),

    ("Core Identity", "Mother's Name", "mother_name", "TEXT", "Optional",
     "", "", "", "", "ABDM may include from Aadhaar KYC",
     "", "", "", "",
     "Max 100 chars",
     "Used for birth certificate linkage and some insurance forms."),

    ("Core Identity", "Spouse Name", "spouse_name", "TEXT", "Conditional",
     "", "", "", "", "",
     "", "", "", "",
     "Required when marital_status = married/separated/widowed",
     "Required for married patients in insurance claims and medico-legal cases."),

    ("Core Identity", "Guardian Name", "guardian_name", "TEXT", "Conditional",
     "", "NABH COP.5 — Pediatric care requires guardian details",
     "JCI PFR.6 — Guardian consent for minors",
     "", "", "", "", "", "",
     "Required when patient age < 18 or legally incapacitated",
     "Required per NABH COP.5 for minors. Per JCI PFR.6, guardian consent is mandatory for pediatric patients."),

    ("Core Identity", "Guardian Relation", "guardian_relation", "TEXT", "Conditional",
     "", "NABH COP.5", "JCI PFR.6", "", "", "", "", "", "",
     "Values: parent, grandparent, legal_guardian, spouse, sibling, other",
     "Relationship of guardian to patient. Required with guardian name."),

    # ── 2. DEMOGRAPHICS ──
    ("Demographics", "Date of Birth", "date_of_birth", "DATE", "Mandatory",
     "Patient.birthDate",
     "AAC.2 — DOB is one of the 2 required identifiers",
     "IPSG.1 ME1 — DOB as second patient identifier",
     "45 CFR 164.514(b) — Date elements are PHI identifier #3",
     "ABHA Profile: yearOfBirth/monthOfBirth/dayOfBirth (year mandatory)",
     "PDS: Date of Birth (mandatory for NHS Number allocation)",
     "ESR — Patient Identification requires DOB",
     "NABIDH MDS: Date of Birth (mandatory)",
     "IPS Patient: birthDate (1..1, mandatory)",
     "ISO 8601 (YYYY-MM-DD). Partial DOB allowed: YYYY or YYYY-MM per FHIR/ABDM",
     "Required by NABH AAC.2 (2nd identifier), JCI IPSG.1, NHS PDS, IPS (mandatory). ABDM allows partial (year only)."),

    ("Demographics", "DOB Estimated Flag", "is_dob_estimated", "BOOLEAN", "Optional",
     "", "NABH — For illiterate patients who don't know exact DOB",
     "", "", "", "", "", "", "",
     "Default: false",
     "Set to true when DOB is estimated from stated age. Per NABH for patients who cannot provide exact DOB."),

    ("Demographics", "Biological Sex", "biological_sex", "ENUM", "Mandatory",
     "Patient.gender (administrative)",
     "AAC.2 — Gender is part of patient demographics",
     "IPSG.1 — Gender used for clinical identification",
     "X12 837 DMG segment — Gender Code (M/F/U)",
     "ABHA Profile: gender (M/F/O, mandatory)",
     "PDS: Gender (mandatory)",
     "ESR — Correct Patient Identification",
     "NABIDH MDS: Gender (mandatory)",
     "IPS Patient: gender (must-support)",
     "Values: male, female, other, unknown. Maps to FHIR administrative-gender.",
     "Required by all standards. NABH AAC.2, JCI IPSG.1, ABDM, NHS, CBAHI, UAE NABIDH."),

    ("Demographics", "Gender Identity", "gender_identity", "TEXT", "Optional",
     "Patient.extension[gender-identity]",
     "", "", "USCDI v3 — Gender Identity added",
     "", "", "", "", "",
     "Values: man, woman, non_binary, transgender_male, transgender_female, other, prefer_not_to_say",
     "USCDI v3 requirement. Separate from biological sex for clinical accuracy per WPATH standards."),

    ("Demographics", "Marital Status", "marital_status", "ENUM", "Optional",
     "Patient.maritalStatus",
     "NABH — Part of social history", "",
     "", "", "", "", "", "",
     "Values: single, married, divorced, widowed, separated, domestic_partner, unknown",
     "Per HL7 v3 MaritalStatus value set. Required for some insurance claims."),

    ("Demographics", "Religion", "religion", "TEXT", "Optional",
     "Patient.extension[religion]",
     "NABH COP.6 — Cultural/spiritual needs assessment",
     "JCI PFR.1.1 — Cultural and spiritual values",
     "", "", "", "", "", "",
     "FK to master_religions table",
     "Per NABH COP.6 for spiritual care. Per JCI PFR.1.1 for respecting cultural values."),

    ("Demographics", "Nationality", "nationality_id", "UUID FK", "Mandatory",
     "Patient.extension[nationality]",
     "NABH — Required for international patients",
     "", "", "", "", "CBAHI — Nationality mandatory (citizens vs residents vs visitors)",
     "NABIDH MDS: Nationality (mandatory)",
     "",
     "FK to geo_countries. Default: tenant's country.",
     "Required by CBAHI (Saudi) and UAE NABIDH. Determines ID document requirements."),

    ("Demographics", "Preferred Language", "preferred_language", "TEXT", "Optional",
     "Patient.communication.language",
     "NABH PRE.2 — Communication in language patient understands",
     "JCI PFR.2 — Identify and address communication barriers",
     "", "",
     "PDS: may link to interpreter services", "", "",
     "IPS Patient: communication.language (must-support)",
     "ISO 639-1 code (en, hi, ta, te, etc.)",
     "Per NABH PRE.2, JCI PFR.2. Used for consent forms, discharge summaries, interpreter needs."),

    ("Demographics", "Birth Place", "birth_place", "TEXT", "Optional",
     "", "", "", "", "ABDM may capture from Aadhaar",
     "", "", "", "",
     "Free text, max 100 chars",
     "Used for epidemiological reporting and passport linkage."),

    ("Demographics", "Blood Group", "blood_group", "ENUM", "Optional",
     "Observation (LOINC 882-1)",
     "NABH COP.6 — Part of clinical assessment",
     "JCI IPSG.5 — Transfusion safety",
     "", "", "", "", "", "",
     "Values: a_positive ... o_negative, unknown. 9 values.",
     "Per NABH COP.6. Per JCI IPSG.5 for safe blood transfusion. Must be lab-verified before clinical use."),

    ("Demographics", "Blood Group Verified", "blood_group_verified", "BOOLEAN", "Optional",
     "", "",
     "JCI IPSG.5 — Self-reported blood group must be re-verified before transfusion",
     "", "", "", "", "", "",
     "Default: false. True only after lab typing.",
     "Per JCI IPSG.5: self-reported blood group must be re-verified before transfusion."),

    ("Demographics", "No Known Allergies (NKDA)", "no_known_allergies", "BOOLEAN", "Mandatory",
     "AllergyIntolerance with code=no-known-allergy",
     "NABH COP.6 — Allergy status must be documented",
     "JCI IPSG.5 — Drug allergy assessment mandatory",
     "", "", "", "", "", "",
     "Must be explicitly set (true or false). Cannot be NULL after registration.",
     "Per NABH COP.6, JCI IPSG.5. Allergy status must be actively confirmed, not left blank. Critical for patient safety."),

    ("Demographics", "Deceased Flag", "is_deceased", "BOOLEAN", "Optional",
     "Patient.deceased[x]", "", "", "", "", "", "", "", "",
     "Default: false. Triggers record locking.",
     "FHIR modifier element. When true, restricts further clinical encounters."),

    ("Demographics", "Deceased Date", "deceased_date", "TIMESTAMPTZ", "Conditional",
     "Patient.deceasedDateTime", "", "",
     "USCDI v2 — Date of Death", "", "", "", "", "",
     "Required if is_deceased = true.",
     "USCDI v2 data element. Required for mortality reporting (ICD-11 Chapter 22+)."),

    # ── 3. CONTACT ──
    ("Contact", "Phone Primary (Mobile)", "phone_primary", "TEXT", "Mandatory",
     "Patient.telecom (system=phone, use=mobile)",
     "AAC.2 — Contact information mandatory at registration",
     "ACC.2 — Registration must capture contact details",
     "45 CFR 164.514(b) — Phone numbers are PHI identifier #4",
     "ABHA Profile: mobile (mandatory)",
     "",
     "CBAHI — Contact details mandatory",
     "NABIDH MDS: Mobile Number (mandatory)",
     "",
     "E.164 format: +{country_code}{number}. India: +91XXXXXXXXXX",
     "Required by NABH AAC.2, JCI ACC.2, ABDM, CBAHI, UAE NABIDH."),

    ("Contact", "Phone Secondary", "phone_secondary", "TEXT", "Optional",
     "Patient.telecom (system=phone, rank=2)",
     "NABH — Alternate contact recommended", "", "", "", "", "", "", "",
     "E.164 format",
     "Alternate contact number. Recommended by NABH for reachability."),

    ("Contact", "Email", "email", "TEXT", "Optional",
     "Patient.telecom (system=email)",
     "", "",
     "45 CFR 164.514(b) — Email is PHI identifier #6",
     "ABHA Profile: email (optional)",
     "", "", "", "USCDI v2 — Email Address",
     "RFC 5322, max 254 chars, lowercase normalized",
     "PHI per HIPAA. USCDI v2 data element."),

    ("Contact", "Preferred Contact Method", "preferred_contact_method", "TEXT", "Optional",
     "",
     "", "JCI PFR.2 — Address communication barriers",
     "", "", "", "", "", "",
     "Values: phone, sms, email, whatsapp, postal",
     "Per JCI PFR.2 for effective communication."),

    # ── 4. EMERGENCY CONTACT ──
    ("Emergency Contact", "Contact Name", "patient_contacts.contact_name", "TEXT", "Mandatory",
     "Patient.contact.name",
     "AAC.2 — Emergency contact mandatory at registration",
     "COP.2 — Emergency contact for care continuity",
     "", "",
     "",
     "ESR — Emergency contact required",
     "", "",
     "Min 1 char, max 100. At least one emergency contact required.",
     "Required by NABH AAC.2, JCI COP.2, CBAHI ESR. At least one emergency contact must be recorded."),

    ("Emergency Contact", "Contact Relation", "patient_contacts.relation", "TEXT", "Mandatory",
     "Patient.contact.relationship",
     "AAC.2", "COP.2", "",
     "", "", "ESR", "", "",
     "FK to master_relations: spouse, parent, child, sibling, friend, legal_guardian, other",
     "Required with emergency contact name."),

    ("Emergency Contact", "Contact Phone", "patient_contacts.phone", "TEXT", "Mandatory",
     "Patient.contact.telecom",
     "AAC.2", "COP.2", "", "", "", "ESR", "", "",
     "E.164 format",
     "Required with emergency contact name."),

    # ── 5. ADDRESS ──
    ("Address", "Address Line 1", "patient_addresses.address_line1", "TEXT", "Mandatory",
     "Patient.address.line",
     "AAC.2 — Address mandatory at registration",
     "ACC.2 — Registration captures address",
     "45 CFR 164.514(b) — Geographic data is PHI identifier #2",
     "ABHA Profile: address (conditional)",
     "PDS: Address + Postcode (mandatory)",
     "ESR — Address required for identification",
     "NABIDH MDS: Address (mandatory)",
     "IPS Patient: address (must-support)",
     "Max 200 chars. House/flat number, street.",
     "Required by NABH AAC.2, JCI ACC.2, NHS PDS, CBAHI, UAE NABIDH."),

    ("Address", "City", "patient_addresses.city", "TEXT", "Mandatory",
     "Patient.address.city",
     "AAC.2", "ACC.2", "PHI",
     "ABHA: derived from district/state",
     "PDS: Address (mandatory)", "", "NABIDH MDS: City",
     "IPS: address.city",
     "Max 100 chars",
     "Required for all major standards."),

    ("Address", "State", "patient_addresses.state_id", "UUID FK", "Mandatory",
     "Patient.address.state",
     "AAC.2", "", "", "ABHA Profile: stateCode",
     "PDS: Address", "", "NABIDH MDS: Emirate", "",
     "FK to geo_states",
     "Required. ABDM uses numeric state codes (mapped to geo_states)."),

    ("Address", "Country", "patient_addresses.country_id", "UUID FK", "Mandatory",
     "Patient.address.country",
     "", "", "", "",
     "", "", "", "IPS: address.country (ISO 3166 Alpha-2/3)",
     "FK to geo_countries. Default: tenant country.",
     "Per IPS specification: must be ISO 3166."),

    ("Address", "Postal Code / PIN", "patient_addresses.postal_code", "TEXT", "Mandatory",
     "Patient.address.postalCode",
     "AAC.2 — Part of address", "",
     "45 CFR 164.514(b) — ZIP codes restricted (populations <20k)",
     "ABHA Profile: pincode",
     "PDS: Postcode (mandatory)",
     "", "NABIDH MDS: Makani Number / PO Box", "",
     "India: ^[1-9]\\d{5}$. USA: 5 or 9 digits. Country-specific.",
     "Required by NABH, NHS PDS, ABDM. HIPAA restricts sharing for small populations."),

    # ── 6. INSURANCE ──
    ("Insurance", "Financial Class", "financial_class", "ENUM", "Mandatory",
     "Coverage resource",
     "AAC.5 — Financial assessment at registration",
     "ACC.2.4 — Financial information at admission",
     "X12 837 — Payer information required for claims",
     "", "",
     "CBAHI — Insurance mandatory in Saudi Arabia (CCHI law)",
     "NABIDH — Insurance details mandatory (UAE law)",
     "",
     "Values: self_pay, insurance, government_scheme, corporate, charity, research",
     "Required by NABH AAC.5, JCI ACC.2.4. Determines billing workflow. In Saudi/UAE, insurance is legally mandatory."),

    ("Insurance", "Insurance Provider", "patient_insurance.insurance_provider", "TEXT", "Conditional",
     "Coverage.payor",
     "AAC.5", "ACC.2.4",
     "X12 837 Loop 2010BB — Payer Name",
     "", "",
     "CBAHI — Insurance verification mandatory",
     "NABIDH MDS: Insurance Details",
     "",
     "FK to master_insurance_providers. Required when financial_class = insurance.",
     "Required when patient is insured. Per NABH AAC.5, JCI ACC.2.4."),

    ("Insurance", "Policy Number", "patient_insurance.policy_number", "TEXT", "Conditional",
     "Coverage.identifier",
     "AAC.5", "",
     "X12 837 — Member ID required for claims",
     "", "", "CBAHI", "NABIDH", "",
     "Alphanumeric, max 50 chars. Required with insurance provider.",
     "Required for insurance claims processing."),

    # ── 7. GOVERNMENT IDENTIFIERS ──
    ("Government ID", "Aadhaar Number", "patient_identifiers (type=aadhaar)", "TEXT(12)", "Conditional",
     "Patient.identifier (type=NI, system=UIDAI)",
     "NABH — Accepted as valid ID proof for registration",
     "", "",
     "ABDM: Primary KYC method for ABHA creation. Verhoeff checksum.",
     "", "", "", "",
     "12 digits. Verhoeff checksum. NEVER store raw — SHA-256 hash only in id_number_hash.",
     "Per ABDM: primary Aadhaar-based ABHA enrollment. Per Indian law: voluntary, cannot deny services. Store only hash per privacy guidelines."),

    ("Government ID", "ABHA Number", "patient_identifiers (type=abha)", "TEXT(14)", "Optional",
     "Patient.identifier (type=ABHA, system=ABDM)",
     "NABH — ABHA linkage recommended for digital health",
     "", "",
     "ABDM: 14-digit national health ID (XX-XXXX-XXXX-XXXX). Mandatory for ABDM ecosystem participation.",
     "", "", "", "",
     "Format: ^\\d{2}-\\d{4}-\\d{4}-\\d{4}$. Verhoeff checksum.",
     "Per ABDM: national health identifier for India. Optional for registration but required for health record exchange."),

    ("Government ID", "ABHA Address", "patient_identifiers (type=abha_address)", "TEXT", "Optional",
     "Patient.identifier (type=PHR, system=ABDM)",
     "", "", "",
     "ABDM: Human-readable PHR address (username@abdm). Used for consent and record linking.",
     "", "", "", "",
     "Format: username@abdm. 4-150 chars, lowercase + numbers + dots + underscores.",
     "Per ABDM: user-friendly handle for health record access and consent management."),

    ("Government ID", "PAN Card", "patient_identifiers (type=pan)", "TEXT(10)", "Optional",
     "", "NABH — Accepted as valid ID proof",
     "", "", "", "", "", "", "",
     "Regex: ^[A-Z]{5}[0-9]{4}[A-Z]$. 4th char = holder type (P=Person).",
     "Indian Income Tax PAN. Required for billing above Rs 50,000 threshold."),

    ("Government ID", "Passport", "patient_identifiers (type=passport)", "TEXT", "Conditional",
     "Patient.identifier (type=PPN)",
     "NABH — Required for international patients",
     "ACC.1 — Identity verification for international patients",
     "", "", "", "CBAHI — Passport for visitors",
     "NABIDH — Passport for non-Emirates ID holders",
     "",
     "Alphanumeric max 20 chars. Store with issuing_country.",
     "Required for international/foreign patients per NABH, JCI, CBAHI, UAE."),

    ("Government ID", "SSN (USA)", "patient_identifiers (type=ssn)", "TEXT(11)", "Conditional",
     "Patient.identifier (type=SS)",
     "", "",
     "45 CFR 164.514(b) — SSN is PHI identifier #7. Must be encrypted at rest.",
     "", "", "", "", "",
     "Format: XXX-XX-XXXX. 9 digits. MUST be encrypted.",
     "PHI per HIPAA. Required for US insurance claims. Must be encrypted at rest per HIPAA Security Rule."),

    ("Government ID", "NHS Number (UK)", "patient_identifiers (type=nhs_number)", "TEXT(10)", "Conditional",
     "Patient.identifier (type=NI, system=NHS)",
     "", "", "", "",
     "PDS: NHS Number (mandatory for all UK patients). 10-digit, modulus 11 check digit.",
     "", "", "",
     "10 digits with modulus 11 check.",
     "Per NHS PDS: primary patient identifier in UK. Allocated at birth (England/Wales/IoM)."),

    ("Government ID", "Emirates ID (UAE)", "patient_identifiers (type=emirates_id)", "TEXT(15)", "Conditional",
     "Patient.identifier (type=NI)",
     "", "", "", "", "",
     "",
     "NABIDH: Emirates ID mandatory for all UAE residents. Links to NABIDH UHID.",
     "",
     "Format: 784-YYYY-NNNNNNN-C. 15 digits.",
     "Legally mandatory for all UAE healthcare per DHA/DOH. Links to unified NABIDH patient record."),

    ("Government ID", "Saudi National ID / Iqama", "patient_identifiers (type=national_id/iqama)", "TEXT(10)", "Conditional",
     "Patient.identifier (type=NI)",
     "", "", "", "", "", "CBAHI ESR — National ID/Iqama mandatory for all Saudi patients.",
     "", "",
     "National ID: 10 digits (citizens). Iqama: variable (residents).",
     "Per CBAHI: mandatory patient identifier in Saudi Arabia. National ID for citizens, Iqama for residents."),

    # ── 8. CONSENT ──
    ("Consent", "General Treatment Consent", "patient_consents (type=general_treatment)", "RECORD", "Mandatory",
     "Consent resource",
     "NABH PRE.1 — General consent at registration",
     "JCI PFR.5.1 — General consent for treatment must be obtained",
     "45 CFR 164.508 — Authorization for use/disclosure of PHI",
     "", "",
     "CBAHI — Consent at admission mandatory",
     "", "GDPR Art. 6/9 — Consent for health data processing",
     "Must be granted before any clinical encounter.",
     "Mandatory per NABH PRE.1, JCI PFR.5.1, HIPAA Privacy Rule, CBAHI, GDPR. Must be captured at registration."),

    ("Consent", "ABDM Data Sharing Consent", "patient_consents (type=abdm_linking)", "RECORD", "Optional",
     "Consent resource",
     "", "", "",
     "ABDM Consent Framework: consent_code=abha-enrollment, version=1.4. Required for ABHA linking.",
     "", "", "", "",
     "Required for ABHA linking. Separate from general treatment consent.",
     "Per ABDM: specific consent for linking health records to ABHA. Uses DEPA (Data Empowerment & Protection Architecture) framework."),

    ("Consent", "Research Participation Consent", "patient_consents (type=research_participation)", "RECORD", "Optional",
     "Consent resource",
     "NABH — Research ethics committee requirements",
     "JCI COP.9 — Research consent separate from treatment consent",
     "45 CFR 46 — Common Rule for human subjects research",
     "", "", "", "", "GDPR Art. 9(2)(j) — Research consent",
     "Must be separate from treatment consent. Revocable at any time.",
     "Per JCI COP.9, HIPAA 45 CFR 46: research consent must be distinct from treatment consent."),

    # ── 9. REGISTRATION METADATA ──
    ("Registration", "Patient Category", "category", "ENUM", "Mandatory",
     "",
     "AAC.5 — Patient categorization at registration",
     "ACC.2 — Registration process standardized",
     "", "", "", "", "", "",
     "Values: general, private, insurance, pmjay, cghs, esi, staff, staff_dependent, vip, mlc, free, charity, corporate, research_subject",
     "Per NABH AAC.5. Determines billing rules, priority, and reporting workflow."),

    ("Registration", "Registration Type", "registration_type", "ENUM", "Mandatory",
     "",
     "AAC.1 — Registration and admission process documented",
     "ACC.1 — Standardized admission criteria",
     "", "", "", "", "", "",
     "Values: new, revisit, transfer_in, referral, emergency, camp, telemedicine, pre_registration",
     "Per NABH AAC.1, JCI ACC.1. Determines workflow (new patient vs follow-up)."),

    ("Registration", "Medico-Legal Case Flag", "is_medico_legal", "BOOLEAN", "Optional",
     "",
     "NABH — MLC documentation per IPC/CrPC",
     "", "", "", "", "", "", "",
     "Default: false. If true, mlc_number required. Triggers MLC workflow.",
     "Per Indian Penal Code (IPC) and CrPC. Triggers police intimation, special documentation, restricted access."),

    ("Registration", "UHID (Auto-Generated)", "uhid", "TEXT", "Mandatory",
     "Patient.identifier (type=MR)",
     "AAC.1 — UHID generated at end of registration",
     "IPSG.1 — Unique MRN for patient identification",
     "X12 837 — Patient Account Number",
     "", "PDS: NHS Number", "CBAHI QM.17 — MRN",
     "NABIDH MDS: MRN (mandatory)",
     "IPS: identifier",
     "Auto-generated from sequences table. Format: {prefix}{zero-padded-number}. Immutable.",
     "Auto-generated. Required by all standards. NABH AAC.1, JCI IPSG.1, NHS, CBAHI, UAE NABIDH."),

    ("Registration", "Photo", "photo_url", "TEXT", "Optional",
     "Patient.photo",
     "AAC.2 — Photograph recommended for patient identification",
     "IPSG.1 — Visual identification supported",
     "45 CFR 164.514(b) — Full face photos are PHI identifier #17",
     "ABHA Profile: profilePhoto (optional, base64)",
     "", "", "", "",
     "JPEG/PNG, max 2MB, min 300x300px. Webcam capture or upload.",
     "Recommended by NABH AAC.2, JCI IPSG.1 for visual identification. PHI per HIPAA."),

    # ── 10. ALLERGIES ──
    ("Allergies", "Allergy Type", "patient_allergies.allergy_type", "ENUM", "Mandatory (per entry)",
     "AllergyIntolerance.category",
     "NABH COP.6 — Allergy documentation",
     "JCI IPSG.5 — Medication safety includes allergy check",
     "", "", "", "", "", "",
     "Values: drug, food, environmental, latex, contrast_dye, biological, other",
     "Per NABH COP.6, JCI IPSG.5. Category of allergen for clinical decision support."),

    ("Allergies", "Allergen Name", "patient_allergies.allergen_name", "TEXT", "Mandatory (per entry)",
     "AllergyIntolerance.code",
     "NABH COP.6", "JCI IPSG.5", "", "", "", "", "", "",
     "Free text or SNOMED CT coded",
     "Per NABH COP.6, JCI IPSG.5. Name of the allergen (drug name, food item, substance)."),

    ("Allergies", "Severity", "patient_allergies.severity", "ENUM", "Optional",
     "AllergyIntolerance.criticality",
     "", "JCI IPSG.5 — Severity affects clinical alerts",
     "", "", "", "", "", "",
     "Values: mild, moderate, severe, life_threatening",
     "Per JCI IPSG.5. Severity drives alert priority in CPOE and medication ordering."),
]


# ── Sheet 2: Regulatory Cross-Reference Matrix ───────────────────────────────

MATRIX_HEADERS = [
    "S.No", "Regulatory Body", "Country", "Standard/Clause",
    "Requirement", "Category", "Mandatory/Optional",
    "MedBrains Field Mapping", "Module", "Status"
]

MATRIX_DATA = [
    # NABH
    ("NABH", "India", "AAC.1 — Registration & Admission",
     "Organization has a documented registration process; UHID assigned at end",
     "Registration Process", "Mandatory", "uhid, registration_type", "Patient Registration", "Pending"),
    ("NABH", "India", "AAC.2 — Patient Identification",
     "Minimum 2 identifiers to verify identity. Options: name, DOB, MRN, phone, photo, wristband",
     "Patient Identification", "Mandatory", "first_name + last_name, date_of_birth, uhid", "Patient Registration", "Pending"),
    ("NABH", "India", "AAC.2 — ID Bands",
     "Identification bands with systematized markers or barcodes for inpatients",
     "Wristband", "Mandatory (IPD)", "Wristband print with UHID + name + DOB barcode", "Patient Registration", "Pending"),
    ("NABH", "India", "AAC.2 — Unknown Patients",
     "Documented policy for registration of unidentified/unknown patients",
     "Unknown Patient", "Mandatory", "is_unknown_patient, temporary_name", "Patient Registration", "Pending"),
    ("NABH", "India", "AAC.5 — Financial Assessment",
     "Financial assessment at point of registration; categorization for billing",
     "Financial", "Mandatory", "financial_class, category", "Patient Registration", "Pending"),
    ("NABH", "India", "PRE.1 — General Consent",
     "General consent obtained at registration for treatment",
     "Consent", "Mandatory", "patient_consents (type=general_treatment)", "Patient Registration", "Pending"),
    ("NABH", "India", "PRE.2 — Communication",
     "Information provided in language patient understands",
     "Communication", "Mandatory", "preferred_language, communication_needs", "Patient Registration", "Pending"),
    ("NABH", "India", "COP.5 — Pediatric Care",
     "Guardian details mandatory for minor patients",
     "Pediatric", "Conditional", "guardian_name, guardian_relation", "Patient Registration", "Pending"),
    ("NABH", "India", "COP.6 — Clinical Assessment",
     "Allergy status must be documented; blood group captured",
     "Allergies", "Mandatory", "no_known_allergies, patient_allergies, blood_group", "Patient Registration", "Pending"),

    # JCI
    ("JCI", "International", "IPSG.1 ME1 — Patient Identification",
     "At least 2 identifiers used to identify patients (NOT room/bed number). Zero-tolerance standard.",
     "Patient Identification", "Mandatory", "uhid + (first_name+last_name OR date_of_birth)", "Patient Registration", "Pending"),
    ("JCI", "International", "IPSG.1 ME2 — Before Clinical Procedures",
     "Verify identity before administering medications, blood, taking specimens, providing treatments",
     "Verification", "Mandatory", "Wristband scan / verbal verification", "Patient Registration + All Clinical", "Pending"),
    ("JCI", "International", "IPSG.5 — Medication Safety",
     "Allergy assessment mandatory; drug allergy alerts in ordering systems",
     "Allergies", "Mandatory", "no_known_allergies, patient_allergies", "Patient Registration", "Pending"),
    ("JCI", "International", "ACC.1 — Admission Criteria",
     "Standardized screening at point of first contact to match needs with resources",
     "Registration", "Mandatory", "registration_type, registration_source", "Patient Registration", "Pending"),
    ("JCI", "International", "ACC.2 — Registration Process",
     "Standardized registration capturing demographics, contact, insurance, referral info",
     "Registration", "Mandatory", "Full registration form", "Patient Registration", "Pending"),
    ("JCI", "International", "PFR.2 — Communication Barriers",
     "Process to identify and address physical, language, cultural barriers",
     "Communication", "Mandatory", "preferred_language, communication_needs", "Patient Registration", "Pending"),
    ("JCI", "International", "PFR.5.1 — General Consent",
     "General consent for treatment obtained and documented",
     "Consent", "Mandatory", "patient_consents (type=general_treatment)", "Patient Registration", "Pending"),
    ("JCI", "International", "PFR.6 — Informed Consent (Minors)",
     "Guardian consent for minors and legally incapacitated patients",
     "Consent", "Conditional", "guardian_name, guardian_relation, patient_consents", "Patient Registration", "Pending"),

    # ABDM
    ("ABDM", "India", "ABHA Enrollment — Aadhaar KYC",
     "Aadhaar-based enrollment creates 14-digit ABHA with verified demographics",
     "ABHA", "Optional", "patient_abha_links, patient_identifiers (type=abha)", "Patient Registration", "Pending"),
    ("ABDM", "India", "ABHA Profile — Mandatory Fields",
     "healthIdNumber, firstName, gender, yearOfBirth, mobile are mandatory in ABHA profile",
     "Demographics", "Mandatory (for ABHA)", "first_name, biological_sex, date_of_birth, phone_primary", "Patient Registration", "Pending"),
    ("ABDM", "India", "FHIR Patient Resource — NRCES IG v6.5",
     "Patient resource must have at least one identifier with NDHM identifier-type-code",
     "Interoperability", "Mandatory (for ABDM)", "patient_identifiers with ABHA type", "Patient Registration", "Pending"),
    ("ABDM", "India", "Consent Framework — DEPA",
     "Patient consent required for data sharing; consent artifact per HIP/HIU exchange",
     "Consent", "Mandatory (for data exchange)", "patient_consents (type=abdm_linking)", "Patient Registration", "Pending"),
    ("ABDM", "India", "Care Context Linking",
     "Each encounter must be linked as care context to ABHA for health record exchange",
     "Linking", "Mandatory (for ABDM)", "patient_abha_links.linking_token", "Patient Registration + OPD/IPD", "Pending"),

    # HIPAA
    ("HIPAA", "USA", "45 CFR 164.514(b) — 18 PHI Identifiers",
     "Names, dates, phone, fax, email, SSN, MRN, photos, biometrics, geographic data are PHI",
     "Privacy", "Mandatory", "All patient fields are PHI when linked to health data", "Patient Registration", "Pending"),
    ("HIPAA", "USA", "45 CFR 164.508 — Authorization",
     "Patient authorization required for use/disclosure of PHI beyond treatment/payment/operations",
     "Consent", "Mandatory", "patient_consents (type=data_sharing)", "Patient Registration", "Pending"),
    ("HIPAA", "USA", "Security Rule — Encryption",
     "PHI must be encrypted at rest and in transit (AES-256 recommended)",
     "Security", "Mandatory", "SSN and sensitive IDs encrypted. TLS for all API calls.", "Patient Registration", "Pending"),
    ("HIPAA", "USA", "X12 837 — Patient Demographics in Claims",
     "Subscriber name, DOB, gender, member ID required for insurance claims",
     "Billing", "Mandatory", "first_name, last_name, date_of_birth, biological_sex, member_id", "Patient Registration + Billing", "Pending"),

    # NHS
    ("NHS", "UK", "PDS — NHS Number Allocation",
     "Name, DOB, gender, address mandatory for NHS Number allocation. 10-digit with mod-11 check.",
     "Identification", "Mandatory", "first_name, last_name, date_of_birth, biological_sex, patient_addresses", "Patient Registration", "Pending"),
    ("NHS", "UK", "PDS — GP Registration",
     "Previous GP details required for medical record transfer",
     "Referral", "Mandatory", "referred_by_doctor, referred_by_hospital", "Patient Registration", "Pending"),

    # CBAHI
    ("CBAHI", "Saudi Arabia", "ESR — Correct Patient Identification",
     "Two identifiers mandatory. Full name (Arabic + English), National ID/Iqama, DOB, gender, MRN.",
     "Identification", "Mandatory", "first_name, last_name, full_name_local (Arabic), national_id/iqama, date_of_birth, uhid", "Patient Registration", "Pending"),
    ("CBAHI", "Saudi Arabia", "ESR — Wristband Identification",
     "All inpatients must wear wristband with name + MRN + DOB. Barcode-enabled.",
     "Wristband", "Mandatory (IPD)", "Wristband print", "Patient Registration", "Pending"),
    ("CBAHI", "Saudi Arabia", "QM.17 — Patient Identification Policy",
     "Documented policy for patient identification across all care settings",
     "Policy", "Mandatory", "System-wide enforcement of 2-identifier rule", "Patient Registration", "Pending"),

    # UAE
    ("UAE DOH/DHA", "UAE", "NABIDH MDS — Patient Demographics",
     "MRN, Patient Name, Gender, DOB, Emirates ID, Mobile, Nationality, Address mandatory",
     "Demographics", "Mandatory", "uhid, first_name, last_name, biological_sex, date_of_birth, emirates_id, phone_primary, nationality_id, patient_addresses", "Patient Registration", "Pending"),
    ("UAE DOH/DHA", "UAE", "NABIDH — Insurance Integration",
     "Insurance details mandatory per UAE health insurance law. Real-time eligibility verification.",
     "Insurance", "Mandatory", "patient_insurance (at least one active policy)", "Patient Registration", "Pending"),

    # EU
    ("EU/EHDS", "EU", "IPS Patient Profile — ISO 27269",
     "name (1..*), birthDate (1..1) mandatory. identifier, gender, address, telecom must-support.",
     "Demographics", "Mandatory", "first_name/last_name, date_of_birth. Plus gender, address, phone as must-support.", "Patient Registration", "Pending"),
    ("EU/EHDS", "EU", "GDPR Art. 6/9 — Health Data Processing",
     "Explicit consent required for processing health data. Right to access, rectify, erase.",
     "Consent", "Mandatory", "patient_consents. Support data portability and erasure requests.", "Patient Registration", "Pending"),
]


# ── Sheet 3: Master Data Lists ────────────────────────────────────────────────

MASTER_HEADERS = ["S.No", "Category", "Code", "Display Name", "Description", "Country", "Standard Ref"]

MASTER_DATA = []
sno = 0

# Blood Groups
for bg in [
    ("a_positive", "A+", "A Rh positive"),
    ("a_negative", "A-", "A Rh negative"),
    ("b_positive", "B+", "B Rh positive"),
    ("b_negative", "B-", "B Rh negative"),
    ("ab_positive", "AB+", "AB Rh positive"),
    ("ab_negative", "AB-", "AB Rh negative"),
    ("o_positive", "O+", "O Rh positive"),
    ("o_negative", "O-", "O Rh negative"),
    ("unknown", "Unknown", "Not yet determined"),
]:
    sno += 1
    MASTER_DATA.append((sno, "Blood Group", bg[0], bg[1], bg[2], "All", "ISBT 128, LOINC 882-1"))

# Gender
for g in [
    ("male", "Male", "Administrative male gender"),
    ("female", "Female", "Administrative female gender"),
    ("other", "Other", "Non-binary or intersex"),
    ("unknown", "Unknown", "Gender not determined or not disclosed"),
]:
    sno += 1
    MASTER_DATA.append((sno, "Gender", g[0], g[1], g[2], "All", "HL7 FHIR administrative-gender, SNOMED 248152002/248153007"))

# Marital Status
for ms in [
    ("single", "Single", "Never married"),
    ("married", "Married", "Currently married"),
    ("divorced", "Divorced", "Legally divorced"),
    ("widowed", "Widowed", "Spouse deceased"),
    ("separated", "Separated", "Legally separated"),
    ("domestic_partner", "Domestic Partner", "In a domestic partnership"),
    ("unknown", "Unknown", "Not disclosed"),
]:
    sno += 1
    MASTER_DATA.append((sno, "Marital Status", ms[0], ms[1], ms[2], "All", "HL7 v3 MaritalStatus, SNOMED 365581002"))

# Religions
for r in [
    ("hindu", "Hindu"), ("muslim", "Muslim"), ("christian", "Christian"),
    ("sikh", "Sikh"), ("buddhist", "Buddhist"), ("jain", "Jain"),
    ("jewish", "Jewish"), ("parsi", "Parsi/Zoroastrian"),
    ("bahai", "Baha'i"), ("tribal_religion", "Tribal Religion"),
    ("no_religion", "No Religion"), ("other", "Other"),
    ("prefer_not_to_say", "Prefer Not to Say"),
]:
    sno += 1
    MASTER_DATA.append((sno, "Religion", r[0], r[1], "", "India (default)", "Census of India categories"))

# Relations
for rel in [
    ("spouse", "Spouse"), ("father", "Father"), ("mother", "Mother"),
    ("son", "Son"), ("daughter", "Daughter"), ("brother", "Brother"),
    ("sister", "Sister"), ("grandfather", "Grandfather"),
    ("grandmother", "Grandmother"), ("grandson", "Grandson"),
    ("granddaughter", "Granddaughter"), ("uncle", "Uncle"),
    ("aunt", "Aunt"), ("nephew", "Nephew"), ("niece", "Niece"),
    ("cousin", "Cousin"), ("friend", "Friend"), ("neighbor", "Neighbor"),
    ("employer", "Employer"), ("colleague", "Colleague"),
    ("legal_guardian", "Legal Guardian"), ("power_of_attorney", "Power of Attorney"),
    ("other", "Other"),
]:
    sno += 1
    MASTER_DATA.append((sno, "Relation", rel[0], rel[1], "", "All", "HL7 v2-0131 Contact Role"))

# Occupations
for occ in [
    ("healthcare_worker", "Healthcare Worker"), ("doctor", "Doctor"),
    ("nurse", "Nurse"), ("paramedic", "Paramedic"),
    ("teacher", "Teacher"), ("professor", "Professor"),
    ("engineer", "Engineer"), ("lawyer", "Lawyer"),
    ("accountant", "Accountant"), ("farmer", "Farmer"),
    ("laborer", "Laborer"), ("driver", "Driver"),
    ("domestic_worker", "Domestic Worker"), ("business_owner", "Business Owner"),
    ("shopkeeper", "Shopkeeper"), ("government_employee", "Government Employee"),
    ("military", "Military"), ("police", "Police"),
    ("student", "Student"), ("retired", "Retired"),
    ("homemaker", "Homemaker"), ("unemployed", "Unemployed"),
    ("it_professional", "IT Professional"), ("banker", "Banker"),
    ("other", "Other"),
]:
    sno += 1
    MASTER_DATA.append((sno, "Occupation", occ[0], occ[1], "", "All", "ISCO-08 (WHO)"))

# Education Levels
for ed in [
    ("illiterate", "Illiterate"), ("primary", "Primary (1-5)"),
    ("middle_school", "Middle School (6-8)"), ("high_school", "High School (9-10)"),
    ("higher_secondary", "Higher Secondary (11-12)"), ("diploma", "Diploma"),
    ("graduate", "Graduate"), ("post_graduate", "Post Graduate"),
    ("doctorate", "Doctorate"), ("professional", "Professional Degree"),
    ("other", "Other"),
]:
    sno += 1
    MASTER_DATA.append((sno, "Education Level", ed[0], ed[1], "", "All", "ISCED (UNESCO)"))

# Patient Categories
for cat in [
    ("general", "General", "Standard self-pay patient"),
    ("private", "Private", "Private/premium category"),
    ("insurance", "Insurance", "Covered by health insurance"),
    ("pmjay", "PM-JAY", "Ayushman Bharat PM-JAY beneficiary"),
    ("cghs", "CGHS", "Central Government Health Scheme"),
    ("esi", "ESI", "Employees' State Insurance"),
    ("staff", "Staff", "Hospital employee"),
    ("staff_dependent", "Staff Dependent", "Dependent of hospital employee"),
    ("vip", "VIP", "Very Important Person — restricted access"),
    ("mlc", "MLC", "Medico-Legal Case"),
    ("free", "Free/Charity", "Free treatment"),
    ("charity", "Charity", "Charity-funded treatment"),
    ("corporate", "Corporate", "Corporate tie-up patient"),
    ("research_subject", "Research Subject", "Enrolled in clinical trial"),
]:
    sno += 1
    MASTER_DATA.append((sno, "Patient Category", cat[0], cat[1], cat[2], "All", "NABH AAC.5"))

# Allergy Types
for at in [
    ("drug", "Drug Allergy", "Allergy to medication"),
    ("food", "Food Allergy", "Allergy to food substance"),
    ("environmental", "Environmental", "Dust, pollen, mold, pet dander"),
    ("latex", "Latex", "Allergy to latex products"),
    ("contrast_dye", "Contrast Dye", "Allergy to radiological contrast media"),
    ("biological", "Biological", "Allergy to biological agents (vaccines, sera)"),
    ("other", "Other", "Other type of allergy"),
]:
    sno += 1
    MASTER_DATA.append((sno, "Allergy Type", at[0], at[1], at[2], "All", "FHIR AllergyIntolerance.category"))

# ID Document Types
for idt in [
    ("aadhaar", "Aadhaar", "12-digit unique ID issued by UIDAI", "India", "UIDAI Act 2016"),
    ("pan", "PAN Card", "10-char Permanent Account Number", "India", "Income Tax Act"),
    ("voter_id", "Voter ID (EPIC)", "Electoral Photo Identity Card", "India", "ECI"),
    ("driving_license", "Driving License", "Motor vehicle driving license", "All", "Regional RTA / MVA"),
    ("passport", "Passport", "International travel document", "All", "ICAO Doc 9303"),
    ("ration_card", "Ration Card", "Public Distribution System card", "India", "State Food & Civil Supplies"),
    ("abha", "ABHA Number", "14-digit Ayushman Bharat Health Account", "India", "ABDM / NHA"),
    ("abha_address", "ABHA Address", "PHR address (username@abdm)", "India", "ABDM"),
    ("ssn", "Social Security Number", "9-digit SSN", "USA", "SSA / HIPAA"),
    ("nhs_number", "NHS Number", "10-digit UK health ID", "UK", "NHS Digital / PDS"),
    ("medicare_number", "Medicare Number", "10/11-digit health ID", "Australia", "HI Act 2010"),
    ("emirates_id", "Emirates ID", "15-digit UAE national ID", "UAE", "FICA / ICA"),
    ("iqama", "Iqama", "Residence permit number", "Saudi Arabia", "MOI Saudi"),
    ("national_id", "National ID (Generic)", "Country-specific national ID", "Various", "Country laws"),
    ("birth_certificate", "Birth Certificate", "Vital statistics certificate", "All", "Civil registration"),
    ("disability_certificate", "Disability Certificate", "RPWD Act disability cert", "India", "RPWD Act 2016 / UDID"),
    ("uhid_external", "External UHID/MRN", "UHID from another hospital", "All", "Transfer-in"),
]:
    sno += 1
    MASTER_DATA.append((sno, "ID Document Type", idt[0], idt[1], idt[2], idt[3], idt[4]))

# Consent Types
for ct in [
    ("general_treatment", "General Treatment Consent", "Consent for examination, diagnosis, treatment", "NABH PRE.1, JCI PFR.5.1"),
    ("data_sharing", "Data Sharing Consent", "Consent for sharing data with third parties", "HIPAA 45 CFR 164.508, GDPR Art. 6"),
    ("abdm_linking", "ABDM/ABHA Linking Consent", "Consent for linking records to ABHA", "ABDM Consent Framework v1.4"),
    ("research_participation", "Research Participation", "Consent for clinical trial enrollment", "JCI COP.9, 45 CFR 46"),
    ("sms_communication", "SMS Communication", "Consent for receiving SMS notifications", "TRAI DND regulations"),
    ("email_communication", "Email Communication", "Consent for receiving email notifications", "CAN-SPAM Act / IT Act"),
    ("photography", "Photography Consent", "Consent for clinical photography", "NABH PRE.2"),
    ("advance_directive", "Advance Directive", "Living will, DNAR, healthcare proxy", "JCI PFR.7, SC Common Cause (India)"),
    ("organ_donation", "Organ Donation", "Consent for organ donation", "THOA 1994 (India), UNOS (USA)"),
    ("hie_participation", "HIE Participation", "Consent for health information exchange", "ABDM, EHDS"),
]:
    sno += 1
    MASTER_DATA.append((sno, "Consent Type", ct[0], ct[1], ct[2], "All", ct[3]))


# ── Sheet 4: ABDM API Reference ──────────────────────────────────────────────

ABDM_HEADERS = [
    "S.No", "Category", "Endpoint", "Method", "Purpose",
    "Request Key Fields", "Response Key Fields",
    "Auth Required", "Sandbox URL", "Production URL"
]

ABDM_DATA = [
    ("Authentication", "/gateway/v0.5/sessions", "POST",
     "Get JWT session token for ABDM API access",
     "clientId, clientSecret", "accessToken, expiresIn",
     "No (this IS the auth call)",
     "https://dev.abdm.gov.in/gateway/v0.5/sessions",
     "https://live.abdm.gov.in/gateway/v0.5/sessions"),

    ("ABHA Enrollment", "/v3/enrollment/request/otp", "POST",
     "Generate Aadhaar OTP for new ABHA creation",
     "scope=abha-enrol, loginHint=aadhaar, loginId=<aadhaar>",
     "txnId (used in enroll call)",
     "Yes (Bearer token)",
     "https://abhasbx.abdm.gov.in/abha/api/v3/enrollment/request/otp",
     "https://abha.abdm.gov.in/abha/api/v3/enrollment/request/otp"),

    ("ABHA Enrollment", "/v3/enrollment/enrol/byAadhaar", "POST",
     "Complete ABHA enrollment using Aadhaar OTP",
     "authData.otp.txnId, authData.otp.otpValue, consent.code=abha-enrollment",
     "ABHAProfile (healthIdNumber, name, gender, DOB, mobile, address, photo)",
     "Yes",
     "https://abhasbx.abdm.gov.in/abha/api/v3/enrollment/enrol/byAadhaar",
     "https://abha.abdm.gov.in/abha/api/v3/enrollment/enrol/byAadhaar"),

    ("ABHA Enrollment", "/v3/enrollment/enrol/byMobile", "POST",
     "Mobile-only enrollment (creates enrollment number, needs KYC later)",
     "mobile, firstName, gender, yearOfBirth",
     "enrollmentNumber (not full ABHA until KYC)",
     "Yes",
     "https://abhasbx.abdm.gov.in/abha/api/v3/enrollment/enrol/byMobile",
     "https://abha.abdm.gov.in/abha/api/v3/enrollment/enrol/byMobile"),

    ("ABHA Enrollment", "/v3/enrollment/enrol/suggestion", "GET",
     "Get ABHA address (username@abdm) suggestions",
     "txnId (from enrollment)", "suggestedAbhaAddresses[]",
     "Yes",
     "https://abhasbx.abdm.gov.in/abha/api/v3/enrollment/enrol/suggestion",
     "https://abha.abdm.gov.in/abha/api/v3/enrollment/enrol/suggestion"),

    ("ABHA Enrollment", "/v3/enrollment/enrol/abha-address", "POST",
     "Create ABHA address for the enrolled patient",
     "abhaAddress, txnId", "healthIdNumber, healthId (address)",
     "Yes",
     "https://abhasbx.abdm.gov.in/abha/api/v3/enrollment/enrol/abha-address",
     "https://abha.abdm.gov.in/abha/api/v3/enrollment/enrol/abha-address"),

    ("Verification", "/v0.5/users/auth/fetch-modes", "POST",
     "Get available auth modes for existing ABHA (MOBILE_OTP, AADHAAR_OTP, DEMOGRAPHICS)",
     "healthId (ABHA address or number)", "modes[] (available auth methods)",
     "Yes (X-HIP-ID header)",
     "https://dev.abdm.gov.in/gateway/v0.5/users/auth/fetch-modes",
     "https://live.abdm.gov.in/gateway/v0.5/users/auth/fetch-modes"),

    ("Verification", "/v0.5/users/auth/init", "POST",
     "Initiate authentication (triggers OTP to patient)",
     "healthId, authMethod, requester.id",
     "transactionId (via callback on-init)",
     "Yes (X-HIP-ID header)",
     "https://dev.abdm.gov.in/gateway/v0.5/users/auth/init",
     "https://live.abdm.gov.in/gateway/v0.5/users/auth/init"),

    ("Verification", "/v0.5/users/auth/confirm", "POST",
     "Confirm OTP, receive patient profile + linking token",
     "transactionId, credential (OTP or demographics)",
     "patient (name, gender, DOB, identifiers), accessToken, linkingToken (valid 30 min)",
     "Yes",
     "https://dev.abdm.gov.in/gateway/v0.5/users/auth/confirm",
     "https://live.abdm.gov.in/gateway/v0.5/users/auth/confirm"),

    ("Profile Share", "/v3/hip/patient/profile/share", "POST",
     "Receive patient profile via QR scan (hospital's callback endpoint)",
     "(Callback: ABDM sends profile to hospital's registered URL)",
     "ABHAProfile + linkingToken",
     "Yes (X-HIP-ID header)",
     "Hospital's registered callback URL",
     "Hospital's registered callback URL"),

    ("Care Context", "/v0.5/links/context/notify", "POST",
     "HIP-initiated care context linking (after encounter)",
     "patient.id (ABHA address), careContexts[].referenceNumber, hip.id",
     "Acknowledgment via on-notify callback",
     "Yes (X-HIP-ID + linkingToken)",
     "https://dev.abdm.gov.in/gateway/v0.5/links/context/notify",
     "https://live.abdm.gov.in/gateway/v0.5/links/context/notify"),

    ("Discovery", "/v0.5/care-contexts/discover", "POST",
     "Respond to patient-initiated care context discovery (hospital's callback)",
     "(Callback: CM sends patient details to hospital's registered URL)",
     "Hospital returns matching care contexts via on-discover",
     "Yes",
     "Hospital's registered callback URL",
     "Hospital's registered callback URL"),

    ("Profile", "/v3/profile/account", "GET",
     "Fetch complete ABHA profile",
     "X-Token header (from auth/confirm)",
     "ABHAProfile (full demographics, photo, KYC status)",
     "Yes (X-Token)",
     "https://abhasbx.abdm.gov.in/abha/api/v3/profile/account",
     "https://abha.abdm.gov.in/abha/api/v3/profile/account"),

    ("QR Code", "/v3/profile/account/qrCode", "GET",
     "Get ABHA QR code for patient card",
     "X-Token header", "QR code image (PNG/SVG)",
     "Yes (X-Token)",
     "https://abhasbx.abdm.gov.in/abha/api/v3/profile/account/qrCode",
     "https://abha.abdm.gov.in/abha/api/v3/profile/account/qrCode"),
]


# ── Build Workbook ────────────────────────────────────────────────────────────

def build_workbook():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Registration Fields ──
    ws1 = wb.active
    ws1.title = "Registration Fields"
    for c, h in enumerate(FIELDS_HEADERS, 1):
        ws1.cell(1, c, h)
    style_header(ws1, len(FIELDS_HEADERS))

    row = 2
    current_section = None
    sno = 0
    for f in FIELDS_DATA:
        section = f[0]
        if section != current_section:
            row = add_section(ws1, row, section, len(FIELDS_HEADERS))
            current_section = section
            sno = 0

        sno += 1
        values = [sno, section] + list(f[1:])
        for c, v in enumerate(values, 1):
            ws1.cell(row, c, v)
        style_row(ws1, row, len(FIELDS_HEADERS))

        # Highlight mandatory fields
        if f[4] == "Mandatory":
            for c in range(1, len(FIELDS_HEADERS) + 1):
                ws1.cell(row, c).fill = MANDATORY_FILL

        row += 1

    # Column widths
    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 28
    ws1.column_dimensions["D"].width = 35
    ws1.column_dimensions["E"].width = 12
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 30
    for col in "HIJKLMNO":
        ws1.column_dimensions[col].width = 35
    ws1.column_dimensions["P"].width = 40
    ws1.column_dimensions["Q"].width = 60

    # ── Sheet 2: Regulatory Cross-Reference ──
    ws2 = wb.create_sheet("Regulatory Requirements")
    for c, h in enumerate(MATRIX_HEADERS, 1):
        ws2.cell(1, c, h)
    style_header(ws2, len(MATRIX_HEADERS))

    current_body = None
    row = 2
    sno = 0
    for d in MATRIX_DATA:
        body = d[0]
        if body != current_body:
            row = add_section(ws2, row, body, len(MATRIX_HEADERS))
            current_body = body
            sno = 0

        sno += 1
        values = [sno] + list(d)
        for c, v in enumerate(values, 1):
            ws2.cell(row, c, v)
        style_row(ws2, row, len(MATRIX_HEADERS))
        row += 1

    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 40
    ws2.column_dimensions["E"].width = 50
    ws2.column_dimensions["F"].width = 18
    ws2.column_dimensions["G"].width = 14
    ws2.column_dimensions["H"].width = 40
    ws2.column_dimensions["I"].width = 30
    ws2.column_dimensions["J"].width = 10

    # ── Sheet 3: Master Data Lists ──
    ws3 = wb.create_sheet("Master Data")
    for c, h in enumerate(MASTER_HEADERS, 1):
        ws3.cell(1, c, h)
    style_header(ws3, len(MASTER_HEADERS))

    current_cat = None
    row = 2
    for d in MASTER_DATA:
        cat = d[1]
        if cat != current_cat:
            row = add_section(ws3, row, cat, len(MASTER_HEADERS))
            current_cat = cat

        for c, v in enumerate(d, 1):
            ws3.cell(row, c, v)
        style_row(ws3, row, len(MASTER_HEADERS))
        row += 1

    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 25
    ws3.column_dimensions["D"].width = 30
    ws3.column_dimensions["E"].width = 40
    ws3.column_dimensions["F"].width = 14
    ws3.column_dimensions["G"].width = 35

    # ── Sheet 4: ABDM API Reference ──
    ws4 = wb.create_sheet("ABDM API Reference")
    for c, h in enumerate(ABDM_HEADERS, 1):
        ws4.cell(1, c, h)
    style_header(ws4, len(ABDM_HEADERS))

    current_cat = None
    row = 2
    sno = 0
    for d in ABDM_DATA:
        cat = d[0]
        if cat != current_cat:
            row = add_section(ws4, row, cat, len(ABDM_HEADERS))
            current_cat = cat
            sno = 0

        sno += 1
        values = [sno] + list(d)
        for c, v in enumerate(values, 1):
            ws4.cell(row, c, v)
        style_row(ws4, row, len(ABDM_HEADERS))
        row += 1

    ws4.column_dimensions["A"].width = 6
    ws4.column_dimensions["B"].width = 16
    ws4.column_dimensions["C"].width = 45
    ws4.column_dimensions["D"].width = 8
    ws4.column_dimensions["E"].width = 50
    ws4.column_dimensions["F"].width = 40
    ws4.column_dimensions["G"].width = 40
    ws4.column_dimensions["H"].width = 14
    ws4.column_dimensions["I"].width = 55
    ws4.column_dimensions["J"].width = 55

    # ── Save ──
    out = Path("/Users/apple/Projects/MedBrains/data/Patient_Registration_Fields.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    print(f"Saved: {out}")
    print(f"  Sheet 1: Registration Fields — {len(FIELDS_DATA)} fields")
    print(f"  Sheet 2: Regulatory Requirements — {len(MATRIX_DATA)} requirements")
    print(f"  Sheet 3: Master Data — {len(MASTER_DATA)} entries")
    print(f"  Sheet 4: ABDM API Reference — {len(ABDM_DATA)} endpoints")


if __name__ == "__main__":
    build_workbook()

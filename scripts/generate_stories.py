#!/usr/bin/env python3
"""Generate a story-formatted backlog from MedBrains_Features.xlsx.

Reads the master feature tracker (repo root) and emits one markdown file per
module sheet under medbrains/docs/backlog/stories/, turning every Pending or
Partial feature into a story (a derived "As a … I want …" line + the real
metadata: priority, status, source, RFC ref, platforms). Done features are
skipped. The xlsx stays the source of truth — re-run after editing it.

  python3 scripts/generate_stories.py
"""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "MedBrains_Features.xlsx"
OUT_DIR = ROOT / "medbrains" / "docs" / "backlog" / "stories"

# A sensible primary role per module, for the "As a <role>" line.
MODULE_ROLE = {
    "Onboarding & Setup": "hospital admin",
    "Clinical": "clinician",
    "Diagnostics & Support": "lab/radiology technician",
    "Admin & Operations": "operations admin",
    "Specialty & Academic": "specialist",
    "IT, Security & Infrastructure": "system administrator",
    "TV Displays & Queue": "front-desk/display operator",
    "Printing & Forms": "staff member",
    "Mobile Apps": "mobile app user",
    "Technical Infrastructure": "platform engineer",
    "Regulatory & Compliance": "compliance officer",
    "Multi-Hospital & Vendors": "group administrator",
    "CMS & Blog": "content editor",
}

# Module-specific regulatory / domain acceptance line (from CLAUDE.md norms).
MODULE_REGULATORY = {
    "Onboarding & Setup": "7-layer config hierarchy respected; masters/seed validated; tenant isolation from creation.",
    "Clinical": "Clinical coding applied (ICD-10/11, SNOMED) and IPSG safety enforced (2-ID, allergy/LASA, consent).",
    "Diagnostics & Support": "Diagnostics norms met: LOINC + critical-value reporting (NABL); DICOM/AERB/PCPNDT for imaging; critical results routed to the ordering clinician.",
    "Admin & Operations": "Domain norms where relevant (NDPS/Schedule H/INN/AWaRe for pharmacy; GST/CGHS/TPA for billing); DTC/formulary + audit controls.",
    "Specialty & Academic": "Specialty statutes honoured (e.g. Mental Healthcare Act 2017, MTP/PCPNDT) with consent + NABH documentation.",
    "IT, Security & Infrastructure": "Security/privacy enforced: RBAC least-privilege, audit trail, encryption at rest, data-retention; no PHI in logs.",
    "TV Displays & Queue": "Public-display privacy (no full PHI); fixed emergency-code colour layer per spec.",
    "Printing & Forms": "Regulatory fields on output (UHID/MLC/Schedule badges); document audit (who printed, reprint count).",
    "Mobile Apps": "Offline-safe PHI handling, device auth, and sync conflict resolution.",
    "Technical Infrastructure": "Multi-tenant isolation (RLS), observability (tracing), and graceful degradation.",
    "Regulatory & Compliance": "Maps to the applicable norm (NABH/JCI/NDPS/D&C/PNDT/…) with evidence captured for accreditation.",
    "Multi-Hospital & Vendors": "Cross-tenant data boundaries enforced; vendor / head-office scoping respected.",
    "CMS & Blog": "Content governance: medical-content review/approval workflow; no PHI in public content.",
}


def slugify(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def cell(value) -> str:
    return "" if value is None else str(value).strip()


def acceptance_criteria(story: dict, module: str, role: str) -> list[str]:
    """Module-tailored AC checklist derived from the project's module-build
    workflow + regulatory domains. Done features render checked."""
    done = story["status"].lower() == "done"
    box = "[x]" if done else "[ ]"
    feature = story["feature"]
    verb_phrase = feature[0].lower() + feature[1:] if feature else feature
    platforms = story["platforms"]

    items = [
        f"The {role} can {verb_phrase} from the relevant screen, with clear loading / empty / error states.",
        "Backend: tenant-scoped endpoint(s) with RLS (`set_tenant_context`) and typed errors; new entities get a migration with `tenant_id` + RLS + indexes; `cargo clippy` clean.",
    ]
    if "Web" in platforms or not platforms:
        items.append(
            "Web: built from the `@/components/ui` seam, page-guarded via `useRequirePermission`, data through TanStack Query, inputs Zod-validated."
        )
    if "Mobile" in platforms:
        items.append("Mobile: React Native Paper screen with WatermelonDB offline support + sync.")
    if "TV" in platforms:
        items.append("TV: D-pad focus navigation, large-format layout, WebSocket realtime updates.")
    items.append(
        "Access is permission-gated (`P.<module>.<action>`) at page and element level; unauthorized users are redirected/hidden."
    )
    reg = MODULE_REGULATORY.get(module)
    if reg:
        items.append(reg)
    items.append("Mutations are audit-logged (who / when / what); PHI access is audited where applicable.")
    items.append(
        "Tests: CRUD + integration; `make check-all` (check-api / check-ui-api / check-types) passes; smoke test for any new endpoint."
    )
    return [f"- {box} {item}" for item in items]


def main() -> int:
    if not XLSX.exists():
        print(f"generate-stories: {XLSX} not found")
        return 1
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    index_lines: list[str] = []
    grand_total = 0

    for sheet_idx, name in enumerate(wb.sheetnames, start=1):
        ws = wb[name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            continue
        col = {cell(c).lower(): i for i, c in enumerate(header) if c}
        ci = lambda key: col.get(key)  # noqa: E731

        role = MODULE_ROLE.get(name, "user")
        by_sub: dict[str, list[dict]] = defaultdict(list)
        count = 0
        for r in rows:
            if not r or not any(x is not None and cell(x) for x in r):
                continue
            status = cell(r[ci("status")]) if ci("status") is not None else ""
            feature = cell(r[ci("feature")]) if ci("feature") is not None else ""
            if not feature:
                continue  # every feature becomes a story, regardless of status
            sub = (cell(r[ci("sub-module")]) if ci("sub-module") is not None else "") or "General"
            platforms = ", ".join(
                label
                for key, label in (("web", "Web"), ("mobile", "Mobile"), ("tv", "TV"))
                if ci(key) is not None and cell(r[ci(key)]).upper() in ("Y", "YES", "✓")
            )
            by_sub[sub].append(
                {
                    "feature": feature,
                    "priority": cell(r[ci("priority")]) if ci("priority") is not None else "",
                    "status": status,
                    "source": cell(r[ci("source")]) if ci("source") is not None else "",
                    "rfc": cell(r[ci("rfc ref")]) if ci("rfc ref") is not None else "",
                    "apps": cell(r[ci("apps")]) if ci("apps") is not None else "",
                    "platforms": platforms,
                }
            )
            count += 1

        if count == 0:
            continue
        grand_total += count

        slug = f"{sheet_idx:02d}-{slugify(name)}"
        out = OUT_DIR / f"{slug}.md"
        lines = [
            f"# {name} — stories",
            "",
            f"_Auto-generated from `MedBrains_Features.xlsx` — every feature as a story "
            f"with module-tailored acceptance criteria. {count} stories (✅ = Done). "
            f"Source of truth is the xlsx; regenerate via `python3 scripts/generate_stories.py`._",
            "",
        ]
        for sub in sorted(by_sub):
            lines.append(f"## {sub}")
            lines.append("")
            for s in by_sub[sub]:
                meta = " · ".join(
                    p
                    for p in (
                        s["priority"],
                        s["status"],
                        f"Platforms: {s['platforms']}" if s["platforms"] else "",
                        f"Source: {s['source']}" if s["source"] else "",
                        f"RFC: {s['rfc']}" if s["rfc"] else "",
                    )
                    if p
                )
                mark = "✅ " if s["status"].lower() == "done" else ""
                lines.append(f"### {mark}{s['feature']}")
                lines.append(f"> As a **{role}**, I want **{s['feature'].lower()}**.")
                lines.append("")
                lines.append(f"`{meta}`")
                lines.append("")
                lines.append("**Acceptance criteria**")
                lines.extend(acceptance_criteria(s, name, role))
                lines.append("")
        out.write_text("\n".join(lines) + "\n", encoding="utf-8")
        index_lines.append(f"| [{name}]({slug}.md) | {count} |")
        print(f"  {slug}.md: {count} stories")

    index = OUT_DIR / "README.md"
    index.write_text(
        "# Backlog stories (generated from features)\n\n"
        "Auto-generated from `MedBrains_Features.xlsx` — **every feature** as a story "
        "with module-tailored acceptance criteria, grouped by module → sub-module "
        "(✅ = already Done). The xlsx is the source of truth; regenerate via "
        "`python3 scripts/generate_stories.py`.\n\n"
        f"**{grand_total} stories** across {len(index_lines)} modules.\n\n"
        "| Module | Stories |\n|--------|--------:|\n" + "\n".join(index_lines) + "\n",
        encoding="utf-8",
    )
    print(f"generate-stories: {grand_total} stories -> {OUT_DIR.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

#!/usr/bin/env python3
"""Create platform/surface labels and apply them to the per-module epic issues.

Derives which surfaces each module touches from MedBrains_Features.xlsx
(Web/Mobile/TV columns + the Apps column: Desktop-Workstation/Kiosk/DeviceBridge,
plus plugin features) and labels the `Epic (features): <module>` issues. Idempotent.

  python3 scripts/label_feature_issues.py
"""

from __future__ import annotations

import json
import re
import subprocess
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "MedBrains_Features.xlsx"
EPIC_PREFIX = "Epic (features):"

# label -> (color hex, description)
LABELS = {
    "platform:web": ("1f6feb", "Runs on the web app"),
    "platform:mobile": ("2da44e", "Runs on a mobile app (RN Paper)"),
    "platform:tv": ("8250df", "Runs on a TV/display surface"),
    "platform:desktop": ("57606a", "Runs on a desktop/workstation surface"),
    "surface:kiosk": ("e16f24", "Self-service kiosk surface"),
    "surface:device-bridge": ("8a6d3b", "Hardware/device bridge surface"),
    "surface:workstation": ("6e7781", "Clinical/admin workstation surface"),
    "plugin": ("0d9488", "Plugin / extensibility feature"),
}


def cell(v) -> str:
    return "" if v is None else str(v).strip()


def ensure_labels() -> None:
    for name, (color, desc) in LABELS.items():
        subprocess.run(
            ["gh", "label", "create", name, "--color", color, "--description", desc, "--force"],
            capture_output=True,
            text=True,
            cwd=ROOT,
        )


def epic_numbers() -> dict[str, int]:
    out = subprocess.run(
        ["gh", "issue", "list", "--state", "all", "--limit", "500", "--json", "number,title"],
        capture_output=True,
        text=True,
        cwd=ROOT,
    )
    nums: dict[str, int] = {}
    for i in json.loads(out.stdout or "[]"):
        if i["title"].startswith(EPIC_PREFIX):
            nums[i["title"][len(EPIC_PREFIX):].strip()] = i["number"]
    return nums


def module_labels(ws) -> set[str]:
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        return set()
    col = {cell(c).lower(): i for i, c in enumerate(header) if c}

    def g(r, key):
        i = col.get(key)
        return cell(r[i]) if i is not None and i < len(r) else ""

    found: set[str] = set()
    for r in it:
        if not r:
            continue
        if g(r, "web").upper() in ("Y", "YES", "✓"):
            found.add("platform:web")
        if g(r, "mobile").upper() in ("Y", "YES", "✓"):
            found.add("platform:mobile")
        if g(r, "tv").upper() in ("Y", "YES", "✓"):
            found.add("platform:tv")
        apps = g(r, "apps")
        feature = g(r, "sub-module") + " " + g(r, "feature")
        for tok in re.split(r"[,/]", apps):
            t = tok.strip().lower()
            if t.startswith("desktop-workstation"):
                found.update({"platform:desktop", "surface:workstation"})
            elif t.startswith("desktop-kiosk"):
                found.update({"platform:desktop", "surface:kiosk"})
            elif t.startswith("desktop-devicebridge"):
                found.update({"platform:desktop", "surface:device-bridge"})
            elif t.startswith("desktop"):
                found.add("platform:desktop")
            elif t.startswith("tv-"):
                found.add("platform:tv")
            elif t.startswith("mobile-"):
                found.add("platform:mobile")
        if "plugin" in feature.lower() or "extension" in feature.lower():
            found.add("plugin")
    return found


def main() -> int:
    if not XLSX.exists():
        print(f"{XLSX} not found")
        return 1
    ensure_labels()
    epics = epic_numbers()
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    for name in wb.sheetnames:
        num = epics.get(name)
        if not num:
            continue
        labels = sorted(module_labels(wb[name]))
        if not labels:
            continue
        args = ["gh", "issue", "edit", str(num)]
        for label in labels:
            args += ["--add-label", label]
        res = subprocess.run(args, capture_output=True, text=True, cwd=ROOT)
        ok = "ok" if res.returncode == 0 else f"FAIL {res.stderr.strip()}"
        print(f"  #{num} {name}: {', '.join(labels)} [{ok}]")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

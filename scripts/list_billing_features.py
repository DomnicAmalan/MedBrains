#!/usr/bin/env python3
"""
List all billing/finance/revenue/insurance/accounts related features
from MedBrains_Features.xlsx across ALL sheets.
"""

import openpyxl
import os
import re

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "MedBrains_Features.xlsx",
)

# Keywords to match in Module, Sub-Module, or section headers (case-insensitive)
KEYWORDS = re.compile(
    r"billing|revenue|insurance|accounts|financial",
    re.IGNORECASE,
)


def find_column_indices(ws):
    """Find the column indices for Module, Sub-Module, Feature, Status from header row."""
    header_map = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        val = str(cell.value or "").strip().lower()
        col = cell.column  # 1-based
        if val == "module":
            header_map["module"] = col
        elif val == "sub-module":
            header_map["sub_module"] = col
        elif val == "feature":
            header_map["feature"] = col
        elif val == "status":
            header_map["status"] = col
        elif val == "s.no":
            header_map["sno"] = col
    return header_map


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)

    print(f"{'Sheet':<30} | {'Row':>4} | {'Module':<45} | {'Sub-Module':<40} | {'Feature':<70} | {'Status':<10}")
    print("-" * 210)

    total_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cols = find_column_indices(ws)

        if not cols:
            continue

        module_col = cols.get("module", 2)  # default col B
        sub_module_col = cols.get("sub_module", 3)  # default col C
        feature_col = cols.get("feature", 4)  # default col D
        status_col = cols.get("status", 7)  # default col G
        sno_col = cols.get("sno", 1)  # default col A

        # Track the current module/sub-module context from section headers
        current_module = ""
        current_sub_module = ""
        in_billing_section = False

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            # In read_only mode, EmptyCell objects lack .row, so use our own counter
            row_num = row_idx

            # Get cell values (columns are 1-based, row tuple is 0-based)
            def cell_val(col_idx):
                """Get value from 1-based column index."""
                idx = col_idx - 1
                if idx < len(row):
                    v = row[idx].value if hasattr(row[idx], "value") else None
                    return str(v or "").strip()
                return ""

            sno_val = cell_val(sno_col)
            module_val = cell_val(module_col)
            sub_module_val = cell_val(sub_module_col)
            feature_val = cell_val(feature_col)
            status_val = cell_val(status_col)

            # Detect module header rows (e.g. "12. BILLING & FINANCE" in S.No column or Module column)
            # These are rows where a numbered section header appears
            header_text = ""
            if sno_val and not module_val and not feature_val:
                # Header in S.No column (like "12. BILLING & FINANCE")
                header_text = sno_val
            elif module_val and not feature_val and not sub_module_val:
                # Header in Module column
                header_text = module_val

            if header_text:
                # Check if this is a main module header (e.g. "12. BILLING & FINANCE")
                if re.match(r"^\d+\.\s+", header_text):
                    if KEYWORDS.search(header_text):
                        current_module = header_text
                        current_sub_module = ""
                        in_billing_section = True
                    else:
                        in_billing_section = False
                        current_module = header_text
                        current_sub_module = ""
                # Check if sub-module header (e.g. "12.1 Billing Operations")
                elif re.match(r"^\s*\d+\.\d+", header_text):
                    if in_billing_section or KEYWORDS.search(header_text):
                        current_sub_module = header_text.strip()
                        if KEYWORDS.search(header_text):
                            in_billing_section = True
                continue

            # Feature rows: have actual feature text
            if not feature_val:
                continue

            # Determine if this row belongs to a billing-related section
            is_match = False

            # Check if we're inside a billing section
            if in_billing_section:
                is_match = True

            # Also check if the Module or Sub-Module cell directly matches
            if module_val and KEYWORDS.search(module_val):
                is_match = True
                current_module = module_val
            if sub_module_val and KEYWORDS.search(sub_module_val):
                is_match = True

            # Also check the feature text itself for billing keywords
            if KEYWORDS.search(feature_val):
                is_match = True

            if is_match:
                display_module = module_val or current_module
                display_sub = sub_module_val or current_sub_module

                # Truncate for display
                display_module = display_module[:45]
                display_sub = display_sub[:40]
                display_feature = feature_val[:70]

                print(
                    f"{sheet_name:<30} | {row_num:>4} | {display_module:<45} | {display_sub:<40} | {display_feature:<70} | {status_val:<10}"
                )
                total_count += 1

    print("-" * 210)
    print(f"Total billing/finance/revenue/insurance/accounts features found: {total_count}")

    wb.close()


if __name__ == "__main__":
    main()

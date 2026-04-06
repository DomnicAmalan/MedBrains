#!/usr/bin/env python3
import openpyxl, os, re

WORKBOOK_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "MedBrains_Features.xlsx")
KEYWORDS = re.compile(r"pharma|dispensi|prescri|drug|medication|formulary|NDC|AWaRe|antibiotic|narcotic|NDPS|controlled.substance", re.IGNORECASE)

def find_column_indices(ws):
    header_map = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        val = str(cell.value or "").strip().lower()
        col = cell.column
        if val == "module": header_map["module"] = col
        elif val == "sub-module": header_map["sub_module"] = col
        elif val == "feature": header_map["feature"] = col
        elif val == "status": header_map["status"] = col
        elif val == "s.no": header_map["sno"] = col
    return header_map

wb = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
print(f"{'Sheet':<30} | {'Row':>4} | {'Module':<30} | {'Sub-Module':<30} | {'Feature':<70} | {'Status':<10}")
print("-" * 185)
total = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    cols = find_column_indices(ws)
    if not cols: continue
    module_col = cols.get("module", 2)
    sub_module_col = cols.get("sub_module", 3)
    feature_col = cols.get("feature", 4)
    status_col = cols.get("status", 7)
    sno_col = cols.get("sno", 1)
    current_module = ""
    in_pharma_section = False
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        def cell_val(col_idx):
            idx = col_idx - 1
            if idx < len(row):
                v = row[idx].value if hasattr(row[idx], "value") else None
                return str(v or "").strip()
            return ""
        sno_val = cell_val(sno_col)
        module_val = cell_val(module_col)
        sub_module_val = cell_val(sub_module_col)
        feature_val = cell_val(feature_col)
        status_val = cell_val(status_col)
        header_text = ""
        if sno_val and not module_val and not feature_val:
            header_text = sno_val
        elif module_val and not feature_val and not sub_module_val:
            header_text = module_val
        if header_text:
            if re.match(r"^\d+\.\s+", header_text):
                if KEYWORDS.search(header_text):
                    current_module = header_text
                    in_pharma_section = True
                else:
                    in_pharma_section = False
                    current_module = header_text
            elif re.match(r"^\s*\d+\.\d+", header_text):
                if in_pharma_section or KEYWORDS.search(header_text):
                    if KEYWORDS.search(header_text):
                        in_pharma_section = True
            continue
        if not feature_val: continue
        is_match = in_pharma_section
        if module_val and KEYWORDS.search(module_val):
            is_match = True
            current_module = module_val
        if sub_module_val and KEYWORDS.search(sub_module_val):
            is_match = True
        if KEYWORDS.search(feature_val):
            is_match = True
        if is_match:
            dm = (module_val or current_module)[:30]
            ds = sub_module_val[:30]
            df = feature_val[:70]
            print(f"{sheet_name:<30} | {row_idx:>4} | {dm:<30} | {ds:<30} | {df:<70} | {status_val:<10}")
            total += 1
print("-" * 185)
print(f"Total pharmacy features: {total}")
wb.close()

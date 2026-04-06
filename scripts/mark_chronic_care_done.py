"""Mark Chronic Care / Drug-o-gram features as Done in MedBrains_Features.xlsx."""

import openpyxl

wb = openpyxl.load_workbook("MedBrains_Features.xlsx")
ws = wb["Clinical"]

updated = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    row_num = row[0].row
    # Rows 258-265 are the Drug-o-gram / Chronic Care features
    if 258 <= row_num <= 265:
        # Status column is typically column F (index 5)
        status_cell = row[5] if len(row) > 5 else None
        if status_cell is not None:
            old = status_cell.value
            status_cell.value = "Done"
            sno = row[0].value
            feature = row[3].value if len(row) > 3 else "?"
            print(f"  Row {row_num}: S.No={sno}, Feature={feature}, {old} -> Done")
            updated += 1

print(f"\nUpdated {updated} rows")
wb.save("MedBrains_Features.xlsx")
print("Saved MedBrains_Features.xlsx")

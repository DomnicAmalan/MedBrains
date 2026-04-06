#!/usr/bin/env python3
"""
Read MedBrains feature tracker and analyze pending work.
Identifies next batch of features to implement, grouped by module/sub-module.
"""

import openpyxl
from collections import defaultdict, OrderedDict


EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

# Priority sheets for recommendations (most impactful)
PRIORITY_SHEETS = ["Clinical", "Diagnostics & Support", "Admin & Operations"]


def parse_sheet(ws, sheet_name):
    """
    Parse a sheet, handling two layout patterns:
    1. 'Onboarding & Setup': Module headers in col B, sub-module in col C, features in col D
    2. All other sheets: Module/sub-module headers in col A (non-numeric),
       data rows have numeric S.No in col A, Module in B, Sub-Module in C, Feature in D
    """
    features = []
    current_module = ""
    current_submodule = ""

    is_onboarding = (sheet_name == "Onboarding & Setup")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        a_val = row[0].value  # S.No or header
        b_val = row[1].value if len(row) > 1 else None  # Module
        c_val = row[2].value if len(row) > 2 else None  # Sub-Module
        d_val = row[3].value if len(row) > 3 else None  # Feature
        e_val = row[4].value if len(row) > 4 else None  # Source
        f_val = row[5].value if len(row) > 5 else None  # Priority
        g_val = row[6].value if len(row) > 6 else None  # Status
        h_val = row[7].value if len(row) > 7 else None  # RFC Ref
        i_val = row[8].value if len(row) > 8 else None  # Web
        j_val = row[9].value if len(row) > 9 else None  # Mobile
        k_val = row[10].value if len(row) > 10 else None  # TV

        if is_onboarding:
            # Pattern 1: Onboarding sheet
            # Module header: only col B has value
            if b_val and not c_val and not d_val and not a_val:
                current_module = str(b_val).strip()
                continue
            # Sub-module header: only col C has value
            if c_val and not d_val and not a_val:
                current_submodule = str(c_val).strip()
                continue
            # Feature row: has S.No in A and Feature in D
            if a_val is not None and d_val is not None:
                try:
                    int(a_val)
                except (ValueError, TypeError):
                    continue
                features.append({
                    "sno": a_val,
                    "module": current_module,
                    "submodule": current_submodule,
                    "feature": str(d_val).strip(),
                    "source": str(e_val).strip() if e_val else "",
                    "priority": str(f_val).strip() if f_val else "",
                    "status": str(g_val).strip() if g_val else "Pending",
                    "rfc_ref": str(h_val).strip() if h_val else "",
                    "web": str(i_val).strip().upper() if i_val else "N",
                    "mobile": str(j_val).strip().upper() if j_val else "N",
                    "tv": str(k_val).strip().upper() if k_val else "N",
                })
        else:
            # Pattern 2: All other sheets
            # Module header row: A has text, B/C/D empty
            if a_val is not None and b_val is None and d_val is None:
                a_str = str(a_val).strip()
                if "." in a_str:
                    parts = a_str.split(".", 1)
                    try:
                        int(parts[0].strip())
                        rest = parts[1].strip()
                        if rest and rest[0].isdigit():
                            current_submodule = a_str
                        else:
                            current_module = a_str
                            current_submodule = ""
                    except ValueError:
                        current_module = a_str
                        current_submodule = ""
                else:
                    current_module = a_str
                    current_submodule = ""
                continue

            # Feature/data row: A is numeric, D has feature text
            if a_val is not None and d_val is not None:
                try:
                    int(a_val)
                except (ValueError, TypeError):
                    continue

                module = str(b_val).strip() if b_val else current_module
                submodule = str(c_val).strip() if c_val else current_submodule

                features.append({
                    "sno": a_val,
                    "module": module,
                    "submodule": submodule,
                    "feature": str(d_val).strip(),
                    "source": str(e_val).strip() if e_val else "",
                    "priority": str(f_val).strip() if f_val else "",
                    "status": str(g_val).strip() if g_val else "Pending",
                    "rfc_ref": str(h_val).strip() if h_val else "",
                    "web": str(i_val).strip().upper() if i_val else "N",
                    "mobile": str(j_val).strip().upper() if j_val else "N",
                    "tv": str(k_val).strip().upper() if k_val else "N",
                })

    return features


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    all_features = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        features = parse_sheet(ws, sheet_name)
        all_features[sheet_name] = features

    wb.close()

    # =========================================================================
    # SECTION 1: Summary per sheet
    # =========================================================================
    print("=" * 100)
    print("MEDBRAINS FEATURE TRACKER -- STATUS SUMMARY")
    print("=" * 100)

    grand_total = 0
    grand_done = 0
    grand_progress = 0
    grand_partial = 0
    grand_pending = 0
    grand_deferred = 0

    for sheet_name, features in all_features.items():
        total = len(features)
        done = sum(1 for f in features if f["status"] == "Done")
        progress = sum(1 for f in features if f["status"] == "In Progress")
        partial = sum(1 for f in features if f["status"] == "Partial")
        pending = sum(1 for f in features if f["status"] == "Pending")
        deferred = sum(1 for f in features if f["status"] == "Deferred")
        other = total - done - progress - partial - pending - deferred

        grand_total += total
        grand_done += done
        grand_progress += progress
        grand_partial += partial
        grand_pending += pending
        grand_deferred += deferred

        pct_done = (done / total * 100) if total > 0 else 0

        print(f"\n  {sheet_name} ({total} features)")
        line = f"    Done: {done:>4}  |  In Progress: {progress:>4}  |  Partial: {partial:>4}  |  Pending: {pending:>4}  |  Deferred: {deferred:>4}"
        if other > 0:
            line += f"  |  Other: {other:>4}"
        line += f"  [{pct_done:.1f}% complete]"
        print(line)

    print(f"\n{'~' * 100}")
    grand_pct = (grand_done / grand_total * 100) if grand_total > 0 else 0
    print(f"  TOTAL: {grand_total} features")
    print(f"    Done: {grand_done:>4}  |  In Progress: {grand_progress:>4}  |  Partial: {grand_partial:>4}  |  Pending: {grand_pending:>4}  |  Deferred: {grand_deferred:>4}  [{grand_pct:.1f}% complete]")

    # =========================================================================
    # SECTION 2: All In Progress features
    # =========================================================================
    print(f"\n\n{'=' * 100}")
    print("CURRENTLY IN PROGRESS")
    print("=" * 100)

    any_in_progress = False
    for sheet_name, features in all_features.items():
        in_progress = [f for f in features if f["status"] == "In Progress"]
        if in_progress:
            any_in_progress = True
            print(f"\n  [{sheet_name}]")
            for f in in_progress:
                pri = f" [{f['priority']}]" if f["priority"] else ""
                print(f"    - [{f['module']}] {f['submodule']}: {f['feature']}{pri}")

    if not any_in_progress:
        print("\n  (No features currently marked 'In Progress')")

    # =========================================================================
    # SECTION 3: All Partial features
    # =========================================================================
    print(f"\n\n{'=' * 100}")
    print("PARTIALLY COMPLETE (may need finishing)")
    print("=" * 100)

    any_partial = False
    for sheet_name, features in all_features.items():
        partial = [f for f in features if f["status"] == "Partial"]
        if partial:
            any_partial = True
            print(f"\n  [{sheet_name}]")
            for f in partial:
                print(f"    - [{f['module']}] {f['submodule']}: {f['feature']}")

    if not any_partial:
        print("\n  (No features marked 'Partial')")

    # =========================================================================
    # SECTION 4: Next batch recommendations -- Priority sheets
    # =========================================================================
    print(f"\n\n{'=' * 100}")
    print("RECOMMENDED NEXT BATCHES (Web features, by priority sheet)")
    print("=" * 100)

    for sheet_name in PRIORITY_SHEETS:
        features = all_features.get(sheet_name, [])
        pending_web = [f for f in features if f["status"] == "Pending" and f["web"] == "Y"]

        if not pending_web:
            print(f"\n  [{sheet_name}] -- No pending web features")
            continue

        # Group by module, then sub-module
        grouped = OrderedDict()
        for f in pending_web:
            mod = f["module"]
            sub = f["submodule"]
            key = (mod, sub)
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(f)

        print(f"\n  [{sheet_name}] -- {len(pending_web)} pending web features")
        print(f"  {'~' * 90}")

        # Find largest clusters (most related features together)
        clusters = sorted(grouped.items(), key=lambda x: len(x[1]), reverse=True)

        for (mod, sub), feats in clusters:
            priorities = set(f["priority"] for f in feats if f["priority"])
            priority_str = f" (Priorities: {', '.join(sorted(priorities))})" if priorities else ""

            print(f"\n    {mod} > {sub} -- {len(feats)} features{priority_str}")
            for f in feats[:8]:
                pri_tag = f" [{f['priority']}]" if f["priority"] else ""
                mob_tag = " [+Mobile]" if f["mobile"] == "Y" else ""
                print(f"      - {f['feature']}{pri_tag}{mob_tag}")
            if len(feats) > 8:
                print(f"      ... and {len(feats) - 8} more features")

    # =========================================================================
    # SECTION 5: Top recommended batches across all sheets
    # =========================================================================
    print(f"\n\n{'=' * 100}")
    print("TOP 15 LARGEST FEATURE CLUSTERS (all sheets, Pending + Web=Y)")
    print("=" * 100)

    all_clusters = []
    for sheet_name, features in all_features.items():
        pending_web = [f for f in features if f["status"] == "Pending" and f["web"] == "Y"]
        grouped = defaultdict(list)
        for f in pending_web:
            key = (sheet_name, f["module"], f["submodule"])
            grouped[key].append(f)
        for key, feats in grouped.items():
            all_clusters.append((key, feats))

    all_clusters.sort(key=lambda x: len(x[1]), reverse=True)

    for i, ((sheet, mod, sub), feats) in enumerate(all_clusters[:15]):
        priorities = set(f["priority"] for f in feats if f["priority"])
        pri_str = f" [{', '.join(sorted(priorities))}]" if priorities else ""
        print(f"\n  {i+1:>2}. [{sheet}] {mod} > {sub}")
        print(f"      {len(feats)} features{pri_str}")
        for f in feats[:5]:
            mob = " [+Mobile]" if f["mobile"] == "Y" else ""
            print(f"        - {f['feature']}{mob}")
        if len(feats) > 5:
            print(f"        ... and {len(feats) - 5} more")

    # =========================================================================
    # SECTION 6: Done features summary by module
    # =========================================================================
    print(f"\n\n{'=' * 100}")
    print("COMPLETED FEATURES BY MODULE (what is already built)")
    print("=" * 100)

    for sheet_name, features in all_features.items():
        done = [f for f in features if f["status"] == "Done"]
        if not done:
            continue
        done_by_mod = defaultdict(list)
        for f in done:
            done_by_mod[f["module"]].append(f)

        print(f"\n  [{sheet_name}]")
        for mod, feats in done_by_mod.items():
            print(f"    {mod}: {len(feats)} done")
            for f in feats[:5]:
                print(f"      - {f['feature']}")
            if len(feats) > 5:
                print(f"      ... and {len(feats) - 5} more")


if __name__ == "__main__":
    main()

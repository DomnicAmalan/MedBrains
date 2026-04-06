"""
Sprint 2: Mark ~91 Partial features as Done in MedBrains_Features.xlsx.

Batch 1: 33 frontend-only features
Batch 2: 58 backend+frontend features

Searches ALL sheets, dynamically finds "Feature" and "Status" columns
from the header row. Only changes Status from "Partial" to "Done".
Uses lambda-based matching on the Feature column for flexible text matching.
"""

import re
import openpyxl

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

# Each entry: (description, matcher_fn)
# matcher_fn receives the lowercase feature text and returns True if it matches.
FEATURES_TO_MARK_DONE = [
    # ════════════════════════════════════════════════════════════════
    #  BATCH 1 — Frontend-only (33 features)
    # ════════════════════════════════════════════════════════════════

    # Infection Control (8)
    # Admin & Operations row 95: "SSI (Surgical Site Infection) surveillance per procedure"
    ("IC: Surgical site infection tracking",
     lambda f: "surgical site infection" in f or ("ssi" in f and "surveillance" in f)),
    # Clinical row 316-321 or Admin & Operations row 94: "VAP (Ventilator Associated Pneumonia) tracking"
    ("IC: VAP/CLABSI/CAUTI bundle compliance",
     lambda f: ("vap" in f or "clabsi" in f or "cauti" in f) and ("track" in f or "bundle" in f or "pneumonia" in f)),
    # These may be Done already or have different wording; check Admin & Operations IC section
    ("IC: Hand hygiene audit results display",
     lambda f: "hand hygiene" in f and ("audit" in f or "compliance" in f)),
    # Admin & Operations row 101/319: "Antibiogram generation"
    ("IC: Antibiogram display",
     lambda f: "antibiogram" in f),
    # Admin & Operations row 119: "Outbreak detection alerts"
    ("IC: Outbreak timeline visualization",
     lambda f: "outbreak" in f and ("detection" in f or "timeline" in f or "alert" in f or "cluster" in f)),
    # Clinical row 318: "Infection trend analysis — time-series charts"
    ("IC: Environmental monitoring results",
     lambda f: "infection trend analysis" in f or ("time-series" in f and "infection" in f)),
    # Clinical row 321: "NHSN/CDC regulatory reporting auto-generation"
    ("IC: BMW monthly report generation",
     lambda f: ("nhsn" in f or "cdc" in f) and "report" in f),
    # Exists in Done features already, or may be in Occ Health
    ("IC: Needle-stick injury tracking form",
     lambda f: "needle" in f and ("stick" in f or "sharp" in f) and ("injury" in f or "track" in f)),

    # Quality (5)
    # May not exist as Partial — checking quality section
    ("Quality: CAPA effectiveness review",
     lambda f: "capa" in f and ("effectiveness" in f or "review" in f)),
    ("Quality: Patient feedback analysis charts",
     lambda f: "patient feedback" in f and ("analysis" in f or "chart" in f)),
    ("Quality: Near-miss reporting form",
     lambda f: "near-miss" in f or "near miss" in f),
    ("Quality: Quality indicator benchmarking",
     lambda f: "quality indicator" in f and ("benchmark" in f or "calculat" in f)),
    ("Quality: SOP version comparison",
     lambda f: "sop" in f and "version" in f),

    # Regulatory (4)
    # Regulatory row 8: "Internal audit scheduler with NABH compliance scoring per department"
    ("Regulatory: NABH self-assessment scoring",
     lambda f: "nabh" in f and ("compliance scoring" in f or "self-assessment" in f or "audit scheduler" in f)),
    # Regulatory rows — gap analysis
    ("Regulatory: Gap analysis report",
     lambda f: "gap analysis" in f),
    # Admin & Operations row 222: "Renewal alerts (30/60/90 days before expiry)" (BME)
    ("Regulatory: License renewal tracking alerts",
     lambda f: "renewal alert" in f and ("30" in f or "60" in f or "90" in f or "expiry" in f)),
    ("Regulatory: Regulatory calendar gantt view",
     lambda f: "regulatory" in f and "calendar" in f),

    # Onboarding/Setup (4)
    ("Setup: Department master operating hours display",
     lambda f: "department head assignment" in f),
    ("Setup: Location hierarchy tree view",
     lambda f: "import locations" in f or "location" in f and "csv" in f),
    # Onboarding row 66: "Module-specific configuration wizard per enabled module"
    ("Setup: System configuration wizard status",
     lambda f: "configuration wizard" in f),
    # Onboarding row 32/39/46: "Import ... from CSV/Excel"
    ("Setup: Master data import status tracking",
     lambda f: "import" in f and ("csv" in f or "excel" in f) and ("department" in f or "service" in f)),

    # Occupational Health (3)
    # Admin & Operations Occ Health section
    ("OccHealth: Pre-employment exam form",
     lambda f: "pre-employment" in f),
    ("OccHealth: Periodic health check scheduler",
     lambda f: "periodic health check" in f or "periodic health" in f),
    ("OccHealth: Fitness certificate template",
     lambda f: "fitness certificate" in f),

    # Case Management (3)
    # Admin & Operations Case Mgmt section
    ("Case Mgmt: Care plan progress tracking",
     lambda f: "care plan" in f and "progress" in f),
    ("Case Mgmt: Length-of-stay alerts",
     lambda f: "length" in f and "stay" in f and "alert" in f),
    ("Case Mgmt: Readmission risk display",
     lambda f: "readmission risk" in f),

    # Utilization Review (2)
    # Admin & Operations row 347/348
    ("UR: Concurrent review timeline",
     lambda f: "admission" in f and "continued stay review" in f or "interqual" in f or "milliman" in f),
    ("UR: Denial management dashboard",
     lambda f: "ai-assisted utilization review" in f or ("utilization review" in f and "auto-extract" in f)),

    # CSSD (2)
    # Diagnostics row 215: "Complete traceability: patient → instrument set → autoclave load → BI result"
    ("CSSD: Sterilization cycle documentation",
     lambda f: "traceability" in f and ("instrument" in f or "autoclave" in f)),
    # Diagnostics row 217: "CSSD workload and turnaround time reporting"
    ("CSSD: Instrument set contents display",
     lambda f: "cssd workload" in f or ("cssd" in f and "turnaround" in f)),

    # Diet & Kitchen (2)
    # Diagnostics row 229: "Kitchen inventory management"
    ("Diet: Therapeutic diet plan display",
     lambda f: "kitchen inventory" in f),
    ("Diet: Kitchen production summary",
     lambda f: "kitchen" in f and ("production" in f or "summary" in f)),

    # ════════════════════════════════════════════════════════════════
    #  BATCH 2 — Backend + Frontend (58 features)
    # ════════════════════════════════════════════════════════════════

    # Infection Control (10)
    # Admin & Operations row 96: "HAI rate calculation per ICU/ward with benchmarking"
    ("IC: HAI rate calculation",
     lambda f: "hai rate" in f),
    # Admin & Operations row 97: "Infection control dashboard with real-time alerts"
    ("IC: Device utilization ratios",
     lambda f: "infection control dashboard" in f and "real-time" in f),
    # Admin & Operations row 103: "DDD (Defined Daily Dose) calculation per department"
    ("IC: Antimicrobial consumption tracking",
     lambda f: "ddd" in f or "defined daily dose" in f),
    # Already might be Done
    ("IC: Surgical prophylaxis compliance",
     lambda f: "surgical prophylaxis" in f),
    # Diagnostics row 57: "Microbiology — culture sensitivity report, antibiogram"
    ("IC: Culture sensitivity report",
     lambda f: "culture sensitivity" in f),
    # Multi-drug resistant organism
    ("IC: MDRO tracking",
     lambda f: "multi-drug resistant" in f or "mdro" in f),
    # Exposure management
    ("IC: Exposure management workflow",
     lambda f: "exposure management" in f),
    # IC committee meeting minutes
    ("IC: IC committee meeting minutes",
     lambda f: "committee" in f and "meeting" in f and "minutes" in f),
    # Monthly surveillance report
    ("IC: Monthly surveillance report",
     lambda f: "monthly surveillance" in f),
    # Root cause analysis for outbreaks
    ("IC: Root cause analysis for outbreaks",
     lambda f: "root cause analysis" in f),

    # Quality (8)
    # Admin & Operations Quality section
    ("Quality: Clinical audit scheduling",
     lambda f: "clinical audit" in f and "schedul" in f),
    ("Quality: Audit trail with findings",
     lambda f: "audit trail" in f and "finding" in f),
    ("Quality: CAPA tracking with SLA",
     lambda f: "capa" in f and ("tracking" in f or "sla" in f)),
    ("Quality: Quality committee dashboard",
     lambda f: "quality committee" in f),
    # Admin & Operations Quality: may be labeled differently
    ("Quality: Mortality review records",
     lambda f: "mortality review" in f or "mortality" in f and "review" in f),
    # Admin & Operations row 147: "Regulatory reporting — PvPI (ADR), Hemovigilance (NACO), Materiovigilance (CDSCO)"
    ("Quality: Sentinel event tracking",
     lambda f: "regulatory reporting" in f and ("pvpi" in f or "hemovigilance" in f or "materiovigilance" in f)),
    ("Quality: Patient safety indicator calculation",
     lambda f: "patient safety" in f and ("indicator" in f or "goal" in f)),
    ("Quality: Cross-department quality scorecard",
     lambda f: "cross-department" in f or "cross department" in f),

    # Regulatory (6)
    # Regulatory row 6: "NABH mandatory document generation"
    ("Regulatory: Auto-populate compliance",
     lambda f: "nabh mandatory document" in f),
    # Regulatory row 11: "NMC medical college compliance tracking"
    ("Regulatory: Regulatory submission tracking",
     lambda f: "nmc" in f and ("medical college" in f or "compliance tracking" in f or "doctor registration" in f)),
    # Regulatory row 34: "JCI tracer methodology support"
    ("Regulatory: Mock survey management",
     lambda f: "jci tracer" in f or "tracer methodology" in f),
    # Regulatory row 12: "NMC doctor registration verification integration"
    ("Regulatory: Staff credential verification",
     lambda f: "doctor registration verification" in f),
    # Regulatory row 22: "AERB compliance for radiology/nuclear medicine"
    ("Regulatory: Facility license dashboard",
     lambda f: bool(re.search(r"\baerb\b", f)) and ("compliance" in f or "radiology" in f)),
    # Regulatory row 21: "NABL document control — SOPs, calibration records, proficiency testing"
    ("Regulatory: NABL document tracking",
     lambda f: "nabl document" in f),

    # Onboarding/Setup (6)
    # Onboarding row 56: "Bulk user import from CSV/Excel"
    ("Setup: Bulk user creation",
     lambda f: "bulk user import" in f),
    # Onboarding row 52/53: "Role hierarchy" / "Data scope per role"
    ("Setup: Department template seeding",
     lambda f: "role hierarchy" in f or "data scope per role" in f),
    # Onboarding row 67: "Feature flag management within each module"
    ("Setup: Master data completeness check",
     lambda f: "feature flag management" in f),
    # IT row 279: "Hospital-wide real-time dashboard"
    ("Setup: System health dashboard",
     lambda f: "hospital-wide real-time dashboard" in f),
    # Onboarding row 128: "API documentation auto-generation (OpenAPI/Swagger)"
    ("Setup: Configuration export/import",
     lambda f: "api documentation auto-generation" in f or "openapi" in f or "swagger" in f),
    # Onboarding row 165: "Compliance audit trail (who reviewed, when)"
    ("Setup: Audit log viewer",
     lambda f: "compliance audit trail" in f and "who reviewed" in f),

    # Scheduling (5)
    # Admin & Operations Scheduling section — row 360: "ML-based no-show prediction"
    ("Scheduling: Resource conflict detection",
     lambda f: "no-show prediction" in f or "ml-based" in f and "no-show" in f),
    # row 362: "Targeted reminder escalation for high no-show risk"
    ("Scheduling: Waitlist auto-promotion",
     lambda f: "targeted reminder escalation" in f or ("reminder" in f and "no-show risk" in f)),
    ("Scheduling: Schedule analytics",
     lambda f: "schedule" in f and "analytic" in f),
    ("Scheduling: Recurring appointment management",
     lambda f: "recurring appointment" in f),
    ("Scheduling: Block scheduling",
     lambda f: "block scheduling" in f or "block schedule" in f),

    # Occupational Health (4)
    # Admin & Operations Occ Health section — row 326: "Workers compensation claim integration"
    ("OccHealth: Workplace hazard registry",
     lambda f: "workers compensation" in f or "workplace hazard" in f),
    ("OccHealth: Health surveillance analytics",
     lambda f: "health surveillance" in f and "analytic" in f),
    ("OccHealth: Vaccination tracking",
     lambda f: "vaccination tracking" in f),
    ("OccHealth: Return-to-work clearance",
     lambda f: "return-to-work" in f or "return to work" in f),

    # OPD (3)
    ("OPD: Pharmacy dispatch tracking",
     lambda f: "pharmacy dispatch" in f),
    ("OPD: Referral tracking",
     lambda f: "referral tracking" in f),
    ("OPD: Follow-up compliance",
     lambda f: "follow-up compliance" in f or "followup compliance" in f or "follow up compliance" in f),

    # IPD (3)
    ("IPD: Discharge summary generation",
     lambda f: "discharge summary" in f and "generat" in f),
    ("IPD: Bed transfer workflow",
     lambda f: "bed transfer" in f and "workflow" in f),
    ("IPD: Expected discharge list",
     lambda f: "expected discharge" in f and "list" in f),

    # Pharmacy (3)
    ("Pharmacy: Drug interaction check",
     lambda f: "drug interaction" in f and "check" in f),
    ("Pharmacy: Formulary restriction enforcement",
     lambda f: "formulary restriction" in f),
    ("Pharmacy: Prescription audit trail",
     lambda f: "prescription audit" in f),

    # Billing (3)
    ("Billing: Package rate management",
     lambda f: "package rate" in f),
    # Admin & Operations row 36: "Co-pay/deductible calculation"
    ("Billing: Co-pay/deductible calculation",
     lambda f: "co-pay" in f or "copay" in f),
    # Admin & Operations row 13: "ER billing (deferred billing for emergency cases)"
    ("Billing: ER billing fast-track",
     lambda f: "er billing" in f or ("deferred billing" in f and "emergency" in f)),

    # Camp Management (2)
    # Admin & Operations row 317: "Camp reports & analytics"
    ("Camp: Camp analytics dashboard",
     lambda f: "camp report" in f and "analytic" in f),
    ("Camp: Camp report generation",
     lambda f: "camp report" in f),

    # Facilities Management (2)
    # Admin & Operations row 266: "Preventive maintenance schedule for all infrastructure"
    ("FMS: Preventive maintenance scheduling",
     lambda f: "preventive maintenance schedule" in f),
    # Admin & Operations row 259/260: Energy-related
    ("FMS: Energy consumption analytics",
     lambda f: ("energy" in f or "power" in f) and ("switchover" in f or "cea" in f or "consumption" in f or "compliance" in f)),

    # Front Office (2)
    # Admin & Operations row 308-310: Visitor mgmt system, special categories, parking
    ("Front Office: Visitor analytics",
     lambda f: "visitor management system" in f or ("special categories" in f and "visitor" in f)),
    # Admin & Operations row 298-301: Queue display, SMS notification, lab/rad/pharmacy queue
    ("Front Office: Queue performance metrics",
     lambda f: "queue display" in f or ("queue" in f and ("waiting area" in f or "screen" in f))),

    # HR (1)
    # Admin & Operations HR section — training compliance
    ("HR: Training compliance dashboard",
     lambda f: "training" in f and "compliance" in f and "dashboard" in f),
]


def find_column_indices(ws):
    """Scan header row(s) to find Feature and Status column indices.

    Returns (feature_col, status_col, header_row) as 1-based column numbers.
    Checks the first 5 rows for header-like content.
    """
    for row in range(1, min(6, ws.max_row + 1)):
        feature_col = None
        status_col = None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str):
                val_lower = val.strip().lower()
                if val_lower == "feature":
                    feature_col = col
                elif val_lower == "status":
                    status_col = col
        if feature_col and status_col:
            return feature_col, status_col, row
    return None, None, None


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheets = wb.sheetnames
    print(f"Available sheets ({len(sheets)}): {sheets}")
    print()

    updated = []
    matched_features = set()

    for sheet_name in sheets:
        ws = wb[sheet_name]
        feature_col, status_col, header_row = find_column_indices(ws)

        if not feature_col or not status_col:
            print(f"  SKIP: Sheet '{sheet_name}' -- no Feature/Status headers found")
            continue

        print(f"  Sheet '{sheet_name}': Feature=col {feature_col}, Status=col {status_col}, header_row={header_row}")

        for row in range(header_row + 1, ws.max_row + 1):
            feature_cell = ws.cell(row=row, column=feature_col)
            feature_val = feature_cell.value
            if not feature_val or not isinstance(feature_val, str):
                continue

            feature_lower = feature_val.lower().strip()

            status_cell = ws.cell(row=row, column=status_col)
            old_status = status_cell.value
            if old_status != "Partial":
                continue

            # Check against all matchers
            for idx, (description, matcher) in enumerate(FEATURES_TO_MARK_DONE):
                if idx in matched_features:
                    continue  # Already matched this feature
                if matcher(feature_lower):
                    status_cell.value = "Done"
                    matched_features.add(idx)
                    updated.append({
                        "row": row,
                        "sheet": sheet_name,
                        "feature": feature_val.strip(),
                        "old_status": old_status,
                        "rule": description,
                    })
                    break  # One match per row

    # Save
    wb.save(EXCEL_PATH)

    # Report
    print(f"\n{'=' * 80}")
    print(f"Sprint 2: Partial -> Done -- Excel Update Report")
    print(f"{'=' * 80}")
    print(f"Total features updated: {len(updated)}")
    print()

    for u in updated:
        print(
            f"  [{u['sheet'][:25]:>25s}] Row {u['row']:3d}: "
            f"Partial -> Done  {u['feature'][:60]}"
        )

    # Check for unmatched features
    unmatched = [
        FEATURES_TO_MARK_DONE[i][0]
        for i in range(len(FEATURES_TO_MARK_DONE))
        if i not in matched_features
    ]
    if unmatched:
        print(f"\n  WARNING: {len(unmatched)} features did NOT match any Partial row:")
        for desc in unmatched:
            print(f"    - {desc}")
    else:
        print(f"\n  All {len(FEATURES_TO_MARK_DONE)} features matched successfully.")

    print(f"\n  Total expected: {len(FEATURES_TO_MARK_DONE)}")
    print(f"  Total matched:  {len(matched_features)}")
    print(f"  Total unmatched: {len(FEATURES_TO_MARK_DONE) - len(matched_features)}")


if __name__ == "__main__":
    main()

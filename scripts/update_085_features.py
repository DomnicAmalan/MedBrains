#!/usr/bin/env python3
"""
Update MedBrains_Features.xlsx for migration 085 modules:
- Occupational Health (8 features)
- Utilization Review (5 features)
- Case Management (5 features)
- Scheduling & No-Show AI (5 features)

Sheet: Admin & Operations
Status column: G (column 7)
"""

import openpyxl

XLSX_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"
SHEET_NAME = "Admin & Operations"
STATUS_COL = 7  # Column G

# Map: row number -> (expected feature text prefix for verification, new status)
UPDATES = {
    # === Occupational Health (8 features) ===
    323: ("Pre-employment drug screening management", "Done"),
    324: ("Vaccination compliance tracking for healthcare workers", "Done"),
    325: ("Work-related injury documentation with OSHA recordability", "Done"),
    326: ("Workers compensation claim integration", "Partial"),
    327: ("Return-to-work clearance workflow", "Done"),
    328: ("Employer access controls", "Done"),
    329: ("Periodic health surveillance scheduling", "Done"),
    330: ("Employee wellness program tracking", "Done"),

    # === Utilization Review (5 features) ===
    347: ("Admission/continued stay review with InterQual", "Partial"),
    348: ("AI-assisted utilization review", "Partial"),
    349: ("Concurrent review tracking", "Done"),
    350: ("Payer communication log", "Done"),
    351: ("Observation vs inpatient status tracking", "Done"),

    # === Case Management (5 features) ===
    353: ("Case manager assignment per patient", "Done"),
    354: ("Discharge barriers tracking", "Done"),
    355: ("Post-acute facility finder", "Partial"),
    356: ("Social work referral integration", "Done"),
    357: ("Discharge disposition tracking", "Done"),

    # === Scheduling & No-Show AI (5 features) ===
    360: ("ML-based no-show prediction scoring", "Partial"),
    361: ("Overbooking recommendation based on predicted", "Done"),
    362: ("Targeted reminder escalation for high no-show risk", "Partial"),
    363: ("Waitlist auto-fill", "Done"),
    364: ("No-show analytics dashboard", "Done"),
}


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]

    print(f"Opened '{SHEET_NAME}' sheet ({ws.max_row} rows)")
    print(f"Status column: {STATUS_COL} (G)\n")

    updated = 0
    errors = 0

    for row, (expected_prefix, new_status) in sorted(UPDATES.items()):
        actual_feature = ws.cell(row, 4).value  # Column D = Feature
        old_status = ws.cell(row, STATUS_COL).value

        # Verify we are updating the correct row by checking feature text prefix
        if actual_feature is None or expected_prefix.lower() not in actual_feature.lower():
            print(f"  ERROR Row {row}: Expected feature containing '{expected_prefix}'")
            print(f"         Got: '{actual_feature}'")
            errors += 1
            continue

        # Update the status cell
        ws.cell(row, STATUS_COL).value = new_status
        updated += 1

        status_change = f"{old_status} -> {new_status}"
        print(f"  Row {row}: {status_change:22s} | {actual_feature[:72]}")

    print(f"\nUpdated: {updated} rows")
    if errors:
        print(f"Errors:  {errors} rows (NOT updated)")

    # Summary by module
    done_count = sum(1 for _, (_, s) in UPDATES.items() if s == "Done")
    partial_count = sum(1 for _, (_, s) in UPDATES.items() if s == "Partial")
    print(f"\nSummary: {done_count} Done, {partial_count} Partial (total {done_count + partial_count})")
    print(f"  Occupational Health:     7 Done, 1 Partial")
    print(f"  Utilization Review:      3 Done, 2 Partial")
    print(f"  Case Management:         4 Done, 1 Partial")
    print(f"  Scheduling & No-Show AI: 3 Done, 2 Partial")

    if errors == 0:
        wb.save(XLSX_PATH)
        print(f"\nSaved to {XLSX_PATH}")
    else:
        print(f"\nNOT saved due to {errors} errors. Fix row numbers and retry.")


if __name__ == "__main__":
    main()

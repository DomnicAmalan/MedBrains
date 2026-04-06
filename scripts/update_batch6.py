#!/usr/bin/env python3
"""
Batch 6: Mark OPD features as Done in MedBrains_Features.xlsx (Clinical sheet).

Features to mark Done (matched by partial text in the Feature column):
1. Queue filtering by department         -> "Doctor queue/worklist" (row with "queue" + "worklist")
2. Queue filtering by doctor             -> same row covers this (queue/worklist includes filters)
3. My patients toggle / filter           -> same row covers this (queue filtering includes my patients)
4. Walk-in token generation              -> "Walk-in token generation"
5. Visit type selection                  -> "Walk-in token generation" covers visit type (walk_in is a visit type)
6. Consultation templates / template picker -> "Specialty-specific casesheet templates"
7. Specialty-based templates             -> "Specialty-specific casesheet templates"
8. Current medications display / banner  -> "Auto-load previous history, allergies, current medications"
9. Auto-load current medications         -> "Auto-load previous history, allergies, current medications"

Note: Some requested features map to the same row because the Excel tracks them
as broader feature descriptions. Each matching row is updated once.
"""

import openpyxl
from pathlib import Path

EXCEL_PATH = Path("/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx")

# Partial-match patterns for features to mark as Done.
# Each tuple is (search_substring_lowercase, description_for_logging).
# Multiple patterns can match the same row; each row is only updated once.
FEATURE_PATTERNS = [
    ("queue/worklist", "Queue filtering by department/doctor + My patients toggle"),
    ("doctor queue", "Queue filtering (alternate match)"),
    ("walk-in token generation", "Walk-in token generation / visit type selection"),
    ("auto-load previous history", "Auto-load current medications / current medications display"),
    ("specialty-specific casesheet templates", "Consultation templates / specialty-based templates"),
]


def count_done_all_sheets(wb: openpyxl.Workbook) -> int:
    """Count total 'Done' features across all sheets."""
    total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Find Status column index from header row
        status_col = None
        for col_idx, cell in enumerate(ws[1], 0):
            if cell.value and str(cell.value).strip().lower() == "status":
                status_col = col_idx
                break
        if status_col is None:
            continue
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            if len(row) > status_col:
                val = row[status_col].value
                if val and str(val).strip() == "Done":
                    total += 1
    return total


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # Count Done before update
    done_before = count_done_all_sheets(wb)
    print(f"Total 'Done' features across ALL sheets BEFORE update: {done_before}")

    ws = wb["Clinical"]

    # Find column indices from header row
    headers = {}
    for col_idx, cell in enumerate(ws[1], 0):
        if cell.value:
            headers[str(cell.value).strip().lower()] = col_idx

    feature_col = headers.get("feature")
    status_col = headers.get("status")

    if feature_col is None or status_col is None:
        print(f"ERROR: Could not find required columns. Headers found: {headers}")
        return

    print(f"\nFeature column index: {feature_col}, Status column index: {status_col}")
    print(f"\nSearching Clinical sheet for OPD features to mark as Done...\n")

    updated_rows = set()
    updated_count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        feature_val = row[feature_col].value
        status_cell = row[status_col]

        if not feature_val or not isinstance(feature_val, str):
            continue

        feature_lower = feature_val.lower()
        row_num = row[0].row

        # Skip if already updated this row
        if row_num in updated_rows:
            continue

        for pattern, description in FEATURE_PATTERNS:
            if pattern in feature_lower:
                old_status = status_cell.value
                if old_status == "Done":
                    print(f"  SKIP (already Done) Row {row_num}: {feature_val}")
                else:
                    status_cell.value = "Done"
                    updated_rows.add(row_num)
                    updated_count += 1
                    print(f"  UPDATED Row {row_num}: '{feature_val}'")
                    print(f"    Status: {old_status} -> Done")
                    print(f"    Covers: {description}")
                break

    print(f"\nTotal rows updated: {updated_count}")

    # Save
    wb.save(EXCEL_PATH)
    print(f"\nSaved to {EXCEL_PATH}")

    # Reload and count Done after update
    wb2 = openpyxl.load_workbook(EXCEL_PATH)
    done_after = count_done_all_sheets(wb2)
    print(f"\nTotal 'Done' features across ALL sheets AFTER update: {done_after}")
    print(f"Net change: +{done_after - done_before}")


if __name__ == "__main__":
    main()

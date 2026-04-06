"""
Update MedBrains_Features.xlsx — Mark 16 Billing Phase 3 features as Done or Partial.

Targets the "Admin & Operations" sheet only.
Uses case-insensitive matching on column D (Feature).
Updates column G (Status).
"""

import openpyxl
import re

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"
SHEET_NAME = "Admin & Operations"
FEATURE_COL = 4  # Column D
STATUS_COL = 7   # Column G

# Billing section rows: 2 (module header "12. BILLING & FINANCE") through 58
# Only update rows within this range to avoid false matches in other modules.
BILLING_ROW_START = 2
BILLING_ROW_END = 58

# Rules: list of (description, match function, new_status)
RULES = [
    # --- Done ---
    (
        "OT billing / OT auto-billing",
        lambda f: bool(re.search(r"ot\s+billing|ot\s+auto[- ]?billing", f)),
        "Done",
    ),
    (
        "Bill print with GST",
        lambda f: "bill print" in f and "gst" in f,
        "Done",
    ),
    (
        "Multi-currency",
        lambda f: "multi-currency" in f or "multi currency" in f,
        "Done",
    ),
    (
        "Billing threshold",
        lambda f: "billing threshold" in f,
        "Done",
    ),
    (
        "Reimbursement claim documentation",
        lambda f: "reimbursement claim documentation" in f,
        "Done",
    ),
    (
        "Dual insurance",
        lambda f: "dual insurance" in f,
        "Done",
    ),
    (
        "Credit patient management",
        lambda f: "credit patient management" in f,
        "Done",
    ),
    (
        "TDS management / TDS deduction",
        lambda f: bool(re.search(r"\btds management\b|\btds deduction\b", f)),
        "Done",
    ),
    (
        "Financial MIS",
        lambda f: "financial mis" in f,
        "Done",
    ),
    (
        "Accounting journal / Journal entries / double-entry",
        lambda f: "accounting journal" in f or "journal entries" in f or "double-entry" in f,
        "Done",
    ),
    (
        "Bank reconciliation",
        lambda f: "bank reconciliation" in f,
        "Done",
    ),
    (
        "P&L by department / Profit and loss",
        lambda f: "p&l by department" in f
        or "profit & loss" in f
        or "profit and loss" in f
        or "profit & loss by department" in f,
        "Done",
    ),
    # --- Partial ---
    (
        "CGHS / ECHS / ESI / AB integration",
        lambda f: bool(
            re.search(r"\bcghs\b|\bechs\b|\besi\b|ayushman bharat", f)
        ),
        "Partial",
    ),
    (
        "NHCX",
        lambda f: "nhcx" in f,
        "Partial",
    ),
    (
        "GST compliance / GST return",
        lambda f: "gst compliance" in f or "gst return" in f,
        "Partial",
    ),
    (
        "ERP integration / ERP export",
        lambda f: "erp integration" in f or "erp export" in f,
        "Partial",
    ),
]


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # --- Revert accidental changes from previous run ---
    # Row 64 was incorrectly changed from "Done" to "Partial" (false ESI match).
    row_64_status = ws.cell(row=64, column=STATUS_COL).value
    if row_64_status == "Partial":
        ws.cell(row=64, column=STATUS_COL).value = "Done"
        print(f"REVERTED Row 64: Partial -> Done (false match from previous run)")

    updated = []
    matched_rules = set()

    for row in range(BILLING_ROW_START, BILLING_ROW_END + 1):
        feature_cell = ws.cell(row=row, column=FEATURE_COL)
        feature = feature_cell.value
        if not feature or not isinstance(feature, str):
            continue

        feature_lower = feature.lower()

        for idx, (description, matcher, new_status) in enumerate(RULES):
            if matcher(feature_lower):
                status_cell = ws.cell(row=row, column=STATUS_COL)
                old_status = status_cell.value
                status_cell.value = new_status
                matched_rules.add(idx)
                updated.append(
                    {
                        "row": row,
                        "feature": feature,
                        "old_status": old_status,
                        "new_status": new_status,
                        "rule": description,
                    }
                )
                break  # One rule per row

    # Save
    wb.save(EXCEL_PATH)

    # Report
    print(f"\n{'='*80}")
    print(f"Billing Phase 3 — Excel Update Report")
    print(f"{'='*80}")
    print(f"Sheet: {SHEET_NAME}")
    print(f"Total rows updated: {len(updated)}")
    print()

    done_count = sum(1 for u in updated if u["new_status"] == "Done")
    partial_count = sum(1 for u in updated if u["new_status"] == "Partial")
    print(f"  Done:    {done_count}")
    print(f"  Partial: {partial_count}")
    print()

    for u in updated:
        marker = "*" if u["old_status"] != u["new_status"] else " "
        print(
            f"  {marker} Row {u['row']:3d}: [{u['old_status']:>7s}] -> [{u['new_status']:>7s}]  "
            f"{u['feature'][:70]}"
        )

    # Check for unmatched rules
    unmatched = [
        RULES[i][0] for i in range(len(RULES)) if i not in matched_rules
    ]
    if unmatched:
        print(f"\nWARNING: {len(unmatched)} rule(s) did not match any row:")
        for desc in unmatched:
            print(f"  - {desc}")

    print(f"\nFile saved: {EXCEL_PATH}")
    print()


if __name__ == "__main__":
    main()

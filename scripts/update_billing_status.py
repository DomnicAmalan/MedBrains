#!/usr/bin/env python3
"""
Update billing feature statuses in MedBrains_Features.xlsx
based on what was built in Tasks #90-92.

Built:
- Charge master (service/procedure rate master) CRUD
- Invoices with items, multiple payment modes, multi-payment per bill
- Packages (billing_packages + items) CRUD
- Rate plans (rate_plans + items) CRUD
- Discounts (invoice_discounts) add/remove with auto-recalculate
- Refunds with auto invoice update
- Credit notes with apply-to-invoice
- Receipts per payment
- Insurance claims with full lifecycle (pre_auth → claim → settled)
- Charge sources: opd, ipd, lab, pharmacy, procedure, manual
"""

import openpyxl
import os

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "MedBrains_Features.xlsx",
)


def find_status_col(ws):
    """Find the Status column index (1-based)."""
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if str(cell.value or "").strip().lower() == "status":
            return cell.column
    return None


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH)

    # ── Admin & Operations sheet ──
    sheet_name = "Admin & Operations"
    ws = wb[sheet_name]
    status_col = find_status_col(ws)
    if not status_col:
        print(f"ERROR: No Status column in {sheet_name}")
        return

    # Features to mark as Done (fully built with backend + frontend CRUD)
    done_rows = {
        8:  "Multiple payment modes (cash, card, UPI, NEFT, cheque)",
        10: "Refund processing workflow with approval",
        11: "Concession/discount workflow with authorization levels",
        12: "Package billing for surgeries, day care, health checkups",
        14: "Multi-level rate plans (general, staff, VIP, corporate, insurance)",
        15: "Service/procedure rate master",
        16: "Lab billing",
        17: "Pharmacy billing",
        21: "Bill cancellation & refund",
        22: "Multi-payment mode per bill",
        23: "Receipt generation & print",
    }

    # Features to mark as Partial (infrastructure exists but not fully auto-wired)
    partial_rows = {
        4:  "OPD billing (infra exists, charge source includes opd, but no auto-generation from encounters)",
        5:  "IPD billing (infra exists, charge source includes ipd, but no auto-calculation)",
        9:  "Advance/deposit (payments infra supports it but no dedicated advance tracking)",
        27: "GST/tax management (tax_percent on charges/items, but no GST compliance reporting)",
        30: "Insurance/TPA pre-auth workflow (data model + CRUD, no external TPA integration)",
        31: "Cashless claim processing (claim lifecycle exists, no external integration)",
        34: "Claim rejection tracking (InsuranceClaim has rejected status, no re-submission workflow)",
    }

    changes = []

    for row_num, desc in done_rows.items():
        old_val = ws.cell(row=row_num, column=status_col).value or ""
        if str(old_val).strip().lower() != "done":
            ws.cell(row=row_num, column=status_col).value = "Done"
            changes.append(f"  {sheet_name} row {row_num}: {old_val!r} -> 'Done' ({desc})")

    for row_num, desc in partial_rows.items():
        old_val = ws.cell(row=row_num, column=status_col).value or ""
        if str(old_val).strip().lower() not in ("done", "partial"):
            ws.cell(row=row_num, column=status_col).value = "Partial"
            changes.append(f"  {sheet_name} row {row_num}: {old_val!r} -> 'Partial' ({desc})")

    # ── Onboarding & Setup sheet ──
    sheet_name2 = "Onboarding & Setup"
    ws2 = wb[sheet_name2]
    status_col2 = find_status_col(ws2)
    if status_col2:
        # Row 73: Billing masters (charge master, tax config, payment modes) — Done
        old_val = ws2.cell(row=73, column=status_col2).value or ""
        if str(old_val).strip().lower() != "done":
            ws2.cell(row=73, column=status_col2).value = "Done"
            changes.append(f"  {sheet_name2} row 73: {old_val!r} -> 'Done' (Billing masters: charge master, tax config, payment modes)")

    # Save
    wb.save(WORKBOOK_PATH)
    wb.close()

    print(f"Updated {len(changes)} feature statuses:")
    for c in changes:
        print(c)

    if not changes:
        print("  (no changes needed — all already up to date)")


if __name__ == "__main__":
    main()

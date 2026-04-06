#!/usr/bin/env python3
"""
Update Blood Bank feature statuses in MedBrains_Features.xlsx
based on what was built.

Blood Bank features are in "Diagnostics & Support" sheet, rows 175-204.
Section headers (merged rows) at 173, 174, 181, 189, 196 are skipped.

Feature rows:
  175-180: 8.1 Donor Management
  182-188: 8.2 Blood Processing & Storage
  190-195: 8.3 Cross-Match & Issue
  197-204: 8.4 Transfusion & Hemovigilance

Done (12 features):
- Row 175: Donor registration with demographics, blood group, ID verification
- Row 176: Donor eligibility/deferral criteria enforcement
- Row 177: Donation record (volume, date, type)
- Row 182: Component preparation (PRBC, FFP, platelets, cryo) tracking
- Row 183: Blood group and Rh typing with double-check
- Row 187: Expiry management with FEFO enforcement
- Row 188: Stock dashboard by component and blood group
- Row 190: Cross-match request from IPD/OT/ER
- Row 192: Cross-match result entry and compatibility determination
- Row 197: Bedside transfusion checklist (nurse documentation)
- Row 198: Transfusion reaction monitoring and reporting
- Row 202: Blood utilization tracking

Partial (8 features):
- Row 178: Donor adverse reaction documentation (recording exists, no full adverse event module)
- Row 180: Donor camp management (camp_name field, no full camp module)
- Row 184: TTI/infectious disease screening result tracking (tti_status field, no lab integration)
- Row 185: Bag labeling with barcode/ISBT 128 (unique constraint, no barcode scanning)
- Row 191: ABO/Rh compatibility verification (manual, no auto-compatibility engine)
- Row 193: Blood issue with double-identity verification (component_id on crossmatch, no auto-matching)
- Row 199: Hemovigilance report to NACO (reaction recording, no full SHOT/NACO reporting)
- Row 203: Blood discard management (status tracking, no analytics dashboard)

Remaining stay Pending:
- Row 179: Repeat donor tracking and recruitment management
- Row 186: Cold chain monitoring (refrigerator/freezer temperature IoT alerts)
- Row 194: Unused blood return workflow with temperature verification
- Row 195: Maximum Surgical Blood Order Schedule (MSBOS) enforcement
- Row 200: Lookback/traceback for post-donation infection detection
- Row 201: SBTC compliance reporting
- Row 204: Blood bank billing
"""

import openpyxl
import os

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "MedBrains_Features.xlsx",
)


def find_status_col(ws):
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if str(cell.value or "").strip().lower() == "status":
            return cell.column
    return None


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    changes = []

    ws = wb["Diagnostics & Support"]
    sc = find_status_col(ws)
    if not sc:
        print("ERROR: No Status column")
        return

    # Skip merged section header rows: 173, 174, 181, 189, 196

    # ── 8.1 Donor Management (rows 175-180) ──
    done_rows = {
        175: "Donor registration with demographics and medical history screening",
        176: "Donor deferral criteria enforcement (auto-check against history)",
        177: "Donation record (volume, date, type — whole blood/apheresis)",
    }

    partial_rows = {
        178: "Donor adverse reaction documentation (recording exists, no full adverse event module)",
        180: "Donor camp management (camp_name field, no full camp module)",
    }
    # Row 179 stays Pending (repeat donor tracking and recruitment management)

    # ── 8.2 Blood Processing & Storage (rows 182-188) ──
    done_rows.update({
        182: "Component preparation (PRBC, FFP, platelets, cryo) tracking",
        183: "Blood group and Rh typing with double-check",
        187: "Expiry management with FEFO enforcement",
        188: "Stock dashboard by component and blood group",
    })

    partial_rows.update({
        184: "Infectious disease screening result tracking (tti_status field, no lab integration)",
        185: "Bag labeling with barcode/ISBT 128 (unique constraint, no barcode scanning)",
    })
    # Row 186 stays Pending (cold chain monitoring)

    # ── 8.3 Cross-Match & Issue (rows 190-195) ──
    done_rows.update({
        190: "Cross-match request from IPD/OT/ER with electronic ordering",
        192: "Cross-match result entry and compatibility determination",
    })

    partial_rows.update({
        191: "ABO/Rh compatibility verification (manual, no auto-compatibility engine)",
        193: "Blood issue with double-identity verification (component_id on crossmatch, no auto-matching)",
    })
    # Rows 194-195 stay Pending (unused blood return, MSBOS enforcement)

    # ── 8.4 Transfusion & Hemovigilance (rows 197-204) ──
    done_rows.update({
        197: "Bedside transfusion checklist (nurse documentation)",
        198: "Transfusion reaction monitoring and reporting",
        202: "Blood utilization tracking",
    })

    partial_rows.update({
        199: "Hemovigilance report to NACO (reaction recording, no full SHOT/NACO reporting)",
        203: "Blood discard management (status tracking, no analytics dashboard)",
    })
    # Rows 200, 201, 204 stay Pending (lookback/traceback, SBTC compliance, blood bank billing)

    for row_num, desc in done_rows.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() != "done":
            ws.cell(row=row_num, column=sc).value = "Done"
            changes.append(f"  Row {row_num}: {old_val!r} -> 'Done' ({desc})")

    for row_num, desc in partial_rows.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() not in ("done", "partial"):
            ws.cell(row=row_num, column=sc).value = "Partial"
            changes.append(f"  Row {row_num}: {old_val!r} -> 'Partial' ({desc})")

    wb.save(WORKBOOK_PATH)
    wb.close()

    print(f"Updated {len(changes)} blood bank feature statuses:")
    for c in changes:
        print(c)

    if not changes:
        print("  (no changes needed)")


if __name__ == "__main__":
    main()

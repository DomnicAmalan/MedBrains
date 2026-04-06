#!/usr/bin/env python3
"""Update BME/CMMS feature statuses in MedBrains_Features.xlsx."""

import openpyxl

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"
SHEET_NAME = "Admin & Operations"

# Keywords to match features -> Done
DONE_KEYWORDS = [
    "comprehensive equipment database",
    "equipment categorization (critical/non-critical",
    "installation and commissioning records",
    "warranty tracking with vendor details",
    "equipment lifecycle tracking",
    "pm schedule generation",
    "pm task checklist per equipment type",
    "work order generation and assignment",
    "pm completion documentation with technician sign-off",
    "pm history per equipment",
    "calibration schedule per equipment",
    "calibration certificate storage",
    "third-party calibration vendor management",
    "amc/cmc contract tracking with validity",
    "vendor response time tracking (sla)",
    "vendor performance evaluation",
    "breakdown reporting",
    "priority classification (critical equipment",
    "downtime tracking per equipment",
    "repair documentation (spare parts used",
]

# Keywords to match features -> Partial
PARTIAL_KEYWORDS = [
    "asset tagging with qr/barcode",
    "pm compliance dashboard",
    "calibration due alert system",
    "out-of-calibration equipment auto-lock",
    "renewal alerts (30/60/90 days",
    "contract cost vs breakdown cost analysis",
    "mtbf tracking",
    "equipment uptime percentage reporting for nabh",
]


def match_feature(feature_text, keywords):
    """Return the matched keyword if feature_text contains any keyword."""
    lower = feature_text.lower()
    for kw in keywords:
        if kw in lower:
            return kw
    return None


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # Verify headers
    headers = [ws.cell(row=1, column=c).value for c in range(1, 12)]
    print(f"Headers: {headers}")
    assert headers[0] == "S.No", f"Expected 'S.No' in col 1, got '{headers[0]}'"
    assert headers[3] == "Feature", f"Expected 'Feature' in col 4, got '{headers[3]}'"
    assert headers[6] == "Status", f"Expected 'Status' in col 7, got '{headers[6]}'"

    STATUS_COL = 7
    FEATURE_COL = 4
    MODULE_COL = 2

    done_count = 0
    partial_count = 0
    done_matched = set()
    partial_matched = set()

    for row_idx in range(2, ws.max_row + 1):
        module_val = ws.cell(row=row_idx, column=MODULE_COL).value
        feature_val = ws.cell(row=row_idx, column=FEATURE_COL).value

        if not module_val or "BME/CMMS" not in str(module_val):
            continue
        if not feature_val:
            continue

        feature_str = str(feature_val)

        # Check for Done match
        kw = match_feature(feature_str, DONE_KEYWORDS)
        if kw:
            old_status = ws.cell(row=row_idx, column=STATUS_COL).value
            ws.cell(row=row_idx, column=STATUS_COL, value="Done")
            done_count += 1
            done_matched.add(kw)
            print(f"  Row {row_idx}: '{feature_str[:70]}' -> Done (was {old_status})")
            continue

        # Check for Partial match
        kw = match_feature(feature_str, PARTIAL_KEYWORDS)
        if kw:
            old_status = ws.cell(row=row_idx, column=STATUS_COL).value
            ws.cell(row=row_idx, column=STATUS_COL, value="Partial")
            partial_count += 1
            partial_matched.add(kw)
            print(f"  Row {row_idx}: '{feature_str[:70]}' -> Partial (was {old_status})")
            continue

        # No match - leave as-is
        current = ws.cell(row=row_idx, column=STATUS_COL).value
        print(f"  Row {row_idx}: '{feature_str[:70]}' -> UNCHANGED ({current})")

    # Check for unmatched keywords
    unmatched_done = set(DONE_KEYWORDS) - done_matched
    unmatched_partial = set(PARTIAL_KEYWORDS) - partial_matched
    if unmatched_done:
        print(f"\nWARNING: {len(unmatched_done)} Done keywords did not match any feature:")
        for kw in sorted(unmatched_done):
            print(f"  - {kw}")
    if unmatched_partial:
        print(f"\nWARNING: {len(unmatched_partial)} Partial keywords did not match any feature:")
        for kw in sorted(unmatched_partial):
            print(f"  - {kw}")

    print(f"\nSummary: {done_count} marked Done, {partial_count} marked Partial")

    wb.save(EXCEL_PATH)
    print(f"Saved to {EXCEL_PATH}")


if __name__ == "__main__":
    main()

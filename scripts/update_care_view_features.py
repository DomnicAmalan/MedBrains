"""
Update MedBrains_Features.xlsx — mark Care View / Ward Dashboard features as Done.

Rows 277-284 in the "Clinical" sheet (1-indexed, row 1 = header).
Status column is G (column 7).
"""

import openpyxl

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"
SHEET_NAME = "Clinical"
STATUS_COL = 7  # Column G
FEATURE_COL = 4  # Column D
ROWS = range(277, 285)  # 277-284 inclusive

EXPECTED_FEATURES = [
    "Ward-level patient grid",
    "Color-coded task urgency",
    "Medication administration pending list per nurse",
    "Vitals collection checklist",
    "Patient handover summary generation",
    "Real-time bed board with acuity scores and fall risk",
    "Discharge readiness tracker",
    "Non-medication task tracking",
]

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
ws = wb[SHEET_NAME]

# Verify header
header_status = ws.cell(row=1, column=STATUS_COL).value
assert header_status == "Status", f"Expected 'Status' in column G header, got '{header_status}'"
print(f"Header check passed: column G = '{header_status}'")

# Verify features and print before values
print("\nBefore update:")
for i, row in enumerate(ROWS):
    feature = ws.cell(row=row, column=FEATURE_COL).value
    status = ws.cell(row=row, column=STATUS_COL).value
    assert EXPECTED_FEATURES[i] in (feature or ""), (
        f"Row {row}: expected feature containing '{EXPECTED_FEATURES[i]}', got '{feature}'"
    )
    print(f"  Row {row}: Feature='{feature}' | Status='{status}'")

# Update status to Done
for row in ROWS:
    ws.cell(row=row, column=STATUS_COL).value = "Done"

# Print after values
print("\nAfter update:")
for row in ROWS:
    feature = ws.cell(row=row, column=FEATURE_COL).value
    status = ws.cell(row=row, column=STATUS_COL).value
    print(f"  Row {row}: Feature='{feature}' | Status='{status}'")

wb.save(EXCEL_PATH)
print(f"\nSaved {EXCEL_PATH} -- 8 features marked Done.")

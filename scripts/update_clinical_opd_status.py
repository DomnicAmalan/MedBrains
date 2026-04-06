#!/usr/bin/env python3
"""Update Clinical sheet OPD features status in MedBrains_Features.xlsx."""

import openpyxl
from pathlib import Path

EXCEL_PATH = Path(__file__).resolve().parent.parent / "MedBrains_Features.xlsx"

# Features to set as "Done" — keyword fragments to match in Feature column
DONE_KEYWORDS = [
    "drug interaction alerts, allergy cross-check, formulary enforcement",
    "Pre-authorization request (insurance patients)",
    "Auto-generation of billing for consultation + ordered investigations",
    "Drug-drug interaction alerts at prescription time",
    "Renal dose adjustment alerts based on creatinine/GFR",
    "Critical value alerts from lab directly on doctor",
    "Clinical protocol/guideline integration",
    "Antibiotic stewardship",
    "Review & reminder management",
    "PG logbook auto-entry",
    "Intern case logging with supervisor verification",
    "Medical student view-only access with anonymization",
    "Supervision hierarchy enforcement",
    "Co-signature workflow for resident orders",
]

# Features to set as "Partial" — keyword fragments
PARTIAL_KEYWORDS = [
    "SNOMED CT coded findings",
    "Multi-doctor appointment",
    "Voice-to-text option",
]

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Clinical"]

    # Verify column structure
    assert ws.cell(row=1, column=4).value == "Feature", "Column 4 should be Feature"
    assert ws.cell(row=1, column=7).value == "Status", "Column 7 should be Status"

    feature_col = 4
    status_col = 7
    changes = []

    for row in range(2, ws.max_row + 1):
        feature_text = ws.cell(row=row, column=feature_col).value
        if not feature_text:
            continue

        old_status = ws.cell(row=row, column=status_col).value

        # Check Done keywords
        for kw in DONE_KEYWORDS:
            if kw.lower() in feature_text.lower():
                ws.cell(row=row, column=status_col).value = "Done"
                changes.append((row, old_status, "Done", feature_text[:80]))
                break

        # Check Partial keywords
        for kw in PARTIAL_KEYWORDS:
            if kw.lower() in feature_text.lower():
                ws.cell(row=row, column=status_col).value = "Partial"
                changes.append((row, old_status, "Partial", feature_text[:80]))
                break

    # Print changes
    print(f"{'Row':<6} {'Old Status':<14} {'New Status':<12} Feature")
    print("-" * 110)
    for row, old, new, feat in changes:
        print(f"{row:<6} {str(old):<14} {new:<12} {feat}")

    print(f"\nTotal changes: {len(changes)}")

    # Count total Done in Clinical sheet
    done_count = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=status_col).value == "Done":
            done_count += 1

    print(f"Total 'Done' features in Clinical sheet: {done_count}")

    # Count across all sheets
    total_done = 0
    total_features = 0
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in range(2, sheet.max_row + 1):
            feat = sheet.cell(row=row, column=feature_col).value
            status = sheet.cell(row=row, column=status_col).value
            if feat and sheet.cell(row=row, column=1).value is not None:
                # Only count actual feature rows (with S.No or numeric value)
                sno = sheet.cell(row=row, column=1).value
                if isinstance(sno, (int, float)):
                    total_features += 1
                    if status == "Done":
                        total_done += 1

    print(f"Total 'Done' features across all sheets: {total_done} / {total_features}")

    wb.save(EXCEL_PATH)
    print(f"\nSaved: {EXCEL_PATH}")

if __name__ == "__main__":
    main()

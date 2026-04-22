#!/usr/bin/env python3
"""
Update CMS & Blog module feature statuses in MedBrains_Features.xlsx.

Based on implementation:
- Database migration 101_cms_blog.sql with comprehensive schema
- Rust core types in cms.rs
- 62 new API endpoints for CMS management
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "MedBrains_Features.xlsx")

# Features to mark as Done (S.No -> new status)
# Based on actual API endpoints implemented
DONE_FEATURES = {
    # Post Management
    1: "Done",   # Create/edit blog posts with markdown editor
    2: "Done",   # Post status workflow (draft/review/approved/published/archived)
    3: "Done",   # Content types (article, quick post, opinion, case study, news, event)
    4: "Done",   # Featured posts management
    5: "Done",   # Feature image upload with alt text and caption
    6: "Done",   # Auto-generate reading time
    7: "Done",   # Post revision history and version tracking
    8: "Done",   # Post slug auto-generation and editing
    9: "Done",   # Post excerpt/summary field
    10: "Done",  # Medical review workflow for clinical content
    11: "Done",  # Submit post for review
    12: "Done",  # Approve/reject post review
    13: "Done",  # Publish post with workflow enforcement
    14: "Done",  # Schedule post for future publication
    15: "Done",  # Archive/unarchive posts

    # Categories
    16: "Done",  # Category CRUD with hierarchical parent-child
    17: "Done",  # Category slug, description, color, icon
    18: "Done",  # Medical review requirement flag per category
    19: "Done",  # Category sort order management

    # Tags
    20: "Done",  # Tag CRUD with slug
    21: "Done",  # Tag description field
    22: "Done",  # Bulk tag management

    # Media Library
    23: "Done",  # Media upload (images, documents)
    24: "Done",  # Media gallery with grid/list view
    25: "Done",  # Image preview and metadata display
    26: "Done",  # Alt text management for accessibility
    27: "Done",  # Media delete with orphan check

    # Authors
    28: "Done",  # Author profile management (bio, credentials)
    29: "Done",  # Author social links (website, Twitter, LinkedIn)
    30: "Done",  # Author slug for public URLs
    31: "Done",  # Map CMS authors to HMS user accounts
    32: "Done",  # Author role assignment (admin, editor, author)

    # Subscribers
    33: "Done",  # Newsletter subscriber list
    34: "Done",  # Email subscription with confirmation flow
    35: "Done",  # Unsubscribe handling
    36: "Done",  # Export subscribers to CSV
    37: "Done",  # Subscriber management (delete, filter)

    # Analytics
    38: "Done",  # Dashboard with total views, posts, subscribers
    39: "Done",  # Per-post view analytics
    40: "Done",  # Top posts ranking

    # Public API
    41: "Done",  # Public API for post listing
    42: "Done",  # Public API for post detail
    43: "Done",  # Public API for featured posts
    44: "Done",  # Public API pagination
    45: "Done",  # Filter by category/tag/author

    # SEO
    46: "Done",  # Meta title and description per post
    47: "Done",  # Open Graph image support
    48: "Done",  # Canonical URL field
    49: "Done",  # Google Analytics integration field

    # Pages
    50: "Done",  # Static pages CRUD
    51: "Done",  # Page templates
    52: "Done",  # Page sort order

    # Site Settings
    53: "Done",  # Site title, tagline, description
    54: "Done",  # Logo and favicon URL
    55: "Done",  # Social media links (Twitter, Facebook, Instagram, YouTube, LinkedIn)
    56: "Done",  # Contact information (email, phone, address)
    57: "Done",  # Custom CSS/JS injection
    58: "Done",  # Posts per page setting

    # Menus
    59: "Done",  # Menu management
    60: "Done",  # Menu items with hierarchy
}

# Features to mark as Partial
PARTIAL_FEATURES = {
    61: "Partial",  # WYSIWYG editor - needs frontend implementation
    62: "Partial",  # Image optimization/resizing - needs backend processing
    63: "Partial",  # RSS feed generation - needs endpoint
    64: "Partial",  # Social sharing buttons - frontend only
    65: "Partial",  # Comment system - schema exists but disabled
    66: "Partial",  # Email newsletter sending - needs email service
    67: "Partial",  # Search within blog - needs full-text search
    68: "Partial",  # Related posts - needs algorithm
    69: "Partial",  # Author dashboard - needs dedicated page
    70: "Partial",  # Content scheduling queue - needs background job
    71: "Partial",  # Multi-language support - fields exist but no i18n
}

# Standard styling
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def main():
    wb = load_workbook(EXCEL_PATH)
    ws = wb["CMS & Blog"]

    done_count = 0
    partial_count = 0

    # Scan rows and update status
    for row_idx in range(2, ws.max_row + 1):
        sno_cell = ws.cell(row=row_idx, column=1)
        status_cell = ws.cell(row=row_idx, column=7)
        feature_cell = ws.cell(row=row_idx, column=4)

        sno = sno_cell.value

        # Check if this is a feature row (numeric S.No)
        if isinstance(sno, (int, float)) and sno == int(sno):
            sno_int = int(sno)
            old_status = status_cell.value or "Pending"
            new_status = None

            if sno_int in DONE_FEATURES:
                new_status = DONE_FEATURES[sno_int]
            elif sno_int in PARTIAL_FEATURES:
                # Only upgrade to Partial if not already Done
                if old_status != "Done":
                    new_status = PARTIAL_FEATURES[sno_int]

            if new_status and new_status != old_status:
                status_cell.value = new_status
                feature_desc = (feature_cell.value or "")[:50]
                print(f"  S.No {sno_int} (row {row_idx}): {old_status!r} -> {new_status!r}  [{feature_desc}]")

                if new_status == "Done":
                    done_count += 1
                elif new_status == "Partial":
                    partial_count += 1

                # Apply styling
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = THIN_BORDER
                    cell.alignment = WRAP_ALIGNMENT

    print(f"\nUpdated {done_count} features to Done")
    print(f"Updated {partial_count} features to Partial")

    if done_count > 0 or partial_count > 0:
        wb.save(EXCEL_PATH)
        print(f"Saved to {os.path.abspath(EXCEL_PATH)}")
    else:
        print("No changes to save.")

    # Calculate final stats
    total_features = 0
    final_done = 0
    final_partial = 0
    final_pending = 0

    for row_idx in range(2, ws.max_row + 1):
        sno = ws.cell(row=row_idx, column=1).value
        status = ws.cell(row=row_idx, column=7).value

        if isinstance(sno, (int, float)) and sno == int(sno):
            total_features += 1
            if status == "Done":
                final_done += 1
            elif status == "Partial":
                final_partial += 1
            else:
                final_pending += 1

    completion_pct = (final_done / total_features * 100) if total_features > 0 else 0

    print(f"\n=== CMS & Blog Module Status ===")
    print(f"Total features: {total_features}")
    print(f"Done: {final_done}")
    print(f"Partial: {final_partial}")
    print(f"Pending: {final_pending}")
    print(f"Completion: {completion_pct:.1f}%")


if __name__ == "__main__":
    main()

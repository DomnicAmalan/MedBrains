"""
Update CSSD module features status in MedBrains_Features.xlsx.

Marks 11 features as Done and 2 features as Partial based on
what was built in the CSSD module implementation.
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "MedBrains_Features.xlsx")

# Mapping from Excel feature text (substring) to target status.
# 11 Done + 2 Partial = 13 total (matching 13 CSSD rows in Excel).
#
# Done features:
#   - Instrument tracking & lifecycle management
#   - Instrument set management
#   - Sterilization load management (decontamination docs)
#   - Sterilization cycle monitoring (autoclave cycle logging)
#   - Chemical indicator recording
#   - Biological indicator recording
#   - Pack issuance tracking
#   - Recall management
#   - Flash sterilization documentation
#   - Instrument lifecycle and repair/sharpening tracking
#   - Sterilizer maintenance log
#
# Partial features:
#   - Complete traceability (barcode field exists, no scanner integration)
#   - CSSD workload and turnaround time reporting (no reporting dashboard yet)

CSSD_STATUS = {
    "Individual instrument tracking with barcode/RFID": "Done",
    "Instrument set composition management": "Done",
    "Decontamination documentation": "Done",
    "Autoclave cycle logging": "Done",
    "Chemical indicator tracking per load": "Done",
    "Biological indicator result tracking with lot linkage": "Done",
    "Sterile pack issuance to OT/ward with expiry tracking": "Done",
    "Recall capability if sterilization failure detected": "Done",
    "Flash sterilization documentation": "Done",
    "Complete traceability": "Partial",
    "Instrument lifecycle and repair/sharpening tracking": "Done",
    "CSSD workload and turnaround time reporting": "Partial",
    "Sterilizer maintenance log": "Done",
}


def find_column_by_header(ws, header_name):
    """Find the column index (1-based) by header name (case-insensitive)."""
    for cell in ws[1]:
        if cell.value and isinstance(cell.value, str) and cell.value.strip().lower() == header_name.lower():
            return cell.column
    return None


def match_feature(feature_text, status_map):
    """Find the status for a feature by matching against the map keys."""
    if not feature_text or not isinstance(feature_text, str):
        return None
    for key, status in status_map.items():
        if key.lower() in feature_text.lower():
            return status
    return None


def main():
    print(f"Loading workbook: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)

    updated = 0
    skipped = 0

    # CSSD features are in "Diagnostics & Support" sheet
    sheet_name = "Diagnostics & Support"
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found!")
        print(f"Available sheets: {wb.sheetnames}")
        return

    ws = wb[sheet_name]

    # Find required columns
    status_col = find_column_by_header(ws, "Status")
    module_col = find_column_by_header(ws, "Module")
    feature_col = find_column_by_header(ws, "Feature")

    if not status_col:
        print("ERROR: Could not find 'Status' column!")
        return
    if not module_col:
        print("ERROR: Could not find 'Module' column!")
        return
    if not feature_col:
        print("ERROR: Could not find 'Feature' column!")
        return

    print(f"Sheet: {sheet_name}")
    print(f"Module column: {module_col}, Feature column: {feature_col}, Status column: {status_col}")
    print("-" * 90)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        module_cell = row[module_col - 1]
        feature_cell = row[feature_col - 1]
        status_cell = row[status_col - 1]

        module_val = module_cell.value
        feature_val = feature_cell.value

        # Check if this row is a CSSD feature row
        if not module_val or not isinstance(module_val, str):
            continue
        if "cssd" not in module_val.lower():
            continue

        # Try to match the feature
        new_status = match_feature(feature_val, CSSD_STATUS)
        if new_status is None:
            print(f"  WARNING: Row {module_cell.row} - No match for feature: {feature_val}")
            skipped += 1
            continue

        old_status = status_cell.value or "None"
        status_cell.value = new_status

        # Apply color formatting for visual distinction
        if new_status == "Done":
            status_cell.font = Font(color="006100")        # Dark green text
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif new_status == "Partial":
            status_cell.font = Font(color="9C6500")        # Dark amber text
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        print(f"  Row {module_cell.row}: [{old_status} -> {new_status}] {feature_val}")
        updated += 1

    print("-" * 90)
    print(f"\nSummary: {updated} updated, {skipped} skipped")

    done_count = sum(1 for v in CSSD_STATUS.values() if v == "Done")
    partial_count = sum(1 for v in CSSD_STATUS.values() if v == "Partial")
    print(f"Expected: {done_count} Done, {partial_count} Partial = {done_count + partial_count} total")

    if updated > 0:
        wb.save(EXCEL_PATH)
        print(f"\nSaved to: {EXCEL_PATH}")
    else:
        print("\nNo changes to save.")


if __name__ == "__main__":
    main()

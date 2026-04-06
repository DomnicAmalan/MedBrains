#!/usr/bin/env python3
"""
Update Diet & Kitchen features status in MedBrains_Features.xlsx

Script updates rows 220-232 in the "Diagnostics & Support" sheet:
- Marks specified rows as "Done" with green formatting
- Marks specified rows as "Partial" with amber formatting
- Skips row 230 (already marked as Done)
"""

import openpyxl
from openpyxl.styles import PatternFill

# File path
FEATURES_FILE = '/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx'
SHEET_NAME = 'Diagnostics & Support'
STATUS_COL = 'G'  # Status column is G

# Define colors
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
AMBER_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Amber (using red-orange)

# Actually, let me use proper colors
# Green for Done: D4EDDA (light green)
# Amber for Partial: FFF3CD (light amber/yellow)
GREEN_FILL = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
AMBER_FILL = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')

# Rows to mark as "Done"
DONE_ROWS = [220, 221, 222, 223, 224, 225, 226, 227, 228, 231, 232]

# Rows to mark as "Partial"
PARTIAL_ROWS = [229]

# Load the workbook
print("Loading workbook...")
wb = openpyxl.load_workbook(FEATURES_FILE)
sheet = wb[SHEET_NAME]

# Update "Done" rows
print("\nUpdating 'Done' rows...")
for row_num in DONE_ROWS:
    cell = sheet[f'{STATUS_COL}{row_num}']
    feature = sheet[f'D{row_num}'].value
    cell.value = 'Done'
    cell.fill = GREEN_FILL
    print(f"  Row {row_num}: ✓ {feature[:60]}...")

# Update "Partial" rows
print("\nUpdating 'Partial' rows...")
for row_num in PARTIAL_ROWS:
    cell = sheet[f'{STATUS_COL}{row_num}']
    feature = sheet[f'D{row_num}'].value
    cell.value = 'Partial'
    cell.fill = AMBER_FILL
    print(f"  Row {row_num}: ◐ {feature[:60]}...")

# Save the workbook
print(f"\nSaving workbook...")
wb.save(FEATURES_FILE)
print(f"✓ Successfully updated {FEATURES_FILE}")

# Verify the changes
print("\nVerifying changes:")
wb_verify = openpyxl.load_workbook(FEATURES_FILE)
sheet_verify = wb_verify[SHEET_NAME]

for row_num in DONE_ROWS + PARTIAL_ROWS:
    status = sheet_verify[f'{STATUS_COL}{row_num}'].value
    feature = sheet_verify[f'D{row_num}'].value
    print(f"  Row {row_num}: Status = '{status}' | {feature[:50]}...")


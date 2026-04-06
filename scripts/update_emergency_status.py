"""
Update MedBrains_Features.xlsx — mark Emergency module features as Done or Partial.

Emergency module (Specialty & Academic sheet):
  - Triage: 4 Done, 2 Partial
  - Clinical Workflow: 2 Done, 4 Partial
  - MLC Mgmt: 3 Done, 3 Partial
  - Mass Casualty: 3 Done, 2 Partial
  Total: 12 Done, 11 Partial
"""

import openpyxl

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb["Specialty & Academic"]

# Find header row and column indices
header_row = None
for row in ws.iter_rows(min_row=1, max_row=5):
    for cell in row:
        if cell.value and str(cell.value).strip().lower() == "module":
            header_row = cell.row
            break
    if header_row:
        break

if not header_row:
    raise RuntimeError("Could not find header row with 'Module' column")

headers = {}
for cell in ws[header_row]:
    if cell.value:
        headers[str(cell.value).strip().lower()] = cell.column

mod_col = headers.get("module", 2)
sub_col = headers.get("sub-module", 3) or headers.get("sub module", 3)
feat_col = headers.get("feature", 4)
status_col = headers.get("status", 7)

print(f"Header row: {header_row}")
print(f"Columns — module={mod_col}, sub-module={sub_col}, feature={feat_col}, status={status_col}")
print()

# Keywords to match features — order matters (first match wins)
done_features = [
    "minimal data registration",
    "temporary uhid",
    "triage scoring",
    "mlc flag",
    "resuscitation bay allocation",
    "real-time medication",
    "mlc register",
    "wound certificate",
    "police intimation tracking",
    "code yellow",
    "batch registration",
    "resource tracking",
]

partial_features = [
    "auto-notification to casualty",
    "timer tracking",
    "code blue activation",
    "crash cart",
    "handover documentation",
    "seamless transition",
    "age estimation",
    "sexual assault",
    "court summons",
    "auto-notification to all",
    "triage tagging",
]

updated_done = 0
updated_partial = 0
skipped = []

for row in ws.iter_rows(min_row=header_row + 1):
    mod_val = str(row[mod_col - 1].value or "").strip().lower()
    feat_val = str(row[feat_col - 1].value or "").strip()

    if "emergency" not in mod_val:
        continue

    feat_lower = feat_val.lower()
    status = None

    for kw in done_features:
        if kw in feat_lower:
            status = "Done"
            break

    if not status:
        for kw in partial_features:
            if kw in feat_lower:
                status = "Partial"
                break

    if status:
        row[status_col - 1].value = status
        if status == "Done":
            updated_done += 1
        else:
            updated_partial += 1
        print(f"  [{status:7s}] {feat_val}")
    else:
        skipped.append(feat_val)

print()
if skipped:
    print(f"Skipped {len(skipped)} Emergency features (no keyword match):")
    for f in skipped:
        print(f"  - {f}")
    print()

print(f"Updated: {updated_done} Done, {updated_partial} Partial ({updated_done + updated_partial} total)")

wb.save(EXCEL_PATH)
print(f"Saved to {EXCEL_PATH}")

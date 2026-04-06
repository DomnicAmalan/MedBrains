#!/usr/bin/env python3
"""
Update MedBrains_Features.xlsx to mark recently completed features as Done/Partial.

Completed features (March 2026):
1. Navigation/Access Control — data-driven nav, per-tab settings permissions, i18n labels, spotlight from nav config
2. Clinical Masters — religions, occupations, relations, insurance providers CRUD
3. OPD Appointments — scheduling, booking, slot management, check-in, complete, cancel, no-show
4. Doctor Schedule Management — weekly schedule CRUD, schedule exceptions (holidays/leave)
5. Live Dashboard — real-time stats (patients, OPD queue, lab pending, revenue, appointments, IPD), recent activity feed
6. Patient Visit History/Timeline — patient detail page with tabs (overview, visits, lab orders, billing, appointments)
7. Patient Detail Page — demographics, allergies, visit summary, flags
"""

import openpyxl
from datetime import datetime

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

# Column indices (1-based)
COL_SNO = 1
COL_MODULE = 2
COL_SUBMOD = 3
COL_FEATURE = 4
COL_SOURCE = 5
COL_PRIORITY = 6
COL_STATUS = 7
COL_RFC = 8
COL_WEB = 9
COL_MOBILE = 10
COL_TV = 11


def get_row_text(ws, row_idx):
    """Get combined lowercase text from S.No, Module, Sub-Module, Feature columns."""
    parts = []
    for col in [COL_SNO, COL_MODULE, COL_SUBMOD, COL_FEATURE]:
        val = ws.cell(row=row_idx, column=col).value
        if val is not None:
            parts.append(str(val))
    return " ".join(parts).lower()


def get_feature_text(ws, row_idx):
    """Get the feature column text."""
    val = ws.cell(row=row_idx, column=COL_FEATURE).value
    return str(val).lower() if val else ""


def get_submod_text(ws, row_idx):
    """Get the sub-module column text."""
    val = ws.cell(row=row_idx, column=COL_SUBMOD).value
    return str(val).lower() if val else ""


def get_module_text(ws, row_idx):
    """Get the module column text."""
    val = ws.cell(row=row_idx, column=COL_MODULE).value
    return str(val).lower() if val else ""


def get_feature_raw(ws, row_idx):
    """Get the raw feature column text."""
    val = ws.cell(row=row_idx, column=COL_FEATURE).value
    return str(val) if val else ""


def is_feature_row(ws, row_idx):
    """Check if this row is a feature row (has a feature description and numeric S.No)."""
    feature = ws.cell(row=row_idx, column=COL_FEATURE).value
    sno = ws.cell(row=row_idx, column=COL_SNO).value
    if feature and sno is not None and isinstance(sno, (int, float)):
        return True
    return False


def update_status(ws, row_idx, new_status, reason):
    """Update the status cell and return a description of what was changed."""
    old_status = ws.cell(row=row_idx, column=COL_STATUS).value
    feature = get_feature_raw(ws, row_idx)
    module = ws.cell(row=row_idx, column=COL_MODULE).value or ""
    submod = ws.cell(row=row_idx, column=COL_SUBMOD).value or ""

    if old_status == new_status:
        return None  # Already correct

    ws.cell(row=row_idx, column=COL_STATUS).value = new_status

    desc = (
        f"  Row {row_idx}: [{old_status}] -> [{new_status}] "
        f"| {module}/{submod}: {feature} "
        f"({reason})"
    )
    return desc


def main():
    print(f"Opening {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    changes = []
    skipped_already_done = 0

    # =========================================================================
    # SHEET: Clinical
    # =========================================================================
    ws = wb["Clinical"]
    print(f"\nProcessing sheet: Clinical ({ws.max_row} rows)")

    for row_idx in range(2, ws.max_row + 1):
        if not is_feature_row(ws, row_idx):
            continue

        text = get_row_text(ws, row_idx)
        feature = get_feature_text(ws, row_idx)
        submod = get_submod_text(ws, row_idx)
        module = get_module_text(ws, row_idx)
        current_status = ws.cell(row=row_idx, column=COL_STATUS).value

        if current_status == "Done":
            skipped_already_done += 1
            continue

        # --- OPD Appointments ---
        if "opd" in module or "opd" in text:
            if "appointment" in submod:
                # Appointment booking (date, time, doctor, department)
                if "appointment booking" in feature and "date" in feature:
                    result = update_status(ws, row_idx, "Done", "OPD appointment booking implemented")
                    if result:
                        changes.append(result)

                # Doctor availability calendar (slot-based)
                elif "doctor availability calendar" in feature or "slot-based" in feature:
                    result = update_status(ws, row_idx, "Done", "Doctor availability calendar with slot management")
                    if result:
                        changes.append(result)

                # Appointment rescheduling
                elif "appointment rescheduling" in feature:
                    result = update_status(ws, row_idx, "Partial", "Reschedule via cancel+rebook, no direct reschedule UI yet")
                    if result:
                        changes.append(result)

                # Appointment cancellation with reason
                elif "appointment cancellation" in feature or "cancellation with reason" in feature:
                    result = update_status(ws, row_idx, "Done", "Cancel with reason implemented")
                    if result:
                        changes.append(result)

                # No-show tracking
                elif "no-show tracking" in feature:
                    result = update_status(ws, row_idx, "Done", "No-show status tracking implemented")
                    if result:
                        changes.append(result)

                # Walk-in token generation
                elif "walk-in token" in feature:
                    result = update_status(ws, row_idx, "Partial", "Walk-in patients supported, token system partial")
                    if result:
                        changes.append(result)

            # --- OPD Consultation ---
            if "consultation" in submod:
                # Patient history view (previous visits, labs, prescriptions)
                if "patient history view" in feature and "previous visits" in feature:
                    result = update_status(ws, row_idx, "Partial", "Visit history tab built in patient detail, consultation-inline view pending")
                    if result:
                        changes.append(result)

                # Patient timeline view during consultation
                elif "patient timeline view during consultation" in feature:
                    result = update_status(ws, row_idx, "Partial", "Timeline exists in patient detail, not yet inline in consultation")
                    if result:
                        changes.append(result)

                # Doctor queue/worklist
                elif "doctor queue" in feature or "worklist" in feature:
                    result = update_status(ws, row_idx, "Partial", "OPD queue via appointments, dedicated worklist UI pending")
                    if result:
                        changes.append(result)

                # Auto-load previous history, allergies
                elif "auto-load previous history" in feature and "allergies" in feature:
                    result = update_status(ws, row_idx, "Partial", "Allergies displayed in patient detail, auto-load in casesheet pending")
                    if result:
                        changes.append(result)

            # --- Clinical Decision Support ---
            if "clinical decision support" in submod:
                # Allergy alerts prominently displayed
                if "allergy alerts" in feature and "prominently displayed" in feature:
                    result = update_status(ws, row_idx, "Partial", "Allergies shown in patient detail/overview, real-time alerts pending")
                    if result:
                        changes.append(result)

        # --- Patient Management ---
        if "patient" in module:
            if "registration" in submod:
                # Patient history timeline (all visits, labs, prescriptions)
                if "patient history timeline" in feature:
                    result = update_status(ws, row_idx, "Done", "Patient detail page with visits/labs/billing/appointments tabs")
                    if result:
                        changes.append(result)

    # =========================================================================
    # SHEET: Onboarding & Setup
    # =========================================================================
    ws = wb["Onboarding & Setup"]
    print(f"\nProcessing sheet: Onboarding & Setup ({ws.max_row} rows)")

    for row_idx in range(2, ws.max_row + 1):
        if not is_feature_row(ws, row_idx):
            continue

        feature = get_feature_text(ws, row_idx)
        current_status = ws.cell(row=row_idx, column=COL_STATUS).value

        if current_status == "Done":
            skipped_already_done += 1
            continue

        # Timezone and locale configuration
        if "timezone and locale" in feature:
            result = update_status(ws, row_idx, "Done", "i18n locale store + timezone config implemented")
            if result:
                changes.append(result)

        # Currency and number format
        elif "currency and number format" in feature:
            result = update_status(ws, row_idx, "Partial", "Currency in country presets, full number formatting pending")
            if result:
                changes.append(result)

        # OPD masters: visit types, queue config, token format
        elif "opd masters" in feature and "visit types" in feature:
            result = update_status(ws, row_idx, "Partial", "Visit types exist in OPD, queue/token config pending")
            if result:
                changes.append(result)

        # Multi-language support setup (i18n configuration)
        elif "multi-language support setup" in feature or "i18n configuration" in feature:
            result = update_status(ws, row_idx, "Done", "react-i18next foundation, locale store, 11 namespaces")
            if result:
                changes.append(result)

        # Doctor profile setup (registration number, specialization, qualification)
        elif "doctor profile setup" in feature:
            result = update_status(ws, row_idx, "Partial", "Basic doctor info exists, full profile (reg no, qualifications) pending")
            if result:
                changes.append(result)

    # =========================================================================
    # SHEET: IT, Security & Infrastructure
    # =========================================================================
    ws = wb["IT, Security & Infrastructure"]
    print(f"\nProcessing sheet: IT, Security & Infrastructure ({ws.max_row} rows)")

    for row_idx in range(2, ws.max_row + 1):
        if not is_feature_row(ws, row_idx):
            continue

        feature = get_feature_text(ws, row_idx)
        module = get_module_text(ws, row_idx)
        submod = get_submod_text(ws, row_idx)
        current_status = ws.cell(row=row_idx, column=COL_STATUS).value

        if current_status == "Done":
            skipped_already_done += 1
            continue

        # --- Dashboard / Analytics features ---
        if "dashboards" in module or "dashboard" in module:
            # Admin dashboard (hospital overview)
            if "admin dashboard" in feature and "hospital overview" in feature:
                result = update_status(ws, row_idx, "Done", "Live dashboard: real-time stats (patients, OPD queue, lab, revenue, appointments, IPD)")
                if result:
                    changes.append(result)

            # OPD footfall analytics
            elif "opd footfall" in feature:
                result = update_status(ws, row_idx, "Partial", "OPD stats on live dashboard, detailed analytics pending")
                if result:
                    changes.append(result)

            # Revenue dashboard (daily/weekly/monthly)
            elif "revenue dashboard" in feature and "daily" in feature:
                result = update_status(ws, row_idx, "Partial", "Revenue stat on live dashboard, detailed revenue dashboard pending")
                if result:
                    changes.append(result)

            # Bed occupancy dashboard
            elif "bed occupancy dashboard" in feature:
                result = update_status(ws, row_idx, "Partial", "IPD bed stats on live dashboard, detailed bed dashboard pending")
                if result:
                    changes.append(result)

        # --- Admin Configuration ---
        if "admin" in module and "configuration" in submod:
            # System settings (date format, currency, timezone)
            if "system settings" in feature and ("date format" in feature or "currency" in feature or "timezone" in feature):
                result = update_status(ws, row_idx, "Partial", "Settings page with tabs, per-tab permissions, full config pending")
                if result:
                    changes.append(result)

        # --- Command Center ---
        if "command center" in module and "visibility" in submod:
            if "hospital-wide real-time dashboard" in feature:
                result = update_status(ws, row_idx, "Partial", "Live dashboard covers some metrics, full command center pending")
                if result:
                    changes.append(result)

        # --- Real-time KPI tiles ---
        if "command center" in module and "analytics" in submod:
            if "real-time kpi tiles" in feature:
                result = update_status(ws, row_idx, "Partial", "Live dashboard has real-time stat tiles, full KPI suite pending")
                if result:
                    changes.append(result)

    # =========================================================================
    # SHEET: Admin & Operations
    # =========================================================================
    ws = wb["Admin & Operations"]
    print(f"\nProcessing sheet: Admin & Operations ({ws.max_row} rows)")

    for row_idx in range(2, ws.max_row + 1):
        if not is_feature_row(ws, row_idx):
            continue

        feature = get_feature_text(ws, row_idx)
        module = get_module_text(ws, row_idx)
        submod = get_submod_text(ws, row_idx)
        current_status = ws.cell(row=row_idx, column=COL_STATUS).value

        if current_status == "Done":
            skipped_already_done += 1
            continue

        # --- Front Office / Queue ---
        if "front office" in module:
            # Real-time doctor availability display
            if "real-time doctor availability" in feature:
                result = update_status(ws, row_idx, "Done", "Doctor schedule with weekly availability, slot management")
                if result:
                    changes.append(result)

        # --- HR / Scheduling ---
        if "hr" in module:
            # Duty roster and shift swap management
            if "duty roster" in feature:
                result = update_status(ws, row_idx, "Partial", "Doctor weekly schedule CRUD built, full duty roster/shift swap pending")
                if result:
                    changes.append(result)

            # Leave management
            if "leave management" in feature:
                result = update_status(ws, row_idx, "Partial", "Doctor schedule exceptions (leave/holiday) built, full HR leave workflow pending")
                if result:
                    changes.append(result)

        # --- Scheduling / No-Show AI ---
        if "scheduling" in module:
            if "no-show analytics dashboard" in feature:
                result = update_status(ws, row_idx, "Partial", "No-show tracking in appointments, analytics dashboard pending")
                if result:
                    changes.append(result)

    # =========================================================================
    # Print summary
    # =========================================================================
    print("\n" + "=" * 80)
    print(f"FEATURE STATUS UPDATES -- {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 80)

    if changes:
        done_count = sum(1 for c in changes if "-> [Done]" in c)
        partial_count = sum(1 for c in changes if "-> [Partial]" in c)
        print(f"\nTotal updates: {len(changes)} ({done_count} Done, {partial_count} Partial)")
        print(f"Skipped (already Done): {skipped_already_done}")
        print()

        print("--- MARKED DONE ---")
        for change in changes:
            if "-> [Done]" in change:
                print(change)

        print("\n--- MARKED PARTIAL ---")
        for change in changes:
            if "-> [Partial]" in change:
                print(change)

        # Save
        print(f"\nSaving to {EXCEL_PATH}...")
        wb.save(EXCEL_PATH)
        print("Saved successfully.")
    else:
        print("\nNo changes needed -- all features already up to date.")

    # Final count
    print("\n" + "-" * 80)
    print("OVERALL STATUS COUNTS:")
    print("-" * 80)
    total_done = 0
    total_partial = 0
    total_pending = 0
    total_features = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_done = 0
        sheet_partial = 0
        sheet_pending = 0
        sheet_total = 0

        for row_idx in range(2, ws.max_row + 1):
            status = ws.cell(row=row_idx, column=COL_STATUS).value
            feature = ws.cell(row=row_idx, column=COL_FEATURE).value
            sno = ws.cell(row=row_idx, column=COL_SNO).value

            if feature and isinstance(sno, (int, float)):
                sheet_total += 1
                if status == "Done":
                    sheet_done += 1
                elif status == "Partial":
                    sheet_partial += 1
                elif status == "Pending":
                    sheet_pending += 1

        if sheet_total > 0:
            print(f"  {sheet_name:40s} Done={sheet_done:3d}  Partial={sheet_partial:3d}  Pending={sheet_pending:3d}  Total={sheet_total:3d}")
            total_done += sheet_done
            total_partial += sheet_partial
            total_pending += sheet_pending
            total_features += sheet_total

    print(f"  {'TOTAL':40s} Done={total_done:3d}  Partial={total_partial:3d}  Pending={total_pending:3d}  Total={total_features:3d}")
    pct = (total_done / total_features * 100) if total_features > 0 else 0
    print(f"\n  Completion: {total_done}/{total_features} features Done ({pct:.1f}%)")
    pct_with_partial = ((total_done + total_partial) / total_features * 100) if total_features > 0 else 0
    print(f"  Including Partial: {total_done + total_partial}/{total_features} ({pct_with_partial:.1f}%)")


if __name__ == "__main__":
    main()

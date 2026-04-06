#!/usr/bin/env python3
"""
Update MedBrains feature tracker — OPD features (March 2026).

Marks specific OPD features as Done or Partial based on implementation progress.
Uses substring matching to find features in the Clinical sheet.
"""

import openpyxl

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"
SHEET_NAME = "Clinical"

# Each rule: (substring to match in Feature column, target Status, expected Sub-Module or None)
# Sub-Module is used to disambiguate when a substring could match multiple features.
UPDATES = [
    {
        "match": "Investigation ordering",
        "sub_module": "Orders & Referrals",
        "new_status": "Done",
        "description": "Investigation ordering (lab/radiology)",
    },
    {
        "match": "Follow-up scheduling",
        "sub_module": "Orders & Referrals",
        "new_status": "Done",
        "description": "Follow-up scheduling with reminders",
    },
    {
        "match": "E-prescription with drug search",
        "sub_module": "Prescription",
        "new_status": "Partial",
        "description": "E-prescription with drug search",
    },
    {
        "match": "Drug formulary master",
        "sub_module": "Prescription",
        "new_status": "Partial",
        "description": "Drug formulary master",
    },
    {
        "match": "Prescription print",
        "sub_module": "Prescription",
        "new_status": "Done",
        "description": "Prescription print",
    },
    {
        "match": "Visit summary print",
        "sub_module": "Post-Consultation",
        "new_status": "Done",
        "description": "Visit summary print",
    },
    {
        # Drug interaction alerts -- confirm still Pending (no change, just verify)
        "match": "drug interaction alerts",
        "sub_module": "Clinical Decision Support",
        "new_status": None,  # None means verify only, don't change
        "description": "Drug-drug interaction alerts (verify Pending)",
    },
    {
        # Favourite prescriptions -- confirm still Pending (no change, just verify)
        "match": "Favourite prescriptions",
        "sub_module": "Prescription",
        "new_status": None,  # None means verify only, don't change
        "description": "Favourite prescriptions (verify Pending)",
    },
]


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        return

    ws = wb[SHEET_NAME]

    # Find header row and column indices
    header_row = 1
    col_indices = {}
    for cell in ws[header_row]:
        if cell.value:
            col_indices[cell.value.strip()] = cell.column - 1  # 0-based index

    required_cols = ["Feature", "Status", "Sub-Module"]
    for col_name in required_cols:
        if col_name not in col_indices:
            print(f"ERROR: Column '{col_name}' not found. Found: {list(col_indices.keys())}")
            return

    feat_idx = col_indices["Feature"]
    status_idx = col_indices["Status"]
    submod_idx = col_indices["Sub-Module"]

    print(f"Loaded '{SHEET_NAME}' sheet -- {ws.max_row} rows")
    print(f"Column indices: Feature={feat_idx}, Status={status_idx}, Sub-Module={submod_idx}")
    print()

    changes = []
    verifications = []

    for update in UPDATES:
        match_str = update["match"].lower()
        target_submod = update["sub_module"]
        new_status = update["new_status"]
        found = False

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            feature_val = row[feat_idx].value
            submod_val = row[submod_idx].value
            status_val = row[status_idx].value
            row_num = row[0].row

            if not feature_val or not isinstance(feature_val, str):
                continue

            # Substring match (case-insensitive)
            if match_str not in feature_val.lower():
                continue

            # Sub-module match (if specified)
            if target_submod and (not submod_val or target_submod.lower() not in submod_val.lower()):
                continue

            found = True

            if new_status is None:
                # Verify only -- report current status
                verifications.append({
                    "row": row_num,
                    "feature": feature_val,
                    "current_status": status_val,
                    "description": update["description"],
                })
            elif status_val != new_status:
                old_status = status_val
                row[status_idx].value = new_status
                changes.append({
                    "row": row_num,
                    "feature": feature_val,
                    "old_status": old_status,
                    "new_status": new_status,
                    "description": update["description"],
                })
            else:
                verifications.append({
                    "row": row_num,
                    "feature": feature_val,
                    "current_status": status_val,
                    "description": f"{update['description']} (already {new_status})",
                })
            break  # Only match first occurrence per rule

        if not found:
            print(f"WARNING: No match found for '{update['description']}' "
                  f"(substring='{update['match']}', sub_module='{target_submod}')")

    # Summary
    print("=" * 70)
    print("CHANGES MADE")
    print("=" * 70)
    if changes:
        for c in changes:
            print(f"  Row {c['row']:3d}: {c['old_status'] or 'None':10s} -> {c['new_status']:10s}  "
                  f"{c['feature'][:60]}")
    else:
        print("  (no changes)")

    print()
    print("=" * 70)
    print("VERIFICATIONS (no change needed)")
    print("=" * 70)
    if verifications:
        for v in verifications:
            print(f"  Row {v['row']:3d}: Status={v['current_status'] or 'None':10s}  "
                  f"{v['description']}")
    else:
        print("  (none)")

    if changes:
        wb.save(EXCEL_PATH)
        print()
        print(f"Saved {len(changes)} change(s) to {EXCEL_PATH}")
    else:
        print()
        print("No changes to save.")


if __name__ == "__main__":
    main()

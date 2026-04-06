#!/usr/bin/env python3
"""Mark Facilities Management (FMS) features in MedBrains_Features.xlsx."""

import openpyxl

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb["Admin & Operations"]

# Find the Status column index
status_col = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value and "Status" in str(ws.cell(row=1, column=col).value):
        status_col = col
        break

if not status_col:
    raise RuntimeError("Could not find Status column")

print(f"Status column found at index {status_col}")

# Done features (22) — rows in "Admin & Operations" sheet
done_rows = [
    235,  # PSA O2 plant monitoring (purity, pressure, flow, temperature)
    236,  # LMO tank level monitoring
    237,  # Cylinder manifold status monitoring
    238,  # Zone valve status per floor/department
    239,  # Pipeline pressure monitoring at key points
    240,  # Daily O2 consumption per department
    242,  # PESO compliance documentation and Drug License tracking
    244,  # Fire equipment inventory with QR code tagging
    245,  # Fire equipment inspection schedule and compliance tracking
    246,  # Mock drill management (quarterly per NABH)
    247,  # Fire NOC validity tracking with renewal alerts
    250,  # Water quality testing schedule and result tracking
    252,  # Legionella prevention
    253,  # Water tank cleaning schedule (6-monthly per NABH)
    254,  # STP treated water quality and reuse tracking
    256,  # Grid power, DG, UPS, solar monitoring dashboard
    257,  # DG runtime, fuel consumption, load percentage tracking
    258,  # UPS battery health and runtime monitoring
    262,  # Department maintenance request submission
    263,  # Work order assignment and tracking
    264,  # Completion documentation with sign-off
    265,  # Vendor service report management
]

# Partial features (6) — IoT/real-time/formal report deferred
partial_rows = [
    241,  # Gas purity analyzer integration — readings stored manually, IoT deferred
    248,  # Code Red activation workflow — drill type stored, real-time notification deferred
    251,  # Dialysis RO water monitoring — RO readings stored, IoT sensor deferred
    259,  # Power switchover test tracking — time stored, auto-monitoring deferred
    260,  # CEA safety compliance — readings available, formal report deferred
    266,  # Preventive maintenance schedule — via work orders, dedicated PM engine deferred
]

# Also mark the Regulatory & Compliance fire safety feature
ws_reg = wb["Regulatory & Compliance"]
reg_status_col = None
for col in range(1, ws_reg.max_column + 1):
    if ws_reg.cell(row=1, column=col).value and "Status" in str(ws_reg.cell(row=1, column=col).value):
        reg_status_col = col
        break

done_count = 0
partial_count = 0

for row in done_rows:
    old = ws.cell(row=row, column=status_col).value
    ws.cell(row=row, column=status_col).value = "Done"
    feature = ws.cell(row=row, column=4).value or ws.cell(row=row, column=5).value or ""
    print(f"  Row {row}: {old} -> Done  ({feature[:60]})")
    done_count += 1

for row in partial_rows:
    old = ws.cell(row=row, column=status_col).value
    ws.cell(row=row, column=status_col).value = "Partial"
    feature = ws.cell(row=row, column=4).value or ws.cell(row=row, column=5).value or ""
    print(f"  Row {row}: {old} -> Partial  ({feature[:60]})")
    partial_count += 1

# Mark fire safety compliance in Regulatory sheet as Done
if reg_status_col:
    old = ws_reg.cell(row=29, column=reg_status_col).value
    ws_reg.cell(row=29, column=reg_status_col).value = "Done"
    feature = ws_reg.cell(row=29, column=4).value or ws_reg.cell(row=29, column=5).value or ""
    print(f"  Regulatory Row 29: {old} -> Done  ({feature[:60]})")
    done_count += 1

wb.save(EXCEL_PATH)
print(f"\nDone: {done_count} features marked Done, {partial_count} features marked Partial")
print("Saved to", EXCEL_PATH)

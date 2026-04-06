"""
Update Infection Control module features status in MedBrains_Features.xlsx.

Marks features as Done or Partial based on implementation status.

Done: HAI surveillance (CLABSI/CAUTI), device-associated infection tracking,
  antibiotic stewardship request/review, hand hygiene audits, environmental
  culture surveillance, biomedical waste tracking (BMW Rules), needle stick
  injury reporting, outbreak management with contact tracing.

Partial: automated alerts/notifications, dashboard/analytics/reporting,
  real-time integration with lab/microbiology, antibiogram generation,
  antimicrobial resistance tracking, SSI bundle compliance, ventilator
  bundle compliance.

Covers rows across multiple sheets:
  - Admin & Operations: main Infection Control section (rows ~90-123)
  - Clinical: Infection Surveillance & Epidemiology (rows ~314-323),
             ICU bundle compliance, OT SSI correlation
  - Diagnostics & Support: microbiology antibiogram, antibiotic stewardship
  - Printing & Forms: IC & Safety Forms
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "MedBrains_Features.xlsx")

# ---- Admin & Operations: Infection Ctrl module (rows ~92-123) ----
ADMIN_OPS_IC_STATUS = {
    # === 14.1 HAI Surveillance ===
    # Done: CLABSI tracking with device-days
    "CLABSI (Central Line Associated BSI) tracking with device-days calculation": "Done",
    # Done: CAUTI tracking
    "CAUTI (Catheter Associated UTI) tracking": "Done",
    # Partial: VAP tracking — ventilator bundle compliance deferred
    "VAP (Ventilator Associated Pneumonia) tracking": "Partial",
    # Partial: SSI surveillance — SSI bundle compliance deferred
    "SSI (Surgical Site Infection) surveillance per procedure type": "Partial",
    # Partial: HAI rate calculation — benchmarking analytics deferred
    "HAI rate calculation per ICU/ward with benchmarking": "Partial",
    # Partial: dashboard with real-time alerts — alert engine deferred
    "Infection control dashboard with real-time alerts": "Partial",

    # === 14.2 Antibiotic Stewardship ===
    # Done: restricted antibiotic approval workflow
    "Restricted antibiotic approval workflow (prescriber": "Done",
    # Done: antibiotic escalation/de-escalation tracking
    "Antibiotic escalation/de-escalation tracking": "Done",
    # Partial: antibiogram generation — needs real-time lab integration
    "Antibiogram generation from microbiology data": "Partial",
    # Done: duration of therapy monitoring
    "Duration of antibiotic therapy monitoring": "Done",
    # Partial: DDD calculation — needs pharmacy integration for automated calc
    "DDD (Defined Daily Dose) calculation per department": "Partial",

    # === 14.3 Biomedical Waste (BMW Rules 2016) ===
    # Done: color-coded waste segregation tracking
    "Color-coded waste segregation tracking": "Done",
    # Done: weight-based waste generation monitoring
    "Weight-based waste generation monitoring": "Done",
    # Done: manifest system for waste transporter
    "Manifest system for waste transporter (challan generation)": "Done",
    # Done: sharp container deployment and fill-level tracking
    "Sharp container deployment and fill-level tracking": "Done",
    # Done: needle stick injury reporting
    "Needle stick injury reporting linked to specific container/ward": "Done",
    # Done: SPCB quarterly return Form IV generation
    "SPCB quarterly return (Form IV) generation": "Done",
    # Done: mercury management tracking
    "Mercury management and phased replacement tracking": "Done",

    # === 14.4 Hand Hygiene & Disinfection ===
    # Done: hand hygiene compliance monitoring (5 moments)
    "Hand hygiene compliance monitoring (5 moments of hand hygiene)": "Done",
    # Done: audit tool for hand hygiene observation
    "Audit tool for hand hygiene observation": "Done",
    # Done: endoscope reprocessing cycle tracking
    "Endoscope reprocessing (HLD) cycle tracking per scope serial number": "Done",
    # Done: scope culture surveillance result tracking
    "Scope culture surveillance result tracking": "Done",
    # Done: environmental culture tracking
    "Environmental culture tracking": "Done",

    # === 14.5 Outbreak Management ===
    # Partial: outbreak detection alerts — automated alert engine deferred
    "Outbreak detection alerts (unusual pathogen clustering)": "Partial",
    # Done: line listing generation for outbreak investigation
    "Line listing generation for outbreak investigation": "Done",
    # Done: contact tracing capability
    "Contact tracing capability": "Done",
    # Done: outbreak mode activation
    "Outbreak mode activation (visitor restriction, enhanced precautions)": "Done",
    # Done: HICC meeting trigger
    "HICC (Hospital Infection Control Committee) meeting trigger": "Done",
}

# ---- Clinical sheet: Infection Surveillance & Epidemiology (rows ~316-323) ----
CLINICAL_IC_STATUS = {
    # Partial: automated infection surveillance — needs real-time lab integration
    "Automated infection surveillance": "Partial",
    # Done: nosocomial infection classification
    "Nosocomial infection classification (CAUTI, CLABSI, SSI, VAP)": "Done",
    # Partial: infection trend analysis — analytics/reporting deferred
    "Infection trend analysis": "Partial",
    # Partial: antibiogram generation — needs hospital-wide micro data
    "Antibiogram generation": "Partial",
    # Partial: outbreak detection algorithm — automated alerts deferred
    "Outbreak detection algorithm": "Partial",
    # Partial: NHSN/CDC regulatory reporting — auto-generation deferred
    "NHSN/CDC regulatory reporting auto-generation": "Partial",
    # Done: line-day/device-day tracking
    "Line-day/device-day tracking for denominator calculation": "Done",
    # Done: contact tracing workflow
    "Contact tracing workflow": "Done",
    # Partial: SSI correlation with OT environment data — real-time integration deferred
    "Surgical site infection correlation with OT environment data": "Partial",
    # Partial: device-related infection rate auto-calculation — analytics deferred
    "Device-related infection rate auto-calculation": "Partial",
}

# ---- Printing & Forms: IC & Safety Forms (rows ~116-120) ----
PRINTING_IC_STATUS = {
    # Done: hand hygiene audit form
    "Hand Hygiene Audit Form (5 Moments observation)": "Done",
    # Done: BMW segregation log form
    "BMW Segregation Log (daily ward-wise weight)": "Done",
    # Done: BMW manifest/challan form
    "BMW Manifest / Challan (for transporter)": "Done",
    # Done: needle stick injury report form
    "Needle Stick Injury Report Form": "Done",
    # Done: HAI surveillance form
    "HAI Surveillance Form": "Done",
    # Done: restricted antibiotic request form
    "Restricted Antibiotic Request Form": "Done",
}


# --- Multi-sheet features that reference IC tangentially ---
# These are rows in other modules where Infection Control features are mentioned
# but belong to a different module. We update them only if clearly IC-related.
OTHER_SHEET_STATUS = {
    # Diagnostics & Support - Laboratory microbiology antibiogram
    ("Diagnostics & Support", "Laboratory", "culture sensitivity report, antibiogram"): "Partial",
    # Admin & Operations - Occ Health needle stick
    ("Admin & Operations", "Occ Health", "Needle stick / blood exposure incident protocol"): "Done",
    # Admin & Operations - Front Office outbreak mode
    ("Admin & Operations", "Front Office", "Outbreak mode"): "Done",
}


def find_column_by_header(ws, header_name):
    """Find the column index (1-based) by header name (case-insensitive)."""
    for cell in ws[1]:
        if cell.value and isinstance(cell.value, str) and cell.value.strip().lower() == header_name.lower():
            return cell.column
    return None


def match_feature(feature_text, status_map):
    """Find the status for a feature by matching against the map keys (substring)."""
    if not feature_text or not isinstance(feature_text, str):
        return None
    for key, status in status_map.items():
        if key.lower() in feature_text.lower():
            return status
    return None


def apply_status_formatting(status_cell, new_status):
    """Apply color formatting to status cell."""
    if new_status == "Done":
        status_cell.font = Font(color="006100")        # Dark green text
        status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif new_status == "Partial":
        status_cell.font = Font(color="9C6500")        # Dark amber text
        status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def update_module_rows(ws, module_col, feature_col, status_col, module_match, status_map, label,
                       submodule_col=None, submodule_match=None):
    """Update status for rows matching a module name in a given sheet.

    If submodule_col and submodule_match are provided, also filter by sub-module.
    """
    updated = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        module_cell = row[module_col - 1]
        feature_cell = row[feature_col - 1]
        status_cell = row[status_col - 1]

        module_val = module_cell.value
        feature_val = feature_cell.value

        if not module_val or not isinstance(module_val, str):
            continue

        # Check if the row's module matches what we're looking for
        module_lower = module_val.strip().lower()
        if not any(m.lower() in module_lower for m in module_match):
            continue

        # Optional sub-module filter
        if submodule_col is not None and submodule_match is not None:
            submod_val = row[submodule_col - 1].value
            if not submod_val or not isinstance(submod_val, str):
                continue
            submod_lower = submod_val.strip().lower()
            if not any(sm.lower() in submod_lower for sm in submodule_match):
                continue

        new_status = match_feature(feature_val, status_map)
        if new_status is None:
            print(f"  WARNING: Row {module_cell.row} - No match for: {feature_val}")
            skipped += 1
            continue

        old_status = status_cell.value or "None"
        status_cell.value = new_status
        apply_status_formatting(status_cell, new_status)

        marker = "*" if old_status != new_status else "="
        print(f"  {marker} Row {module_cell.row}: [{old_status} -> {new_status}] {feature_val}")
        updated += 1

    return updated, skipped


def update_other_sheet_rows(wb):
    """Update specific rows in other modules that reference IC features."""
    updated = 0

    for (sheet_name, module_match, feature_match), new_status in OTHER_SHEET_STATUS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet '{sheet_name}' not found for other-module update")
            continue

        ws = wb[sheet_name]
        module_col = find_column_by_header(ws, "Module")
        feature_col = find_column_by_header(ws, "Feature")
        status_col = find_column_by_header(ws, "Status")

        if not all([module_col, feature_col, status_col]):
            continue

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            mod_val = row[module_col - 1].value
            feat_val = row[feature_col - 1].value
            status_cell = row[status_col - 1]

            if not mod_val or not isinstance(mod_val, str):
                continue
            if not feat_val or not isinstance(feat_val, str):
                continue

            if module_match.lower() in mod_val.strip().lower() and feature_match.lower() in feat_val.lower():
                old_status = status_cell.value or "None"
                # Don't downgrade: if already Done, don't change to Partial
                if old_status == "Done" and new_status == "Partial":
                    print(f"  = Row {row[0].row} [{sheet_name}]: Already Done, skipping: {feat_val}")
                    continue
                status_cell.value = new_status
                apply_status_formatting(status_cell, new_status)
                marker = "*" if old_status != new_status else "="
                print(f"  {marker} Row {row[0].row} [{sheet_name}]: [{old_status} -> {new_status}] {feat_val}")
                updated += 1

    return updated


def main():
    print(f"Loading workbook: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)

    total_updated = 0
    total_skipped = 0

    # ---- 1. Admin & Operations: Infection Ctrl (main section) ----
    print("\n" + "=" * 100)
    print("SECTION 1: Admin & Operations - Infection Control (14.x)")
    print("=" * 100)

    sheet_name = "Admin & Operations"
    ws = wb[sheet_name]
    module_col = find_column_by_header(ws, "Module")
    feature_col = find_column_by_header(ws, "Feature")
    status_col = find_column_by_header(ws, "Status")

    if not all([module_col, feature_col, status_col]):
        print(f"ERROR: Missing columns in {sheet_name}!")
        return

    print(f"  Columns: Module={module_col}, Feature={feature_col}, Status={status_col}")

    u, s = update_module_rows(
        ws, module_col, feature_col, status_col,
        module_match=["Infection Ctrl"],
        status_map=ADMIN_OPS_IC_STATUS,
        label="Admin Ops IC",
    )
    total_updated += u
    total_skipped += s

    # ---- 2. Clinical: Infection Surveillance & Epidemiology ----
    print("\n" + "=" * 100)
    print("SECTION 2: Clinical - Infection Surveillance & Epidemiology (14.x)")
    print("=" * 100)

    sheet_name = "Clinical"
    ws = wb[sheet_name]
    module_col = find_column_by_header(ws, "Module")
    feature_col = find_column_by_header(ws, "Feature")
    status_col = find_column_by_header(ws, "Status")

    if not all([module_col, feature_col, status_col]):
        print(f"ERROR: Missing columns in {sheet_name}!")
        return

    print(f"  Columns: Module={module_col}, Feature={feature_col}, Status={status_col}")

    u, s = update_module_rows(
        ws, module_col, feature_col, status_col,
        module_match=["Infection Surv"],
        status_map=CLINICAL_IC_STATUS,
        label="Clinical IC Surv",
    )
    total_updated += u
    total_skipped += s

    # Also update ICU bundle compliance row and OT SSI row in Clinical
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        mod_val = row[module_col - 1].value
        feat_val = row[feature_col - 1].value
        status_cell = row[status_col - 1]

        if not feat_val or not isinstance(feat_val, str):
            continue
        if not mod_val or not isinstance(mod_val, str):
            continue

        new_status = None
        # ICU bundle compliance: device-related infection rate
        if mod_val.strip() == "ICU" and "Device-related infection rate" in feat_val:
            new_status = "Partial"
        # OT: SSI correlation
        elif mod_val.strip() == "IPD" and "Surgical site infection correlation" in feat_val:
            new_status = "Partial"

        if new_status:
            old_status = status_cell.value or "None"
            status_cell.value = new_status
            apply_status_formatting(status_cell, new_status)
            marker = "*" if old_status != new_status else "="
            print(f"  {marker} Row {row[0].row}: [{old_status} -> {new_status}] {feat_val}")
            total_updated += 1

    # ---- 3. Printing & Forms: IC & Safety Forms ----
    print("\n" + "=" * 100)
    print("SECTION 3: Printing & Forms - IC & Safety Forms")
    print("=" * 100)

    sheet_name = "Printing & Forms"
    ws = wb[sheet_name]
    module_col = find_column_by_header(ws, "Module")
    feature_col = find_column_by_header(ws, "Feature")
    status_col = find_column_by_header(ws, "Status")

    if not all([module_col, feature_col, status_col]):
        print(f"ERROR: Missing columns in {sheet_name}!")
        return

    submod_col = find_column_by_header(ws, "Sub-Module")
    print(f"  Columns: Module={module_col}, Sub-Module={submod_col}, Feature={feature_col}, Status={status_col}")

    u, s = update_module_rows(
        ws, module_col, feature_col, status_col,
        module_match=["Printing"],
        status_map=PRINTING_IC_STATUS,
        label="Printing IC Forms",
        submodule_col=submod_col,
        submodule_match=["IC & Safety"],
    )
    total_updated += u
    total_skipped += s

    # ---- 4. Other sheets: tangential IC references ----
    print("\n" + "=" * 100)
    print("SECTION 4: Other sheets - tangential IC references")
    print("=" * 100)

    u = update_other_sheet_rows(wb)
    total_updated += u

    # ---- Summary ----
    print("\n" + "=" * 100)
    print("SUMMARY")
    print("=" * 100)

    # Count expected
    all_maps = [ADMIN_OPS_IC_STATUS, CLINICAL_IC_STATUS, PRINTING_IC_STATUS]
    done_count = sum(1 for m in all_maps for v in m.values() if v == "Done")
    partial_count = sum(1 for m in all_maps for v in m.values() if v == "Partial")
    other_done = sum(1 for v in OTHER_SHEET_STATUS.values() if v == "Done")
    other_partial = sum(1 for v in OTHER_SHEET_STATUS.values() if v == "Partial")

    print(f"Total rows updated: {total_updated}")
    print(f"Total rows skipped (no match): {total_skipped}")
    print(f"IC core features: {done_count} Done, {partial_count} Partial")
    print(f"Other-module IC refs: {other_done} Done, {other_partial} Partial")
    print(f"Grand total expected: {done_count + partial_count + other_done + other_partial}")

    if total_updated > 0:
        wb.save(EXCEL_PATH)
        print(f"\nSaved to: {EXCEL_PATH}")
    else:
        print("\nNo changes to save.")


if __name__ == "__main__":
    main()

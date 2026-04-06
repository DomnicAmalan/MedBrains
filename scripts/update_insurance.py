#!/usr/bin/env python3
"""
Update Insurance & TPA module feature statuses in MedBrains_Features.xlsx.

Features are in the "Admin & Operations" sheet, rows 334-344 (Excel rows).
S.No values 282-291 in column A. Status column is G (column 7).

Row 338 is a sub-module header ("23.2 Prior Authorization Automation"), not a feature row.
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "MedBrains_Features.xlsx")

# Status updates keyed by S.No (column A integer value)
STATUS_UPDATES = {
    282: "Partial",   # Eligibility verification — no X12 270/271 external API
    283: "Done",      # Auto-verify at 3 touchpoints
    284: "Done",      # Benefits breakdown
    285: "Done",      # Policy status & coverage dates
    286: "Done",      # PA requirement detection
    287: "Partial",   # Electronic PA submission — no payer portal API
    288: "Done",      # PA status dashboard
    289: "Done",      # Auto-attach clinical docs
    290: "Done",      # Denial appeal workflow
    291: "Partial",   # PA TAT tracking — tracked but no automated escalation alerts
}

# Standard styling
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

MODULE_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBMODULE_HEADER_FILL = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")
BOLD_FONT = Font(bold=True)


def main():
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Admin & Operations"]

    updated_count = 0

    # Scan all rows to find matching S.No values
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        cell_a = row[0]  # Column A — S.No
        sno = cell_a.value

        # Check if this is a feature row with a matching S.No
        if isinstance(sno, int) and sno in STATUS_UPDATES:
            new_status = STATUS_UPDATES[sno]
            status_cell = ws.cell(row=cell_a.row, column=7)  # Column G = Status
            old_status = status_cell.value

            # Update status
            status_cell.value = new_status

            # Apply standard feature row styling to all columns
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=cell_a.row, column=col_idx)
                cell.border = THIN_BORDER
                cell.alignment = WRAP_ALIGNMENT

            feature_desc = ws.cell(row=cell_a.row, column=4).value or ""
            print(f"  S.No {sno} (row {cell_a.row}): {old_status!r} -> {new_status!r}  [{feature_desc[:60]}]")
            updated_count += 1

        # Also ensure module and sub-module headers have correct styling
        elif isinstance(sno, str) and "INSURANCE" in sno.upper():
            # Module header row (e.g., "23. INSURANCE VERIFICATION...")
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=cell_a.row, column=col_idx)
                cell.fill = MODULE_HEADER_FILL
                cell.font = BOLD_FONT
            print(f"  Header row {cell_a.row}: styled module header")

        elif isinstance(sno, str) and sno.strip().startswith("23."):
            # Sub-module header row (e.g., "  23.1 Electronic Insurance Verification")
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=cell_a.row, column=col_idx)
                cell.fill = SUBMODULE_HEADER_FILL
                cell.font = BOLD_FONT
            print(f"  Header row {cell_a.row}: styled sub-module header")

    print(f"\nUpdated {updated_count} feature rows.")

    if updated_count != len(STATUS_UPDATES):
        print(f"WARNING: Expected {len(STATUS_UPDATES)} updates but only found {updated_count}!")

    # Save
    wb.save(EXCEL_PATH)
    print(f"Saved to {os.path.abspath(EXCEL_PATH)}")


if __name__ == "__main__":
    main()

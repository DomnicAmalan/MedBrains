#!/usr/bin/env python3
"""
Update Inventory Phase 2 feature statuses in MedBrains_Features.xlsx.

24 software-buildable Partial features → Done in "Diagnostics & Support" sheet.
Row 153 (cold chain IoT) remains Partial.
"""

import openpyxl
import os

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "MedBrains_Features.xlsx",
)


def find_status_col(ws):
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if str(cell.value or "").strip().lower() == "status":
            return cell.column
    return None


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    changes = []

    ws = wb["Diagnostics & Support"]
    sc = find_status_col(ws)
    if not sc:
        print("ERROR: No Status column in 'Diagnostics & Support' sheet")
        wb.close()
        return

    # Inventory Phase 2 features to mark Done
    # Row 153 (cold chain IoT) intentionally excluded — remains Partial
    done_rows = {
        142: "Consumption analysis / department-wise issue tracking",
        143: "Dead stock identification",
        145: "Purchase vs consumption analysis",
        147: "Inventory valuation report",
        152: "ABC-VED-FSN analysis",
        154: "Compliance reporting (NMC/NABH checklist)",
        155: "Patient-level consumable tracking",
        156: "Patient billing integration for consumables",
        157: "Department-wise stock issue",
        158: "Reorder level alerts",
        159: "Consignment stock tracking",
        160: "Barcode/location-based stock management",
        161: "Stock return workflow",
        162: "Implant registry / tracking",
        163: "Equipment condemnation workflow",
        164: "Vendor performance tracking",
        165: "Vendor comparison tool",
        166: "Emergency purchase workflow",
        167: "Supplier payment tracking",
        168: "Rate contract management enhancements",
        169: "Batch/serial number tracking",
        170: "FEFO issue logic",
        171: "Supplier payment reconciliation",
        172: "Store catalog enhancements (VED class, bin location, HSN)",
    }

    for row_num, desc in done_rows.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() != "done":
            ws.cell(row=row_num, column=sc).value = "Done"
            changes.append(f"  Row {row_num}: {old_val!r} -> 'Done' ({desc})")

    wb.save(WORKBOOK_PATH)
    wb.close()

    print(f"Updated {len(changes)} Inventory Phase 2 feature statuses:")
    for c in changes:
        print(c)

    if not changes:
        print("  (no changes needed — all already marked Done)")


if __name__ == "__main__":
    main()

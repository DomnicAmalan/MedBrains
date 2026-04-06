#!/usr/bin/env python3
"""
Update IPD Clinical and OT (Operation Theatre) feature statuses in MedBrains_Features.xlsx.

This script marks features as Done or Partial based on the completed implementation
of IPD clinical features (progress notes, assessments, MAR, care plans, etc.) and
OT features (booking, scheduling, WHO checklist, anesthesia records, etc.).

Only updates features in sheets that have IPD or OT sub-module rows.
Does NOT downgrade features already marked as Done.
"""

import os
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

# ──────────────────────────────────────────────────────────────────────────────
# Feature matching rules
# Each rule is a dict with:
#   "keywords": list of keyword groups (ANY group match = feature matches)
#                each group is a list of strings that must ALL be present (AND)
#   "submodule_filter": optional — only match if Sub-Module column contains this
#   "module_filter": optional — only match if Module column contains this
# ──────────────────────────────────────────────────────────────────────────────

# IPD Clinical features to mark as DONE
IPD_DONE_RULES = [
    # Progress Notes / SOAP documentation
    {
        "keywords": [
            ["progress note"],
            ["daily progress"],
            ["doctor rounds note"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    # Clinical assessments / scoring — Morse Fall Scale
    {
        "keywords": [
            ["morse fall"],
            ["fall risk assessment"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    # Braden Scale
    {
        "keywords": [
            ["braden scale"],
            ["pressure ulcer risk"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    # Glasgow Coma Scale
    {
        "keywords": [
            ["glasgow coma"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    # Pain assessment scoring
    {
        "keywords": [
            ["pain assessment"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    # Medication Administration Record (MAR) / 5 Rights
    {
        "keywords": [
            ["medication administration record"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    # Intake/Output chart
    {
        "keywords": [
            ["intake/output"],
            ["i/o", "chart"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    # Nursing assessment on admission (head-to-toe)
    {
        "keywords": [
            ["initial nursing assessment"],
            ["nursing assessment", "admission"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Care plans (nursing diagnosis, goals, interventions)
    {
        "keywords": [
            ["nursing care plan"],
            ["care plan creation"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Handover/Handoff reports (ISBAR/SBAR)
    {
        "keywords": [
            ["sbar"],
            ["handover", "format"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Discharge checklist
    {
        "keywords": [
            ["discharge checklist"],
        ],
        "submodule_filter": "Discharge",
    },
    # Bed management dashboard / bed status tracking
    {
        "keywords": [
            ["bed occupancy dashboard"],
            ["real-time bed occupancy"],
        ],
        "submodule_filter": "Ward & Bed",
    },
    # Digital case sheet entry
    {
        "keywords": [
            ["digital case sheet"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    # Vitals monitoring (periodic charting) with graphical trending
    {
        "keywords": [
            ["vitals monitoring"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    # Medication administration timestamp with nurse ID
    {
        "keywords": [
            ["medication administration timestamp"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Nursing notes
    {
        "keywords": [
            ["nursing notes"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
]

# IPD features to mark as PARTIAL
IPD_PARTIAL_RULES = [
    # Automated alerts/notifications (deferred)
    {
        "keywords": [
            ["critical patient flagging"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Advanced bed management (housekeeping integration)
    {
        "keywords": [
            ["bed cleaning status"],
        ],
        "submodule_filter": "Ward & Bed",
    },
    {
        "keywords": [
            ["bed turnaround"],
        ],
        "submodule_filter": "Ward & Bed",
    },
    {
        "keywords": [
            ["isolation bed"],
        ],
        "submodule_filter": "Ward & Bed",
    },
    {
        "keywords": [
            ["bed blocking"],
            ["bed reservation"],
        ],
        "submodule_filter": "Ward & Bed",
    },
    # Clinical decision support / high-alert medication
    {
        "keywords": [
            ["high-alert medication"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Barcode medication verification (hardware integration)
    {
        "keywords": [
            ["barcode scan", "medication"],
            ["barcode scanning", "medication"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Investigation result inline display
    {
        "keywords": [
            ["investigation result", "inline"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    # Diet chart management
    {
        "keywords": [
            ["diet chart management"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    # Wound care documentation
    {
        "keywords": [
            ["wound care documentation"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    # PRN medication documentation
    {
        "keywords": [
            ["prn medication"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Missed dose documentation
    {
        "keywords": [
            ["missed dose"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Central line / catheter care checklists (partial via ICU bundle checks)
    {
        "keywords": [
            ["central line care"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    {
        "keywords": [
            ["catheter care checklist"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    # Restraint documentation
    {
        "keywords": [
            ["restraint documentation"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    # Transfusion documentation
    {
        "keywords": [
            ["transfusion documentation"],
        ],
        "submodule_filter": "Clinical Inpatient",
    },
    # Nurse call system (deferred)
    {
        "keywords": [
            ["nurse call system"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Escalation if nurse call unresponded
    {
        "keywords": [
            ["escalation", "nurse call"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Patient identification wristband verification
    {
        "keywords": [
            ["patient identification wristband"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Consent verification before procedure
    {
        "keywords": [
            ["consent verification before procedure"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Blood transfusion checklist
    {
        "keywords": [
            ["blood transfusion checklist"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Incident reporting (anonymous)
    {
        "keywords": [
            ["incident reporting", "anonymous"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Restraint monitoring log
    {
        "keywords": [
            ["restraint monitoring log"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Elopement risk assessment
    {
        "keywords": [
            ["elopement risk"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # OT nursing — surgical safety checklist
    {
        "keywords": [
            ["ot nursing"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Dialysis nursing
    {
        "keywords": [
            ["dialysis nursing"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Endoscopy nursing / sedation / Aldrete
    {
        "keywords": [
            ["endoscopy nursing"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Chemotherapy administration
    {
        "keywords": [
            ["chemotherapy administration"],
        ],
        "submodule_filter": "Nursing Services",
    },
    # Admission checklist validation
    {
        "keywords": [
            ["admission checklist validation"],
        ],
        "submodule_filter": "Admission",
    },
    # Attender/next-of-kin capture
    {
        "keywords": [
            ["attender", "next-of-kin"],
        ],
        "submodule_filter": "Admission",
    },
    # Admission slip/form print
    {
        "keywords": [
            ["admission slip"],
        ],
        "submodule_filter": "Admission",
    },
    # Estimated cost for admission
    {
        "keywords": [
            ["estimated cost", "admission"],
        ],
        "submodule_filter": "Admission",
    },
    # Advance payment collection
    {
        "keywords": [
            ["advance payment"],
        ],
        "submodule_filter": "Admission",
    },
    # Pre-authorization for insurance
    {
        "keywords": [
            ["pre-authorization", "insurance"],
        ],
        "submodule_filter": "Admission",
    },
    # MLC registration at admission
    {
        "keywords": [
            ["mlc registration"],
        ],
        "submodule_filter": "Admission",
    },
    # Nursery management
    {
        "keywords": [
            ["nursery management"],
        ],
        "submodule_filter": "Ward & Bed",
    },
    # TV display for bed status
    {
        "keywords": [
            ["tv display", "bed status"],
        ],
        "submodule_filter": "Ward & Bed",
    },
    # Death summary auto-generation
    {
        "keywords": [
            ["death summary"],
        ],
        "submodule_filter": "Discharge",
    },
    # Death certificate generation
    {
        "keywords": [
            ["death certificate"],
        ],
        "submodule_filter": "Discharge",
    },
    # Birth certificate generation
    {
        "keywords": [
            ["birth certificate"],
        ],
        "submodule_filter": "Discharge",
    },
    # Dept-wise billing at discharge
    {
        "keywords": [
            ["dept-wise billing"],
        ],
        "submodule_filter": "Discharge",
    },
    # Billing threshold control
    {
        "keywords": [
            ["billing threshold"],
        ],
        "submodule_filter": "Discharge",
    },
]

# OT features to mark as DONE
OT_DONE_RULES = [
    # OT scheduling calendar with slot management
    {
        "keywords": [
            ["ot scheduling calendar"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Emergency OT booking
    {
        "keywords": [
            ["emergency ot booking"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Pre-operative assessment checklist
    {
        "keywords": [
            ["pre-operative assessment"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Pre-anesthesia check-up (PAC)
    {
        "keywords": [
            ["pre-anesthesia check"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # WHO Surgical Safety Checklist
    {
        "keywords": [
            ["who surgical safety"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Anesthesia record
    {
        "keywords": [
            ["anesthesia record"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Surgeon's operative note template
    {
        "keywords": [
            ["operative note"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Post-Anesthesia Care Unit (PACU) monitoring (Aldrete)
    {
        "keywords": [
            ["pacu"],
            ["aldrete"],
            ["post-anesthesia care"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # OT preparation checklist
    {
        "keywords": [
            ["ot preparation checklist"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Post-operative order set generation
    {
        "keywords": [
            ["post-operative order"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Handover from OT to ward/ICU
    {
        "keywords": [
            ["handover from ot"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Surgical site marking verification
    {
        "keywords": [
            ["surgical site marking"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Blood reservation request to blood bank
    {
        "keywords": [
            ["blood reservation request"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Instrument and sponge count
    {
        "keywords": [
            ["instrument", "sponge count"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Specimen collection documentation
    {
        "keywords": [
            ["specimen collection"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Blood product usage documentation
    {
        "keywords": [
            ["blood product usage"],
        ],
        "submodule_filter": "Operation Theatre",
    },
]

# OT features to mark as PARTIAL
OT_PARTIAL_RULES = [
    # OT utilization report
    {
        "keywords": [
            ["ot utilization"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Surgeon-wise case load analysis
    {
        "keywords": [
            ["surgeon-wise"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Implant/prosthesis tracking
    {
        "keywords": [
            ["implant", "tracking"],
            ["implant", "prosthesis"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # OT consumable tracking
    {
        "keywords": [
            ["ot consumable"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # OT turnaround time tracking
    {
        "keywords": [
            ["ot turnaround"],
        ],
        "submodule_filter": "Operation Theatre",
    },
    # Anesthesia complication tracking
    {
        "keywords": [
            ["anesthesia complication"],
        ],
        "submodule_filter": "Operation Theatre",
    },
]


def normalize(text):
    """Lowercase and strip a string for comparison."""
    if text is None:
        return ""
    return str(text).lower().strip()


def matches_keywords(feature_text, keyword_groups):
    """
    Check if feature_text matches ANY of the keyword groups.
    Each group is a list of strings that must ALL appear in the text (AND within group).
    Groups are OR'd together.
    """
    feat_lower = normalize(feature_text)
    if not feat_lower:
        return False

    for group in keyword_groups:
        if all(kw.lower() in feat_lower for kw in group):
            return True
    return False


def matches_filter(cell_value, filter_value):
    """Check if a cell value contains the filter string (case-insensitive)."""
    if filter_value is None:
        return True
    return normalize(filter_value) in normalize(cell_value)


def find_column(ws, name):
    """Find a column index (1-based) by header name."""
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
        for cell in row:
            if cell.value and normalize(cell.value) == normalize(name):
                return cell.column
    return None


def apply_rules(ws, rules, target_status, module_col, submod_col, feature_col,
                status_col, report, sheet_name):
    """Apply a list of matching rules to a worksheet, updating status where matched."""
    for row_idx in range(2, ws.max_row + 1):
        module_val = ws.cell(row=row_idx, column=module_col).value
        submod_val = ws.cell(row=row_idx, column=submod_col).value
        feature_val = ws.cell(row=row_idx, column=feature_col).value
        status_cell = ws.cell(row=row_idx, column=status_col)
        current_status = normalize(status_cell.value)

        # Skip empty rows
        if not feature_val:
            continue

        # Never downgrade Done to Partial
        if current_status == "done" and target_status == "Partial":
            continue

        # Don't re-set something already at the target status
        if current_status == normalize(target_status):
            continue

        for rule in rules:
            # Check module filter
            module_filter = rule.get("module_filter")
            if module_filter and not matches_filter(module_val, module_filter):
                continue

            # Check sub-module filter
            submod_filter = rule.get("submodule_filter")
            if submod_filter and not matches_filter(submod_val, submod_filter):
                continue

            # Check feature keywords
            if matches_keywords(str(feature_val), rule["keywords"]):
                old_status = status_cell.value or "None"
                status_cell.value = target_status
                report.append({
                    "sheet": sheet_name,
                    "row": row_idx,
                    "module": module_val,
                    "sub_module": submod_val,
                    "feature": str(feature_val)[:80],
                    "old_status": old_status,
                    "new_status": target_status,
                })
                break  # Only match first rule per row


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: File not found: {EXCEL_PATH}")
        sys.exit(1)

    print(f"Loading {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    report = []

    # Process ALL sheets — IPD/OT features might appear in multiple sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Find column indices
        status_col = find_column(ws, "Status")
        module_col = find_column(ws, "Module")
        submod_col = find_column(ws, "Sub-Module")
        feature_col = find_column(ws, "Feature")

        if not all([status_col, module_col, submod_col, feature_col]):
            # Some sheets may not have standard headers — skip
            continue

        # Apply Done rules first (higher priority), then Partial
        # IPD Done
        apply_rules(ws, IPD_DONE_RULES, "Done",
                     module_col, submod_col, feature_col, status_col,
                     report, sheet_name)
        # OT Done
        apply_rules(ws, OT_DONE_RULES, "Done",
                     module_col, submod_col, feature_col, status_col,
                     report, sheet_name)
        # IPD Partial
        apply_rules(ws, IPD_PARTIAL_RULES, "Partial",
                     module_col, submod_col, feature_col, status_col,
                     report, sheet_name)
        # OT Partial
        apply_rules(ws, OT_PARTIAL_RULES, "Partial",
                     module_col, submod_col, feature_col, status_col,
                     report, sheet_name)

    # Print report
    print()
    print("=" * 110)
    print("UPDATE REPORT -- IPD Clinical & OT Features")
    print("=" * 110)

    if not report:
        print("\nNo features were updated.")
    else:
        done_count = sum(1 for r in report if r["new_status"] == "Done")
        partial_count = sum(1 for r in report if r["new_status"] == "Partial")

        print(f"\nTotal updated: {len(report)} features ({done_count} Done, {partial_count} Partial)")

        # Group by status
        for status_label in ["Done", "Partial"]:
            entries = [r for r in report if r["new_status"] == status_label]
            if not entries:
                continue

            print(f"\n--- Marked as {status_label} ({len(entries)} features) ---")
            print(f"{'Sheet':<25} {'Row':<5} {'Sub-Module':<25} {'Feature':<60} {'Was':<10}")
            print("-" * 125)
            for r in entries:
                print(
                    f"{r['sheet']:<25} {r['row']:<5} "
                    f"{str(r['sub_module'] or ''):<25} "
                    f"{r['feature']:<60} "
                    f"{r['old_status']:<10}"
                )

    # Save
    print(f"\nSaving to {EXCEL_PATH}...")
    wb.save(EXCEL_PATH)
    print("File saved successfully.")

    # Summary
    print(f"\nSummary: {len(report)} features updated in MedBrains_Features.xlsx")


if __name__ == "__main__":
    main()

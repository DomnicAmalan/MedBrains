#!/usr/bin/env python3
"""
Update IPD Phase 2b feature statuses in MedBrains_Features.xlsx.

IPD Phase 2b (migration 076) built:
- IP type configuration & ward type management
- Admission checklist validation
- Bed blocking/reservation system
- Bed turnaround time tracking
- Isolation bed flagging
- Clinical documentation (wound care, central line, catheter, drain tracking)
- Restraint monitoring (30-min intervals)
- PRN/missed dose documentation
- High-alert medication double-check flag
- Critical patient flagging
- Inter-department transfer with clinical summary
- Death summary & certificate generation
- Birth record & certificate generation
- Discharge TAT tracking
- OT consumable tracking
- OT utilization reports
- Surgical site infection tracking
"""

import openpyxl
import os

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "MedBrains_Features.xlsx",
)


def find_status_col(ws):
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if str(cell.value or "").strip().lower() == "status":
            return cell.column
    return None


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    changes = []

    ws = wb["Clinical"]
    sc = find_status_col(ws)
    if not sc:
        print("ERROR: No Status column in Clinical sheet")
        return

    # IPD Phase 2b features to mark Done (were Partial)
    done_rows = {
        # Admission
        117: "OPD/ER admission with bed selection (available beds API + admit-from-OPD)",
        118: "Auto-transfer of OPD history to IPD (vitals/diagnoses/prescriptions copy)",
        119: "Admission checklist validation (admission_checklists table + CRUD)",
        122: "IP type configuration & ward type management (ip_type_configurations + wards)",
        # Ward & Bed Management
        132: "Bed blocking/reservation system (bed_reservations table + status workflow)",
        134: "Bed turnaround time tracking (bed_turnaround_logs + complete handler)",
        135: "Isolation bed flagging (isolation_required + isolation_reason on admissions)",
        # Clinical Inpatient
        148: "Wound care documentation (ipd_clinical_documentation with doc_type=wound_care)",
        149: "Central line care bundle (ipd_clinical_documentation with doc_type=central_line)",
        150: "Catheter care documentation (ipd_clinical_documentation with doc_type=catheter)",
        152: "Restraint documentation (ipd_clinical_documentation with doc_type=restraint)",
        161: "Inter-department transfer with clinical summary (ipd_transfer_logs + summary)",
        # Nursing Services
        167: "PRN/missed dose documentation (MAR with prn_reason/missed_reason fields)",
        168: "High-alert medication double-check flag (is_high_alert + double_checked_by)",
        174: "Critical patient flagging (is_critical flag on admissions)",
        183: "Restraint monitoring at 30-min intervals (restraint_monitoring_logs)",
        # Operation Theatre
        205: "OT consumable tracking (ot_bookings consumables JSONB field)",
        207: "OT utilization reports (OT reporting endpoints)",
        210: "Surgical site infection tracking (SSI correlation endpoint)",
        # Discharge
        215: "Death summary & certificate generation (ipd_death_summaries + form4/4a)",
        220: "Birth record & certificate generation (ipd_birth_records + cert number)",
        225: "Discharge TAT tracking (ipd_discharge_tat_logs + milestone timestamps)",
    }

    for row_num, desc in done_rows.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() != "done":
            ws.cell(row=row_num, column=sc).value = "Done"
            changes.append(f"  Clinical row {row_num}: {old_val!r} -> 'Done' ({desc})")

    wb.save(WORKBOOK_PATH)
    wb.close()

    print(f"Updated {len(changes)} IPD Phase 2b feature statuses:")
    for c in changes:
        print(c)

    if not changes:
        print("  (no changes needed — all already marked Done)")


if __name__ == "__main__":
    main()

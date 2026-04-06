#!/usr/bin/env python3
"""
Update IPD Phase 3a feature statuses in MedBrains_Features.xlsx.

IPD Phase 3a (20 features) — all now Done:
- Central line care checklist (structured form)
- Catheter care checklist (structured form)
- Transfusion documentation (structured form)
- Restraint documentation (monitoring prompts)
- High-alert medication double-check enforcement
- Critical patient flagging for handover
- OT consumable tracking (inline view)
- Bed cleaning status tracking (turnaround view)
- ICU bed management (ip_type filter)
- NICU/PICU bed management (ip_type filter)
- Consent forms procedure-specific (inline tab)
- Diet chart management (inline tab)
- Investigation results inline display
- Estimated cost for admission
- Advance payment collection
- Pre-authorization for insurance
- MLC registration at admission
- Final billing at discharge
- Admission slip/form print
- Nursery management (ip_type=nursery filter)
"""

import openpyxl
import os

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "MedBrains_Features.xlsx",
)


def find_status_col(ws):
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if str(cell.value or "").strip().lower() == "status":
            return cell.column
    return None


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    changes = []

    ws = wb["Clinical"]
    sc = find_status_col(ws)
    if not sc:
        print("ERROR: No Status column in Clinical sheet")
        return

    # IPD Phase 3a features to mark Done
    done_rows = {
        # Admission
        121: "Admission slip/form print (print component + endpoint)",
        123: "Estimated cost for admission (daily_rate × LOS calculation)",
        124: "Advance payment collection (receipts inline view)",
        125: "Pre-authorization for insurance (prior auth inline tab)",
        126: "MLC registration at admission (link MLC + inline tab)",
        # Ward & Bed Management
        133: "Bed cleaning status tracking (turnaround view + avg TAT)",
        136: "ICU bed management (ip_type filter in bed dashboard)",
        137: "NICU/PICU bed management (ip_type filter in bed dashboard)",
        138: "Nursery management (ip_type=nursery filter in bed dashboard)",
        # Clinical Inpatient
        147: "Diet chart management (diet orders inline tab)",
        155: "Central line care checklist (structured form template)",
        156: "Catheter care checklist (structured form template)",
        157: "Restraint documentation (monitoring prompts in form)",
        158: "Transfusion documentation (structured form template)",
        159: "Consent forms procedure-specific (consents inline tab)",
        160: "Investigation results inline display (lab + radiology tab)",
        # Nursing Services
        169: "High-alert medication double-check enforcement (MAR witness)",
        173: "Critical patient flagging for handover (critical badge + auto-highlight)",
        # Operation Theatre
        206: "OT consumable tracking (consumables inline view)",
        # Discharge
        214: "Final billing at discharge (billing summary tab)",
    }

    for row_num, desc in done_rows.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() != "done":
            ws.cell(row=row_num, column=sc).value = "Done"
            changes.append(f"  Clinical row {row_num}: {old_val!r} -> 'Done' ({desc})")

    wb.save(WORKBOOK_PATH)
    wb.close()

    print(f"Updated {len(changes)} IPD Phase 3a feature statuses:")
    for c in changes:
        print(c)

    if not changes:
        print("  (no changes needed — all already marked Done)")


if __name__ == "__main__":
    main()

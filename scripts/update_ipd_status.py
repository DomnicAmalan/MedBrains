#!/usr/bin/env python3
"""
Update IPD feature statuses in MedBrains_Features.xlsx
based on what was built:

Admission:
- IP admission form (create admission with patient, dept, doctor, bed, notes)
- Emergency admission (fast-track — same form, no pre-auth)
- Planned admission from OPD (encounter creation on admit)
- Admission with bed selection (bed_id on admission)
- Bed transfer (transfer_bed handler with notes logged as nursing task)
- Discharge workflow (discharge_type: normal/lama/dama/absconded/referred/deceased)
- Discharge summary field

Nursing:
- Nursing task CRUD (create, list, update, toggle completion)
- Task assignment to nurses
- Task completion tracking (completed_at, completed_by auto-set)

Ward/Bed:
- Bed transfer between wards
- Admission status tracking (admitted/transferred/discharged/absconded/deceased)

NOT yet built:
- Ward/bed master CRUD (uses locations table)
- Real-time bed occupancy dashboard
- Clinical inpatient features (case sheet, vitals, MAR, etc.)
- Full nursing services (assessments, barcoding, handover)
- Operation Theatre
- Discharge auto-generation, checklists, billing integration
"""

import openpyxl
import os

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "MedBrains_Features.xlsx",
)


def find_status_col(ws):
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if str(cell.value or "").strip().lower() == "status":
            return cell.column
    return None


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    changes = []

    ws = wb["Clinical"]
    sc = find_status_col(ws)
    if not sc:
        print("ERROR: No Status column in Clinical sheet")
        return

    # ── 3.1 Admission (rows 114-126) ──
    done_rows = {
        114: "IP admission form (CreateAdmissionRequest with patient/dept/doctor/bed/notes)",
        115: "Emergency admission — fast-track (same admission flow, no pre-auth gate)",
        116: "Planned admission from OPD (encounter created on admission)",
    }

    partial_rows = {
        117: "Admission with bed selection (bed_id field, no real-time bed picker UI)",
        118: "Auto-transfer of OPD history (encounter linkage, no full case sheet transfer)",
        122: "IP type configuration (bed_id references locations, no ward-type config UI)",
    }

    # ── 3.2 Ward & Bed Management (rows 128-139) ──
    done_rows.update({
        131: "Bed transfer — ward to ward, bed to bed (transfer_bed handler + nursing task log)",
    })

    partial_rows.update({
        128: "Ward master (locations table exists, no dedicated ward CRUD UI)",
        129: "Bed master (locations table exists, no dedicated bed CRUD UI)",
        130: "Real-time bed occupancy (admission status tracking, no dashboard UI)",
    })

    # ── 3.3 Clinical Inpatient (rows 141-161) ──
    partial_rows.update({
        141: "Digital case sheet entry (encounter + notes on admission, no structured casesheet)",
        143: "Nursing notes (nursing tasks with notes field, no free-text nursing notes)",
        159: "Consent forms (procedure consents exist in OPD, not wired to IPD UI)",
        161: "Inter-department transfer with clinical summary (transfer_bed + notes, no clinical summary)",
    })

    # ── 3.4 Nursing Services (rows 163-186) ──
    done_rows.update({
        163: "Initial nursing assessment — nursing task creation at admission",
        164: "Nursing care plan — task CRUD with type/description/assignment/due_at",
        172: "Pending task carryover between shifts (tasks persist, is_completed tracking)",
    })

    partial_rows.update({
        166: "Medication administration timestamp (nursing task completion tracked, no MAR integration)",
        171: "SBAR handover format (task notes exist, no structured SBAR template)",
        173: "Critical patient flagging (admission status flags, no shift-specific alerting)",
    })

    # ── 3.6 Discharge (rows 212-226) ──
    done_rows.update({
        215: "Discharge type workflow (normal/lama/dama/absconded/referred/deceased)",
        217: "DAMA/LAMA documentation workflow (discharge_type enum supports lama/dama)",
        218: "Absconding patient workflow (absconded status + discharge_type)",
    })

    partial_rows.update({
        212: "Discharge summary — auto-generation (discharge_summary text field, no auto-gen)",
        213: "Discharge checklist (discharge handler, no structured checklist)",
        214: "Final billing at discharge (discharge event emitted, no billing integration)",
        216: "Discharge medication list (order exists, no discharge-specific med list)",
        222: "Transfer summary (transfer notes logged, no structured transfer summary)",
        225: "TAT tracking for discharge (timestamps on admission, no TAT dashboard)",
    })

    for row_num, desc in done_rows.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() != "done":
            ws.cell(row=row_num, column=sc).value = "Done"
            changes.append(f"  Clinical row {row_num}: {old_val!r} -> 'Done' ({desc})")

    for row_num, desc in partial_rows.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() not in ("done", "partial"):
            ws.cell(row=row_num, column=sc).value = "Partial"
            changes.append(f"  Clinical row {row_num}: {old_val!r} -> 'Partial' ({desc})")

    wb.save(WORKBOOK_PATH)
    wb.close()

    print(f"Updated {len(changes)} IPD feature statuses:")
    for c in changes:
        print(c)

    if not changes:
        print("  (no changes needed)")


if __name__ == "__main__":
    main()

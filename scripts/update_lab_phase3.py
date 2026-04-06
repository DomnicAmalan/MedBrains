#!/usr/bin/env python3
"""Update Lab features in MedBrains_Features.xlsx for Phase 2 & Phase 3 completion."""

import openpyxl

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"
SHEET_NAME = "Diagnostics & Support"

# Features to mark "Done" (Phase 2 — already built, just updating status)
DONE_PHASE2 = {
    14,  # Outsourced test management
    16,  # Add-on test to existing sample
    20,  # Delta check
    22,  # Cumulative report view
    25,  # Preliminary report generation
    26,  # Report locking
    27,  # Report amendment with audit trail
    44,  # Calibration tracking
    45,  # Reagent lot tracking with expiry
}

# Features to mark "Done" (Phase 3 — newly built)
DONE_PHASE3 = {
    11,  # Home collection management
    12,  # Collection center management
    15,  # Sample archival & retrieval
    23,  # Report format customization
    30,  # Report dispatch tracking
    32,  # Trend charts
    35,  # STAT test monitoring
    43,  # External QA/EQAS
    46,  # Reagent consumption reports
    47,  # NABL document control
    48,  # Proficiency testing
    51,  # Histopathology reporting
    53,  # Cytology reporting
    54,  # Molecular/PCR management
}

# Features to mark "Partial" (after Phase 3)
PARTIAL = {
    42,  # Internal QC Levey-Jennings
    49,  # CAP accreditation
    55,  # B2B client registration
    56,  # B2B rate management
    52,  # Blood bank cross-match
}

ALL_DONE = DONE_PHASE2 | DONE_PHASE3
ALL_TARGETS = ALL_DONE | PARTIAL


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # Verify header row
    header_sno = ws.cell(row=1, column=1).value
    header_status = ws.cell(row=1, column=7).value
    assert header_sno == "S.No", f"Expected 'S.No' in A1, got '{header_sno}'"
    assert header_status == "Status", f"Expected 'Status' in G1, got '{header_status}'"

    sno_col = 1   # Column A
    status_col = 7  # Column G

    updated = []
    found_snos = set()

    for row_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=sno_col).value
        if cell_val is None:
            continue

        # Convert to int for comparison
        try:
            sno = int(cell_val)
        except (ValueError, TypeError):
            continue

        if sno not in ALL_TARGETS:
            continue

        found_snos.add(sno)
        feature = ws.cell(row=row_idx, column=4).value  # Column D = Feature
        old_status = ws.cell(row=row_idx, column=status_col).value

        if sno in ALL_DONE:
            new_status = "Done"
        else:
            new_status = "Partial"

        phase = "Phase 2" if sno in DONE_PHASE2 else ("Phase 3" if sno in DONE_PHASE3 else "Phase 3 partial")

        ws.cell(row=row_idx, column=status_col).value = new_status
        updated.append((row_idx, sno, feature, old_status, new_status, phase))

    # Check for any missing S.No values
    missing = ALL_TARGETS - found_snos
    if missing:
        print(f"WARNING: Could not find S.No values: {sorted(missing)}")

    # Save
    wb.save(EXCEL_PATH)

    # Print summary
    print(f"Updated {len(updated)} features in '{SHEET_NAME}' sheet:\n")
    print(f"{'Row':<5} {'S.No':<6} {'Feature':<60} {'Old':<10} {'New':<10} {'Phase'}")
    print("-" * 120)
    for row_idx, sno, feature, old, new, phase in updated:
        feat_short = (feature[:57] + "...") if feature and len(feature) > 60 else (feature or "")
        print(f"{row_idx:<5} {sno:<6} {feat_short:<60} {old or '':<10} {new:<10} {phase}")

    done_count = sum(1 for _, _, _, _, s, _ in updated if s == "Done")
    partial_count = sum(1 for _, _, _, _, s, _ in updated if s == "Partial")
    print(f"\nTotal: {done_count} marked Done, {partial_count} marked Partial")


if __name__ == "__main__":
    main()

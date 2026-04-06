#!/usr/bin/env python3
"""
Update Lab/LIS feature statuses in MedBrains_Features.xlsx
based on what was built:
- Test catalog CRUD (code, name, sample_type, normal_range, unit, price, tat_hours)
- Lab orders (electronic ordering, priority, status transitions)
- Sample collection with timestamp + collector
- Sample rejection with reason documentation
- Result entry (parameter, value, unit, range, flag, notes)
- Result authorization (verify workflow)
- Test panels/profiles (group tests, CRUD)
- Order cancellation
- Integration events (lab.completed)
"""

import openpyxl
import os

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "MedBrains_Features.xlsx",
)


def find_status_col(ws):
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if str(cell.value or "").strip().lower() == "status":
            return cell.column
    return None


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH)

    changes = []

    # ── Diagnostics & Support sheet — Lab features rows 4-70 ──
    ws = wb["Diagnostics & Support"]
    sc = find_status_col(ws)
    if not sc:
        print("ERROR: No Status column")
        return

    # Done features
    done_rows = {
        4:  "Lab test master (test name, sample type, container, department)",
        5:  "Lab panel/profile master (CBC = Hb+WBC+Plt+...)",
        6:  "Electronic test ordering from OPD/IPD/ER",
        8:  "Sample collection acknowledgment with timestamp and collector ID",
        9:  "Sample rejection criteria enforcement with reason documentation",
        11: "Emergency/STAT sample handling",
        21: "Manual result entry with normal range display",
        25: "Result authorization workflow (technician → pathologist verification)",
    }

    # Partial features (infrastructure exists but not fully auto-wired)
    partial_rows = {
        7:  "Barcode label generation (infra exists, no barcode generation yet)",
        10: "Sample tracking (status transitions exist, no granular tracking UI)",
        22: "Auto-validation (flag field exists, no auto-validation rules)",
        23: "Abnormal/critical value flagging (flag enum exists, no auto-alerting)",
        28: "Auto-delivery of results to doctor dashboard (integration event emitted)",
        38: "TAT tracking (tat_hours on catalog, no SLA dashboard)",
    }

    for row_num, desc in done_rows.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() != "done":
            ws.cell(row=row_num, column=sc).value = "Done"
            changes.append(f"  Diagnostics & Support row {row_num}: {old_val!r} -> 'Done' ({desc})")

    for row_num, desc in partial_rows.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() not in ("done", "partial"):
            ws.cell(row=row_num, column=sc).value = "Partial"
            changes.append(f"  Diagnostics & Support row {row_num}: {old_val!r} -> 'Partial' ({desc})")

    wb.save(WORKBOOK_PATH)
    wb.close()

    print(f"Updated {len(changes)} feature statuses:")
    for c in changes:
        print(c)

    if not changes:
        print("  (no changes needed)")


if __name__ == "__main__":
    main()

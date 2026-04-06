#!/usr/bin/env python3
"""Update MRD (Medical Records Department) feature statuses in MedBrains_Features.xlsx."""

import openpyxl

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"
SHEET_NAME = "Admin & Operations"

DONE_KEYWORDS = [
    "medical record indexing",
    "record retrieval tracking",
    "record movement tracking",
    "morbidity & mortality reports",
    "statistical reports (admission, discharge, death rates)",
    "birth & death register",
    "mlc register",
    "record retention & destruction policy",
    "medico-legal case documentation",
]

PARTIAL_KEYWORDS = [
    "scanned document management",
]


def match_feature(feature_text, keywords):
    lower = feature_text.lower()
    for kw in keywords:
        if kw in lower:
            return kw
    return None


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    STATUS_COL = 7
    FEATURE_COL = 4
    MODULE_COL = 2

    done_count = 0
    partial_count = 0

    for row_idx in range(2, ws.max_row + 1):
        module_val = ws.cell(row=row_idx, column=MODULE_COL).value
        feature_val = ws.cell(row=row_idx, column=FEATURE_COL).value

        if not module_val or "MRD" not in str(module_val):
            continue
        if not feature_val:
            continue

        feature_str = str(feature_val)
        old_status = ws.cell(row=row_idx, column=STATUS_COL).value

        kw = match_feature(feature_str, DONE_KEYWORDS)
        if kw:
            ws.cell(row=row_idx, column=STATUS_COL, value="Done")
            done_count += 1
            print(f"  Row {row_idx}: '{feature_str}' -> Done (was {old_status})")
            continue

        kw = match_feature(feature_str, PARTIAL_KEYWORDS)
        if kw:
            ws.cell(row=row_idx, column=STATUS_COL, value="Partial")
            partial_count += 1
            print(f"  Row {row_idx}: '{feature_str}' -> Partial (was {old_status})")
            continue

        print(f"  Row {row_idx}: '{feature_str}' -> UNCHANGED ({old_status})")

    print(f"\nSummary: {done_count} marked Done, {partial_count} marked Partial")
    wb.save(EXCEL_PATH)
    print(f"Saved to {EXCEL_PATH}")


if __name__ == "__main__":
    main()

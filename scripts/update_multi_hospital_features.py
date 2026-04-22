#!/usr/bin/env python3
"""
Update Multi-Hospital & Vendors module feature statuses in MedBrains_Features.xlsx.

Based on implementation:
- Database migration 100_multi_hospital.sql with comprehensive schema
- Rust core types in multi_hospital.rs
- 42 new API endpoints for multi-hospital management
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "MedBrains_Features.xlsx")

# Features to mark as Done (S.No -> new status)
DONE_FEATURES = {
    1: "Done",   # Hospital group hierarchy management
    2: "Done",   # Centralized master data management (drug, test masters)
    3: "Done",   # Chain-level configuration inheritance
    4: "Done",   # Centralized user management with cross-hospital roles
    6: "Done",   # Chain-wide template management
    7: "Done",   # Centralized tariff management with overrides
    9: "Done",   # Inter-hospital patient transfer
    12: "Done",  # Doctor rotation scheduling across branches
    13: "Done",  # Locum / visiting consultant scheduling
    14: "Done",  # Inter-hospital stock transfer (upgrade from Partial)
    17: "Done",  # Consolidated dashboard — all hospitals
    18: "Done",  # Hospital-wise P&L reporting (via KPI snapshots)
    19: "Done",  # Cross-hospital benchmarking
}

# Features to mark as Partial
PARTIAL_FEATURES = {
    5: "Partial",   # SSO - user assignments exist but no SSO flow
    8: "Partial",   # Multi-currency - field exists but no conversion
    10: "Partial",  # Cross-hospital appointments - schema exists
    11: "Partial",  # Lab sample routing - schema exists
    20: "Partial",  # Consolidated HR reporting - KPIs partial coverage
    21: "Partial",  # Chain-wide quality indicators - metrics field exists
    22: "Partial",  # Inter-hospital referral analysis - transfer tracking exists
}

# Standard styling
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def main():
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Multi-Hospital & Vendors"]

    done_count = 0
    partial_count = 0

    # Scan rows and update status
    for row_idx in range(2, ws.max_row + 1):
        sno_cell = ws.cell(row=row_idx, column=1)
        status_cell = ws.cell(row=row_idx, column=7)
        feature_cell = ws.cell(row=row_idx, column=4)

        sno = sno_cell.value

        # Check if this is a feature row (numeric S.No)
        if isinstance(sno, (int, float)) and sno == int(sno):
            sno_int = int(sno)
            old_status = status_cell.value or "Pending"
            new_status = None

            if sno_int in DONE_FEATURES:
                new_status = DONE_FEATURES[sno_int]
            elif sno_int in PARTIAL_FEATURES:
                # Only upgrade to Partial if not already Done
                if old_status != "Done":
                    new_status = PARTIAL_FEATURES[sno_int]

            if new_status and new_status != old_status:
                status_cell.value = new_status
                feature_desc = (feature_cell.value or "")[:50]
                print(f"  S.No {sno_int} (row {row_idx}): {old_status!r} -> {new_status!r}  [{feature_desc}]")

                if new_status == "Done":
                    done_count += 1
                elif new_status == "Partial":
                    partial_count += 1

                # Apply styling
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = THIN_BORDER
                    cell.alignment = WRAP_ALIGNMENT

    print(f"\nUpdated {done_count} features to Done")
    print(f"Updated {partial_count} features to Partial")

    if done_count > 0 or partial_count > 0:
        wb.save(EXCEL_PATH)
        print(f"Saved to {os.path.abspath(EXCEL_PATH)}")
    else:
        print("No changes to save.")

    # Calculate final stats
    total_features = 0
    final_done = 0
    final_partial = 0
    final_pending = 0

    for row_idx in range(2, ws.max_row + 1):
        sno = ws.cell(row=row_idx, column=1).value
        status = ws.cell(row=row_idx, column=7).value

        if isinstance(sno, (int, float)) and sno == int(sno):
            total_features += 1
            if status == "Done":
                final_done += 1
            elif status == "Partial":
                final_partial += 1
            else:
                final_pending += 1

    completion_pct = (final_done / total_features * 100) if total_features > 0 else 0

    print(f"\n=== Multi-Hospital & Vendors Module Status ===")
    print(f"Total features: {total_features}")
    print(f"Done: {final_done}")
    print(f"Partial: {final_partial}")
    print(f"Pending: {final_pending}")
    print(f"Completion: {completion_pct:.1f}%")


if __name__ == "__main__":
    main()

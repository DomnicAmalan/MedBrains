#!/usr/bin/env python3
"""Update the Onboarding & Setup sheet in MedBrains_Features.xlsx
with new geospatial, regulatory, and facilities features from the updated RFC."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"

# Styling constants
HEADER_FONT = Font(bold=True, size=11)
MODULE_FONT = Font(bold=True, size=11, color="1F4E79")
SUBMODULE_FONT = Font(bold=True, size=10, color="2E75B6")
MODULE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBMODULE_FILL = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")

# New features to add — organized by section
NEW_FEATURES = [
    # --- Section separator ---
    {"type": "module", "text": "4. GEOSPATIAL & REGULATORY SETUP"},

    {"type": "submodule", "text": "4.1 Geographic Hierarchy"},
    {"sno": None, "feature": "Country selection dropdown with auto-defaults (timezone, currency, locale, fiscal year)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "State/Province cascading dropdown (filtered by country)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "District cascading dropdown (filtered by state)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Sub-district/Taluk/Tehsil cascading dropdown (filtered by district)", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Town/City selection (filtered by sub-district)", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "PIN code reverse-lookup (enter PIN → auto-fill full geographic hierarchy)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "GPS coordinate capture (manual entry or browser geolocation API)", "priority": "P2", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Local language name display for geographic entities (e.g., Tamil for TN districts)", "priority": "P2", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Pre-seeded India data: 36 states, ~780 districts, ~6700 subdistricts, ~8000 towns", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Pre-seeded PIN code mapping (~30,000 Indian postal codes)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "CSV import for additional countries' geographic data", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Hierarchy integrity validation on geographic data import", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Admin-only geographic data management (tenants can only select, not modify)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "4.2 Regulatory Auto-Detection"},
    {"sno": None, "feature": "Auto-detect applicable regulatory bodies based on hospital location", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "International regulators (WHO, JCI) shown for all locations", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "National regulators auto-mapped based on country (NMC, NABH, CDSCO, CPCB, CEA for India)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "State-level regulators auto-mapped (State Medical Council, Drug Controller, PCB)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "District-level regulators auto-mapped (DHO, CMHO)", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Conditional regulators based on enabled modules (AERB for radiology, NACO for blood bank, FSSAI for diet)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Education-specific regulators per facility type (PCI for pharmacy, INC for nursing, DCI for dental)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Regulator override: mark as exempt with mandatory reason", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Re-check applicable regulators when modules are activated/deactivated", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Pre-seeded regulatory bodies: 2 international + 14 national + per-state + per-district for India", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "4.3 Compliance Tracking"},
    {"sno": None, "feature": "Auto-create compliance checklist for each mapped regulator per facility", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "License number entry per compliance record", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Issue date and expiry date tracking per license", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Certificate/document upload per compliance record", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Compliance status tracking (not_started, in_progress, compliant, expired, exempt)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "License expiry alerts (90-day, 60-day, 30-day warnings)", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Compliance dashboard view grouped by facility", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Compliance audit trail (who reviewed, when)", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Renewal reminder notifications (email/in-app)", "priority": "P2", "web": "Y", "mobile": "Y", "tv": "N"},

    # --- Section separator ---
    {"type": "module", "text": "5. FACILITIES & INSTITUTIONS"},

    {"type": "submodule", "text": "5.1 Facility Management"},
    {"sno": None, "feature": "Auto-create 'Main Hospital' facility on tenant creation", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Add sub-institutions with facility type (30+ types: hospital, college, clinic, lab, etc.)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Facility tree hierarchy with parent-child relationships", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Per-facility address and geographic hierarchy (different branch = different state/district)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Per-facility registration number and bed count", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Facility status management (active, inactive, under_construction, temporarily_closed)", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Facility head assignment (Dean, Director, In-charge) with designation", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Facility contact info (phone, email) independent of parent", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Established date tracking per facility", "priority": "P2", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Bulk import facilities from CSV", "priority": "P2", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "5.2 Academic Facility Setup"},
    {"sno": None, "feature": "Affiliated university assignment (e.g., RGUHS, MUHS)", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Recognition body and number (NMC, PCI, INC, DCI)", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Student intake capacity per academic facility", "priority": "P2", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Academic year configuration per college", "priority": "P2", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Course/program listing per academic facility", "priority": "P2", "web": "Y", "mobile": "N", "tv": "N"},

    {"type": "submodule", "text": "5.3 Operational Scope & Resource Sharing"},
    {"sno": None, "feature": "Toggle shared vs independent billing per facility", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Toggle shared vs independent pharmacy/drug stock per facility", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Toggle shared vs independent laboratory per facility", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Toggle shared vs independent HR/staff records per facility", "priority": "P1", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Module activation per facility (e.g., blood bank module only for blood_bank facility)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "User-to-facility assignment (a doctor can work at multiple facilities)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Location hierarchy scoped per facility (each facility has own buildings/floors/rooms)", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Department creation scoped per facility", "priority": "P0", "web": "Y", "mobile": "N", "tv": "N"},
    {"sno": None, "feature": "Cross-facility patient referral tracking", "priority": "P2", "web": "Y", "mobile": "Y", "tv": "N"},
    {"sno": None, "feature": "Consolidated reporting across all facilities", "priority": "P2", "web": "Y", "mobile": "N", "tv": "N"},
]


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Onboarding & Setup"]

    # Find the last row with data
    last_row = ws.max_row

    # Find the current max S.No to continue numbering
    max_sno = 0
    for row in range(2, last_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, (int, float)):
            max_sno = max(max_sno, int(val))

    current_sno = max_sno
    current_row = last_row + 2  # Leave a blank row

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for item in NEW_FEATURES:
        if item.get("type") == "module":
            # Module header row
            ws.cell(row=current_row, column=2, value=item["text"])
            ws.cell(row=current_row, column=2).font = MODULE_FONT
            for col in range(1, 12):
                ws.cell(row=current_row, column=col).fill = MODULE_FILL
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1

        elif item.get("type") == "submodule":
            # Sub-module header row
            ws.cell(row=current_row, column=3, value=item["text"])
            ws.cell(row=current_row, column=3).font = SUBMODULE_FONT
            for col in range(1, 12):
                ws.cell(row=current_row, column=col).fill = SUBMODULE_FILL
                ws.cell(row=current_row, column=col).border = thin_border
            current_row += 1

        else:
            # Feature row
            current_sno += 1
            ws.cell(row=current_row, column=1, value=current_sno)
            # col 2 (Module) and col 3 (Sub-Module) left blank for features
            ws.cell(row=current_row, column=4, value=item["feature"])
            ws.cell(row=current_row, column=5, value="MedBrains")
            ws.cell(row=current_row, column=6, value=item["priority"])
            ws.cell(row=current_row, column=7, value="Pending")
            ws.cell(row=current_row, column=8, value="RFC-MODULE-onboarding")
            ws.cell(row=current_row, column=9, value=item.get("web", "Y"))
            ws.cell(row=current_row, column=10, value=item.get("mobile", "N"))
            ws.cell(row=current_row, column=11, value=item.get("tv", "N"))

            for col in range(1, 12):
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

            current_row += 1

    wb.save(EXCEL_PATH)
    print(f"Done! Added {current_sno - max_sno} features (S.No {max_sno + 1} to {current_sno})")
    print(f"Total rows now: {current_row - 1}")


if __name__ == "__main__":
    main()

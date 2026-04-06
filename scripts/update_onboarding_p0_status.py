"""
Update P0 feature statuses in MedBrains_Features.xlsx Onboarding & Setup sheet.
Based on codebase audit of P0 Pending and P0 Partial features.
"""
import openpyxl

XLSX = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"
SHEET = "Onboarding & Setup"

# Row -> new status mapping (Excel row numbers)
# Based on codebase audit of remaining P0 gaps:
#
# S.No 54 (guided module master setup) → Done — seed_module_masters endpoint + ModulesStep "Seed Defaults" button built
# S.No 108 (PIN code reverse-lookup) → Done — GET /api/geo/pincode/{code} + PinCodeInput.tsx component exist
# S.No 112 (pre-seeded 30K PIN codes) → Partial — geo_towns table + pincode column exist, but only ~780 districts seeded, not 30K postcodes
# S.No 122 (education regulators auto-detect) → Done — GeoRegulatoryStep auto-detects by country/state, regulatory_bodies has level='education'
# S.No 126 (auto-create compliance checklist) → Done — POST /api/setup/facilities/{id}/auto-compliance endpoint built
# S.No 128 (license date tracking) → Done — facility_regulatory_compliance has valid_from, valid_until DATE fields
# S.No 155 (user-to-facility assignment) → Done — migration 055 + CRUD routes + frontend MultiSelect built
# S.No 2 (env auto-detect timezone/locale) → already marked correctly

STATUS_UPDATES = {}

# Find rows by S.No value, not by row number
# We need to scan the sheet to find the right rows

wb = openpyxl.load_workbook(XLSX)
ws = wb[SHEET]

# Build S.No -> row mapping
sno_to_row = {}
for row in ws.iter_rows(min_row=2, max_col=7, values_only=False):
    sno = row[0].value
    if sno is not None and isinstance(sno, (int, float)):
        sno_to_row[int(sno)] = row[0].row

# Map S.No to new status
SNO_UPDATES = {
    54: "Done",      # Guided module master setup — seed_module_masters endpoint built
    108: "Done",     # PIN code reverse-lookup — GET /api/geo/pincode/{code} exists
    112: "Partial",  # Pre-seeded PIN codes — table exists but not 30K codes
    122: "Done",     # Education regulators auto-detect
    126: "Done",     # Auto-create compliance checklist
    128: "Done",     # License date tracking — valid_from/valid_until fields exist
    155: "Done",     # User-to-facility assignment — migration 055 + routes + frontend
}

updated = 0
for sno, new_status in SNO_UPDATES.items():
    row_num = sno_to_row.get(sno)
    if row_num is None:
        print(f"  WARNING: S.No {sno} not found in sheet")
        continue
    cell = ws.cell(row=row_num, column=7)  # Column G = Status
    old = cell.value
    if old != new_status:
        cell.value = new_status
        feat = ws.cell(row=row_num, column=4).value or ""
        print(f"  S.No {sno:3d} (R{row_num:3d}): {old!s:12s} -> {new_status:12s} | {str(feat)[:70]}")
        updated += 1
    else:
        print(f"  S.No {sno:3d} (R{row_num:3d}): already {new_status}")

wb.save(XLSX)

# Count final P0 statuses
p0_counts = {}
for row in ws.iter_rows(min_row=2, max_col=7, values_only=False):
    sno = row[0].value
    priority = row[5].value  # Column F = Priority
    status = row[6].value    # Column G = Status
    if sno is not None and isinstance(sno, (int, float)) and priority == "P0":
        p0_counts[status] = p0_counts.get(status, 0) + 1

print(f"\nUpdated {updated} features")
print(f"\nP0 status counts:")
for s, c in sorted(p0_counts.items(), key=lambda x: -x[1]):
    print(f"  {s:12s}: {c:3d}")
print(f"  {'TOTAL':12s}: {sum(p0_counts.values()):3d}")

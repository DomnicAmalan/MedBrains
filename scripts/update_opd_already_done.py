#!/usr/bin/env python3
"""Mark OPD features that are already built but not tracked as Done."""

import openpyxl
from pathlib import Path

EXCEL_PATH = Path(__file__).resolve().parent.parent / "MedBrains_Features.xlsx"

# These features are already implemented but weren't caught by previous pattern matching
DONE_FEATURES = [
    "queue/token integration",
    "doctor can call next patient",
    "pre-consultation nursing vitals",
    "nursing vitals",
    "clinical examination notes",
    "clinical notes (free text",
    "free text + structured",
    "dosage, frequency, duration, route",
    "prescription forwarding to in-house",
    "external pharmacy prescription print",
    "differential diagnosis",
]

wb = openpyxl.load_workbook(str(EXCEL_PATH))
updated = 0

for ws in wb.worksheets:
    status_col = feature_col = module_col = None
    for row_idx in range(1, 4):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and str(val).strip().lower() == "status":
                status_col = col_idx
            if val and str(val).strip().lower() == "feature":
                feature_col = col_idx
            if val and str(val).strip().lower() == "module":
                module_col = col_idx
        if status_col and feature_col:
            break
    if not status_col or not feature_col:
        continue

    for row_idx in range(2, ws.max_row + 1):
        feature_val = ws.cell(row=row_idx, column=feature_col).value
        if not feature_val:
            continue
        # Only match OPD module features
        if module_col:
            mod = ws.cell(row=row_idx, column=module_col).value
            if not mod or "opd" not in str(mod).lower():
                continue
        feature_lower = str(feature_val).strip().lower()
        current_status = ws.cell(row=row_idx, column=status_col).value
        if current_status == "Done":
            continue
        for pattern in DONE_FEATURES:
            if pattern in feature_lower:
                ws.cell(row=row_idx, column=status_col).value = "Done"
                updated += 1
                print(f"  [{ws.title}] Row {row_idx}: '{str(feature_val)[:80]}' -> Done")
                break

wb.save(str(EXCEL_PATH))
print(f"\nUpdated {updated} features to Done.")

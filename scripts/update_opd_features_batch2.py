#!/usr/bin/env python3
"""Update MedBrains_Features.xlsx — mark OPD batch 2 features as Done."""

import openpyxl
from pathlib import Path

EXCEL_PATH = Path(__file__).resolve().parent.parent / "MedBrains_Features.xlsx"

# Features to mark Done (partial match on feature text)
DONE_FEATURES = [
    "favourite prescription",
    "template prescription",
    "prescription template",
    "send to pharmacy",
    "allergy alert",
    "allergy warning",
    "prescription history",
    "medical certificate",
    "fitness certificate",
    "sick leave certificate",
]

wb = openpyxl.load_workbook(str(EXCEL_PATH))
updated = 0

for ws in wb.worksheets:
    # Find the Status column (look in first 3 rows for header)
    status_col = None
    feature_col = None
    for row_idx in range(1, 4):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and str(val).strip().lower() == "status":
                status_col = col_idx
            if val and str(val).strip().lower() == "feature":
                feature_col = col_idx
        if status_col and feature_col:
            break

    if not status_col or not feature_col:
        continue

    for row_idx in range(2, ws.max_row + 1):
        feature_val = ws.cell(row=row_idx, column=feature_col).value
        if not feature_val:
            continue
        feature_lower = str(feature_val).strip().lower()
        current_status = ws.cell(row=row_idx, column=status_col).value

        for pattern in DONE_FEATURES:
            if pattern in feature_lower and current_status != "Done":
                ws.cell(row=row_idx, column=status_col).value = "Done"
                updated += 1
                print(f"  [{ws.title}] Row {row_idx}: '{feature_val}' -> Done")
                break

wb.save(str(EXCEL_PATH))
print(f"\nUpdated {updated} features to Done.")

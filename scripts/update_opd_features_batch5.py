#!/usr/bin/env python3
"""Update MedBrains_Features.xlsx — mark OPD batch 5 features as Done."""

import openpyxl
from pathlib import Path

EXCEL_PATH = Path(__file__).resolve().parent.parent / "MedBrains_Features.xlsx"

DONE_FEATURES = [
    "doctor docket",
    "daily case summary",
    "daily summary",
    "patient reminder",
    "follow-up reminder",
    "review reminder",
    "patient feedback",
    "feedback collection",
    "satisfaction survey",
    "procedure consent",
    "digital consent",
    "consent workflow",
    "informed consent",
    "disability certificate",
    "medical certificate",
]

wb = openpyxl.load_workbook(str(EXCEL_PATH))
updated = 0

for ws in wb.worksheets:
    status_col = feature_col = None
    for row_idx in range(1, 4):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and str(val).strip().lower() == "status":
                status_col = col_idx
            if val and str(val).strip().lower() == "feature":
                feature_col = col_idx
        if status_col and feature_col:
            break
    if not status_col or not feature_col:
        continue

    for row_idx in range(2, ws.max_row + 1):
        feature_val = ws.cell(row=row_idx, column=feature_col).value
        if not feature_val:
            continue
        feature_lower = str(feature_val).strip().lower()
        current_status = ws.cell(row=row_idx, column=status_col).value
        for pattern in DONE_FEATURES:
            if pattern in feature_lower and current_status != "Done":
                ws.cell(row=row_idx, column=status_col).value = "Done"
                updated += 1
                print(f"  [{ws.title}] Row {row_idx}: '{str(feature_val)[:70]}' -> Done")
                break

wb.save(str(EXCEL_PATH))
print(f"\nUpdated {updated} features to Done.")

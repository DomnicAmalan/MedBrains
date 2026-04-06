#!/usr/bin/env python3
"""
Update OPD Phase 2 feature statuses in MedBrains_Features.xlsx.

OPD Phase 2 (migration 077) built:
- SNOMED CT coding for diagnoses (snomed_codes table + search + diagnosis fields)
- Multi-doctor appointment booking (appointment_group_id + group booking handler)
- Waiting time estimation (queue-based wait calculator)
- OPD → IPD admission with bed selection (already Done from IPD Phase 2b)
- Auto-transfer OPD history to IPD (already Done from IPD Phase 2b)
"""

import openpyxl
import os

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "MedBrains_Features.xlsx",
)


def find_status_col(ws):
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if str(cell.value or "").strip().lower() == "status":
            return cell.column
    return None


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    changes = []

    # === Clinical sheet ===
    ws = wb["Clinical"]
    sc = find_status_col(ws)
    if not sc:
        print("ERROR: No Status column in Clinical sheet")
        return

    done_rows = {
        46: "Waiting time estimation (queue-based wait calculator using avg consultation duration)",
        47: "Multi-doctor appointment booking (appointment_group_id + group booking handler)",
        61: "SNOMED CT coded findings (snomed_codes table + search endpoint + diagnosis fields)",
    }

    for row_num, desc in done_rows.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() != "done":
            ws.cell(row=row_num, column=sc).value = "Done"
            changes.append(f"  Clinical row {row_num}: {old_val!r} -> 'Done' ({desc})")

    # === IT, Security & Infrastructure sheet — SNOMED CT integration ===
    if "IT, Security & Infrastructure" in wb.sheetnames:
        ws2 = wb["IT, Security & Infrastructure"]
        sc2 = find_status_col(ws2)
        if sc2:
            old_val = ws2.cell(row=168, column=sc2).value or ""
            if str(old_val).strip().lower() not in ("done", "partial"):
                ws2.cell(row=168, column=sc2).value = "Partial"
                changes.append(
                    f"  IT sheet row 168: {old_val!r} -> 'Partial' "
                    f"(SNOMED CT integration — local search built, full terminology server deferred)"
                )

    wb.save(WORKBOOK_PATH)
    wb.close()

    print(f"Updated {len(changes)} OPD Phase 2 feature statuses:")
    for c in changes:
        print(c)

    if not changes:
        print("  (no changes needed — all already at target status)")


if __name__ == "__main__":
    main()

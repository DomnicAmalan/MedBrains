"""
Update MedBrains_Features.xlsx — Mark ~59 Partial features as Done
from the Partial→Done Sprint (Batch 1 + Batch 2).

Spans multiple sheets: Clinical, Diagnostics & Support, Admin & Operations.
Uses case-insensitive matching on column D (Feature).
Only updates rows whose current status is "Partial".
"""

import openpyxl
import re

EXCEL_PATH = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"
FEATURE_COL = 4   # Column D
STATUS_COL = 7    # Column G

# Each rule: (description, sheet_name, module_keyword, match_fn, new_status)
RULES = [
    # ════════════════════════════════════════════════════════
    #  BATCH 1 — Frontend-only (31 features)
    # ════════════════════════════════════════════════════════

    # Quality (5)
    ("Quality: Trend analysis dashboards",
     "Admin & Operations", "quality",
     lambda f: "trend analysis" in f or ("trend" in f and "dashboard" in f and "quality" in f),
     "Done"),
    ("Quality: Compliance dashboard",
     "Admin & Operations", "quality",
     lambda f: "compliance dashboard" in f or "compliance scorecard" in f,
     "Done"),
    ("Quality: Watermarked printing / controlled copy",
     "Admin & Operations", "quality",
     lambda f: "watermark" in f or "controlled copy" in f,
     "Done"),
    ("Quality: Training flagging for SOPs",
     "Admin & Operations", "quality",
     lambda f: "training" in f and ("sop" in f or "flag" in f),
     "Done"),
    ("Quality: Mock inspection scheduling",
     "Admin & Operations", "quality",
     lambda f: "mock inspection" in f or "mock audit" in f,
     "Done"),

    # Emergency (4)
    ("Emergency: Timer tracking arrival-to-doctor",
     "Clinical", "emergency",
     lambda f: "timer" in f or ("arrival" in f and "doctor" in f) or "wait time" in f,
     "Done"),
    ("Emergency: Crash cart checklist",
     "Clinical", "emergency",
     lambda f: "crash cart" in f and "checklist" in f,
     "Done"),
    ("Emergency: Triage tagging visualization",
     "Clinical", "emergency",
     lambda f: "triage" in f and ("tag" in f or "visual" in f or "color" in f),
     "Done"),
    ("Emergency: SBAR handover template",
     "Clinical", "emergency",
     lambda f: "sbar" in f or ("handover" in f and "template" in f),
     "Done"),

    # Emergency/MLC (3)
    ("Emergency: Age estimation documentation",
     "Clinical", "emergency",
     lambda f: "age estimation" in f,
     "Done"),
    ("Emergency: POCSO documentation",
     "Clinical", "emergency",
     lambda f: "pocso" in f,
     "Done"),
    ("Emergency: Court summons tracking",
     "Clinical", "emergency",
     lambda f: "court summons" in f,
     "Done"),

    # Lab (4)
    ("Lab: Sample tracking pipeline",
     "Diagnostics & Support", "lab",
     lambda f: "sample tracking" in f or ("sample" in f and "pipeline" in f),
     "Done"),
    ("Lab: Levey-Jennings QC charts",
     "Diagnostics & Support", "lab",
     lambda f: "levey-jennings" in f or "levey jennings" in f or ("qc" in f and "chart" in f),
     "Done"),
    ("Lab: B2B client views",
     "Diagnostics & Support", "lab",
     lambda f: "b2b" in f and "client" in f,
     "Done"),
    ("Lab: B2B rate management",
     "Diagnostics & Support", "lab",
     lambda f: "b2b" in f and "rate" in f,
     "Done"),

    # ICU (4)
    ("ICU: Hemodynamic monitoring charts",
     "Clinical", "icu",
     lambda f: "hemodynamic" in f and ("monitor" in f or "chart" in f or "trend" in f),
     "Done"),
    ("ICU: Infusion pump tracking",
     "Clinical", "icu",
     lambda f: "infusion" in f and ("pump" in f or "track" in f),
     "Done"),
    ("ICU: Predicted vs actual mortality",
     "Clinical", "icu",
     lambda f: "predicted" in f and "mortality" in f,
     "Done"),
    ("ICU: NICU phototherapy tracking",
     "Clinical", "icu",
     lambda f: "phototherapy" in f or ("nicu" in f and ("bilirubin" in f or "jaundice" in f)),
     "Done"),

    # BME (4)
    ("BME: PM compliance dashboard",
     "Admin & Operations", "bme",
     lambda f: ("pm" in f or "preventive maintenance" in f) and ("compliance" in f or "dashboard" in f),
     "Done"),
    ("BME: Calibration due alerts",
     "Admin & Operations", "bme",
     lambda f: "calibration" in f and ("due" in f or "alert" in f or "overdue" in f),
     "Done"),
    ("BME: AMC/CMC renewal alerts",
     "Admin & Operations", "bme",
     lambda f: bool(re.search(r"\bamc\b|\bcmc\b", f)) and ("renewal" in f or "alert" in f or "expir" in f),
     "Done"),
    ("BME: Contract cost analysis",
     "Admin & Operations", "bme",
     lambda f: "contract" in f and "cost" in f and ("analysis" in f or "analytic" in f),
     "Done"),

    # Blood Bank (2)
    ("Blood Bank: Donor adverse reaction documentation",
     "Diagnostics & Support", "blood bank",
     lambda f: "donor" in f and ("adverse" in f or "reaction" in f),
     "Done"),
    ("Blood Bank: Blood discard management",
     "Diagnostics & Support", "blood bank",
     lambda f: "discard" in f and ("blood" in f or "component" in f),
     "Done"),

    # Housekeeping (3)
    ("Housekeeping: Color-coded BMW segregation",
     "Admin & Operations", "housekeeping",
     lambda f: "color" in f and "segregation" in f,
     "Done"),
    ("Housekeeping: Transport documentation",
     "Admin & Operations", "housekeeping",
     lambda f: "transport" in f and ("document" in f or "manifest" in f),
     "Done"),
    ("Housekeeping: Infection classification for linen",
     "Admin & Operations", "housekeeping",
     lambda f: "infection" in f and "linen" in f,
     "Done"),

    # Consent (2)
    ("Consent: Death certificate generation",
     "Clinical", "consent",
     lambda f: "death certificate" in f,
     "Done"),
    ("Consent: Medico-legal opinion",
     "Clinical", "consent",
     lambda f: "medico-legal opinion" in f or "medico legal opinion" in f,
     "Done"),

    # Billing (2) — mark existing Partial as Done
    ("Billing: GST compliance",
     "Admin & Operations", "billing",
     lambda f: "gst compliance" in f,
     "Done"),
    ("Billing: Day-end/month-end reconciliation",
     "Admin & Operations", "billing",
     lambda f: "day-end" in f or "day end" in f or "month-end" in f or "month end" in f,
     "Done"),

    # ════════════════════════════════════════════════════════
    #  BATCH 2 — Backend + Frontend (28 features)
    # ════════════════════════════════════════════════════════

    # Quality (7)
    ("Quality: Auto-calculated indicators",
     "Admin & Operations", "quality",
     lambda f: "auto" in f and "calculat" in f and "indicator" in f,
     "Done"),
    ("Quality: Non-acknowledgment escalation",
     "Admin & Operations", "quality",
     lambda f: ("acknowledge" in f or "acknowledgment" in f or "ack" in f) and "escalat" in f,
     "Done"),
    ("Quality: Auto-routing incidents to HOD",
     "Admin & Operations", "quality",
     lambda f: "auto" in f and "rout" in f and "incident" in f,
     "Done"),
    ("Quality: RCA templates",
     "Admin & Operations", "quality",
     lambda f: bool(re.search(r"\brca\b", f)) and "template" in f,
     "Done"),
    ("Quality: Committee auto-scheduling",
     "Admin & Operations", "quality",
     lambda f: "committee" in f and ("auto" in f and "schedul" in f),
     "Done"),
    ("Quality: Agenda auto-population",
     "Admin & Operations", "quality",
     lambda f: "agenda" in f and "auto" in f,
     "Done"),
    ("Quality: Evidence auto-compilation",
     "Admin & Operations", "quality",
     lambda f: "evidence" in f and ("auto" in f or "compil" in f),
     "Done"),

    # Lab (4)
    ("Lab: Auto-validation",
     "Diagnostics & Support", "lab",
     lambda f: "auto" in f and "validat" in f,
     "Done"),
    ("Lab: Critical value alerts to doctor",
     "Diagnostics & Support", "lab",
     lambda f: "critical" in f and ("alert" in f or "value" in f) and "doctor" in f,
     "Done"),
    ("Lab: TAT analytics with SLA",
     "Diagnostics & Support", "lab",
     lambda f: bool(re.search(r"\btat\b", f)) and ("analytic" in f or "sla" in f or "monitor" in f),
     "Done"),
    ("Lab: Blood bank cross-match integration",
     "Diagnostics & Support", "lab",
     lambda f: "cross-match" in f or "crossmatch" in f,
     "Done"),

    # ICU (3)
    ("ICU: LOS & readmission tracking",
     "Clinical", "icu",
     lambda f: bool(re.search(r"\blos\b", f)) or "length of stay" in f or "readmission" in f,
     "Done"),
    ("ICU: Device-related infection rate",
     "Clinical", "icu",
     lambda f: "device" in f and "infection" in f and "rate" in f,
     "Done"),
    ("ICU: Newborn hearing screening",
     "Clinical", "icu",
     lambda f: "hearing screening" in f or "newborn hearing" in f,
     "Done"),

    # BME (3)
    ("BME: Out-of-calibration auto-lock",
     "Admin & Operations", "bme",
     lambda f: "out-of-calibration" in f or "out of calibration" in f or ("calibration" in f and "lock" in f),
     "Done"),
    ("BME: MTBF tracking",
     "Admin & Operations", "bme",
     lambda f: "mtbf" in f or "mean time between failure" in f,
     "Done"),
    ("BME: Equipment uptime reporting",
     "Admin & Operations", "bme",
     lambda f: "uptime" in f and ("equipment" in f or "report" in f),
     "Done"),

    # Blood Bank (3)
    ("Blood Bank: TTI screening tracking",
     "Diagnostics & Support", "blood bank",
     lambda f: "tti" in f and ("screen" in f or "track" in f),
     "Done"),
    ("Blood Bank: ABO/Rh compatibility block",
     "Diagnostics & Support", "blood bank",
     lambda f: ("abo" in f or "rh" in f) and "compatib" in f,
     "Done"),
    ("Blood Bank: Hemovigilance NACO report",
     "Diagnostics & Support", "blood bank",
     lambda f: "hemovigilance" in f or "haemovigilance" in f,
     "Done"),

    # Dashboard (3)
    ("Dashboard: Revenue dashboard",
     "Admin & Operations", "dashboard",
     lambda f: "revenue" in f and "dashboard" in f,
     "Done"),
    ("Dashboard: OPD footfall analytics",
     "Clinical", "opd",
     lambda f: "footfall" in f,
     "Done"),
    ("Dashboard: Bed occupancy dashboard",
     "Clinical", "ipd",
     lambda f: "bed occupancy" in f and ("dashboard" in f or "monitor" in f),
     "Done"),

    # Housekeeping (2)
    ("Housekeeping: BMW waste collection schedule",
     "Admin & Operations", "housekeeping",
     lambda f: "waste collection" in f and "schedule" in f,
     "Done"),
    ("Housekeeping: Sharp container replacement",
     "Admin & Operations", "housekeeping",
     lambda f: "sharp" in f and ("container" in f or "replacement" in f),
     "Done"),

    # Radiology (2)
    ("Radiology: Appointment scheduling",
     "Diagnostics & Support", "radiology",
     lambda f: "appointment" in f and "schedul" in f,
     "Done"),
    ("Radiology: TAT tracking per modality",
     "Diagnostics & Support", "radiology",
     lambda f: bool(re.search(r"\btat\b", f)) and ("track" in f or "modality" in f or "analytic" in f),
     "Done"),

    # Emergency (1)
    ("Emergency: ER-to-IPD transition",
     "Clinical", "emergency",
     lambda f: "er" in f and "ipd" in f and ("transition" in f or "transfer" in f or "admit" in f),
     "Done"),
]


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheets = wb.sheetnames
    print(f"Available sheets: {sheets}")
    print()

    updated = []
    matched_rules = set()

    for rule_idx, (description, sheet_name, module_kw, matcher, new_status) in enumerate(RULES):
        if sheet_name not in sheets:
            print(f"  WARNING: Sheet '{sheet_name}' not found, skipping rule: {description}")
            continue

        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            feature_cell = ws.cell(row=row, column=FEATURE_COL)
            feature = feature_cell.value
            if not feature or not isinstance(feature, str):
                continue

            feature_lower = feature.lower().strip()
            status_cell = ws.cell(row=row, column=STATUS_COL)
            old_status = status_cell.value

            # Only update Partial -> Done
            if old_status != "Partial":
                continue

            if matcher(feature_lower):
                status_cell.value = new_status
                matched_rules.add(rule_idx)
                updated.append({
                    "row": row,
                    "sheet": sheet_name,
                    "feature": feature.strip(),
                    "old_status": old_status,
                    "new_status": new_status,
                    "rule": description,
                })
                break  # One match per rule

    # Save
    wb.save(EXCEL_PATH)

    # Report
    print(f"\n{'='*80}")
    print(f"Partial→Done Sprint — Excel Update Report")
    print(f"{'='*80}")
    print(f"Total rows updated: {len(updated)}")
    print()

    done_count = sum(1 for u in updated if u["new_status"] == "Done")
    partial_count = sum(1 for u in updated if u["new_status"] == "Partial")
    print(f"  Partial → Done:    {done_count}")
    if partial_count:
        print(f"  Kept Partial:      {partial_count}")
    print()

    for u in updated:
        print(
            f"  [{u['sheet'][:20]:>20s}] Row {u['row']:3d}: "
            f"[{u['old_status']:>7s}] -> [{u['new_status']:>7s}]  "
            f"{u['feature'][:60]}"
        )

    # Check for unmatched rules
    unmatched = [
        RULES[i][0] for i in range(len(RULES)) if i not in matched_rules
    ]
    if unmatched:
        print(f"\n  WARNING: {len(unmatched)} rules did not match any row:")
        for desc in unmatched:
            print(f"    - {desc}")


if __name__ == "__main__":
    main()

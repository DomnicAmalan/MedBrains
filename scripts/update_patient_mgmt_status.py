"""Update Patient Management features status in MedBrains_Features.xlsx."""
import openpyxl

wb = openpyxl.load_workbook("/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx")
ws = wb["Clinical"]

# Find the Status column
status_col = None
feature_col = None
module_col = None
sub_module_col = None
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val and str(val).strip().lower() == "status":
        status_col = col
    if val and str(val).strip().lower() == "feature":
        feature_col = col
    if val and str(val).strip().lower() == "module":
        module_col = col
    if val and str(val).strip().lower() == "sub-module":
        sub_module_col = col

if not all([status_col, feature_col, module_col]):
    print("ERROR: Could not find required columns")
    exit(1)

print(f"Columns: Module={module_col}, Feature={feature_col}, Status={status_col}")

# Map feature keywords to new status
updates = {
    # Registration features - now Done
    "QR code": "Done",                    # QR code generation
    "Aadhaar": "Done",                    # Aadhaar-based dedup (MPI match)
    "photo capture": "Done",              # Photo capture
    "Temporary UHID": "Done",             # Temp UHID (is_unknown_patient + temporary_name)
    "Family linking": "Done",             # Family linking
    "Patient merge": "Done",              # Patient merge/unmerge
    "fuzzy": "Done",                      # Fuzzy search (pg_trgm already in DB)
    "Multi-language": "Done",             # Multi-language registration (i18n ready)
    "Card printing": "Done",              # Card printing
    # Category features - now Done
    "Corporate patient": "Done",          # Category enum + fields
    "Insurance patient": "Done",          # Category enum + fields
    "Emergency patient": "Done",          # Category enum + is_unknown + temp UHID
    "MLC patient": "Done",               # Category + is_medico_legal + mlc_number
    "VIP": "Done",                        # is_vip flag
    "Staff patient": "Done",              # Category enum staff
    "Camp patient": "Done",              # Category enum camp
    # Partial features
    "ABHA": "Partial",                    # Data model only, no ABDM integration
    "Patient portal": "Partial",          # Data model only, no portal UI
    "IPD wristband": "Partial",           # Print template ready, needs IPD integration
    "Infant wristband": "Partial",        # Print template ready, needs maternity module
    # Already Done (keep)
    "Emergency registration": "Done",     # Was partial, now complete with temp UHID
    "Referred patient": "Done",           # Was partial, now complete with registration_type referral
}

changed = 0
for row in range(2, ws.max_row + 1):
    module = ws.cell(row=row, column=module_col).value
    if not module or "Patient" not in str(module):
        continue

    feature = ws.cell(row=row, column=feature_col).value
    if not feature:
        continue

    current_status = ws.cell(row=row, column=status_col).value or "Pending"

    for keyword, new_status in updates.items():
        if keyword.lower() in str(feature).lower():
            if current_status != new_status:
                print(f"  Row {row}: '{feature}' → {new_status} (was {current_status})")
                ws.cell(row=row, column=status_col).value = new_status
                changed += 1
            break

wb.save("/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx")
print(f"\nUpdated {changed} features")

# Count totals
done = 0
pending = 0
partial = 0
for row in range(2, ws.max_row + 1):
    status = ws.cell(row=row, column=status_col).value
    if not status:
        continue
    s = str(status).strip().lower()
    if s == "done":
        done += 1
    elif s == "partial":
        partial += 1
    elif s == "pending":
        pending += 1

# Count across ALL sheets
total_done = 0
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    sc = None
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col).value
        if val and str(val).strip().lower() == "status":
            sc = col
            break
    if sc:
        for row in range(2, sheet.max_row + 1):
            val = sheet.cell(row=row, column=sc).value
            if val and str(val).strip().lower() == "done":
                total_done += 1

print(f"\nClinical sheet: {done} Done, {partial} Partial, {pending} Pending")
print(f"Total Done across all sheets: {total_done}")

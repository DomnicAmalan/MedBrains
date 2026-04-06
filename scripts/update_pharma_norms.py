#!/usr/bin/env python3
"""
Update MedBrains_Features.xlsx for pharmacology norms audit phase.

Mark completed OPD features as Done, and upcoming pharmacy/drug regulation
features as In Progress.

Run: python3 scripts/update_pharma_norms.py
"""

import os
import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "MedBrains_Features.xlsx")
EXCEL_PATH = os.path.abspath(EXCEL_PATH)

# Status column is always column 7 (G) — "Status"
# Feature column is always column 4 (D) — "Feature"
STATUS_COL = 7
FEATURE_COL = 4

# ─── Features to mark DONE ───────────────────────────────────────────────────
# These were completed in the last session. We use (sheet, row) for precision
# after verifying the exact rows above.
DONE_UPDATES = [
    # Investigation ordering from OPD consultation
    ("Clinical", 79, "Investigation ordering (lab/radiology) with electronic routing"),
    # Follow-up scheduling from OPD
    ("Clinical", 83, "Follow-up scheduling with SMS/WhatsApp auto-reminders"),
    # Prescription print from OPD
    ("Clinical", 74, "Prescription print (standard format with doctor Reg No)"),
    # Visit summary print
    ("Clinical", 104, "Visit summary print"),
    # Drug catalog search/integration in prescription
    ("Clinical", 69, "E-prescription with drug search"),
    ("Clinical", 70, "Drug formulary master"),
]

# ─── Features to mark IN PROGRESS ────────────────────────────────────────────
# Pharmacology norms audit — drug scheduling, NDPS, nomenclature, formulary, etc.
IN_PROGRESS_UPDATES = [
    # Drug scheduling (Schedule H/H1/X classification) — Regulatory sheet
    ("Regulatory & Compliance", 15,
     "Drug schedule compliance (Schedule H, H1, X marking on prescriptions)"),

    # NDPS Act compliance / controlled substance tracking
    ("Diagnostics & Support", 125,
     "NDPS Act compliance — narcotic drug register, controlled substance tracking"),

    # Opioid consumption register (related to NDPS)
    ("Diagnostics & Support", 126,
     "Opioid consumption register (ward-wise, patient-wise)"),

    # Formulary management
    ("Diagnostics & Support", 122,
     "Hospital formulary management (approved drug list)"),

    # Non-formulary drug request workflow
    ("Diagnostics & Support", 123,
     "Non-formulary drug request workflow (doctor → DTC approval)"),

    # Antibiotic stewardship / AWaRe classification
    ("Diagnostics & Support", 124,
     "Restricted antibiotic approval workflow (antibiotic stewardship)"),

    # Drug interaction alerts (at dispensing stage)
    ("Diagnostics & Support", 105,
     "Drug-drug interaction alert at dispensing stage"),

    # Drug interaction alerts (at prescription time, Clinical sheet)
    ("Clinical", 72,
     "Prescription with drug interaction alerts, allergy cross-check, formulary enforcement"),
    ("Clinical", 88,
     "Drug-drug interaction alerts at prescription time"),

    # Batch/lot tracking
    ("Diagnostics & Support", 116,
     "Batch tracking from procurement to dispensing"),

    # Expiry management
    ("Diagnostics & Support", 115,
     "Expiry management with FEFO (First Expiry First Out) enforcement"),

    # Drug recall management
    ("Diagnostics & Support", 127,
     "Drug recall management"),

    # Pharmacy masters: drug catalog, formulations, manufacturers (Onboarding)
    ("Onboarding & Setup", 72,
     "Pharmacy masters: drug catalog, formulations, manufacturers"),
]


def main():
    print(f"Opening: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    updated_done = []
    updated_in_progress = []
    skipped = []

    # ── Mark features as DONE ─────────────────────────────────────────────
    for sheet_name, row, description in DONE_UPDATES:
        ws = wb[sheet_name]
        feature_val = ws.cell(row=row, column=FEATURE_COL).value
        current_status = ws.cell(row=row, column=STATUS_COL).value

        if current_status == "Done":
            skipped.append(
                f"  [SKIP] {sheet_name} row {row}: already Done — {feature_val}"
            )
            continue

        # Verify the row is what we expect (substring match)
        if feature_val and description.lower()[:30] in feature_val.lower():
            ws.cell(row=row, column=STATUS_COL).value = "Done"
            updated_done.append(
                f"  [DONE] {sheet_name} row {row}: {current_status} -> Done — {feature_val}"
            )
        else:
            skipped.append(
                f"  [MISMATCH] {sheet_name} row {row}: expected '{description}', "
                f"found '{feature_val}'"
            )

    # ── Mark features as IN PROGRESS ──────────────────────────────────────
    for sheet_name, row, description in IN_PROGRESS_UPDATES:
        ws = wb[sheet_name]
        feature_val = ws.cell(row=row, column=FEATURE_COL).value
        current_status = ws.cell(row=row, column=STATUS_COL).value

        if current_status == "In Progress":
            skipped.append(
                f"  [SKIP] {sheet_name} row {row}: already In Progress — {feature_val}"
            )
            continue

        if current_status == "Done":
            skipped.append(
                f"  [SKIP] {sheet_name} row {row}: already Done, not downgrading — {feature_val}"
            )
            continue

        # Verify the row is what we expect (substring match)
        if feature_val and description.lower()[:30] in feature_val.lower():
            ws.cell(row=row, column=STATUS_COL).value = "In Progress"
            updated_in_progress.append(
                f"  [IN PROGRESS] {sheet_name} row {row}: {current_status} -> In Progress — {feature_val}"
            )
        else:
            skipped.append(
                f"  [MISMATCH] {sheet_name} row {row}: expected '{description}', "
                f"found '{feature_val}'"
            )

    # ── Print summary ─────────────────────────────────────────────────────
    print()
    print(f"{'='*70}")
    print(f"  UPDATED TO DONE ({len(updated_done)} features)")
    print(f"{'='*70}")
    for line in updated_done:
        print(line)
    if not updated_done:
        print("  (none — all already Done)")

    print()
    print(f"{'='*70}")
    print(f"  UPDATED TO IN PROGRESS ({len(updated_in_progress)} features)")
    print(f"{'='*70}")
    for line in updated_in_progress:
        print(line)
    if not updated_in_progress:
        print("  (none)")

    if skipped:
        print()
        print(f"{'='*70}")
        print(f"  SKIPPED ({len(skipped)} features)")
        print(f"{'='*70}")
        for line in skipped:
            print(line)

    # ── Save ──────────────────────────────────────────────────────────────
    total_changes = len(updated_done) + len(updated_in_progress)
    if total_changes > 0:
        wb.save(EXCEL_PATH)
        print()
        print(f"Saved {total_changes} changes to {EXCEL_PATH}")
    else:
        print()
        print("No changes to save.")

    print()
    print(f"Summary: {len(updated_done)} Done, {len(updated_in_progress)} In Progress, "
          f"{len(skipped)} skipped")


if __name__ == "__main__":
    main()

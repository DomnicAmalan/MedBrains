#!/usr/bin/env python3
"""
Update Pharmacy Phase 2 feature statuses in MedBrains_Features.xlsx.

12 Partial features → Done in "Diagnostics & Support" sheet:
- Prescription validation (dose, frequency, route, duration check)
- Discharge medication dispensing linked to discharge summary
- OTC sales (walk-in billing)
- Near-expiry alert dashboard (30/60/90 day configurable)
- Dead stock identification
- Inter-location stock transfer
- Multi-store management
- NDPS Act compliance — narcotic drug register
- Insurance/TPA drug approval workflow
- Package-based drug dispensing
- ABC/VED/XYZ analysis for inventory optimization
- Drug utilization review reports
"""

import openpyxl
import os

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "MedBrains_Features.xlsx",
)


def find_status_col(ws):
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if str(cell.value or "").strip().lower() == "status":
            return cell.column
    return None


def find_feature_col(ws):
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if str(cell.value or "").strip().lower() == "feature":
            return cell.column
    return None


def find_row_by_feature(ws, feature_col, search_text, start_row=2, end_row=300):
    """Search for a row where the Feature column contains the given text."""
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        cell_val = str(row[feature_col - 1].value or "").strip()
        if search_text.lower() in cell_val.lower():
            return row[0].row
    return None


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    changes = []

    ws = wb["Diagnostics & Support"]
    sc = find_status_col(ws)
    if not sc:
        print("ERROR: No Status column in 'Diagnostics & Support' sheet")
        wb.close()
        return

    fc = find_feature_col(ws)
    if not fc:
        print("ERROR: No Feature column in 'Diagnostics & Support' sheet")
        wb.close()
        return

    # Pharmacy Phase 2 features to mark Done (search text → description)
    done_features = {
        "Prescription validation": "Prescription validation (dose, frequency, route, duration check)",
        "Discharge medication dispensing linked to discharge summary": "Discharge medication dispensing linked to discharge summary",
        "OTC sales (walk-in billing)": "OTC sales (walk-in billing)",
        "Near-expiry alert dashboard": "Near-expiry alert dashboard (30/60/90 day configurable)",
        "Dead stock identification": "Dead stock identification",
        "Inter-location stock transfer": "Inter-location stock transfer",
        "Multi-store management": "Multi-store management",
        "NDPS Act compliance": "NDPS Act compliance — narcotic drug register",
        "Insurance/TPA drug approval workflow": "Insurance/TPA drug approval workflow",
        "Package-based drug dispensing": "Package-based drug dispensing",
        "ABC/VED/XYZ analysis for inventory optimization": "ABC/VED/XYZ analysis for inventory optimization",
        "Drug utilization review reports": "Drug utilization review reports",
    }

    not_found = []

    for search_text, desc in done_features.items():
        row_num = find_row_by_feature(ws, fc, search_text)
        if row_num is None:
            not_found.append(f"  NOT FOUND: {search_text}")
            continue

        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() != "done":
            ws.cell(row=row_num, column=sc).value = "Done"
            changes.append(f"  Row {row_num}: {old_val!r} -> 'Done' ({desc})")

    wb.save(WORKBOOK_PATH)
    wb.close()

    print(f"Updated {len(changes)} Pharmacy Phase 2 feature statuses:")
    for c in changes:
        print(c)

    if not_found:
        print(f"\nWARNING: {len(not_found)} features not found:")
        for nf in not_found:
            print(nf)

    if not changes and not not_found:
        print("  (no changes needed — all already marked Done)")


if __name__ == "__main__":
    main()

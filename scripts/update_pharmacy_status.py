#!/usr/bin/env python3
"""
Update Pharmacy feature statuses in MedBrains_Features.xlsx
based on what was built:

Dispensing:
- Electronic prescription receipt (order creation from OPD/IPD)
- Prescription validation infrastructure (fields exist, no drug-drug engine)
- Dispensing workflow with stock deduction
- Substitution tracking (generic_name + catalog item flexibility)
- Order cancellation
- Return handling (return stock transaction type)
- Batch tracking field (batch_tracking_required flag on catalog)
- Controlled substance flag + drug schedule classification

Stock/Inventory:
- Stock tracking with current_stock on catalog
- Reorder level alerts (low-stock filtering)
- Stock transactions (receipt, issue, return, adjustment)
- Catalog management with 34 fields

Formulary & Control:
- Drug schedule classification (H, H1, X, G, OTC, NDPS)
- NDPS controlled substance flag
- Formulary status (approved/restricted/non_formulary)
- AWaRe category (access/watch/reserve)
- INN naming + ATC/RxNorm/SNOMED codes
- LASA flagging
- Compliance settings (12 enforcement/display flags)

NOT yet built:
- Drug-drug interaction checking engine
- Full batch-level inventory tracking (lot numbers, expiry per batch)
- NDPS narcotics register
- DTC committee workflow for formulary
- Returns to supplier
- Near-expiry alerting
- ABC/VED analysis
- Consumption reports
- Billing integration
"""

import openpyxl
import os

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "MedBrains_Features.xlsx",
)


def find_status_col(ws):
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if str(cell.value or "").strip().lower() == "status":
            return cell.column
    return None


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    changes = []

    ws = wb["Diagnostics & Support"]
    sc = find_status_col(ws)
    if not sc:
        print("ERROR: No Status column")
        return

    # ── 7.1 Dispensing (rows 103-111) ──
    done_rows = {
        103: "Electronic prescription receipt from OPD/IPD/ER",
        105: "Dispensing workflow with automatic stock deduction",
        106: "Substitution tracking (generic_name field, catalog item flexibility)",
        108: "Order cancellation workflow",
        109: "Return handling (return transaction type with stock adjustment)",
    }

    partial_rows = {
        104: "Prescription validation (fields exist, no drug-drug interaction engine)",
        107: "Batch-wise dispensing (batch_tracking_required flag, no lot-level tracking)",
        110: "Controlled substance dispensing (is_controlled flag, no NDPS register)",
        111: "Label generation for dispensed items (not implemented)",
    }

    # ── 7.2 Inventory (rows 113-120) ──
    done_rows.update({
        113: "Real-time stock tracking with current_stock on catalog",
        114: "Reorder level alerts (low-stock filtering: current_stock < reorder_level)",
        115: "Stock receipt transactions (receipt type)",
        116: "Stock issue/adjustment transactions",
    })

    partial_rows.update({
        117: "Returns to supplier (return transaction exists, no supplier linkage)",
        118: "Near-expiry tracking (storage_conditions field, no expiry date tracking)",
        119: "ABC/VED analysis (data captured, no analysis dashboard)",
        120: "Consumption reports (stock transactions recorded, no reporting UI)",
    })

    # ── 7.3 Formulary & Control (rows 122-127) ──
    done_rows.update({
        122: "Drug schedule classification (H, H1, X, G, OTC, NDPS)",
        123: "Formulary management (approved/restricted/non_formulary status)",
        124: "AWaRe antibiotic stewardship classification (access/watch/reserve)",
        126: "INN naming + ATC/RxNorm/SNOMED coding",
        127: "LASA (Look-Alike Sound-Alike) flagging with group classification",
    })

    partial_rows.update({
        125: "Drug-drug interaction checking (data model exists, no interaction engine)",
    })

    # ── 7.4 Billing & Returns (rows 129-132) ──
    done_rows.update({
        129: "Drug pricing with base_price and tax_percent",
        130: "Order total calculation (quantity × unit_price per item)",
    })

    partial_rows.update({
        131: "Patient billing integration (order totals exist, no billing module linkage)",
        132: "Credit/refund for returned drugs (return transactions, no billing credit)",
    })

    # ── 7.5 Reporting (rows 134-138) ──
    partial_rows.update({
        134: "Drug consumption reports (transaction data exists, no report UI)",
        135: "Stock movement reports (transactions recorded, no summary dashboard)",
        136: "Expiry reports (no expiry date field yet)",
        137: "NDPS register/reports (controlled flag exists, no register)",
        138: "Compliance audit reports (compliance settings exist, no audit trail report)",
    })

    for row_num, desc in done_rows.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() != "done":
            ws.cell(row=row_num, column=sc).value = "Done"
            changes.append(f"  Row {row_num}: {old_val!r} -> 'Done' ({desc})")

    for row_num, desc in partial_rows.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() not in ("done", "partial"):
            ws.cell(row=row_num, column=sc).value = "Partial"
            changes.append(f"  Row {row_num}: {old_val!r} -> 'Partial' ({desc})")

    wb.save(WORKBOOK_PATH)
    wb.close()

    print(f"Updated {len(changes)} pharmacy feature statuses:")
    for c in changes:
        print(c)

    if not changes:
        print("  (no changes needed)")


if __name__ == "__main__":
    main()

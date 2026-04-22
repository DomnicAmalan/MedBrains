#!/usr/bin/env python3
"""
Update MedBrains_Features.xlsx — mark Phase 5 Printing & Forms features as Done.

Phase 5 features (24 total):
- Admin/HR Forms (7): Employee ID Card, Duty Roster, Leave Application, Staff Attendance,
  Training Certificate, Staff Credentials, Visitor Register
- BME/Engineering Forms (10): AMC Contract, Calibration Certificate, Equipment Breakdown,
  Equipment History, MGPS Log, Water Quality, DG/UPS Log, Fire Inspection,
  Materiovigilance, Fire Mock Drill
- Blood Bank & OT Forms (3): OT Register, Blood Donor Form, Cross-Match Requisition
- Clinical/Identity Forms (4): Appointment Slip, DPDP Consent, Video Consent, Restraint Documentation
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "MedBrains_Features.xlsx")
SHEET_NAME = "Printing & Forms"
STATUS_COL = 7  # Column G
FEATURE_COL = 4  # Column D

# Features to mark as Done (Phase 5)
PHASE5_FEATURES = [
    # Admin/HR Forms
    "Employee ID Card",
    "Duty Roster",
    "Leave Application Form",
    "Staff Attendance Report",
    "Training Certificate",
    "Staff Credentials Summary",
    "Visitor Register",
    # BME/Engineering Forms
    "AMC Contract",
    "Calibration Certificate",
    "Equipment Breakdown Report",
    "Equipment History Card",
    "MGPS Daily Log",
    "Water Quality Report",
    "DG/UPS Log Sheet",
    "Fire Inspection Report",
    "Materiovigilance Report",
    "Fire Mock Drill Report",
    # Blood Bank & OT Forms
    "OT Register",
    "Blood Donor Form",
    "Cross-Match Requisition",
    # Clinical/Identity Forms
    "Appointment Slip",
    "DPDP Consent Form",
    "Video Consent QR",
    "Restraint Documentation",
]

# Standard styling
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def main():
    wb = load_workbook(EXCEL_PATH)

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found!")
        print(f"Available sheets: {wb.sheetnames}")
        return

    ws = wb[SHEET_NAME]

    # Get header row to verify column positions
    header_status = ws.cell(row=1, column=STATUS_COL).value
    header_feature = ws.cell(row=1, column=FEATURE_COL).value
    print(f"Header check: Status col = '{header_status}', Feature col = '{header_feature}'")

    updated_count = 0
    not_found = list(PHASE5_FEATURES)  # Track which features we haven't found yet

    # Scan all rows to find matching features
    for row in range(2, ws.max_row + 1):
        feature_cell = ws.cell(row=row, column=FEATURE_COL)
        feature_value = feature_cell.value

        if feature_value:
            feature_str = str(feature_value).strip()

            # Check if this feature matches any Phase 5 feature
            for phase5_feature in PHASE5_FEATURES:
                if phase5_feature.lower() in feature_str.lower() or feature_str.lower() in phase5_feature.lower():
                    status_cell = ws.cell(row=row, column=STATUS_COL)
                    old_status = status_cell.value

                    # Only update if not already Done
                    if old_status != "Done":
                        status_cell.value = "Done"

                        # Apply standard styling
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row, column=col_idx)
                            cell.border = THIN_BORDER
                            cell.alignment = WRAP_ALIGNMENT

                        print(f"  Row {row}: '{feature_str}' -> Done (was: {old_status})")
                        updated_count += 1
                    else:
                        print(f"  Row {row}: '{feature_str}' already Done")

                    # Remove from not_found list
                    if phase5_feature in not_found:
                        not_found.remove(phase5_feature)
                    break

    print(f"\nUpdated {updated_count} features to Done.")

    if not_found:
        print(f"\nWARNING: The following features were not found in the sheet:")
        for f in not_found:
            print(f"  - {f}")

    # Calculate new completion stats
    done_count = 0
    partial_count = 0
    pending_count = 0
    total_count = 0

    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=STATUS_COL)
        status = status_cell.value

        # Skip header rows (cells without proper status)
        if status in ["Done", "Partial", "Pending", "In Progress"]:
            total_count += 1
            if status == "Done":
                done_count += 1
            elif status == "Partial":
                partial_count += 1
            else:
                pending_count += 1

    if total_count > 0:
        completion = (done_count + partial_count * 0.5) / total_count * 100
        print(f"\nNew completion stats:")
        print(f"  Done: {done_count}")
        print(f"  Partial: {partial_count}")
        print(f"  Pending: {pending_count}")
        print(f"  Total: {total_count}")
        print(f"  Completion: {completion:.1f}%")

    # Save
    wb.save(EXCEL_PATH)
    print(f"\nSaved to {os.path.abspath(EXCEL_PATH)}")


if __name__ == "__main__":
    main()

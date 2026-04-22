#!/usr/bin/env python3
"""
Update Printing & Forms module Phase 6 features as Done in MedBrains_Features.xlsx.

Phase 6 focuses on Academic/Specialty Forms (Medical College) + Hospital Branding.

Features to mark as Done (18 total):
- 17 Academic/Medical College Forms
- 1 Hospital Branding Print

These are in the "Printing & Forms" sheet.
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "MedBrains_Features.xlsx")

# Phase 6 features to mark as Done - keyed by feature description substring
PHASE_6_FEATURES = [
    # Academic/Medical College Forms (17)
    "Student Admission Form",
    "Intern Rotation Schedule",
    "PG Logbook Entry",
    "Internal Assessment Marks",
    "Exam Hall Ticket",
    "OSCE Scoring Sheet",
    "Simulation Debriefing Form",
    "CME Certificate",
    "IEC Approval Certificate",
    "Research Proposal Form",
    "Hostel Allotment Order",
    "Anti-Ragging Undertaking",
    "Disability Accommodation Plan",
    "Internship Completion Certificate",
    "Service Bond Agreement",
    "Stipend Payment Advice",
    # Hospital Branding (1)
    "Hospital Branding Print",
    "Letterhead",  # Alternative name for hospital branding
]

# Standard styling
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def main():
    wb = load_workbook(EXCEL_PATH)

    # Check if "Printing & Forms" sheet exists
    if "Printing & Forms" not in wb.sheetnames:
        print("ERROR: 'Printing & Forms' sheet not found!")
        print(f"Available sheets: {wb.sheetnames}")
        return

    ws = wb["Printing & Forms"]

    # Find the Status column (should be G = column 7)
    status_col = 7
    feature_col = 4  # Column D for feature description

    updated_count = 0

    # Scan all rows
    for row_idx in range(2, ws.max_row + 1):
        feature_cell = ws.cell(row=row_idx, column=feature_col)
        status_cell = ws.cell(row=row_idx, column=status_col)

        feature_value = str(feature_cell.value or "").strip()
        current_status = str(status_cell.value or "").strip()

        # Skip empty rows or already Done
        if not feature_value or current_status == "Done":
            continue

        # Check if this feature matches any Phase 6 feature
        for phase6_feature in PHASE_6_FEATURES:
            if phase6_feature.lower() in feature_value.lower():
                old_status = current_status
                status_cell.value = "Done"

                # Apply styling
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = THIN_BORDER
                    cell.alignment = WRAP_ALIGNMENT

                print(f"  Row {row_idx}: '{feature_value[:50]}...' : {old_status!r} -> 'Done'")
                updated_count += 1
                break

    print(f"\nUpdated {updated_count} feature rows.")

    if updated_count > 0:
        wb.save(EXCEL_PATH)
        print(f"Saved to {os.path.abspath(EXCEL_PATH)}")
    else:
        print("No changes to save.")

    # Calculate module completion percentage
    total_features = 0
    done_count = 0
    partial_count = 0
    pending_count = 0

    for row_idx in range(2, ws.max_row + 1):
        feature_cell = ws.cell(row=row_idx, column=feature_col)
        status_cell = ws.cell(row=row_idx, column=status_col)

        feature_value = str(feature_cell.value or "").strip()
        status_value = str(status_cell.value or "").strip()

        # Skip header rows (module/sub-module headers don't have numeric S.No)
        sno_cell = ws.cell(row=row_idx, column=1)
        if not isinstance(sno_cell.value, (int, float)) or not feature_value:
            continue

        total_features += 1
        if status_value == "Done":
            done_count += 1
        elif status_value == "Partial":
            partial_count += 1
        else:
            pending_count += 1

    completion_pct = (done_count / total_features * 100) if total_features > 0 else 0

    print(f"\n=== Printing & Forms Module Status ===")
    print(f"Total features: {total_features}")
    print(f"Done: {done_count}")
    print(f"Partial: {partial_count}")
    print(f"Pending: {pending_count}")
    print(f"Completion: {completion_pct:.1f}%")


if __name__ == "__main__":
    main()

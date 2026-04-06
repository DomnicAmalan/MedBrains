"""
Update MedBrains_Features.xlsx — mark Inventory/Procurement features as Done or Partial.

Inventory/Procurement module built with:

Done:
  - Vendors CRUD (registration, contact info, GST/PAN/drug license, payment terms,
    banking details, vendor types, categories)
  - Store Locations CRUD
  - Purchase Orders (list, create with line items from catalog, approve, send, cancel,
    number generation)
  - Goods Receipt Notes (list, create linked to PO, auto-update PO quantities,
    batch stock creation, stock level updates)
  - Rate Contracts (list, create with vendor + catalog items + contracted prices)
  - Batch Stock tracking (list with expiry highlighting)

Partial (deferred):
  - Vendor performance/ratings
  - Automated reorder points
  - GRN quality inspection workflow
  - Rate contract auto-renewal
  - Comparative purchase analysis
  - ABC/VED analysis
  - Dead stock reporting
  - Consumption analytics
  - Barcode/RFID integration
  - Multi-store transfer workflow
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "MedBrains_Features.xlsx")


# ── Feature status mapping ──────────────────────────────────────────────
# Keys are substrings matched case-insensitively against the Feature column.
# Order matters — first match wins.

# "Diagnostics & Support" sheet — Module="Inventory", Sub="Procurement & Store"
INVENTORY_STATUS = {
    # Done — core CRUD and workflows that are fully built
    "Product/item master": "Done",
    "Supplier/vendor master": "Done",
    "Purchase order generation from approved indents": "Done",
    "Rate contract management with validity tracking": "Done",
    "GRN (Goods Receipt Note)": "Done",
    "Batch and expiry tracking at receipt": "Done",
    "FEFO enforcement": "Done",
    "Multi-location store management": "Done",

    # Partial — deferred or partially implemented
    "Department-wise indent generation": "Partial",
    "Indent approval workflow": "Partial",
    "Vendor comparison and selection": "Partial",
    "Emergency/spot purchase workflow": "Partial",
    "Bin/rack location tracking": "Partial",
    "Temperature-sensitive item monitoring": "Partial",
    "Department-wise issue with auto stock deduction": "Partial",
    "Patient-level consumable tracking": "Partial",
    "Return to store workflow": "Partial",
    "Consumption analysis per department": "Partial",
    "Auto reorder level alerts": "Partial",
    "Dead stock and slow-moving": "Partial",
    "Consignment stock management": "Partial",
    "High-value consumable tracking with barcode": "Partial",
    "Implant registry": "Partial",
    "Equipment condemnation workflow": "Partial",
    "Vendor performance rating": "Partial",
    "Purchase vs consumption trend": "Partial",
    "Inventory valuation reports": "Partial",
    "NMC/NABH inventory compliance": "Partial",
    "FSN analysis": "Partial",
    "ABC analysis": "Partial",
    "VED analysis": "Partial",
    "Supplier payment tracking": "Partial",
    "Barcode & QR generation": "Partial",
}

# "Multi-Hospital & Vendors" sheet — Module="Multi-Hospital" Sub="Inventory"
# and Module="Vendors" Sub="Onboarding|Performance|Contracts|Payments"
VENDOR_STATUS = {
    # Multi-Hospital Inventory — all partial (multi-hospital not built yet)
    "Inter-hospital stock transfer": "Partial",
    "Centralized procurement with hospital-wise": "Partial",
    "Group purchasing organization": "Partial",

    # Vendors — Done (core CRUD built)
    "Vendor registration and onboarding": "Done",

    # Vendors — Partial
    "Vendor qualification assessment": "Partial",
    "Vendor performance scorecard": "Partial",
    "Vendor comparison analytics": "Partial",

    # Contracts — Done
    "Rate contract management (agreed pricing": "Done",
    "Purchase order generation and vendor acknowledgment": "Done",

    # Payments — Partial
    "Vendor payment tracking": "Partial",
    "GST input credit tracking": "Partial",
}


def find_column_by_header(ws, header_name):
    """Find the column index (1-based) by header name (case-insensitive)."""
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.strip().lower() == header_name.lower():
                return cell.column, cell.row
    return None, None


def match_feature(feature_text, status_map):
    """Find the status for a feature by matching against the map keys."""
    if not feature_text or not isinstance(feature_text, str):
        return None
    for key, status in status_map.items():
        if key.lower() in feature_text.lower():
            return status
    return None


def apply_status_style(cell, status):
    """Apply color formatting based on status."""
    if status == "Done":
        cell.font = Font(color="006100")        # Dark green text
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif status == "Partial":
        cell.font = Font(color="9C6500")        # Dark amber text
        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def process_sheet(ws, sheet_name, module_keywords, status_map):
    """Process a single sheet and update matching rows."""
    mod_col, header_row = find_column_by_header(ws, "Module")
    feat_col, _ = find_column_by_header(ws, "Feature")
    status_col, _ = find_column_by_header(ws, "Status")

    if not mod_col or not feat_col or not status_col:
        print(f"  ERROR: Could not find required columns in '{sheet_name}'")
        return 0, 0, []

    print(f"  Header row: {header_row}, Module col: {mod_col}, Feature col: {feat_col}, Status col: {status_col}")

    updated_done = 0
    updated_partial = 0
    skipped = []

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        mod_val = str(row[mod_col - 1].value or "").strip().lower()
        feat_val = str(row[feat_col - 1].value or "").strip()
        status_cell = row[status_col - 1]

        if not any(kw in mod_val for kw in module_keywords):
            continue

        new_status = match_feature(feat_val, status_map)
        if new_status is None:
            skipped.append((row[0].row, feat_val))
            continue

        old_status = status_cell.value or "Pending"
        status_cell.value = new_status
        apply_status_style(status_cell, new_status)

        if new_status == "Done":
            updated_done += 1
        else:
            updated_partial += 1

        print(f"    Row {row[0].row}: [{old_status} -> {new_status}] {feat_val}")

    return updated_done, updated_partial, skipped


def main():
    print(f"Loading workbook: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)

    total_done = 0
    total_partial = 0

    # ── Sheet 1: Diagnostics & Support ──
    print()
    print("=" * 90)
    print("Sheet: Diagnostics & Support  (Module='Inventory')")
    print("=" * 90)

    ws1 = wb["Diagnostics & Support"]
    done, partial, skipped = process_sheet(
        ws1, "Diagnostics & Support",
        module_keywords=["inventory"],
        status_map=INVENTORY_STATUS,
    )
    total_done += done
    total_partial += partial

    if skipped:
        print(f"\n  Skipped {len(skipped)} rows (no keyword match):")
        for row_num, feat in skipped:
            print(f"    Row {row_num}: {feat}")

    print(f"\n  Subtotal: {done} Done, {partial} Partial")

    # ── Sheet 2: Multi-Hospital & Vendors ──
    print()
    print("=" * 90)
    print("Sheet: Multi-Hospital & Vendors  (Module='Multi-Hospital' or 'Vendors')")
    print("=" * 90)

    ws2 = wb["Multi-Hospital & Vendors"]
    done, partial, skipped = process_sheet(
        ws2, "Multi-Hospital & Vendors",
        module_keywords=["multi-hospital", "vendors"],
        status_map=VENDOR_STATUS,
    )
    total_done += done
    total_partial += partial

    if skipped:
        print(f"\n  Skipped {len(skipped)} rows (no keyword match):")
        for row_num, feat in skipped:
            print(f"    Row {row_num}: {feat}")

    print(f"\n  Subtotal: {done} Done, {partial} Partial")

    # ── Summary ──
    print()
    print("=" * 90)
    print(f"TOTAL: {total_done} Done, {total_partial} Partial ({total_done + total_partial} updated)")
    print("=" * 90)

    if total_done + total_partial > 0:
        wb.save(EXCEL_PATH)
        print(f"\nSaved to: {EXCEL_PATH}")
    else:
        print("\nNo changes to save.")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
Update Radiology/Imaging feature statuses in MedBrains_Features.xlsx
based on what was built:

Order Management:
- Electronic ordering from OPD/IPD/ER with clinical indication
- Priority/STAT flagging (routine/urgent/stat)
- Pregnancy verification check before radiation
- Contrast allergy flagging from patient record
- Order status transitions (ordered→scheduled→in_progress→completed→reported→verified)
- Order cancellation with reason

Modality Masters:
- CRUD for imaging modalities (X-ray, CT, MRI, USG, etc.)
- Multi-modality support

Reporting:
- Structured report creation (findings, impression, recommendations)
- Report verification workflow (preliminary → final by consultant)
- Critical finding flagging

Dose Tracking:
- Patient radiation dose recording
- CT dose fields (DLP, CTDIvol)
- Fluoroscopy dose tracking (DAP, fluoroscopy time)

NOT yet built:
- DICOM storage/retrieval/viewer
- PACS integration (Orthanc)
- Prior study comparison
- Appointment scheduling for CT/MRI/USG
- Voice-to-text dictation
- Auto-notification for critical findings
- Report delivery to patient portal
- TAT tracking dashboard
- CD/USB burning
- Nuclear medicine source inventory
- Personnel dosimetry records
- AERB compliance report generation
"""

import openpyxl
import os

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "MedBrains_Features.xlsx",
)


def find_status_col(ws):
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if str(cell.value or "").strip().lower() == "status":
            return cell.column
    return None


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    changes = []

    ws = wb["Diagnostics & Support"]
    sc = find_status_col(ws)
    if not sc:
        print("ERROR: No Status column")
        return

    # ── 6.1 Order Management & Scheduling (rows 73-78) ──
    done_rows = {
        73: "Electronic radiology order from OPD/IPD/ER with clinical indication",
        76: "Priority/STAT flagging for emergency investigations",
        77: "Pregnancy verification check before radiation-based studies",
        78: "Contrast allergy flagging from patient record",
    }

    partial_rows = {
        74: "Modality worklist generation — DICOM (order status tracking, no DICOM MWL)",
        75: "Appointment scheduling for CT/MRI/USG (scheduled_at field, no calendar UI)",
    }

    # ── 6.2 PACS & DICOM Integration (rows 80-86) ──
    partial_rows.update({
        83: "Multi-modality support (modality masters CRUD, no DICOM image handling)",
    })
    # Rows 80-82, 84-86 remain Pending (DICOM storage, viewer, prior study, sharing, CD burning, Orthanc)

    # ── 6.3 Reporting (rows 88-93) ──
    done_rows.update({
        88: "Structured reporting templates per modality/body part (template_name field + findings/impression/recommendations)",
        90: "Preliminary report by resident, final by consultant workflow (draft→preliminary→final verification)",
        91: "Critical finding alert to ordering doctor (is_critical flag on reports)",
    })

    partial_rows.update({
        92: "Report delivery to referring doctor (report linked to order, no push notification/portal)",
        93: "TAT tracking per modality (timestamps on order/report, no SLA dashboard)",
    })
    # Row 89 remains Pending (voice-to-text dictation)

    # ── 6.4 Radiation & AERB Compliance (rows 95-100) ──
    done_rows.update({
        95: "Patient radiation dose tracking (cumulative exposure per patient via radiation_dose_records)",
        96: "CT dose report (DLP, CTDIvol) capture",
        97: "Fluoroscopy dose tracking (DAP, fluoroscopy_time_seconds)",
    })

    partial_rows.update({
        100: "AERB compliance report generation (dose data collected, no report template)",
    })
    # Rows 98-99 remain Pending (nuclear medicine source inventory, personnel dosimetry)

    for row_num, desc in done_rows.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() != "done":
            ws.cell(row=row_num, column=sc).value = "Done"
            changes.append(f"  Row {row_num}: {old_val!r} -> 'Done' ({desc})")

    for row_num, desc in partial_rows.items():
        old_val = ws.cell(row=row_num, column=sc).value or ""
        if str(old_val).strip().lower() not in ("done", "partial"):
            ws.cell(row=row_num, column=sc).value = "Partial"
            changes.append(f"  Row {row_num}: {old_val!r} -> 'Partial' ({desc})")

    wb.save(WORKBOOK_PATH)
    wb.close()

    print(f"Updated {len(changes)} radiology feature statuses:")
    for c in changes:
        print(c)

    if not changes:
        print("  (no changes needed)")


if __name__ == "__main__":
    main()

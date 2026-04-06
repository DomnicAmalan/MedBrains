"""
Update Regulatory & Compliance sheet — mark features Done/Partial
after building the Regulatory module (migration 072, backend routes,
frontend 5-tab page, 12 permissions).

New tables: compliance_checklists, compliance_checklist_items, adr_reports,
materiovigilance_reports, pcpndt_forms, compliance_calendar
"""
import openpyxl

XLSX = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"
SHEET = "Regulatory & Compliance"

# Row -> new status mapping
# Only upgrade status (Pending→Done/Partial, Partial→Done)
STATUS_UPDATES = {
    # ── 1.1 NABH ──────────────────────────────────────────────
    # R9: NABH readiness dashboard — NOW have compliance dashboard
    #     with accreditation body scoring, department-wise compliance, gap analysis
    9: "Done",

    # ── 1.3 CDSCO ─────────────────────────────────────────────
    # R16: ADR reporting to PvPI — adr_reports table, CRUD + submit-to-PvPI workflow
    16: "Done",
    # R17: Materiovigilance reporting — materiovigilance_reports table, CRUD + submit-to-CDSCO
    17: "Done",

    # ── 1.4 PCPNDT & Others ─────────────────────────────────
    # R24: PCPNDT Form F — pcpndt_forms table, CRUD + quarterly report
    24: "Done",
    # R25: PCPNDT auto-block gender disclosure — gender_disclosure_blocked column, UI enforcement
    25: "Done",
    # R29: Fire safety compliance tracking — FMS module has fire_equipment, fire_inspections,
    #      fire_drills, fire_noc tables. Compliance calendar now aggregates all deadlines.
    29: "Done",

    # ── 2.1 JCI ────────────────────────────────────────────────
    # R35: JCI compliance dashboard — NOW have compliance dashboard with body-specific scoring
    35: "Done",

    # ── Additional upgrades from compliance_checklists + calendar ──
    # R6: NABH mandatory document generation → still Partial (templates exist but no auto-gen)
    # R8: Internal audit scheduler → Partial upgraded: compliance_calendar has audit scheduling
    8: "Partial",  # Already Partial, no change — compliance_calendar adds scheduling but not automated

    # R33: International Patient Safety Goals tracking → already Partial, compliance checklists
    #      can now track IPSG per department
    33: "Partial",
}

wb = openpyxl.load_workbook(XLSX)
ws = wb[SHEET]

updated = 0
for row_num, new_status in STATUS_UPDATES.items():
    cell = ws.cell(row=row_num, column=7)  # Column G = Status
    old = str(cell.value).strip() if cell.value else "Pending"

    # Only upgrade: Pending→Done/Partial, Partial→Done. Never downgrade.
    should_update = False
    if old == "Pending" and new_status in ("Done", "Partial"):
        should_update = True
    elif old == "Partial" and new_status == "Done":
        should_update = True

    if should_update:
        cell.value = new_status
        feat = ws.cell(row=row_num, column=4).value or ""
        print(f"  R{row_num:3d}: {old:12s} -> {new_status:12s} | {feat[:70]}")
        updated += 1

wb.save(XLSX)

# Count final statuses
counts = {}
for row in ws.iter_rows(min_row=2, max_col=7, values_only=False):
    sno = row[0].value
    status = row[6].value
    if sno is not None and isinstance(sno, (int, float)):
        counts[status] = counts.get(status, 0) + 1

print(f"\nUpdated {updated} features")
print(f"\nFinal status counts:")
for s, c in sorted(counts.items(), key=lambda x: -x[1]):
    pct = c / sum(counts.values()) * 100
    print(f"  {s:12s}: {c:3d} ({pct:.1f}%)")
print(f"  {'TOTAL':12s}: {sum(counts.values()):3d}")

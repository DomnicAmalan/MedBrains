"""
Update Regulatory & Compliance sheet in MedBrains_Features.xlsx
Based on comprehensive codebase audit of existing infrastructure.
"""
import openpyxl

XLSX = "/Users/apple/Projects/MedBrains/MedBrains_Features.xlsx"
SHEET = "Regulatory & Compliance"

# Row -> new status mapping
# Based on codebase audit:
# - quality_accreditation_standards table supports NABH/JCI/NMC/NABL/ABDM/NAAC
# - quality_indicators, quality_documents, quality_incidents, quality_capa exist
# - quality_audits with scoring and findings exist
# - quality_accreditation_compliance tracks per-standard compliance status
# - pharmacy_catalog has drug_schedule, is_controlled, inn_name, atc_code, etc.
# - compliance settings (12 enforcement/display toggles) exist
# - infection_control has HAI surveillance, BMW waste tracking, hand hygiene, outbreaks
# - emergency has MLC cases, police intimation, triage, code activations
# - audit_log with hash chain exists
# - regulatory_bodies table with NABH/JCI/ABDM/etc. seeded
# - field_regulatory_links (167 mappings) exist
# - patient_consents and procedure_consents exist
# - geo_countries, geo_states, geo_districts with pre-seeded India data exist
# - regulatory auto-detection in onboarding GeoRegulatoryStep exists
# - RegulatoryManagement.tsx and ComplianceSettings.tsx admin pages exist

STATUS_UPDATES = {
    # ── 1.1 NABH ──────────────────────────────────────────────
    # R4: NABH 5th edition standards mapping — quality_accreditation_standards supports nabh body,
    #      167 field-regulatory links with NABH clause references exist
    4: "Done",
    # R5: NABH entry-level vs full accreditation criteria tracking
    #      quality_accreditation_standards has version field, compliance tracking exists
    5: "Done",
    # R6: NABH mandatory document generation (policies, SOPs)
    #      quality_documents with lifecycle (draft→released) and version control exist
    #      But no auto-generation of specific NABH documents
    6: "Partial",
    # R7: Quality indicator tracking (NABH 32 indicators)
    #      quality_indicators with frequency/thresholds/benchmarks + quality_indicator_values exist
    7: "Done",
    # R8: Internal audit scheduler with NABH compliance scoring
    #      quality_audits with audit_type, findings, score exist
    #      But no automated scheduling
    8: "Partial",
    # R9: NABH readiness dashboard — chapter-wise completion percentage
    #      quality_accreditation_compliance tracks per-standard status
    #      But no dedicated dashboard with chapter-wise rollup
    9: "Partial",

    # ── 1.2 NMC ───────────────────────────────────────────────
    # R11: NMC medical college compliance (teaching hospitals)
    #       quality_accreditation_standards supports nmc body type
    #       But no specific medical college compliance workflows
    11: "Partial",
    # R12: NMC doctor registration verification integration
    #       medical_registration_number field exists on users (doctors)
    #       But no NMC API integration for verification
    12: "Partial",
    # R13: NMC fee structure compliance reporting
    #       No specific implementation
    13: "Pending",

    # ── 1.3 CDSCO ─────────────────────────────────────────────
    # R15: Drug schedule compliance (Schedule H, H1, X marking)
    #       pharmacy_catalog has drug_schedule, enforce_drug_scheduling toggle
    #       ComplianceSettings.tsx has enforcement toggles
    15: "Done",
    # R16: ADR reporting to PvPI
    #       No specific PvPI integration (pharmacy has basic adverse event fields)
    16: "Pending",
    # R17: Materiovigilance reporting (medical device adverse events)
    #       No specific implementation
    17: "Pending",
    # R18: Hemovigilance reporting (blood transfusion adverse events)
    #       transfusion_records with reaction reporting exist in blood bank module
    #       But no CDSCO-format export
    18: "Partial",

    # ── 1.4 NABL, AERB, PCPNDT & Others ──────────────────────
    # R20: NABL accreditation compliance tracking for laboratory
    #       quality_accreditation_standards supports nabl body type
    #       quality_accreditation_compliance tracks status per standard
    20: "Done",
    # R21: NABL document control — SOPs, calibration records
    #       quality_documents system exists but no lab-specific calibration tracking
    21: "Partial",
    # R22: AERB compliance for radiology/nuclear medicine
    #       Radiology module exists, no specific AERB compliance tracking
    22: "Partial",
    # R23: Radiation dose tracking and personnel TLD badge monitoring
    #       radiation_dose_records table exists with DLP, CTDIvol, DAP, fluoroscopy_time
    #       Cumulative patient dose tracking exists
    #       But no personnel TLD badge monitoring
    23: "Partial",
    # R24: PCPNDT Form F mandatory for pregnant ultrasounds
    #       No specific PCPNDT workflow implementation
    24: "Pending",
    # R25: PCPNDT auto-block gender disclosure in fetal imaging
    #       No specific implementation
    25: "Pending",
    # R26: IRDAI compliant claim format generation
    #       insurance_claims table exists in billing module
    #       But no IRDAI-specific format export
    26: "Partial",
    # R27: ABDM — ABHA creation, HIP/HIU integration, consent manager
    #       patient_consents has abdm_linking consent type
    #       regulatory_bodies has ABDM entry
    #       But no ABDM API integration
    27: "Partial",
    # R28: CPCB biomedical waste reporting (BMW Rules 2016)
    #       biowaste_records with 7 BMW waste categories (yellow/red/white/blue/cytotoxic/chemical/radioactive) exist
    #       Full CRUD in infection control module
    28: "Done",
    # R29: Fire safety compliance tracking
    #       No specific fire safety module
    29: "Pending",

    # ── 2.1 JCI ────────────────────────────────────────────────
    # R32: JCI accreditation standards mapping (14 chapters)
    #       quality_accreditation_standards supports jci body type
    #       167 field-regulatory links include JCI clause refs (IPSG.1, PFR.5.1, etc.)
    32: "Done",
    # R33: International Patient Safety Goals tracking
    #       quality_indicators for safety metrics exist
    #       Patient identification (2 identifiers), consent, medication safety built in
    33: "Partial",
    # R34: JCI tracer methodology support
    #       quality_audits support focused audits
    #       But no specific tracer methodology workflow
    34: "Partial",
    # R35: JCI compliance dashboard with gap analysis
    #       quality_accreditation_compliance exists
    #       But no dedicated JCI dashboard
    35: "Partial",
    # R36: JCI quality improvement — PDSA cycle tracking
    #       quality_capa with corrective/preventive actions exist
    #       But no specific PDSA cycle UI
    36: "Partial",

    # ── 2.2 US — HIPAA, CMS, ONC ──────────────────────────────
    # R38: HIPAA Privacy Rule compliance
    #       audit_log with hash chain, patient consent management, access controls exist
    #       But no specific HIPAA minimum necessary standard workflow
    38: "Partial",
    # R39: HIPAA Security Rule — safeguards checklist
    #       Permission system (111 permissions), RLS, encrypted passwords exist
    #       But no specific HIPAA safeguards checklist UI
    39: "Partial",
    # R40: HIPAA breach notification workflow
    #       quality_incidents with regulatory_reported_at field exist
    #       But no specific 60-day notification workflow
    40: "Partial",
    # R41: CMS Conditions of Participation compliance
    #       No specific implementation
    41: "Pending",
    # R42: ONC Health IT certification criteria
    #       No specific implementation
    42: "Pending",

    # ── 2.3 EU — GDPR & CE ────────────────────────────────────
    # R44: GDPR data processing agreements
    #       regulatory_bodies has GDPR entry
    #       patient_consents has data_sharing type
    #       But no DPA management UI
    44: "Partial",
    # R45: GDPR rights (access, erasure, portability)
    #       No specific data subject request workflows
    45: "Pending",
    # R46: DPO dashboard with DPIA
    #       No specific implementation
    46: "Pending",
    # R47: EU MDR medical device regulation compliance
    #       No specific implementation
    47: "Pending",

    # ── 2.4 Middle East ────────────────────────────────────────
    # R49-54: DHA, DOH, MOH-SA, QCHP, MOH-Oman — all region-specific integrations
    #          Framework exists (regulatory_bodies, accreditation_standards)
    #          But no specific integrations
    49: "Pending",
    50: "Pending",
    51: "Pending",
    52: "Pending",
    53: "Pending",
    54: "Pending",

    # ── 2.5 Africa, Asia-Pacific & Others ──────────────────────
    # R56-61: All region-specific integrations — not implemented
    56: "Pending",
    57: "Pending",
    58: "Pending",
    59: "Pending",
    60: "Pending",
    61: "Pending",

    # ── 2.6 ISO & Security Certifications ──────────────────────
    # R63: ISO 27001 ISMS control mapping
    #       quality_accreditation_standards supports "other" body type
    #       But no specific ISO 27001 controls mapped
    63: "Partial",
    # R64: ISO 9001 quality management
    #       quality management module covers QMS concepts
    #       But no specific ISO 9001 mapping
    64: "Partial",
    # R65: SOC 2 Type II evidence collection
    #       audit_log provides evidence trail
    #       But no specific SOC 2 control monitoring
    65: "Partial",
    # R66: HITRUST CSF certification readiness
    #       No specific implementation
    66: "Pending",
    # R67: CERT-In 6-hour incident reporting, 180-day log retention
    #       quality_incidents with regulatory reporting exist
    #       audit_log captures all changes
    #       But no specific CERT-In 6-hour workflow
    67: "Partial",

    # ── 3.1 Multi-Layer Geographic Dataset ─────────────────────
    # R70: Country master — name, ISO code, WHO region, regulatory bodies mapping
    #       geo_countries table with iso_code, default_locale, timezone, etc. (14 countries seeded)
    #       regulatory_bodies has country_id linkage
    70: "Done",
    # R71: Country-level regulation engine — auto-apply rules based on country
    #       GeoRegulatoryStep auto-detects regulators by country/state
    #       But no full "regulation engine" that auto-applies rules
    71: "Partial",
    # R72: State / Province layer
    #       geo_states table with state_code, pre-seeded (36 Indian states)
    72: "Done",
    # R73: State-level health department reporting integration
    #       No specific state health dept API integration
    73: "Pending",
    # R74: District layer with lat/long, boundaries
    #       geo_districts table pre-seeded (~780 Indian districts)
    #       But no lat/long or boundary polygons
    74: "Partial",
    # R75: District health office reporting
    #       No specific implementation
    75: "Pending",
    # R76: Subdistrict / Taluk layer
    #       No subdistrict table exists
    76: "Pending",
    # R77: Town / Village layer
    #       No village table exists
    77: "Pending",
    # R78: Facility layer — hospital coordinates, license number
    #       facilities table has address, registration fields
    #       But no coordinates/license_number specific fields
    78: "Partial",

    # ── 3.2 Coordinate-Based Regulator Auto-Mapping ────────────
    # R80: Geohash-based hospital-to-regulator mapping
    #       No geohash implementation
    80: "Pending",
    # R81: Auto-detect applicable regulatory bodies from coordinates
    #       GeoRegulatoryStep does auto-detection from country/state selection (not coordinates)
    81: "Partial",
    # R82: District-level inspector assignment
    #       No specific implementation
    82: "Pending",
    # R83: GIS map visualization
    #       No GIS/map implementation
    83: "Pending",
    # R84: Multi-jurisdiction compliance view
    #       No specific implementation
    84: "Pending",
    # R85: Regulatory change notification
    #       No specific implementation
    85: "Pending",

    # ── 3.3 Geographic Data Management ─────────────────────────
    # R87: Geographic hierarchy seed data with coordinates
    #       geo_countries + geo_states + geo_districts seeded (migration 030)
    #       But no coordinates (lat/long)
    87: "Partial",
    # R88: Boundary polygon storage (PostGIS/GeoJSON)
    #       No PostGIS/GeoJSON implementation
    88: "Pending",
    # R89: Address geocoding service integration
    #       No geocoding integration
    89: "Pending",
    # R90: Postal code to geographic hierarchy resolution
    #       No PIN code lookup API integration
    90: "Pending",
    # R91: Nearest facility finder
    #       No geospatial proximity search
    91: "Pending",
    # R92: Ambulance routing using geographic data
    #       No routing implementation
    92: "Pending",
    # R93: Catchment area analysis
    #       No heat map / catchment analysis
    93: "Pending",
}

wb = openpyxl.load_workbook(XLSX)
ws = wb[SHEET]

updated = 0
for row_num, new_status in STATUS_UPDATES.items():
    cell = ws.cell(row=row_num, column=7)  # Column G = Status
    old = cell.value
    if old != new_status:
        cell.value = new_status
        feat = ws.cell(row=row_num, column=4).value or ""
        print(f"  R{row_num:3d}: {old!s:12s} -> {new_status:12s} | {feat[:70]}")
        updated += 1

wb.save(XLSX)

# Count final statuses
counts = {}
for row in ws.iter_rows(min_row=2, max_col=7, values_only=False):
    sno = row[0].value
    status = row[6].value
    if sno is not None and isinstance(sno, (int, float)):
        counts[status] = counts.get(status, 0) + 1

print(f"\nUpdated {updated} features")
print(f"\nFinal status counts:")
for s, c in sorted(counts.items(), key=lambda x: -x[1]):
    pct = c / sum(counts.values()) * 100
    print(f"  {s:12s}: {c:3d} ({pct:.1f}%)")
print(f"  {'TOTAL':12s}: {sum(counts.values()):3d}")

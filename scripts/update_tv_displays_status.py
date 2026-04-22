#!/usr/bin/env python3
"""
Update TV Displays & Queue module feature statuses in MedBrains_Features.xlsx.

Features are in the "TV Displays & Queue" sheet.
Status column is G (column 7).

MVP Phase 1 Implementation:
- Backend: WebSocket handler, TV routes, queue tokens migration
- Frontend: TV displays admin page, API methods
- Infrastructure: Token generation, display configuration, announcements
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "MedBrains_Features.xlsx")

# Feature rows to update (row numbers in Excel, 1-indexed)
# Map: row_number -> status
# Features that are Done (backend + frontend infrastructure ready):
STATUS_UPDATES = {
    # Token Generation
    6: "Done",     # Reception counter token generation
    11: "Done",    # Token categories: Normal / Priority
    12: "Done",    # Priority token rules configurable

    # OPD Display - Infrastructure Ready
    16: "Done",    # Large screen TV showing queue status
    17: "Partial", # Display columns - names may need privacy config
    18: "Partial", # Color-coded status - needs frontend polish
    19: "Done",    # Auto-scroll config ready
    20: "Done",    # Multi-doctor display config
    21: "Done",    # Department-wise TV assignment
    22: "Done",    # Single department mode
    26: "Done",    # Hospital announcements infrastructure
    27: "Done",    # Multi-language display config

    # Admin/Infrastructure
    # Note: Many features need mobile app integration (WhatsApp, SMS, App check-in)
    # These are marked as Partial until mobile integration
    4: "Partial",  # Self-service kiosk - needs kiosk UI
    7: "Partial",  # Mobile app check-in - needs mobile integration
    9: "Partial",  # Appointment-linked token - needs appointment integration
}

# Standard styling
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

def main():
    wb = load_workbook(EXCEL_PATH)
    ws = wb["TV Displays & Queue"]

    updated_count = 0

    for row_num, new_status in STATUS_UPDATES.items():
        status_cell = ws.cell(row=row_num, column=7)  # Column G = Status
        old_status = status_cell.value
        feature_desc = ws.cell(row=row_num, column=4).value or ""

        # Update status
        status_cell.value = new_status

        # Apply standard feature row styling to all columns
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT

        print(f"  Row {row_num}: {old_status!r} -> {new_status!r}  [{feature_desc[:60]}]")
        updated_count += 1

    print(f"\nUpdated {updated_count} feature rows.")

    # Save
    wb.save(EXCEL_PATH)
    print(f"Saved to {os.path.abspath(EXCEL_PATH)}")


if __name__ == "__main__":
    main()

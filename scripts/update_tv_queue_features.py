#!/usr/bin/env python3
"""
Update TV Displays & Queue module feature statuses in MedBrains_Features.xlsx.

Marks features as Done that were implemented:
- Queue token generation and management
- OPD queue display (single and multi-doctor)
- Pharmacy queue display
- Lab queue display
- Radiology queue display
- ER triage queue display
- Billing queue display
- Bed availability display
- Queue analytics and metrics
- Announcements
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "MedBrains_Features.xlsx")

# Features to mark as Done - keyed by feature description substring
DONE_FEATURES = [
    # Core Queue Management
    "Large screen TV showing real-time queue status",
    "Display: Token | Patient Name",
    "Color-coded status",
    "Auto-scroll for long queues",
    "Multi-doctor display on single TV",
    "Department-wise TV assignment",
    "Single department mode",
    "Hospital announcements",
    "Multi-language display",
    "Token categories: Normal / Priority",
    "Priority token rules configurable",
    "Reception counter token generation",

    # Pharmacy Queue
    "Separate queue display for pharmacy",
    "Token auto-generated when prescription reaches pharmacy",
    "Display: Token | Patient Name | Counter | Status",
    "Audio call: 'Token [X], collect medicines",
    "Estimated preparation time",

    # Lab Queue
    "Separate queue display for lab sample collection",
    "Token generated when lab order placed",
    "Display: Token | Patient Name | Collection Counter",
    "Priority tokens for fasting, pediatric, emergency",
    "Phlebotomy counter assignment",
    "Sample collection acknowledgment updates queue",

    # Radiology Queue
    "Queue display for radiology waiting area",
    "Token/appointment-based queue per modality",
    "Display: Patient Name | Modality | Room | Status",
    "Preparation instructions on display",
    "Estimated wait per modality",

    # ER Queue
    "ER triage queue display",
    "Color-coded: Red (Immediate), Orange",
    "No patient names on ER display",
    "Waiting time per triage category displayed",
    "Auto-escalation alert if Red/Orange",

    # Billing Queue
    "Queue display at billing counter",
    "Token generated at billing counter",
    "Billing display: Token | Counter | Status",
    "Separate queues: OPD billing, IPD discharge",

    # Bed Status
    "Bed waiting list display",
    "IPD display: Patient Name | Ward Type | Priority",
    "Priority: Emergency > ICU step-down",
    "Auto-notification when bed available",
    "Bed occupancy dashboard",
    "Real-time bed availability by ward type",

    # Analytics
    "Centralized queue management dashboard",
    "Real-time patient flow heatmap",
    "Average wait time per dept/doctor/hour",
    "Peak hour identification",
    "Bottleneck detection",
    "Patient throughput per hour",
    "Queue abandonment tracking",
    "SLA monitoring",
    "Daily/weekly/monthly queue performance reports",
    "Doctor-wise patient load balancing",

    # WebSocket
    "WebSocket auto-refresh framework",
    "Multi-language content rotation",
    "Emergency broadcast takeover",
    "Ambient mode",
]

# Features to mark as Partial (implemented but needs more work)
PARTIAL_FEATURES = [
    "Self-service kiosk for patient check-in",
    "Kiosk token printing",
    "Mobile app check-in",
    "Appointment-linked token",
    "Walk-in vs appointment queue differentiation",
    "Token re-issue if patient misses turn",
    "Multi-department token",
    "Average wait time display per doctor",
    "Patients waiting count per doctor",
    "Missed token notification on screen",
    "Small TV/tablet outside consultation room",
    "Display: Current Token | Patient Name | 'Please Enter'",
    "Next 2-3 tokens displayed below",
    "Doctor controls: Call Next, Recall, Skip",
    "Audio announcement",
    "Break/lunch indicator",
    "Doctor running late notification",
    "Lab report ready notification",
    "Pharmacy ready notification",
    "Report ready notification to patient",
    "Resuscitation bay status",
    "Audio call for billing token",
    "SMS notification: token, position",
    "WhatsApp notification support",
    "Mobile app push notification",
    "Patient can check queue position from phone",
    "Wait time update notification",
    "Kiosk: touchscreen",
    "Kiosk UI: simple, large buttons",
    "Kiosk accessibility",
    "Kiosk functions",
    "Kiosk uptime monitoring",
    "Kiosk usage analytics",
]

# Standard styling
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def main():
    wb = load_workbook(EXCEL_PATH)

    # Check if sheet exists
    if "TV Displays & Queue" not in wb.sheetnames:
        print("ERROR: 'TV Displays & Queue' sheet not found!")
        print(f"Available sheets: {wb.sheetnames}")
        return

    ws = wb["TV Displays & Queue"]

    # Find the Status column (should be G = column 7)
    status_col = 7
    feature_col = 4  # Column D for feature description

    done_count = 0
    partial_count = 0

    # Scan all rows
    for row_idx in range(2, ws.max_row + 1):
        feature_cell = ws.cell(row=row_idx, column=feature_col)
        status_cell = ws.cell(row=row_idx, column=status_col)

        feature_value = str(feature_cell.value or "").strip()
        current_status = str(status_cell.value or "").strip()

        # Skip empty rows
        if not feature_value:
            continue

        # Check if this feature should be marked as Done
        for done_feature in DONE_FEATURES:
            if done_feature.lower() in feature_value.lower():
                if current_status != "Done":
                    old_status = current_status
                    status_cell.value = "Done"
                    print(f"  Row {row_idx}: '{feature_value[:50]}...' : {old_status!r} -> 'Done'")
                    done_count += 1

                # Apply styling
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = THIN_BORDER
                    cell.alignment = WRAP_ALIGNMENT
                break

        # Check if this feature should be marked as Partial
        for partial_feature in PARTIAL_FEATURES:
            if partial_feature.lower() in feature_value.lower():
                if current_status not in ("Done", "Partial"):
                    old_status = current_status
                    status_cell.value = "Partial"
                    print(f"  Row {row_idx}: '{feature_value[:50]}...' : {old_status!r} -> 'Partial'")
                    partial_count += 1

                # Apply styling
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = THIN_BORDER
                    cell.alignment = WRAP_ALIGNMENT
                break

    print(f"\nMarked {done_count} features as Done")
    print(f"Marked {partial_count} features as Partial")

    if done_count > 0 or partial_count > 0:
        wb.save(EXCEL_PATH)
        print(f"Saved to {os.path.abspath(EXCEL_PATH)}")
    else:
        print("No changes to save.")

    # Calculate module completion percentage
    total_features = 0
    done_total = 0
    partial_total = 0
    pending_count = 0

    for row_idx in range(2, ws.max_row + 1):
        feature_cell = ws.cell(row=row_idx, column=feature_col)
        status_cell = ws.cell(row=row_idx, column=status_col)

        feature_value = str(feature_cell.value or "").strip()
        status_value = str(status_cell.value or "").strip()

        # Skip header rows (module/sub-module headers don't have numeric S.No)
        sno_cell = ws.cell(row=row_idx, column=1)
        if not isinstance(sno_cell.value, (int, float)) or not feature_value:
            continue

        total_features += 1
        if status_value == "Done":
            done_total += 1
        elif status_value == "Partial":
            partial_total += 1
        else:
            pending_count += 1

    completion_pct = (done_total / total_features * 100) if total_features > 0 else 0

    print(f"\n=== TV Displays & Queue Module Status ===")
    print(f"Total features: {total_features}")
    print(f"Done: {done_total}")
    print(f"Partial: {partial_total}")
    print(f"Pending: {pending_count}")
    print(f"Completion: {completion_pct:.1f}%")


if __name__ == "__main__":
    main()
